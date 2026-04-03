"""
Workbook inspection utility for APS .xlsx/.xlsm files.

Use this to preview sheet tables, export CSV snapshots, and summarize workbook
contents without needing the live Excel UI.
"""
from __future__ import annotations

import argparse
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


DEFAULT_HEADER_ROW = 3
DEFAULT_TABLE_SHEETS = [
    "Config",
    "SKU_Master",
    "BOM",
    "Inventory",
    "Sales_Orders",
    "Resource_Master",
    "Routing",
    "Campaign_Config",
    "Changeover_Matrix",
    "Queue_Times",
    "Scenarios",
    "CTP_Request",
    "BOM_Output",
    "Capacity_Map",
    "Schedule_Output",
    "Campaign_Schedule",
    "Material_Plan",
    "Equipment_Schedule",
    "Scenario_Output",
    "CTP_Output",
]


def _clean_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame()

    cleaned = df.copy()
    cleaned.columns = [
        "" if (pd.isna(col) or str(col).strip().startswith("Unnamed:")) else str(col).strip()
        for col in cleaned.columns
    ]
    cleaned = cleaned.loc[:, [col != "" for col in cleaned.columns]]
    cleaned = cleaned.dropna(axis=0, how="all").dropna(axis=1, how="all")
    return cleaned.reset_index(drop=True)


def read_sheet_table(path: str | Path, sheet_name: str, *, header_row: int = DEFAULT_HEADER_ROW) -> pd.DataFrame:
    return _clean_dataframe(
        pd.read_excel(path, sheet_name=sheet_name, header=max(int(header_row) - 1, 0), engine="openpyxl")
    )


def workbook_summary(path: str | Path, *, header_row: int = DEFAULT_HEADER_ROW) -> list[dict]:
    workbook_path = Path(path)
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    summaries = []
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            info = {
                "sheet": sheet_name,
                "max_row": int(ws.max_row or 0),
                "max_col": int(ws.max_column or 0),
            }
            try:
                df = read_sheet_table(workbook_path, sheet_name, header_row=header_row)
            except Exception:
                df = pd.DataFrame()
            if not df.empty:
                info["table_rows"] = int(len(df))
                info["table_cols"] = int(len(df.columns))
                info["headers"] = list(df.columns[:8])
            summaries.append(info)
        return summaries
    finally:
        try:
            wb.close()
        except Exception:
            pass


def export_sheets(path: str | Path, export_dir: str | Path, sheets: list[str], *, header_row: int = DEFAULT_HEADER_ROW) -> None:
    out_dir = Path(export_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    for sheet_name in sheets:
        df = read_sheet_table(path, sheet_name, header_row=header_row)
        safe_name = sheet_name.replace(" ", "_")
        df.to_csv(out_dir / f"{safe_name}.csv", index=False)


def _print_summary(path: str | Path, *, header_row: int = DEFAULT_HEADER_ROW) -> None:
    print(f"Workbook: {Path(path).resolve()}")
    for info in workbook_summary(path, header_row=header_row):
        base = f"- {info['sheet']}: used_range={info['max_row']}x{info['max_col']}"
        if "table_rows" in info:
            headers = ", ".join(info.get("headers", []))
            base += f" | table={info['table_rows']}x{info['table_cols']} | headers=[{headers}]"
        print(base)


def _print_preview(path: str | Path, sheet_name: str, rows: int, *, header_row: int = DEFAULT_HEADER_ROW) -> None:
    print(f"\n[{sheet_name}]")
    df = read_sheet_table(path, sheet_name, header_row=header_row)
    if df.empty:
        print("(empty or non-tabular sheet)")
        return
    print(f"rows={len(df)} cols={len(df.columns)}")
    print(df.head(rows).fillna("").to_string(index=False))


def main() -> None:
    parser = argparse.ArgumentParser(description="Inspect APS workbook data directly from .xlsx/.xlsm files.")
    parser.add_argument("--workbook", default="APS_BF_SMS_RM.xlsx", help="Path to workbook file")
    parser.add_argument("--summary", action="store_true", help="Print sheet-level workbook summary")
    parser.add_argument("--sheet", action="append", default=[], help="Sheet to preview; can be repeated")
    parser.add_argument("--rows", type=int, default=8, help="Preview row count per sheet")
    parser.add_argument("--header-row", type=int, default=DEFAULT_HEADER_ROW, help="Header row number for tabular sheets")
    parser.add_argument("--export-dir", help="Optional directory to export selected sheets as CSV")
    parser.add_argument("--all-tables", action="store_true", help="Use the default APS table-sheet list for preview/export")
    args = parser.parse_args()

    workbook_path = Path(args.workbook)
    if not workbook_path.exists():
        raise SystemExit(f"Workbook not found: {workbook_path}")

    selected_sheets = list(dict.fromkeys(args.sheet))
    if args.all_tables:
        selected_sheets = list(dict.fromkeys(selected_sheets + DEFAULT_TABLE_SHEETS))

    if args.summary or not selected_sheets:
        _print_summary(workbook_path, header_row=args.header_row)

    for sheet_name in selected_sheets:
        _print_preview(workbook_path, sheet_name, args.rows, header_row=args.header_row)

    if args.export_dir:
        export_targets = selected_sheets or DEFAULT_TABLE_SHEETS
        export_sheets(workbook_path, args.export_dir, export_targets, header_row=args.header_row)
        print(f"\nExported {len(export_targets)} sheet snapshots to {Path(args.export_dir).resolve()}")


if __name__ == "__main__":
    main()
