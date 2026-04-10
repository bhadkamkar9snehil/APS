"""Workbook-backed configuration helpers for APS runtime."""
from __future__ import annotations

from copy import copy
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Mapping, Optional

import openpyxl
import pandas as pd

ALGORITHM_CONFIG_SHEET = "Algorithm_Config"
SYSTEM_CONFIG_SHEET = "Config"
SYSTEM_CONFIG_HEADER_ROW_0_BASED = 2

# Canonical algorithm keys exposed in Algorithm_Config and the legacy aliases that
# still appear in the workbook Config sheet and older Python call sites.
ALGORITHM_TO_LEGACY_ALIASES: Dict[str, List[str]] = {
    "HEAT_SIZE_MT": ["Default_Batch_Size_MT"],
    "PLANNING_HORIZON_DAYS": ["Planning_Horizon_Days"],
    "SOLVER_TIME_LIMIT_SECONDS": ["Default_Solver_Limit_Sec"],
    "CAMPAIGN_MIN_SIZE_MT": ["Min_Campaign_MT"],
    "CAMPAIGN_MAX_SIZE_MT": ["Max_Campaign_MT"],
    "BYPRODUCT_INVENTORY_MODE": ["Byproduct_Inventory_Mode"],
}

LEGACY_TO_ALGORITHM_ALIASES: Dict[str, str] = {
    alias: canonical
    for canonical, aliases in ALGORITHM_TO_LEGACY_ALIASES.items()
    for alias in aliases
}

WORKBOOK_CONFIG_EXTENSION_ROWS: List[Dict[str, Any]] = [
    {
        "insert_after": "SETUP_TIME_FIRST_HEAT_ONLY",
        "value_from_system_config": "Allow_Scheduler_Default_Masters",
        "values": [
            "Allow_Scheduler_Default_Masters",
            "SCHEDULER",
            "Allow Scheduler Default Masters",
            "N",
            "BOOLEAN",
            None,
            None,
            "Y/N",
            "Allow demo scheduler defaults when routing or resource masters are incomplete.",
            "HIGH",
        ],
    },
    {
        "insert_after": "Allow_Scheduler_Default_Masters",
        "value_from_system_config": "Queue_Enforcement",
        "values": [
            "Queue_Enforcement",
            "SCHEDULER",
            "Default Queue Enforcement",
            "Hard",
            "CHOICE",
            None,
            None,
            "mode",
            "Fallback queue-time enforcement when Queue_Times rows omit an enforcement value.",
            "MEDIUM",
        ],
    },
    {
        "insert_after": "Queue_Enforcement",
        "value_from_system_config": "Campaign_Serialization_Mode",
        "values": [
            "Campaign_Serialization_Mode",
            "SCHEDULER",
            "Campaign Serialization Mode",
            "STRICT_END_TO_END",
            "CHOICE",
            None,
            None,
            "mode",
            "Controls whether campaigns serialize end to end or may overlap after SMS completion.",
            "HIGH",
        ],
    },
    {
        "insert_after": "Campaign_Serialization_Mode",
        "values": [
            "DEFAULT_HEAT_DURATION_HOURS",
            "SCHEDULER",
            "Default Heat Duration (hrs)",
            1.5,
            "QUANTITY",
            0.25,
            24,
            "hours",
            "Default APS heat duration used by planning heat batches and lightweight schedule simulation.",
            "MEDIUM",
        ],
    },
    {
        "insert_after": "DEFAULT_HEAT_DURATION_HOURS",
        "values": [
            "RM_DURATION_FACTOR",
            "SCHEDULER",
            "RM Duration Factor",
            1.2,
            "RATIO",
            0.1,
            10,
            "ratio",
            "Multiplier to estimate RM hours from heat duration in APS planner simulation.",
            "MEDIUM",
        ],
    },
    {
        "insert_after": "RM_DURATION_FACTOR",
        "values": [
            "PLANNING_HORIZON_HOURS",
            "SCHEDULER",
            "Planning Horizon (hours)",
            336,
            "QUANTITY",
            1,
            8760,
            "hours",
            "Explicit APS planner simulation horizon in hours; overrides derived day-based horizon when provided.",
            "MEDIUM",
        ],
    },
    {
        "insert_after": "YIELD_LOSS_DEFAULT_PCT",
        "value_from_system_config": "Batch_Unit_Name",
        "values": [
            "Batch_Unit_Name",
            "SYSTEM",
            "Batch Unit Name",
            "Heat",
            "TEXT",
            None,
            None,
            "label",
            "Human-readable label for the primary production batch unit in workbook-facing views.",
            "LOW",
        ],
    },
    {
        "insert_after": "Batch_Unit_Name",
        "value_from_system_config": "Primary_Batch_Resource_Group",
        "values": [
            "Primary_Batch_Resource_Group",
            "CAMPAIGN",
            "Primary Batch Resource Group",
            "EAF",
            "CHOICE",
            None,
            None,
            "resource group",
            "Resource group that defines the primary batch for BOM tracing and heat estimation.",
            "HIGH",
        ],
    },
    {
        "insert_after": "Primary_Batch_Resource_Group",
        "value_from_system_config": "Campaign_Group_By",
        "values": [
            "Campaign_Group_By",
            "CAMPAIGN",
            "Campaign Group By",
            "Route_Family,Campaign_Group,Grade,Product_Family,Route_Variant",
            "TEXT",
            None,
            None,
            "columns",
            "Comma-separated sales-order attributes used to build compatible campaign families.",
            "HIGH",
        ],
    },
    {
        "insert_after": "Campaign_Group_By",
        "values": [
            "PRIORITY_SEQUENCE",
            "CAMPAIGN",
            "Priority Sequence",
            "URGENT,HIGH,NORMAL,LOW",
            "LIST",
            None,
            None,
            "list",
            "Priority ordering used across campaign building and APS planning.",
            "HIGH",
        ],
    },
    {
        "insert_after": "PRIORITY_SEQUENCE",
        "values": [
            "APS_MAX_LOT_MT",
            "CAMPAIGN",
            "APS Max Lot Size",
            500,
            "QUANTITY",
            1,
            5000,
            "MT",
            "Maximum APS planning-order lot size before splitting into another lot.",
            "HIGH",
        ],
    },
    {
        "insert_after": "APS_MAX_LOT_MT",
        "values": [
            "APS_MAX_HEATS_PER_LOT",
            "CAMPAIGN",
            "APS Max Heats per Lot",
            8,
            "COUNT",
            1,
            100,
            "heats",
            "Maximum number of heats allowed in one APS planning lot.",
            "MEDIUM",
        ],
    },
    {
        "insert_after": "APS_MAX_HEATS_PER_LOT",
        "values": [
            "APS_URGENT_WINDOW_HOURS",
            "CAMPAIGN",
            "APS Urgent Window",
            48,
            "QUANTITY",
            1,
            720,
            "hours",
            "Orders due within this window are treated as urgent during APS lot formation.",
            "HIGH",
        ],
    },
    {
        "insert_after": "APS_URGENT_WINDOW_HOURS",
        "values": [
            "APS_MAX_DUE_SPREAD_DAYS",
            "CAMPAIGN",
            "APS Max Due Spread",
            3,
            "COUNT",
            0,
            60,
            "days",
            "Maximum allowed due-date spread within one APS planning lot.",
            "MEDIUM",
        ],
    },
    {
        "insert_after": "APS_MAX_DUE_SPREAD_DAYS",
        "values": [
            "APS_SECTION_TOLERANCE_MM",
            "CAMPAIGN",
            "APS Section Tolerance",
            0.6,
            "THRESHOLD",
            0,
            10,
            "mm",
            "Maximum section-size difference allowed when grouping compatible APS planning orders.",
            "MEDIUM",
        ],
    },
    {
        "insert_after": "APS_SECTION_TOLERANCE_MM",
        "value_from_system_config": "Allow_Legacy_Primary_Batch_Fallback",
        "values": [
            "Allow_Legacy_Primary_Batch_Fallback",
            "CAMPAIGN",
            "Allow Legacy Primary Batch Fallback",
            "N",
            "BOOLEAN",
            None,
            None,
            "Y/N",
            "Allow legacy heat estimation only for diagnostics when primary-batch BOM tracing fails.",
            "MEDIUM",
        ],
    },
    {
        "insert_after": "Allow_Legacy_Primary_Batch_Fallback",
        "value_from_system_config": "Manual_Campaign_Grouping_Mode",
        "values": [
            "Manual_Campaign_Grouping_Mode",
            "CAMPAIGN",
            "Manual Campaign Grouping Mode",
            "PRESERVE_EXACT",
            "CHOICE",
            None,
            None,
            "mode",
            "Controls whether manually assigned Campaign_ID groups stay intact or split to max size.",
            "HIGH",
        ],
    },
    {
        "insert_after": "Manual_Campaign_Grouping_Mode",
        "value_from_system_config": "Default_Section_Fallback",
        "values": [
            "Default_Section_Fallback",
            "CAMPAIGN",
            "Default Section Fallback",
            6.5,
            "QUANTITY",
            0,
            100,
            "mm",
            "Fallback product section when demand or SKU data is missing section information.",
            "MEDIUM",
        ],
    },
    {
        "insert_after": "Default_Section_Fallback",
        "value_from_system_config": "ROLLING_MODE_DEFAULT",
        "values": [
            "ROLLING_MODE_DEFAULT",
            "CAMPAIGN",
            "Rolling Mode Default",
            "HOT",
            "CHOICE",
            None,
            None,
            "mode",
            "Default rolling mode for APS planning orders when order rows do not specify one.",
            "MEDIUM",
        ],
    },
    {
        "insert_after": "ZERO_TOLERANCE_THRESHOLD",
        "value_from_system_config": "BOM_Structure_Error_Mode",
        "values": [
            "BOM_Structure_Error_Mode",
            "BOM",
            "BOM Structure Error Mode",
            "RAISE",
            "CHOICE",
            None,
            None,
            "mode",
            "Controls whether BOM cycles and excessive depth hard-fail or are recorded into holds.",
            "HIGH",
        ],
    },
    {
        "insert_after": "CTP_MERGE_PENALTY",
        "value_from_system_config": "Require_Authoritative_CTP_Inventory",
        "values": [
            "Require_Authoritative_CTP_Inventory",
            "CTP",
            "Require Authoritative CTP Inventory",
            "Y",
            "BOOLEAN",
            None,
            None,
            "Y/N",
            "Block CTP promises when committed inventory lineage is not authoritative.",
            "HIGH",
        ],
    },
    {
        "insert_after": "Require_Authoritative_CTP_Inventory",
        "values": [
            "CTP_DECISION_PRECEDENCE_SEQUENCE",
            "CTP",
            "CTP Decision Precedence Sequence",
            "PROMISE_CONFIRMED_STOCK_ONLY,PROMISE_CONFIRMED_MERGED,PROMISE_CONFIRMED_NEW_CAMPAIGN,PROMISE_HEURISTIC_ONLY,PROMISE_LATER_DATE,PROMISE_SPLIT_REQUIRED,PROMISE_CONDITIONAL_EXPEDITE,CANNOT_PROMISE_POLICY_ONLY,CANNOT_PROMISE_CAPACITY,CANNOT_PROMISE_MATERIAL,CANNOT_PROMISE_INVENTORY_TRUST,CANNOT_PROMISE_MASTER_DATA,CANNOT_PROMISE_MIXED_BLOCKERS",
            "LIST",
            None,
            None,
            "list",
            "Decision precedence used when ranking competing CTP scenarios and alternatives.",
            "HIGH",
        ],
    },
]

WORKBOOK_SYSTEM_CONFIG_UPDATES: Dict[str, Any] = {
    "Planning_Horizon_Days": 14,
}

_MISSING = object()


def canonicalize_config_key(key: str | None) -> str:
    text = str(key or "").strip()
    if not text:
        return ""
    return LEGACY_TO_ALGORITHM_ALIASES.get(text, text)


def config_key_candidates(key: str | None) -> List[str]:
    canonical = canonicalize_config_key(key)
    if not canonical:
        return []
    candidates = [canonical]
    for alias in ALGORITHM_TO_LEGACY_ALIASES.get(canonical, []):
        if alias not in candidates:
            candidates.append(alias)
    raw_text = str(key or "").strip()
    if raw_text and raw_text not in candidates:
        candidates.insert(0, raw_text)
    return candidates


def _normalize_workbook_path(source: Any) -> Path:
    if isinstance(source, Path):
        return source
    if isinstance(source, str):
        return Path(source.strip().strip('"').strip("'"))
    workbook_attr = getattr(source, "workbook_path", None)
    if workbook_attr is not None:
        return _normalize_workbook_path(workbook_attr)
    raise TypeError(f"Unsupported workbook source: {type(source)!r}")


def _copy_row_style(ws, source_row: int, target_row: int) -> None:
    for col in range(1, ws.max_column + 1):
        source_cell = ws.cell(source_row, col)
        if source_cell.has_style:
            ws.cell(target_row, col)._style = copy(source_cell._style)


def _find_row_by_key(ws, key_col: int, key: str, *, start_row: int = 2) -> int | None:
    needle = str(key or "").strip()
    if not needle:
        return None
    for row in range(start_row, ws.max_row + 1):
        if str(ws.cell(row, key_col).value or "").strip() == needle:
            return row
    return None


def _same_value(left: Any, right: Any) -> bool:
    if left is None and right is None:
        return True
    left_bool = resolve_config_bool({"value": left}, "value", default=False)
    right_bool = resolve_config_bool({"value": right}, "value", default=False)
    left_text = str(left).strip().upper() if left is not None else ""
    right_text = str(right).strip().upper() if right is not None else ""
    bool_tokens = {"Y", "YES", "TRUE", "1", "ON", "N", "NO", "FALSE", "0", "OFF"}
    if left_text in bool_tokens and right_text in bool_tokens:
        return left_bool == right_bool
    if isinstance(left, bool) or isinstance(right, bool):
        if left_text in bool_tokens or right_text in bool_tokens:
            return left_bool == right_bool
    if isinstance(left, str) and isinstance(right, str):
        return left.strip().upper() == right.strip().upper()
    try:
        if pd.isna(left) and pd.isna(right):
            return True
    except Exception:
        pass
    return left == right


def _serialize_config_value(value: Any, data_type: str | None = None) -> Any:
    if value is None:
        return None
    normalized_type = str(data_type or "").strip().upper()
    if normalized_type in {"LIST", "SET"}:
        if isinstance(value, (list, tuple, set)):
            return ", ".join(str(item).strip() for item in value if str(item).strip())
    if normalized_type == "BOOLEAN":
        return "TRUE" if bool(value) else "FALSE"
    return value


@dataclass(frozen=True)
class WorkbookConfigSnapshot:
    workbook_path: Path
    system_config: Dict[str, Any]
    algorithm_config: "AlgorithmConfig"
    runtime_config: Dict[str, Any]
    sources: Dict[str, str]
    conflicts: List[Dict[str, Any]]


class AlgorithmConfig:
    """Typed access to Algorithm_Config workbook parameters."""

    def __init__(self, config_df: pd.DataFrame | None = None):
        self.config_dict: Dict[str, Any] = {}
        self.metadata: Dict[str, Dict[str, Any]] = {}

        if config_df is not None and not config_df.empty:
            self._load_from_dataframe(config_df)

    def _load_from_dataframe(self, config_df: pd.DataFrame) -> None:
        for _, row in config_df.iterrows():
            key = canonicalize_config_key(row.get("Config_Key", ""))
            if not key:
                continue

            value = row.get("Current_Value")
            data_type = str(row.get("Data_Type", "")).strip().upper()
            min_val = row.get("Min_Value")
            max_val = row.get("Max_Value")
            converted = self._convert_value(value, data_type)

            if not self._validate_value(key, converted, data_type, min_val, max_val):
                raise ValueError(
                    f"Invalid config value for {key}: {converted} "
                    f"(type={data_type}, min={min_val}, max={max_val})"
                )

            self.config_dict[key] = converted
            self.metadata[key] = {
                "data_type": data_type,
                "min": min_val,
                "max": max_val,
                "category": str(row.get("Category", "")).strip(),
                "description": str(row.get("Description", "")).strip(),
                "parameter_name": str(row.get("Parameter_Name", "")).strip(),
            }

    def _convert_value(self, value: Any, data_type: str) -> Any:
        if pd.isna(value) or value is None:
            return None

        normalized = str(data_type or "").upper()

        if normalized == "BOOLEAN":
            return str(value).strip().upper() in {"TRUE", "YES", "1", "Y", "ON"}
        if normalized in {"DURATION", "COUNT"}:
            return int(float(value))
        if normalized in {"QUANTITY", "PERCENTAGE", "WEIGHT", "RATIO", "THRESHOLD"}:
            return float(value)
        if normalized in {"LIST", "SET"}:
            if isinstance(value, str):
                return [item.strip() for item in value.split(",") if item.strip()]
            return list(value) if value else []
        if normalized == "CHOICE":
            if isinstance(value, (list, tuple, set)):
                for item in value:
                    text = str(item).strip()
                    if text:
                        return text
                return None
            return str(value).strip()
        return str(value).strip()

    def _validate_value(
        self,
        key: str,
        value: Any,
        data_type: str | None,
        min_val: Any,
        max_val: Any,
    ) -> bool:
        if value is None:
            return True

        normalized = str(data_type or "").upper()
        if normalized in {"DURATION", "COUNT", "QUANTITY", "PERCENTAGE", "WEIGHT", "RATIO", "THRESHOLD"}:
            try:
                numeric = float(value)
                if pd.notna(min_val) and numeric < float(min_val):
                    return False
                if pd.notna(max_val) and numeric > float(max_val):
                    return False
            except (TypeError, ValueError):
                return False
        return True

    def get(self, key: str, default: Any = None) -> Any:
        for candidate in config_key_candidates(key):
            if candidate in self.config_dict:
                return self.config_dict[candidate]
        return default

    def get_duration_minutes(self, key: str, default: int = 0) -> int:
        value = self.get(key, default)
        return int(value) if value is not None else default

    def get_percentage(self, key: str, default: float = 0.0) -> float:
        value = self.get(key, default)
        return float(value) if value is not None else default

    def get_weight(self, key: str, default: int = 1) -> int:
        value = self.get(key, default)
        return int(value) if value is not None else default

    def get_float(self, key: str, default: float = 0.0) -> float:
        value = self.get(key, default)
        return float(value) if value is not None else default

    def get_list(self, key: str, default: List[str] | None = None) -> List[str]:
        fallback = default or []
        value = self.get(key, fallback)
        if isinstance(value, list):
            return value
        return fallback

    def get_bool(self, key: str, default: bool = False) -> bool:
        value = self.get(key, default)
        return bool(value) if value is not None else default

    def all_params(self, *, include_aliases: bool = False) -> Dict[str, Any]:
        params = dict(self.config_dict)
        if include_aliases:
            for canonical, aliases in ALGORITHM_TO_LEGACY_ALIASES.items():
                if canonical not in params:
                    continue
                for alias in aliases:
                    params.setdefault(alias, params[canonical])
        return params

    def params_by_category(self, category: str, *, include_aliases: bool = False) -> Dict[str, Any]:
        matched = {
            key: value
            for key, value in self.config_dict.items()
            if self.metadata.get(key, {}).get("category") == category
        }
        if include_aliases:
            for canonical, aliases in ALGORITHM_TO_LEGACY_ALIASES.items():
                if canonical not in matched:
                    continue
                for alias in aliases:
                    matched.setdefault(alias, matched[canonical])
        return matched

    def update(self, key: str, value: Any, user: str = "SYSTEM", reason: str = "") -> bool:
        canonical = canonicalize_config_key(key)
        meta = self.metadata.get(canonical, {})
        converted = self._convert_value(value, meta.get("data_type"))
        if not self._validate_value(
            canonical,
            converted,
            meta.get("data_type"),
            meta.get("min"),
            meta.get("max"),
        ):
            return False
        self.config_dict[canonical] = converted
        return True


_config_instance: Optional[AlgorithmConfig] = None
_loaded_workbook_path: Optional[Path] = None


def load_algorithm_config(config_df: pd.DataFrame) -> AlgorithmConfig:
    global _config_instance
    _config_instance = AlgorithmConfig(config_df)
    return _config_instance


def load_algorithm_config_from_workbook(workbook_path: str | Path | Any) -> AlgorithmConfig:
    global _config_instance, _loaded_workbook_path
    path = _normalize_workbook_path(workbook_path)
    xls = pd.ExcelFile(path)
    if ALGORITHM_CONFIG_SHEET not in xls.sheet_names:
        _config_instance = AlgorithmConfig()
        _loaded_workbook_path = path
        return _config_instance
    config_df = xls.parse(ALGORITHM_CONFIG_SHEET, header=0)
    _config_instance = AlgorithmConfig(config_df)
    _loaded_workbook_path = path
    return _config_instance


def get_config(workbook_path: str | Path | Any | None = None, reload: bool = False) -> AlgorithmConfig:
    global _config_instance, _loaded_workbook_path
    if workbook_path is not None:
        path = _normalize_workbook_path(workbook_path)
        if reload or _config_instance is None or _loaded_workbook_path != path:
            return load_algorithm_config_from_workbook(path)
    if _config_instance is None:
        _config_instance = AlgorithmConfig()
    return _config_instance


def read_system_config(workbook_path: str | Path | Any) -> Dict[str, Any]:
    path = _normalize_workbook_path(workbook_path)
    try:
        df = pd.read_excel(path, sheet_name=SYSTEM_CONFIG_SHEET, header=SYSTEM_CONFIG_HEADER_ROW_0_BASED)
    except Exception:
        return {}
    if df.empty or "Key" not in df.columns:
        return {}
    df = df.dropna(subset=["Key"]).copy()
    df["Key"] = df["Key"].astype(str).str.strip()
    df = df[df["Key"] != ""]
    return {
        str(row["Key"]).strip(): row.get("Value")
        for _, row in df.iterrows()
        if str(row.get("Key", "")).strip()
    }


def resolve_config_value(config: Mapping[str, Any] | AlgorithmConfig | None, key: str, default: Any = None) -> Any:
    if isinstance(config, AlgorithmConfig):
        return config.get(key, default)
    for candidate in config_key_candidates(key):
        if isinstance(config, Mapping) and candidate in config:
            return config[candidate]
    if config is None:
        return get_config().get(key, default)
    return default


def resolve_config_float(config: Mapping[str, Any] | AlgorithmConfig | None, key: str, default: float) -> float:
    value = resolve_config_value(config, key, default)
    numeric = pd.to_numeric(pd.Series([value]), errors="coerce").iloc[0]
    return float(default if pd.isna(numeric) else numeric)


def resolve_config_int(config: Mapping[str, Any] | AlgorithmConfig | None, key: str, default: int) -> int:
    return int(round(resolve_config_float(config, key, float(default))))


def resolve_config_bool(config: Mapping[str, Any] | AlgorithmConfig | None, key: str, default: bool = False) -> bool:
    value = resolve_config_value(config, key, default)
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    return str(value).strip().upper() in {"Y", "YES", "TRUE", "1", "ON"}


def _build_runtime_config(
    system_config: Dict[str, Any],
    algorithm_config: AlgorithmConfig,
) -> tuple[Dict[str, Any], Dict[str, str], List[Dict[str, Any]]]:
    runtime = algorithm_config.all_params(include_aliases=True)
    sources = {key: ALGORITHM_CONFIG_SHEET for key in runtime}
    conflicts: List[Dict[str, Any]] = []

    for key, value in (system_config or {}).items():
        canonical = canonicalize_config_key(key)
        if key in runtime or canonical in algorithm_config.config_dict:
            algorithm_value = runtime.get(key, algorithm_config.get(canonical))
            runtime[key] = algorithm_value
            if key in LEGACY_TO_ALGORITHM_ALIASES and canonical in algorithm_config.config_dict:
                sources[key] = f"{ALGORITHM_CONFIG_SHEET} (legacy alias)"
            elif key == canonical:
                sources[key] = ALGORITHM_CONFIG_SHEET
            else:
                sources[key] = f"{ALGORITHM_CONFIG_SHEET} (canonical alias)"
            if not _same_value(value, algorithm_value):
                conflicts.append(
                    {
                        "key": key,
                        "canonical_key": canonical,
                        "config_sheet_value": value,
                        "algorithm_config_value": algorithm_value,
                        "winning_source": ALGORITHM_CONFIG_SHEET,
                    }
                )
            continue

        runtime[key] = value
        sources[key] = SYSTEM_CONFIG_SHEET

        if canonical != key and canonical not in runtime:
            runtime[canonical] = value
            sources[canonical] = f"{SYSTEM_CONFIG_SHEET} (canonical alias)"

    return runtime, sources, conflicts


def load_workbook_config_snapshot(
    workbook_path: str | Path | Any,
    *,
    reload_algorithm: bool = True,
) -> WorkbookConfigSnapshot:
    path = _normalize_workbook_path(workbook_path)
    algorithm_config = get_config(path, reload=reload_algorithm)
    system_config = read_system_config(path)
    runtime_config, sources, conflicts = _build_runtime_config(system_config, algorithm_config)
    return WorkbookConfigSnapshot(
        workbook_path=path,
        system_config=system_config,
        algorithm_config=algorithm_config,
        runtime_config=runtime_config,
        sources=sources,
        conflicts=conflicts,
    )


def update_algorithm_config_in_workbook(
    workbook_path: str | Path | Any,
    key: str,
    value: Any,
    *,
    user: str = "SYSTEM",
    reason: str = "",
) -> Dict[str, Any]:
    path = _normalize_workbook_path(workbook_path)
    config = get_config(path, reload=True)
    canonical = canonicalize_config_key(key)
    meta = config.metadata.get(canonical)
    if meta is None:
        raise KeyError(f"Parameter {canonical} not found")

    converted = config._convert_value(value, meta.get("data_type"))
    if not config._validate_value(canonical, converted, meta.get("data_type"), meta.get("min"), meta.get("max")):
        raise ValueError(f"Invalid value for {canonical}: {value}")

    wb = openpyxl.load_workbook(path)
    try:
        if ALGORITHM_CONFIG_SHEET not in wb.sheetnames:
            raise KeyError(f"Workbook is missing sheet: {ALGORITHM_CONFIG_SHEET}")

        ws = wb[ALGORITHM_CONFIG_SHEET]
        headers = {
            str(ws.cell(1, col).value).strip(): col
            for col in range(1, ws.max_column + 1)
            if ws.cell(1, col).value is not None
        }
        key_col = headers.get("Config_Key")
        value_col = headers.get("Current_Value")
        notes_col = headers.get("Notes")
        if key_col is None or value_col is None:
            raise KeyError("Algorithm_Config is missing Config_Key or Current_Value column")

        row_idx = None
        for row in range(2, ws.max_row + 1):
            row_key = canonicalize_config_key(ws.cell(row, key_col).value)
            if row_key == canonical:
                row_idx = row
                break
        if row_idx is None:
            raise KeyError(f"Parameter {canonical} not found")

        old_value = config.get(canonical)
        ws.cell(row_idx, value_col).value = _serialize_config_value(converted, meta.get("data_type"))

        if notes_col is not None:
            note_parts = [datetime.now().isoformat(timespec="seconds"), str(user or "SYSTEM").strip()]
            if reason:
                note_parts.append(str(reason).strip())
            note_text = " | ".join(part for part in note_parts if part)
            existing = str(ws.cell(row_idx, notes_col).value or "").strip()
            ws.cell(row_idx, notes_col).value = f"{existing}\n{note_text}".strip()

        if SYSTEM_CONFIG_SHEET in wb.sheetnames:
            config_ws = wb[SYSTEM_CONFIG_SHEET]
            config_headers = {
                str(config_ws.cell(3, col).value).strip(): col
                for col in range(1, config_ws.max_column + 1)
                if config_ws.cell(3, col).value is not None
            }
            config_key_col = config_headers.get("Key")
            config_value_col = config_headers.get("Value")
            if config_key_col is not None and config_value_col is not None:
                sync_keys = [canonical, *ALGORITHM_TO_LEGACY_ALIASES.get(canonical, [])]
                for sync_key in dict.fromkeys(sync_keys):
                    for row in range(4, config_ws.max_row + 1):
                        if str(config_ws.cell(row, config_key_col).value or "").strip() == sync_key:
                            config_ws.cell(row, config_value_col).value = _serialize_config_value(
                                converted,
                                meta.get("data_type"),
                            )
                            break

        wb.save(path)
    finally:
        wb.close()

    refreshed = get_config(path, reload=True)
    return {
        "key": canonical,
        "old_value": old_value,
        "new_value": refreshed.get(canonical),
        "metadata": refreshed.metadata.get(canonical, {}),
    }


def upgrade_workbook_config(
    workbook_path: str | Path | Any,
    *,
    output_path: str | Path | Any | None = None,
) -> Dict[str, Any]:
    source_path = _normalize_workbook_path(workbook_path)
    target_path = _normalize_workbook_path(output_path) if output_path is not None else source_path
    target_path.parent.mkdir(parents=True, exist_ok=True)
    keep_vba = source_path.suffix.lower() == ".xlsm" or target_path.suffix.lower() == ".xlsm"

    wb = openpyxl.load_workbook(source_path, keep_vba=keep_vba)
    inserted_keys: List[str] = []
    updated_system_config_keys: List[str] = []
    try:
        if ALGORITHM_CONFIG_SHEET not in wb.sheetnames:
            raise KeyError(f"Workbook is missing sheet: {ALGORITHM_CONFIG_SHEET}")

        ws = wb[ALGORITHM_CONFIG_SHEET]
        headers = {
            str(ws.cell(1, col).value).strip(): col
            for col in range(1, ws.max_column + 1)
            if ws.cell(1, col).value is not None
        }
        key_col = headers.get("Config_Key")
        if key_col is None:
            raise KeyError("Algorithm_Config is missing Config_Key column")

        pending_rows = []
        existing_keys = {
            str(ws.cell(row, key_col).value or "").strip()
            for row in range(2, ws.max_row + 1)
            if ws.cell(row, key_col).value is not None
        }
        system_config_values = read_system_config(source_path)
        for row_def in WORKBOOK_CONFIG_EXTENSION_ROWS:
            values = list(row_def["values"])
            key = str(values[0]).strip()
            if key in existing_keys:
                continue
            system_value_key = row_def.get("value_from_system_config")
            if system_value_key:
                values[3] = system_config_values.get(system_value_key, values[3])
            anchor_row = _find_row_by_key(ws, key_col, row_def["insert_after"]) or ws.max_row
            pending_rows.append((anchor_row, values))
            existing_keys.add(key)

        for anchor_row, values in sorted(pending_rows, key=lambda item: item[0], reverse=True):
            insert_at = anchor_row + 1
            ws.insert_rows(insert_at, amount=1)
            if anchor_row >= 1:
                _copy_row_style(ws, anchor_row, insert_at)
            for col, value in enumerate(values, start=1):
                ws.cell(insert_at, col).value = value
            inserted_keys.append(str(values[0]).strip())

        if SYSTEM_CONFIG_SHEET in wb.sheetnames:
            config_ws = wb[SYSTEM_CONFIG_SHEET]
            config_headers = {
                str(config_ws.cell(3, col).value).strip(): col
                for col in range(1, config_ws.max_column + 1)
                if config_ws.cell(3, col).value is not None
            }
            config_key_col = config_headers.get("Key")
            config_value_col = config_headers.get("Value")
            if config_key_col is not None and config_value_col is not None:
                for key, value in WORKBOOK_SYSTEM_CONFIG_UPDATES.items():
                    row_idx = _find_row_by_key(config_ws, config_key_col, key, start_row=4)
                    if row_idx is None:
                        continue
                    if not _same_value(config_ws.cell(row_idx, config_value_col).value, value):
                        config_ws.cell(row_idx, config_value_col).value = value
                        updated_system_config_keys.append(key)

        wb.save(target_path)
    finally:
        wb.close()

    return {
        "source_workbook": str(source_path),
        "workbook_path": str(target_path),
        "inserted_keys": inserted_keys,
        "updated_system_config_keys": updated_system_config_keys,
    }
