from __future__ import annotations

from copy import deepcopy
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
import pandas as pd

from .workbook_schema import SHEETS, SheetConfig


class ExcelStore:
    def __init__(self, workbook_path: str | Path):
        self.workbook_path = Path(workbook_path)

    def _open(self):
        if not self.workbook_path.exists():
            raise FileNotFoundError(f"Workbook not found: {self.workbook_path}")
        return openpyxl.load_workbook(self.workbook_path)

    def _clean_value(self, value: Any) -> Any:
        if value is None:
            return None
        if isinstance(value, (datetime, date, pd.Timestamp)):
            try:
                return value.isoformat()
            except Exception:
                return str(value)
        return value

    def _parse_value(self, value: Any) -> Any:
        return None if value == "" else value

    def _headers(self, ws, header_row_1_based: int) -> List[str]:
        headers: List[str] = []
        for cell in ws[header_row_1_based]:
            headers.append(str(cell.value).strip() if cell.value is not None else "")
        while headers and headers[-1] == "":
            headers.pop()
        return headers

    def _iter_records_with_rows(self, cfg: SheetConfig) -> Tuple[List[str], List[Tuple[int, Dict[str, Any]]]]:
        wb = self._open()
        ws = wb[cfg.excel_name]
        headers = self._headers(ws, cfg.header_row_1_based)
        records: List[Tuple[int, Dict[str, Any]]] = []
        for row_idx in range(cfg.header_row_1_based + 1, ws.max_row + 1):
            row_values = [ws.cell(row_idx, col_idx).value for col_idx in range(1, len(headers) + 1)]
            if all(v is None for v in row_values):
                continue
            record = {headers[i]: self._clean_value(row_values[i]) for i in range(len(headers))}
            records.append((row_idx, record))
        wb.close()
        return headers, records

    def list_sheet_configs(self) -> List[Dict[str, Any]]:
        return [
            {
                "api_name": cfg.api_name,
                "excel_name": cfg.excel_name,
                "header_row": cfg.header_row_1_based,
                "key_field": cfg.key_field,
                "read_only": cfg.read_only,
            }
            for cfg in SHEETS.values()
        ]

    def list_rows(
        self,
        api_name: str,
        *,
        search: Optional[str] = None,
        sort_by: Optional[str] = None,
        sort_dir: str = "asc",
        limit: Optional[int] = None,
        offset: int = 0,
        filters: Optional[Dict[str, str]] = None,
    ) -> Dict[str, Any]:
        cfg = SHEETS[api_name]
        headers, rows = self._iter_records_with_rows(cfg)
        items = [deepcopy(r) for _, r in rows]

        if filters:
            def match_filters(item: Dict[str, Any]) -> bool:
                for k, expected in filters.items():
                    if k not in item:
                        return False
                    actual = "" if item[k] is None else str(item[k])
                    if actual != expected:
                        return False
                return True
            items = [x for x in items if match_filters(x)]

        if search:
            needle = search.lower()
            items = [x for x in items if any(needle in str(v).lower() for v in x.values() if v is not None)]

        if sort_by and sort_by in headers:
            items.sort(key=lambda x: (x.get(sort_by) is None, str(x.get(sort_by, ""))), reverse=(sort_dir == "desc"))

        total = len(items)
        if offset:
            items = items[offset:]
        if limit is not None:
            items = items[:limit]

        return {"headers": headers, "items": items, "total": total}

    def get_row(self, api_name: str, key_value: str) -> Optional[Dict[str, Any]]:
        cfg = SHEETS[api_name]
        if not cfg.key_field:
            return None
        _, rows = self._iter_records_with_rows(cfg)
        for _, row in rows:
            if str(row.get(cfg.key_field)) == str(key_value):
                return deepcopy(row)
        return None

    def create_row(self, api_name: str, payload: Dict[str, Any]) -> Dict[str, Any]:
        cfg = SHEETS[api_name]
        if cfg.read_only:
            raise ValueError(f"{api_name} is read-only")
        wb = self._open()
        ws = wb[cfg.excel_name]
        headers = self._headers(ws, cfg.header_row_1_based)

        if cfg.key_field and (payload.get(cfg.key_field) is None or str(payload.get(cfg.key_field)).strip() == ""):
            raise ValueError(f"Missing required key field: {cfg.key_field}")

        if cfg.key_field and self._row_index_by_key_ws(ws, cfg, str(payload[cfg.key_field]), headers) is not None:
            raise ValueError(f"Duplicate key: {payload[cfg.key_field]}")

        row_idx = ws.max_row + 1
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row_idx, col_idx).value = self._parse_value(payload.get(header))
        wb.save(self.workbook_path)
        wb.close()
        return self.get_row(api_name, str(payload[cfg.key_field])) if cfg.key_field else payload

    def update_row(self, api_name: str, key_value: str, payload: Dict[str, Any], *, partial: bool) -> Dict[str, Any]:
        cfg = SHEETS[api_name]
        if cfg.read_only:
            raise ValueError(f"{api_name} is read-only")
        if not cfg.key_field:
            raise ValueError(f"{api_name} does not expose a row key")
        wb = self._open()
        ws = wb[cfg.excel_name]
        headers = self._headers(ws, cfg.header_row_1_based)
        row_idx = self._row_index_by_key_ws(ws, cfg, key_value, headers)
        if row_idx is None:
            wb.close()
            raise KeyError(f"Row not found for {cfg.key_field}={key_value}")

        current = {headers[i - 1]: ws.cell(row_idx, i).value for i in range(1, len(headers) + 1)}
        merged = current.copy() if partial else {h: None for h in headers}
        for k, v in payload.items():
            if k in headers:
                merged[k] = self._parse_value(v)

        if cfg.key_field in payload and str(payload[cfg.key_field]) != str(key_value):
            existing = self._row_index_by_key_ws(ws, cfg, str(payload[cfg.key_field]), headers)
            if existing is not None and existing != row_idx:
                wb.close()
                raise ValueError(f"Duplicate key: {payload[cfg.key_field]}")

        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row_idx, col_idx).value = merged.get(header)
        wb.save(self.workbook_path)
        wb.close()
        new_key = str(merged.get(cfg.key_field, key_value))
        return self.get_row(api_name, new_key)

    def delete_row(self, api_name: str, key_value: str) -> None:
        cfg = SHEETS[api_name]
        if cfg.read_only:
            raise ValueError(f"{api_name} is read-only")
        if not cfg.key_field:
            raise ValueError(f"{api_name} does not expose a row key")
        wb = self._open()
        ws = wb[cfg.excel_name]
        headers = self._headers(ws, cfg.header_row_1_based)
        row_idx = self._row_index_by_key_ws(ws, cfg, key_value, headers)
        if row_idx is None:
            wb.close()
            raise KeyError(f"Row not found for {cfg.key_field}={key_value}")
        ws.delete_rows(row_idx, 1)
        wb.save(self.workbook_path)
        wb.close()

    def bulk_replace(self, api_name: str, items: List[Dict[str, Any]]) -> Dict[str, Any]:
        cfg = SHEETS[api_name]
        if cfg.read_only:
            raise ValueError(f"{api_name} is read-only")
        wb = self._open()
        ws = wb[cfg.excel_name]
        headers = self._headers(ws, cfg.header_row_1_based)
        if ws.max_row > cfg.header_row_1_based:
            ws.delete_rows(cfg.header_row_1_based + 1, ws.max_row - cfg.header_row_1_based)
        row_idx = cfg.header_row_1_based + 1
        for item in items:
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row_idx, col_idx).value = self._parse_value(item.get(header))
            row_idx += 1
        wb.save(self.workbook_path)
        wb.close()
        return {"replaced": len(items)}

    def workbook_snapshot(self) -> Dict[str, Any]:
        snapshot: Dict[str, Any] = {}
        for api_name, cfg in SHEETS.items():
            if cfg.read_only:
                continue
            data = self.list_rows(api_name, limit=5)
            snapshot[api_name] = {
                "headers": data["headers"],
                "sample_count": len(data["items"]),
                "total": data["total"],
            }
        return snapshot

    def _row_index_by_key_ws(self, ws, cfg: SheetConfig, key_value: str, headers: List[str]) -> Optional[int]:
        if not cfg.key_field:
            return None
        key_col_idx = headers.index(cfg.key_field) + 1
        for row_idx in range(cfg.header_row_1_based + 1, ws.max_row + 1):
            value = ws.cell(row_idx, key_col_idx).value
            if value is None:
                continue
            if str(value) == str(key_value):
                return row_idx
        return None
