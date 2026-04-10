"""Workbook master-data usage audit."""
from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd

from engine.config import load_workbook_config_snapshot

MASTERDATA_AUDIT_SPEC: dict[str, dict[str, Any]] = {
    "Algorithm_Config": {
        "role": "Primary runtime config",
        "header_row": 1,
        "used_columns": [
            "Config_Key",
            "Category",
            "Parameter_Name",
            "Current_Value",
            "Data_Type",
            "Min_Value",
            "Max_Value",
            "Description",
        ],
        "consumers": ["engine/config.py", "xaps_application_api.py"],
        "deprecated": False,
    },
    "Config": {
        "role": "Deprecated compatibility config",
        "header_row": 3,
        "used_columns": ["Key", "Value", "Description"],
        "consumers": ["engine/config.py"],
        "deprecated": True,
    },
    "SKU_Master": {
        "role": "SKU master data",
        "header_row": 3,
        "used_columns": [
            "SKU_ID",
            "SKU_Name",
            "Category",
            "Grade",
            "Section_mm",
            "Needs_VD",
            "Route_Variant",
            "Product_Family",
            "Attribute_1",
        ],
        "consumers": ["engine/campaign.py", "engine/ctp.py", "xaps_application_api.py"],
        "deprecated": False,
    },
    "BOM": {
        "role": "Material structure",
        "header_row": 3,
        "used_columns": [
            "Parent_SKU",
            "Child_SKU",
            "Flow_Type",
            "Qty_Per",
            "Scrap_%",
            "Yield_Pct",
            "Level",
        ],
        "consumers": ["engine/bom_explosion.py", "engine/campaign.py", "xaps_application_api.py"],
        "deprecated": False,
    },
    "Inventory": {
        "role": "Available and reserved stock",
        "header_row": 3,
        "used_columns": ["SKU_ID", "Available_Qty", "Reserved_Qty"],
        "consumers": ["engine/bom_explosion.py", "engine/campaign.py", "engine/ctp.py", "xaps_application_api.py"],
        "deprecated": False,
    },
    "Sales_Orders": {
        "role": "Demand input",
        "header_row": 3,
        "used_columns": [
            "SO_ID",
            "SKU_ID",
            "Grade",
            "Section_mm",
            "Order_Qty_MT",
            "Order_Date",
            "Delivery_Date",
            "Priority",
            "Order_Type",
            "Rolling_Mode",
            "Campaign_Group",
            "Campaign_ID",
            "Status",
        ],
        "consumers": ["engine/campaign.py", "engine/aps_planner.py", "xaps_application_api.py", "scenarios/scenario_runner.py"],
        "deprecated": False,
    },
    "Resource_Master": {
        "role": "Equipment/resource master",
        "header_row": 3,
        "used_columns": [
            "Resource_ID",
            "Resource_Name",
            "Plant",
            "Status",
            "Avail_Hours_Day",
            "Efficiency_%",
            "Operation_Group",
            "Default_Cycle_Min",
            "Default_Setup_Min",
        ],
        "consumers": ["engine/scheduler.py", "engine/capacity.py", "xaps_application_api.py"],
        "deprecated": False,
    },
    "Routing": {
        "role": "Process routing and timing",
        "header_row": 3,
        "used_columns": [
            "SKU_ID",
            "Grade",
            "Operation",
            "Resource_Group",
            "Preferred_Resource",
            "Cycle_Time_Min_Heat",
            "Setup_Time_Min",
            "Transfer_Time_Min",
            "Sequence",
            "Op_Seq",
            "Min_Campaign_MT",
            "Max_Campaign_MT",
            "Is_Optional",
            "Optional_Condition",
        ],
        "consumers": ["engine/scheduler.py", "engine/campaign.py", "engine/capacity.py", "engine/ctp.py", "xaps_application_api.py"],
        "deprecated": False,
    },
    "Campaign_Config": {
        "role": "Campaign sequencing master",
        "header_row": 3,
        "used_columns": ["Grade", "Grade_Seq_Order"],
        "consumers": ["engine/campaign.py", "engine/ctp.py", "xaps_application_api.py"],
        "deprecated": False,
    },
    "Changeover_Matrix": {
        "role": "Grade-to-grade changeover matrix",
        "header_row": 3,
        "used_columns": ["From \\ To"],
        "consumers": ["engine/scheduler.py", "engine/capacity.py", "xaps_application_api.py"],
        "deprecated": False,
    },
    "Queue_Times": {
        "role": "Queue constraints between operations",
        "header_row": 3,
        "used_columns": ["From_Operation", "To_Operation", "Min_Queue_Min", "Max_Queue_Min", "Enforcement"],
        "consumers": ["engine/scheduler.py", "xaps_application_api.py"],
        "deprecated": False,
    },
    "Scenarios": {
        "role": "Scenario inputs",
        "header_row": 3,
        "used_columns": ["Parameter", "Value"],
        "consumers": ["scenarios/scenario_runner.py", "xaps_application_api.py"],
        "deprecated": False,
    },
    "CTP_Request": {
        "role": "CTP request input",
        "header_row": 3,
        "used_columns": ["Request_ID"],
        "consumers": ["xaps_application_api.py"],
        "deprecated": False,
    },
    "CTP_Output": {
        "role": "CTP output persistence",
        "header_row": 3,
        "used_columns": ["Request_ID"],
        "consumers": ["xaps_application_api.py"],
        "deprecated": False,
    },
}


def _sheet_dataframe(path: Path, sheet_name: str, header_row: int) -> pd.DataFrame:
    return pd.read_excel(path, sheet_name=sheet_name, header=header_row - 1, dtype=str).dropna(how="all").reset_index(drop=True)


def audit_workbook_masterdata(workbook_path: str | Path) -> dict[str, Any]:
    path = Path(workbook_path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet_names = set(wb.sheetnames)
    finally:
        wb.close()

    sheets: dict[str, Any] = {}
    for sheet_name, spec in MASTERDATA_AUDIT_SPEC.items():
        report = {
            "role": spec["role"],
            "consumers": list(spec["consumers"]),
            "deprecated": bool(spec.get("deprecated", False)),
            "exists": sheet_name in sheet_names,
        }
        if sheet_name not in sheet_names:
            report["missing"] = True
            sheets[sheet_name] = report
            continue

        try:
            df = _sheet_dataframe(path, sheet_name, int(spec["header_row"]))
            columns = [str(col).strip() for col in df.columns]
            used_columns = [col for col in spec["used_columns"] if col in columns]
            unused_columns = [col for col in columns if col not in spec["used_columns"]]
            report.update(
                {
                    "row_count": int(len(df)),
                    "columns": columns,
                    "used_columns_present": used_columns,
                    "used_columns_missing": [col for col in spec["used_columns"] if col not in columns],
                    "unused_columns": unused_columns,
                }
            )
        except Exception as exc:
            report["error"] = str(exc)
        sheets[sheet_name] = report

    config_duplicates: dict[str, Any] = {"conflict_count": 0, "conflicts": []}
    try:
        snapshot = load_workbook_config_snapshot(path)
        config_duplicates = {
            "primary_runtime_sheet": "Algorithm_Config",
            "compatibility_sheet": "Config",
            "conflict_count": len(snapshot.conflicts),
            "conflicts": snapshot.conflicts,
        }
    except Exception as exc:
        config_duplicates["error"] = str(exc)

    return {
        "workbook": str(path),
        "sheet_count": len(sheet_names),
        "master_sheets": sheets,
        "config_duplicates": config_duplicates,
    }
