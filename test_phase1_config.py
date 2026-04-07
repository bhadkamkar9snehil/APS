"""Test Phase 1: Configuration Infrastructure implementation."""

import sys
sys.path.insert(0, ".")

from data.loader import load_all
from engine.config import get_config
import json

def test_algorithm_config_sheet():
    """Test Algorithm_Config sheet was created with all parameters."""
    print("\n[TEST 1] Algorithm_Config Sheet Creation")
    print("-" * 60)

    import openpyxl
    wb = openpyxl.load_workbook("APS_BF_SMS_RM.xlsx")

    assert "Algorithm_Config" in wb.sheetnames, "Algorithm_Config sheet not found"
    ws = wb["Algorithm_Config"]

    # Check headers
    headers = [cell.value for cell in ws[1]]
    expected_headers = [
        "Config_Key", "Category", "Parameter_Name", "Current_Value",
        "Data_Type", "Min_Value", "Max_Value", "Unit", "Description",
        "Impact_Level"
    ]
    assert headers[:10] == expected_headers, f"Headers mismatch: {headers[:10]}"

    # Count data rows (exclude header)
    data_rows = ws.max_row - 1
    print(f"[OK] Sheet created with {data_rows} parameters")
    print(f"[OK] Headers validated: {', '.join(expected_headers)}")
    return True


def test_config_loading():
    """Test configuration loads from Excel sheet."""
    print("\n[TEST 2] Configuration Loading")
    print("-" * 60)

    data = load_all("APS_BF_SMS_RM.xlsx")
    config = get_config()

    params = config.all_params()
    assert len(params) > 0, "Configuration did not load any parameters"
    print(f"[OK] Loaded {len(params)} parameters from Algorithm_Config sheet")

    # Verify all expected categories exist
    categories = set(config.metadata[k]['category'] for k in params)
    expected_categories = {'SCHEDULER', 'CAMPAIGN', 'BOM', 'CTP', 'CAPACITY'}
    assert expected_categories.issubset(categories), f"Missing categories: {expected_categories - categories}"
    print(f"[OK] All 5 categories present: {', '.join(sorted(categories))}")

    return True


def test_config_validation():
    """Test configuration validation and bounds checking."""
    print("\n[TEST 3] Configuration Validation")
    print("-" * 60)

    data = load_all("APS_BF_SMS_RM.xlsx")
    config = get_config()

    # Test numeric bounds
    eaf_time = config.get("CYCLE_TIME_EAF_MIN")
    assert 60 <= eaf_time <= 180, f"EAF time {eaf_time} outside valid range [60, 180]"
    print(f"[OK] CYCLE_TIME_EAF_MIN = {eaf_time} (valid range: 60-180)")

    # Test list parameters
    vd_grades = config.get_list("VD_REQUIRED_GRADES")
    assert len(vd_grades) > 0, "VD_REQUIRED_GRADES is empty"
    assert "1080" in vd_grades, "SAE 1080 not in VD grades"
    print(f"[OK] VD_REQUIRED_GRADES = {vd_grades}")

    # Test boolean parameters
    setup_first = config.get_bool("SETUP_TIME_FIRST_HEAT_ONLY")
    assert isinstance(setup_first, bool), "Setup parameter is not boolean"
    print(f"[OK] SETUP_TIME_FIRST_HEAT_ONLY = {setup_first}")

    return True


def test_getter_methods():
    """Test type-safe getter methods."""
    print("\n[TEST 4] Type-Safe Getter Methods")
    print("-" * 60)

    data = load_all("APS_BF_SMS_RM.xlsx")
    config = get_config()

    # Test duration getter
    eaf_min = config.get_duration_minutes("CYCLE_TIME_EAF_MIN")
    assert isinstance(eaf_min, int), "Duration getter did not return int"
    assert eaf_min == 90, f"Expected 90, got {eaf_min}"
    print(f"[OK] get_duration_minutes('CYCLE_TIME_EAF_MIN') = {eaf_min}")

    # Test weight getter
    queue_weight = config.get_weight("OBJECTIVE_QUEUE_VIOLATION_WEIGHT")
    assert isinstance(queue_weight, int), "Weight getter did not return int"
    assert queue_weight == 500, f"Expected 500, got {queue_weight}"
    print(f"[OK] get_weight('OBJECTIVE_QUEUE_VIOLATION_WEIGHT') = {queue_weight}")

    # Test percentage getter
    ratio = config.get_percentage("OBJECTIVE_SMS_LATENESS_RATIO")
    assert isinstance(ratio, float), "Percentage getter did not return float"
    assert ratio == 0.5, f"Expected 0.5, got {ratio}"
    print(f"[OK] get_percentage('OBJECTIVE_SMS_LATENESS_RATIO') = {ratio}")

    # Test list getter
    grades = config.get_list("LOW_CARBON_BILLET_GRADES")
    assert isinstance(grades, list), "List getter did not return list"
    assert "1008" in grades, "SAE 1008 not in list"
    print(f"[OK] get_list('LOW_CARBON_BILLET_GRADES') = {grades}")

    return True


def test_category_filtering():
    """Test filtering parameters by category."""
    print("\n[TEST 5] Category-Based Filtering")
    print("-" * 60)

    data = load_all("APS_BF_SMS_RM.xlsx")
    config = get_config()

    # Test each category
    categories = {
        'SCHEDULER': 16,  # Expected count
        'CAMPAIGN': 14,
        'BOM': 7,
        'CTP': 6,
        'CAPACITY': 3,
    }

    for category, expected_count in categories.items():
        params = config.params_by_category(category)
        actual_count = len(params)
        assert actual_count == expected_count, f"Category {category}: expected {expected_count}, got {actual_count}"
        print(f"[OK] Category {category}: {actual_count} parameters")

    return True


def test_config_metadata():
    """Test configuration metadata (descriptions, min/max, etc)."""
    print("\n[TEST 6] Configuration Metadata")
    print("-" * 60)

    data = load_all("APS_BF_SMS_RM.xlsx")
    config = get_config()

    # Check metadata for a sample parameter
    meta = config.metadata.get("CYCLE_TIME_EAF_MIN")
    assert meta is not None, "No metadata for CYCLE_TIME_EAF_MIN"
    assert meta.get('data_type') == 'DURATION', f"Wrong data type: {meta.get('data_type')}"
    assert meta.get('min') == 60, f"Wrong min: {meta.get('min')}"
    assert meta.get('max') == 180, f"Wrong max: {meta.get('max')}"
    assert meta.get('category') == 'SCHEDULER', f"Wrong category: {meta.get('category')}"

    print(f"[OK] Parameter metadata complete:")
    print(f"  - Data type: {meta.get('data_type')}")
    print(f"  - Min: {meta.get('min')}, Max: {meta.get('max')}")
    print(f"  - Category: {meta.get('category')}")
    print(f"  - Description: {meta.get('description')[:50]}...")

    return True


def test_all_params_dict():
    """Test all_params() returns complete dictionary."""
    print("\n[TEST 7] All Parameters Dictionary")
    print("-" * 60)

    data = load_all("APS_BF_SMS_RM.xlsx")
    config = get_config()

    all_params = config.all_params()
    assert len(all_params) == 46, f"Expected 46 params, got {len(all_params)}"
    assert "CYCLE_TIME_EAF_MIN" in all_params, "CYCLE_TIME_EAF_MIN not in dict"
    assert "CTP_SCORE_STOCK_ONLY" in all_params, "CTP_SCORE_STOCK_ONLY not in dict"

    print(f"[OK] all_params() returns {len(all_params)} parameters")
    print(f"[OK] Sample keys: {', '.join(list(all_params.keys())[:5])}")

    return True


def main():
    """Run all tests."""
    print("=" * 60)
    print("Phase 1: Configuration Infrastructure — Test Suite")
    print("=" * 60)

    tests = [
        test_algorithm_config_sheet,
        test_config_loading,
        test_config_validation,
        test_getter_methods,
        test_category_filtering,
        test_config_metadata,
        test_all_params_dict,
    ]

    passed = 0
    failed = 0

    for test in tests:
        try:
            if test():
                passed += 1
        except AssertionError as e:
            print(f"[FAIL] FAILED: {e}")
            failed += 1
        except Exception as e:
            print(f"[FAIL] ERROR: {e}")
            failed += 1

    print("\n" + "=" * 60)
    print(f"Results: {passed} passed, {failed} failed")
    print("=" * 60)

    return failed == 0


if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)
