"""
Build APS_BF_SMS_RM.xlsx
Plant: 1 Blast Furnace | 1 SMS (2 EAF + 3 LRF + 1 VD + 2 CCM) | 2 Rolling Mills
Products: Wire Rod Coils, varying grades and cross-sections
EAF heat size: 50 MT

Run once: python build_template_v3.py
"""
import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime, timedelta
import os, random

random.seed(42)
OUTPUT = os.path.join(os.path.dirname(__file__), "APS_BF_SMS_RM.xlsx")

# ── Palette ─────────────────────────────────────────────────────────────────
C_DARK_BLUE   = "1F3864"
C_MID_BLUE    = "2F5496"
C_LIGHT_BLUE  = "BDD7EE"
C_GREEN       = "375623"
C_LIGHT_GREEN = "E2EFDA"
C_AMBER       = "7F6000"
C_LIGHT_AMBER = "FFEB9C"
C_RED         = "9C0006"
C_LIGHT_RED   = "FFC7CE"
C_GREY_HDR    = "F2F2F2"
C_WHITE       = "FFFFFF"
C_ORANGE_TAB  = "ED7D31"
C_GREEN_TAB   = "70AD47"
C_HELP_TAB    = "5B9BD5"
C_DOC_FILL    = "EAF2F8"
C_DOC_BORDER  = "9CC2E5"

def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def hdr_style(ws, row, col, text, bg=C_MID_BLUE, fg=C_WHITE, bold=True, wrap=False):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(bold=bold, color=fg, name="Calibri", size=10)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    c.border = thin_border()
    return c

def data_cell(ws, row, col, value, bold=False, align="center", num_fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, name="Calibri", size=10)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border = thin_border()
    if num_fmt:
        c.number_format = num_fmt
    return c

def write_table(ws, headers, rows, start_row=1, hdr_bg=C_MID_BLUE, col_widths=None):
    for c, h in enumerate(headers, 1):
        hdr_style(ws, start_row, c, h, bg=hdr_bg)
    for r, row in enumerate(rows, start_row + 1):
        for c, val in enumerate(row, 1):
            data_cell(ws, r, c, val)
    # Auto width
    for c, h in enumerate(headers, 1):
        vals = [str(h)] + [str(r[c-1]) for r in rows]
        w = max(len(v) for v in vals) + 4
        if col_widths and c-1 < len(col_widths):
            w = col_widths[c-1]
        ws.column_dimensions[get_column_letter(c)].width = min(w, 35)
    ws.freeze_panes = ws.cell(row=start_row+1, column=1)

def sheet_title(ws, text, subtitle=""):
    ws.cell(row=1, column=1, value=text).font = Font(bold=True, size=14, color=C_DARK_BLUE, name="Calibri")
    if subtitle:
        ws.cell(row=2, column=1, value=subtitle).font = Font(size=10, color="808080", italic=True, name="Calibri")
    return 3 if subtitle else 2

def add_doc_panel(ws, start_col, lines, title="Sheet Guide", width=6):
    end_col = start_col + width - 1

    ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
    head = ws.cell(1, start_col, value=title)
    head.font = Font(bold=True, color=C_DARK_BLUE, size=11, name="Calibri")
    head.fill = PatternFill("solid", fgColor=C_LIGHT_BLUE)
    head.alignment = Alignment(horizontal="left", vertical="center")
    head.border = thin_border()

    for row_idx, line in enumerate(lines, start=2):
        ws.merge_cells(start_row=row_idx, start_column=start_col, end_row=row_idx, end_column=end_col)
        cell = ws.cell(row_idx, start_col, value=line)
        cell.font = Font(size=9, color="4F4F4F", name="Calibri")
        cell.fill = PatternFill("solid", fgColor=C_DOC_FILL)
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell.border = thin_border()

    for col in range(start_col, end_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14


def build_help_sheet(ws, sections):
    ws.sheet_properties.tabColor = C_HELP_TAB
    row = sheet_title(
        ws,
        "APS Help",
        "Central workbook reference. Sheet-level guide panels have been removed so the working sheets stay clean."
    )
    ws.cell(row, 1).value = "Use Control_Panel to run the APS. Use this Help tab for sheet purpose, editing guidance, and output interpretation."
    ws.cell(row, 1).font = Font(size=10, color="808080", italic=True, name="Calibri")
    row += 2

    for sheet_name, (_, lines) in sections.items():
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        head = ws.cell(row, 1, value=sheet_name)
        head.font = Font(bold=True, color=C_WHITE, size=11, name="Calibri")
        head.fill = PatternFill("solid", fgColor=C_DARK_BLUE)
        head.alignment = Alignment(horizontal="left", vertical="center")
        head.border = thin_border()
        row += 1

        for line in lines:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            cell = ws.cell(row, 1, value=line)
            cell.font = Font(size=10, color="2F2F2F", name="Calibri")
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.border = thin_border()
            row += 1

        row += 1

    for col, width in zip(("A", "B", "C", "D"), (28, 28, 28, 28)):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A6"

wb = openpyxl.Workbook()

# ════════════════════════════════════════════════════════════════════════════
# SHEET 1: Config
# ════════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Config"
ws.sheet_properties.tabColor = C_GREEN_TAB
r = sheet_title(ws, "Config", "System key-value parameters that let the APS behave from workbook masters instead of Python constants.")

headers = ["Key", "Value", "Description"]
config_rows = [
    ("Batch_Unit_Name", "Heat", "Planner label for the batch unit used in scheduling and documentation."),
    ("Primary_Batch_Resource_Group", "EAF", "Resource group that defines the primary batch in steel planning."),
    ("Default_Batch_Size_MT", 50.0, "Fallback batch size in MT when no other batch-size rule overrides it."),
    ("Campaign_Group_By", "Route_Family,Campaign_Group,Grade,Product_Family,Route_Variant", "CSV list of sales-order attributes used to build compatible campaigns."),
    ("Planning_Horizon_Days", 14, "Default scheduling horizon in days when Scenarios does not override it."),
    ("Default_Solver_Limit_Sec", 30.0, "Default CP-SAT solver time limit in seconds."),
    ("Min_Campaign_MT", 100.0, "Fallback minimum campaign tonnage."),
    ("Max_Campaign_MT", 500.0, "Fallback maximum campaign tonnage."),
    ("Queue_Enforcement", "Hard", "Default queue-time treatment: Hard or Soft."),
    ("Campaign_Serialization_Mode", "STRICT_END_TO_END", "Production policy: STRICT_END_TO_END. Workbook production runs keep full end-to-end campaign serialization."),
    ("Allow_Scheduler_Default_Masters", "N", "Y/N switch. Default N = fail fast on missing routing/resource/timing masters; workbook production runs require N."),
    ("BOM_Structure_Error_Mode", "RAISE", "Production policy: RAISE. BOM cycles / max-depth errors hard-fail workbook planning runs instead of being recorded quietly."),
    ("Allow_Legacy_Primary_Batch_Fallback", "N", "Y/N switch. Default N = block campaign release when primary-batch BOM tracing fails; Y = legacy heat estimate for diagnostics only."),
    ("Manual_Campaign_Grouping_Mode", "PRESERVE_EXACT", "Production policy: PRESERVE_EXACT. Manual Campaign_ID groupings stay intact instead of being auto-split by max campaign size."),
    ("Byproduct_Inventory_Mode", "DEFERRED", "Production policy: DEFERRED. Track byproducts but do not credit stock immediately in workbook planning runs."),
    ("Require_Authoritative_CTP_Inventory", "Y", "Y/N switch. Default Y = block workbook CTP promises when committed inventory lineage is non-authoritative."),
    ("Default_Section_Fallback", 6.5, "Fallback product section in mm when demand data is missing a section."),
    ("Workbook_Name", "aps_bf_sms_rm", "Workbook identity used by runtime support checks."),
]
write_table(ws, headers, config_rows, start_row=r, col_widths=[30, 26, 72])


# ════════════════════════════════════════════════════════════════════════════
# SHEET 2: Control_Panel
# ════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Control_Panel")
ws.sheet_properties.tabColor = C_DARK_BLUE

for col, width in {
    "A": 2.5, "B": 16, "C": 16, "D": 16, "E": 16,
    "F": 3.5, "G": 13, "H": 13, "I": 13, "J": 13, "K": 13, "L": 13,
    "M": 3.5, "N": 13, "O": 13, "P": 13, "Q": 13,
}.items():
    ws.column_dimensions[col].width = width

for row in range(1, 23):
    ws.row_dimensions[row].height = 20
ws.row_dimensions[1].height = 24
ws.row_dimensions[2].height = 20
for row in (8, 11, 14, 17):
    ws.row_dimensions[row].height = 28

ws.merge_cells("B1:Q1")
ws["B1"] = "APS - Steel Plant (BF | SMS | Rolling Mill)"
ws["B1"].font = Font(bold=True, size=18, color=C_DARK_BLUE, name="Calibri")
ws["B1"].alignment = Alignment(horizontal="left", vertical="center")

ws.merge_cells("B2:Q2")
ws["B2"] = "Advanced Planning & Scheduling | Primary batch = Heat | Standard heat size = 50 MT"
ws["B2"].font = Font(size=11, color="808080", italic=True, name="Calibri")
ws["B2"].alignment = Alignment(horizontal="left", vertical="center")

ws.merge_cells("B3:F3")
ws["B3"] = f"Workbook refreshed: {datetime.now().strftime('%d-%b-%Y %H:%M')}"
ws["B3"].font = Font(size=9, color="AAAAAA", name="Calibri")
ws["B3"].alignment = Alignment(horizontal="left", vertical="center")

control_panel_blocks = [
    ("B5:E5", "Plant Snapshot", C_DARK_BLUE, C_WHITE, True, 11),
    ("B6:E6", "1 Blast Furnace", C_WHITE, C_DARK_BLUE, False, 10),
    ("B7:E7", "2 EAF  |  3 LRF  |  1 VD", "F7F9FC", C_DARK_BLUE, False, 10),
    ("B8:E8", "2 Continuous Casters", C_WHITE, C_DARK_BLUE, False, 10),
    ("B9:E9", "2 Rolling Mills", "F7F9FC", C_DARK_BLUE, False, 10),
    ("B10:E10", "Planning unit: Heat | 50 MT", "FFF4CC", C_AMBER, False, 10),
    ("G5:L5", "APS Actions", C_DARK_BLUE, C_WHITE, True, 11),
    ("G6:L6", "Run the full plan or trigger one focused step.", C_DOC_FILL, "808080", False, 10),
    ("N5:Q5", "Recommended Flow", C_DARK_BLUE, C_WHITE, True, 11),
    ("N6:Q6", "1. Update Sales_Orders, Inventory, or Scenarios.", C_WHITE, C_DARK_BLUE, False, 10),
    ("N7:Q7", "2. Run Schedule for the live plan.", "F7F9FC", C_DARK_BLUE, False, 10),
    ("N8:Q8", "3. Review Campaign_Schedule and Equipment_Schedule.", C_WHITE, C_DARK_BLUE, False, 10),
    ("N9:Q9", "4. Compare cases in Scenario_Output only when needed.", "F7F9FC", C_DARK_BLUE, False, 10),
    ("N10:Q10", "5. Use Help for detailed sheet explanations.", C_WHITE, C_DARK_BLUE, False, 10),
    ("N12:Q12", "Review Outputs", C_DARK_BLUE, C_WHITE, True, 11),
    ("N13:Q13", "Campaign_Schedule: release order and due-date margin.", C_WHITE, C_DARK_BLUE, False, 10),
    ("N14:Q14", "Schedule_Output: master row-level plan across SMS and RM.", "F7F9FC", C_DARK_BLUE, False, 10),
    ("N15:Q15", "Equipment_Schedule: machine-wise dispatch packets.", C_WHITE, C_DARK_BLUE, False, 10),
    ("N16:Q16", "Material_Plan: campaign material commitment and shortages.", "F7F9FC", C_DARK_BLUE, False, 10),
    ("N17:Q17", "Schedule_Gantt: visual resource timeline.", C_WHITE, C_DARK_BLUE, False, 10),
]

for range_ref, text, bg, fg, bold, size in control_panel_blocks:
    ws.merge_cells(range_ref)
    cell = ws[range_ref.split(":")[0]]
    cell.value = text
    cell.font = Font(bold=bold, color=fg, size=size, name="Calibri", italic=(range_ref == "G6:L6"))
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="left" if range_ref.startswith("N") or range_ref == "G6:L6" else "center", vertical="center", wrap_text=True)
    cell.border = thin_border()


# ════════════════════════════════════════════════════════════════════════════
# SHEET 3: SKU_Master
# ════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("SKU_Master")
ws.sheet_properties.tabColor = "4472C4"
r = sheet_title(ws, "SKU Master", "All products: Finished Goods, stage outputs, process intermediates, byproducts, and raw materials")

# Grade colour map (for reference column)
GRADE_COLORS = {
    "SAE 1008": "DDEBF7", "SAE 1018": "BDD7EE", "SAE 1035": "9DC3E6",
    "SAE 1045": "2E75B6", "SAE 1065": "1F4E79", "SAE 1080": "833C00",
    "CHQ 1006": "7030A0", "Cr-Mo 4140": "4472C4",
}

headers = ["SKU_ID", "SKU_Name", "Category", "Grade", "Section_mm", "Coil_Wt_MT",
           "UOM", "Needs_VD", "Lead_Time_Days", "Safety_Stock_MT",
           "Route_Variant", "Product_Family", "Attribute_1"]

def grade_code(grade):
    return str(grade).replace(" ", "").replace("-", "")


def section_code(section_mm):
    return str(section_mm).replace(".", "")


LOW_CARBON_BILLET_GRADES = {"SAE 1008", "SAE 1018", "SAE 1035"}


def product_family_for_grade(grade):
    return "BIL-130" if grade in LOW_CARBON_BILLET_GRADES else "BIL-150"


VD_GRADES = {"SAE 1080", "CHQ 1006", "Cr-Mo 4140"}
rows = []
# ── Finished Goods: Wire Rod Coils ──
coils = [
    # (grade, section, needs_vd)
    ("SAE 1008", 5.5,  "N"), ("SAE 1008", 6.5,  "N"), ("SAE 1008", 8.0,  "N"),
    ("SAE 1018", 5.5,  "N"), ("SAE 1018", 6.5,  "N"), ("SAE 1018", 8.0,  "N"),
    ("SAE 1018", 10.0, "N"), ("SAE 1018", 12.0, "N"),
    ("SAE 1035", 6.5,  "N"), ("SAE 1035", 8.0,  "N"),
    ("SAE 1045", 6.5,  "N"), ("SAE 1045", 8.0,  "N"), ("SAE 1045", 10.0, "N"),
    ("SAE 1065", 5.5,  "N"), ("SAE 1065", 6.5,  "N"),
    ("SAE 1080", 5.5,  "Y"), ("SAE 1080", 6.5,  "Y"),
    ("CHQ 1006", 5.5,  "Y"), ("CHQ 1006", 6.5,  "Y"),
    ("Cr-Mo 4140", 8.0, "Y"), ("Cr-Mo 4140", 10.0, "Y"),
]
unique_grades = list(dict.fromkeys(grade for grade, _, _ in coils))
for grade, sec, vd in coils:
    sku = f"FG-WR-{grade_code(grade)}-{section_code(sec)}"
    name = f"Wire Rod {grade} {sec}mm"
    rows.append((sku, name, "Finished Good", grade, sec, 2.0, "MT", vd, 3, 20,
                 vd, product_family_for_grade(grade), sec))

# ── Stage Outputs / Intermediates ──
rm_output_skus = {}
for grade, sec, vd in coils:
    sku = f"RM-OUT-{grade_code(grade)}-{section_code(sec)}"
    rm_output_skus[(grade, sec)] = sku
    rows.append((sku, f"RM Output {grade} {sec}mm", "Process Intermediate", grade, sec, 2.0, "MT", vd, 1, 0,
                 vd, product_family_for_grade(grade), sec))

billets = [
    ("BIL-130-1008", "Billet 130mm SAE 1008", "SAE 1008", 130, "N"),
    ("BIL-130-1018", "Billet 130mm SAE 1018", "SAE 1018", 130, "N"),
    ("BIL-130-1035", "Billet 130mm SAE 1035", "SAE 1035", 130, "N"),
    ("BIL-150-1045", "Billet 150mm SAE 1045", "SAE 1045", 150, "N"),
    ("BIL-150-1065", "Billet 150mm SAE 1065", "SAE 1065", 150, "N"),
    ("BIL-150-1080", "Billet 150mm SAE 1080", "SAE 1080", 150, "Y"),
    ("BIL-150-CHQ",  "Billet 150mm CHQ 1006", "CHQ 1006", 150, "Y"),
    ("BIL-150-4140", "Billet 150mm Cr-Mo 4140","Cr-Mo 4140",150,"Y"),
]
billet_names = {}
for sku, name, grade, sec, vd in billets:
    billet_names[sku] = name
    rows.append((sku, f"CCM Output {name}", "Semi-Finished", grade, sec, 50.0, "MT", vd, 1, 100,
                 vd, product_family_for_grade(grade), None))

eaf_raw_skus = {}
eaf_out_skus = {}
lrf_out_skus = {}
vd_out_skus = {}
for grade in unique_grades:
    gd = grade_code(grade)
    needs_vd = "Y" if grade in VD_GRADES else "N"
    eaf_raw_skus[grade] = f"EAF-RAW-{gd}"
    eaf_out_skus[grade] = f"EAF-OUT-{gd}"
    lrf_out_skus[grade] = f"LRF-OUT-{gd}"
    product_family = product_family_for_grade(grade)
    rows.append((eaf_raw_skus[grade], f"EAF Raw Material Charge {grade}", "Process Intermediate", grade, "-", 50.0, "MT", needs_vd, 0, 0,
                 needs_vd, product_family, None))
    rows.append((eaf_out_skus[grade], f"EAF Output Liquid Steel {grade}", "Process Intermediate", grade, "-", 50.0, "MT", needs_vd, 0, 0,
                 needs_vd, product_family, None))
    rows.append((lrf_out_skus[grade], f"LRF Output Graded Liquid Steel {grade}", "Process Intermediate", grade, "-", 50.0, "MT", needs_vd, 0, 0,
                 needs_vd, product_family, None))
    if grade in VD_GRADES:
        vd_out_skus[grade] = f"VD-OUT-{gd}"
        rows.append((vd_out_skus[grade], f"VD Output Degassed Steel {grade}", "Process Intermediate", grade, "-", 50.0, "MT", "Y", 0, 0,
                     "Y", product_family, None))

rows.extend([
    ("BF-RAW-MIX", "Blast Furnace Raw Material Mix", "Process Intermediate", "-", "-", 50.0, "MT", "N", 0, 0, "N", "", None),
    ("BF-HM", "Hot Metal from Blast Furnace", "Process Intermediate", "-", "-", 50.0, "MT", "N", 0, 500, "N", "", None),
])

rows.extend([
    ("BF-SLAG", "Blast Furnace Slag", "Byproduct/Waste", "-", "-", 1.0, "MT", "N", 0, 0, "N", "", None),
    ("EAF-SLAG", "EAF Slag", "Byproduct/Waste", "-", "-", 1.0, "MT", "N", 0, 0, "N", "", None),
    ("LRF-WASTE", "LRF Waste / Skull", "Byproduct/Waste", "-", "-", 1.0, "MT", "N", 0, 0, "N", "", None),
    ("VD-WASTE", "VD Waste / Skull", "Byproduct/Waste", "-", "-", 1.0, "MT", "N", 0, 0, "N", "", None),
    ("CCM-CROP", "CCM Crop Loss", "Byproduct/Waste", "-", "-", 1.0, "MT", "N", 0, 0, "N", "", None),
    ("RM-ENDCUT", "Rolling Mill End Cuts", "Byproduct/Waste", "-", "-", 1.0, "MT", "N", 0, 0, "N", "", None),
    ("RM-SCALE", "Rolling Mill Scale / Misc Waste", "Byproduct/Waste", "-", "-", 1.0, "MT", "N", 0, 0, "N", "", None),
])

# ── Raw Materials ──
rms = [
    ("RM-SCRAP", "Steel Scrap",             "Raw Material", "-", "-", 1.0,  "MT", "N", 3,  500, "N", "", None),
    ("RM-LIME",  "Lime",                    "Raw Material", "-", "-", 1.0,  "MT", "N", 2,  300, "N", "", None),
    ("RM-DOLO",  "Dolomite",                "Raw Material", "-", "-", 1.0,  "MT", "N", 2,  200, "N", "", None),
    ("RM-FESI",  "Ferro Silicon (FeSi)",    "Raw Material", "-", "-", 1.0,  "MT", "N", 7,  20, "N", "", None),
    ("RM-FEMN",  "Ferro Manganese (FeMn)",  "Raw Material", "-", "-", 1.0,  "MT", "N", 7,  25, "N", "", None),
    ("RM-FECR",  "Ferro Chrome (FeCr)",     "Raw Material", "-", "-", 1.0,  "MT", "N", 10, 15, "N", "", None),
    ("RM-ELEC",  "Graphite Electrodes",     "Raw Material", "-", "-", 1.0,  "Pcs","N", 14, 30, "N", "", None),
    ("RM-COAL",  "Coal / Coke (BF Fuel)",   "Raw Material", "-", "-", 1.0,  "MT", "N", 5,  1000, "N", "", None),
    ("RM-IRON",  "Iron Ore Pellets",        "Raw Material", "-", "-", 1.0,  "MT", "N", 3,  2000, "N", "", None),
]
for row in rms:
    rows.append(row)

write_table(ws, headers, rows, start_row=r,
            col_widths=[22, 30, 18, 14, 12, 14, 6, 9, 16, 18, 14, 14, 12])

# Conditional formatting by material class
category_colors = {
    "Finished Good": "DEEAF1",
    "Semi-Finished": "E2EFDA",
    "Process Intermediate": "FCE4D6",
    "Raw Material": "FFF2CC",
    "Byproduct/Waste": "EDEDED",
}
for row_idx in range(r+1, r+1+len(rows)):
    cat = ws.cell(row_idx, 3).value
    color = category_colors.get(cat)
    if color:
        for col in range(1, len(headers)+1):
            ws.cell(row_idx, col).fill = PatternFill("solid", fgColor=color)


# ════════════════════════════════════════════════════════════════════════════
# SHEET 4: BOM
# ════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("BOM")
ws.sheet_properties.tabColor = "4472C4"
r = sheet_title(ws, "Bill of Materials", "Stagewise BOM: FG -> RM Output -> CCM -> LRF/VD -> EAF -> BF/raw materials, with byproducts tagged separately.")

headers = ["BOM_ID", "Parent_SKU", "Child_SKU", "Flow_Type", "Qty_Per", "Scrap_%",
           "Yield_Pct", "Level", "UOM", "Note"]

bom_rows = []

coil_to_billet = {
    "SAE 1008": "BIL-130-1008", "SAE 1018": "BIL-130-1018",
    "SAE 1035": "BIL-130-1035", "SAE 1045": "BIL-150-1045",
    "SAE 1065": "BIL-150-1065", "SAE 1080": "BIL-150-1080",
    "CHQ 1006": "BIL-150-CHQ",  "Cr-Mo 4140": "BIL-150-4140",
}


def add_bom_row(parent_sku, child_sku, flow_type, qty_per, scrap_pct, level, uom, note):
    yield_pct = round(max(0.0, 100.0 - float(scrap_pct)), 2) if flow_type == "INPUT" else None
    bom_rows.append((f"BOM-{add_bom_row.bid:03d}", parent_sku, child_sku, flow_type, qty_per, scrap_pct, yield_pct, level, uom, note))
    add_bom_row.bid += 1


add_bom_row.bid = 1


def rm_loss_pct(section_mm):
    return 12.0 if section_mm < 6.5 else (10.5 if section_mm < 8.0 else 9.0)


eaf_charge_recipes = {
    "SAE 1008": [("BF-HM", 0.60, 1.0, "MT"), ("RM-SCRAP", 0.44, 1.0, "MT"), ("RM-LIME", 0.06, 0.0, "MT"), ("RM-DOLO", 0.02, 0.0, "MT"), ("RM-FESI", 0.003, 0.0, "MT"), ("RM-FEMN", 0.008, 0.0, "MT"), ("RM-ELEC", 0.003, 0.0, "Pcs")],
    "SAE 1018": [("BF-HM", 0.58, 1.0, "MT"), ("RM-SCRAP", 0.45, 1.0, "MT"), ("RM-LIME", 0.06, 0.0, "MT"), ("RM-DOLO", 0.02, 0.0, "MT"), ("RM-FESI", 0.004, 0.0, "MT"), ("RM-FEMN", 0.010, 0.0, "MT"), ("RM-ELEC", 0.003, 0.0, "Pcs")],
    "SAE 1035": [("BF-HM", 0.58, 1.0, "MT"), ("RM-SCRAP", 0.45, 1.0, "MT"), ("RM-LIME", 0.06, 0.0, "MT"), ("RM-DOLO", 0.02, 0.0, "MT"), ("RM-FESI", 0.005, 0.0, "MT"), ("RM-FEMN", 0.012, 0.0, "MT"), ("RM-ELEC", 0.003, 0.0, "Pcs")],
    "SAE 1045": [("BF-HM", 0.56, 1.0, "MT"), ("RM-SCRAP", 0.47, 1.0, "MT"), ("RM-LIME", 0.06, 0.0, "MT"), ("RM-DOLO", 0.02, 0.0, "MT"), ("RM-FESI", 0.006, 0.0, "MT"), ("RM-FEMN", 0.015, 0.0, "MT"), ("RM-ELEC", 0.003, 0.0, "Pcs")],
    "SAE 1065": [("BF-HM", 0.55, 1.0, "MT"), ("RM-SCRAP", 0.48, 1.0, "MT"), ("RM-LIME", 0.05, 0.0, "MT"), ("RM-DOLO", 0.02, 0.0, "MT"), ("RM-FESI", 0.007, 0.0, "MT"), ("RM-FEMN", 0.018, 0.0, "MT"), ("RM-ELEC", 0.003, 0.0, "Pcs")],
    "SAE 1080": [("BF-HM", 0.55, 1.0, "MT"), ("RM-SCRAP", 0.48, 1.0, "MT"), ("RM-LIME", 0.05, 0.0, "MT"), ("RM-DOLO", 0.02, 0.0, "MT"), ("RM-FESI", 0.008, 0.0, "MT"), ("RM-FEMN", 0.020, 0.0, "MT"), ("RM-ELEC", 0.003, 0.0, "Pcs")],
    "CHQ 1006": [("BF-HM", 0.60, 1.0, "MT"), ("RM-SCRAP", 0.43, 1.0, "MT"), ("RM-LIME", 0.06, 0.0, "MT"), ("RM-DOLO", 0.02, 0.0, "MT"), ("RM-FESI", 0.003, 0.0, "MT"), ("RM-FEMN", 0.005, 0.0, "MT"), ("RM-ELEC", 0.003, 0.0, "Pcs")],
    "Cr-Mo 4140": [("BF-HM", 0.58, 1.0, "MT"), ("RM-SCRAP", 0.45, 1.0, "MT"), ("RM-LIME", 0.06, 0.0, "MT"), ("RM-DOLO", 0.02, 0.0, "MT"), ("RM-FESI", 0.010, 0.0, "MT"), ("RM-FEMN", 0.012, 0.0, "MT"), ("RM-FECR", 0.014, 0.0, "MT"), ("RM-ELEC", 0.003, 0.0, "Pcs")],
}

# Level 1-2: Final FG -> RM output -> CCM billet, with RM wastage modeled as byproducts.
for grade, sec, vd in coils:
    fg_sku = f"FG-WR-{grade_code(grade)}-{section_code(sec)}"
    rm_out = rm_output_skus[(grade, sec)]
    billet_sku = coil_to_billet[grade]
    loss_pct = rm_loss_pct(sec)
    end_cut_ratio = round((loss_pct / 100.0) * 0.75, 3)
    scale_ratio = round((loss_pct / 100.0) * 0.25, 3)

    add_bom_row(fg_sku, rm_out, "INPUT", 1.0, 0.0, 1, "MT", "Final finished good dispatch after inspection/packing")
    add_bom_row(rm_out, billet_sku, "INPUT", 1.0, loss_pct, 2, "MT", f"Rolling yield {100 - loss_pct:.1f}% from CCM billet")
    add_bom_row(rm_out, "RM-ENDCUT", "BYPRODUCT", end_cut_ratio, 0.0, 2, "MT", "End cuts from crop/shear losses")
    add_bom_row(rm_out, "RM-SCALE", "BYPRODUCT", scale_ratio, 0.0, 2, "MT", "Scale and miscellaneous rolling waste")

# Level 3-6: CCM -> VD/LRF -> EAF, with stage outputs and wastes.
for billet_sku, _, grade, _, _ in billets:
    if grade in vd_out_skus:
        add_bom_row(billet_sku, vd_out_skus[grade], "INPUT", 1.0, 5.0, 3, "MT", "CCM output requires VD-treated steel; caster yield 95%")
    else:
        add_bom_row(billet_sku, lrf_out_skus[grade], "INPUT", 1.0, 5.0, 3, "MT", "CCM output requires graded LRF steel; caster yield 95%")
    add_bom_row(billet_sku, "CCM-CROP", "BYPRODUCT", 0.05, 0.0, 3, "MT", "CCM crop/head-tail loss")

for grade in unique_grades:
    if grade in vd_out_skus:
        add_bom_row(vd_out_skus[grade], lrf_out_skus[grade], "INPUT", 1.0, 0.5, 4, "MT", "VD-treated steel generated from LRF output")
        add_bom_row(vd_out_skus[grade], "VD-WASTE", "BYPRODUCT", 0.005, 0.0, 4, "MT", "VD skull and process waste")
    add_bom_row(lrf_out_skus[grade], eaf_out_skus[grade], "INPUT", 1.0, 0.7, 5, "MT", "LRF output generated from EAF liquid steel")
    add_bom_row(lrf_out_skus[grade], "LRF-WASTE", "BYPRODUCT", 0.007, 0.0, 5, "MT", "LRF slag/skull and process waste")
    add_bom_row(eaf_out_skus[grade], eaf_raw_skus[grade], "INPUT", 1.0, 0.0, 6, "MT", "EAF liquid steel generated from charged EAF raw materials")
    add_bom_row(eaf_out_skus[grade], "EAF-SLAG", "BYPRODUCT", 0.12, 0.0, 6, "MT", "EAF slag and oxidation loss")

# Level 7: EAF raw materials, including hot metal from BF.
for grade, materials in eaf_charge_recipes.items():
    for rm_sku, qty, scrap, uom in materials:
        add_bom_row(eaf_raw_skus[grade], rm_sku, "INPUT", qty, scrap, 7, uom, "EAF charge and alloy recipe")

# Level 8-9: BF raw material mix -> hot metal + BF slag.
add_bom_row("BF-HM", "BF-RAW-MIX", "INPUT", 1.0, 0.0, 8, "MT", "Hot metal tapped from blast furnace burden")
add_bom_row("BF-HM", "BF-SLAG", "BYPRODUCT", 0.22, 0.0, 8, "MT", "Blast furnace slag generation")
add_bom_row("BF-RAW-MIX", "RM-IRON", "INPUT", 1.62, 1.5, 9, "MT", "BF ore burden rate")
add_bom_row("BF-RAW-MIX", "RM-COAL", "INPUT", 0.65, 2.0, 9, "MT", "BF coke/fuel rate")
add_bom_row("BF-RAW-MIX", "RM-LIME", "INPUT", 0.08, 0.0, 9, "MT", "BF flux addition")
add_bom_row("BF-RAW-MIX", "RM-DOLO", "INPUT", 0.03, 0.0, 9, "MT", "BF dolomite addition")

write_table(ws, headers, bom_rows, start_row=r,
            col_widths=[10, 26, 20, 12, 9, 9, 9, 8, 6, 28])


# ════════════════════════════════════════════════════════════════════════════
# SHEET 5: Inventory
# ════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Inventory")
ws.sheet_properties.tabColor = "4472C4"
r = sheet_title(ws, "Inventory", "Current stock levels for FG, stage outputs, raw materials, and waste/byproduct sinks. Available_Qty = unreserved stock on hand.")

headers = ["SKU_ID", "SKU_Name", "Location", "Available_Qty", "Reserved_Qty",
           "UOM", "Shelf_Life_Days", "Last_Updated"]
today = datetime.now().strftime("%Y-%m-%d")

inv_rows = []
# FG coils — some stock
fg_stock = {
    "SAE 1008": [("5.5", 40), ("6.5", 60), ("8.0", 30)],
    "SAE 1018": [("5.5", 20), ("6.5", 35), ("8.0", 25), ("10.0", 15), ("12.0", 10)],
    "SAE 1035": [("6.5", 18), ("8.0", 12)],
    "SAE 1045": [("6.5", 10), ("8.0", 8), ("10.0", 5)],
    "SAE 1065": [("5.5", 6), ("6.5", 4)],
    "SAE 1080": [("5.5", 0), ("6.5", 0)],
    "CHQ 1006": [("5.5", 0), ("6.5", 2)],
    "Cr-Mo 4140": [("8.0", 0), ("10.0", 0)],
}
for grade, sec, vd in coils:
    sku = f"FG-WR-{grade_code(grade)}-{section_code(sec)}"
    name = f"Wire Rod {grade} {sec}mm"
    avail = 0
    for s, q in fg_stock.get(grade, []):
        if float(s) == sec:
            avail = q
    inv_rows.append((sku, name, "WH-FG", avail, 0, "MT", 365, today))

# Rolling output / process intermediates
for grade, sec, vd in coils:
    sku = rm_output_skus[(grade, sec)]
    inv_rows.append((sku, f"RM Output {grade} {sec}mm", "WIP-RM", 0, 0, "MT", 30, today))

# Billets / CCM output
billet_stock = {
    "BIL-130-1008": 250, "BIL-130-1018": 180, "BIL-130-1035": 100,
    "BIL-150-1045": 120, "BIL-150-1065": 60,  "BIL-150-1080": 0,
    "BIL-150-CHQ":  0,   "BIL-150-4140": 0,
}
for sku, name in billet_names.items():
    inv_rows.append((sku, name, "WH-CCM", billet_stock[sku], 0, "MT", 180, today))

for grade in unique_grades:
    inv_rows.append((eaf_raw_skus[grade], f"EAF Raw Material Charge {grade}", "WIP-EAF", 0, 0, "MT", 7, today))
    inv_rows.append((eaf_out_skus[grade], f"EAF Output Liquid Steel {grade}", "WIP-EAF", 0, 0, "MT", 2, today))
    inv_rows.append((lrf_out_skus[grade], f"LRF Output Graded Liquid Steel {grade}", "WIP-LRF", 0, 0, "MT", 2, today))
    if grade in vd_out_skus:
        inv_rows.append((vd_out_skus[grade], f"VD Output Degassed Steel {grade}", "WIP-VD", 0, 0, "MT", 2, today))

inv_rows.append(("BF-RAW-MIX", "Blast Furnace Raw Material Mix", "WIP-BF", 0, 0, "MT", 7, today))
inv_rows.append(("BF-HM", "Hot Metal from Blast Furnace", "BF-Torpedo", 800, 0, "MT", 2, today))

for sku, name in [
    ("BF-SLAG", "Blast Furnace Slag"),
    ("EAF-SLAG", "EAF Slag"),
    ("LRF-WASTE", "LRF Waste / Skull"),
    ("VD-WASTE", "VD Waste / Skull"),
    ("CCM-CROP", "CCM Crop Loss"),
    ("RM-ENDCUT", "Rolling Mill End Cuts"),
    ("RM-SCALE", "Rolling Mill Scale / Misc Waste"),
]:
    inv_rows.append((sku, name, "WASTE-YARD", 0, 0, "MT", 30, today))

# Raw Materials
rm_stock = [
    ("RM-SCRAP", "Steel Scrap",          "SCRAP-YARD", 1200, 0),
    ("RM-LIME",  "Lime",                 "RM-STORE",   600, 0),
    ("RM-DOLO",  "Dolomite",             "RM-STORE",   400, 0),
    ("RM-FESI",  "Ferro Silicon",        "ALLOY-STORE", 28, 0),
    ("RM-FEMN",  "Ferro Manganese",      "ALLOY-STORE", 38, 0),
    ("RM-FECR",  "Ferro Chrome",         "ALLOY-STORE", 12, 0),
    ("RM-ELEC",  "Graphite Electrodes",  "EAF-STORE",  60, 0),
    ("RM-COAL",  "Coal/Coke",            "BF-STORE",  2500, 0),
    ("RM-IRON",  "Iron Ore Pellets",     "BF-STORE",  4000, 0),
]
for sku, name, loc, avail, res in rm_stock:
    inv_rows.append((sku, name, loc, avail, res, "MT", 180, today))

write_table(ws, headers, inv_rows, start_row=r,
            col_widths=[26, 28, 14, 14, 14, 6, 16, 14])

# Highlight zero-stock rows
for row_idx in range(r+1, r+1+len(inv_rows)):
    avail = ws.cell(row_idx, 4).value
    if avail == 0:
        for col in range(1, len(headers)+1):
            ws.cell(row_idx, col).fill = PatternFill("solid", fgColor=C_LIGHT_AMBER)


# ════════════════════════════════════════════════════════════════════════════
# SHEET 6: Sales_Orders
# ════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Sales_Orders")
ws.sheet_properties.tabColor = "4472C4"
r = sheet_title(ws, "Sales Orders", "Open orders from customers. Delivery dates drive the schedule.")

headers = ["SO_ID", "Customer", "Region", "SKU_ID", "Grade", "Section_mm",
           "Order_Qty_MT", "Coils_Count", "Order_Date", "Delivery_Date",
           "Priority", "Campaign_Group", "Campaign_ID", "Status"]

customers = [
    ("Tata Wiron",          "West"),
    ("Usha Martin",         "East"),
    ("JSW Steel",           "South"),
    ("Mesco Steel",         "North"),
    ("Shyam Wire",          "North"),
    ("Bansal Wire",         "West"),
    ("Rathi Bars",          "Central"),
    ("Sunflag Iron",        "West"),
    ("Facor Steels",        "Central"),
    ("Kalyani Carpenter",   "West"),
]

grade_section_pool = [
    ("SAE 1008", 5.5, "N"), ("SAE 1008", 6.5, "N"), ("SAE 1008", 8.0, "N"),
    ("SAE 1018", 5.5, "N"), ("SAE 1018", 6.5, "N"), ("SAE 1018", 8.0, "N"),
    ("SAE 1018", 10.0,"N"), ("SAE 1035", 6.5, "N"), ("SAE 1035", 8.0, "N"),
    ("SAE 1045", 6.5, "N"), ("SAE 1045", 8.0, "N"),
    ("SAE 1065", 5.5, "N"), ("SAE 1065", 6.5, "N"),
    ("SAE 1080", 5.5, "Y"), ("SAE 1080", 6.5, "Y"),
    ("CHQ 1006", 5.5, "Y"), ("CHQ 1006", 6.5, "Y"),
    ("Cr-Mo 4140", 8.0, "Y"), ("Cr-Mo 4140", 10.0, "Y"),
]

base = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
so_rows = []
random.seed(42)
campaign_map = {}  # grade -> campaign letter

for i in range(1, 31):
    cust, region = random.choice(customers)
    grade, sec, vd = random.choice(grade_section_pool)
    grade_code = grade.replace(" ", "").replace("-", "")
    sec_code = str(sec).replace(".", "")
    sku = f"FG-WR-{grade_code}-{sec_code}"

    # Quantity: 20-120 MT (so 10-60 coils)
    qty_mt = random.choice([20, 40, 40, 60, 60, 80, 80, 100, 120])
    coil_count = int(qty_mt / 2.0)

    # Order placed 1-8 days ago
    order_date = base - timedelta(days=random.randint(1, 8))
    # Delivery in 5-16 days from now
    dd_days = random.randint(5, 16)
    delivery_date = base + timedelta(days=dd_days)

    priority = "URGENT" if dd_days <= 6 else ("HIGH" if dd_days <= 10 else "NORMAL")

    # Campaign group = grade (same grade → same campaign)
    camp_key = grade.replace(" ", "-")
    if camp_key not in campaign_map:
        campaign_map[camp_key] = f"CMP-{grade_code[:6]}"

    so_rows.append((
        f"SO-{i:03d}", cust, region, sku, grade, sec,
        qty_mt, coil_count,
        order_date.strftime("%Y-%m-%d"),
        delivery_date.strftime("%Y-%m-%d"),
        priority, campaign_map[camp_key], "", "Open"
    ))

stress_orders = [
    ("Usha Martin", "East", "SAE 1080", 5.5, 140, 2, 4, "URGENT", "MAN-1080-A"),
    ("JSW Steel", "South", "SAE 1080", 6.5, 120, 2, 5, "URGENT", "MAN-1080-A"),
    ("Mesco Steel", "North", "CHQ 1006", 6.5, 80, 3, 6, "HIGH", "MAN-CHQ-A"),
    ("Rathi Bars", "Central", "CHQ 1006", 5.5, 100, 3, 7, "HIGH", "MAN-CHQ-A"),
    ("Facor Steels", "Central", "Cr-Mo 4140", 8.0, 60, 2, 8, "HIGH", "SPECIAL-4140"),
    ("Kalyani Carpenter", "West", "Cr-Mo 4140", 10.0, 80, 2, 9, "HIGH", "SPECIAL-4140"),
    ("Bansal Wire", "West", "SAE 1018", 12.0, 160, 4, 12, "NORMAL", ""),
    ("Sunflag Iron", "West", "SAE 1018", 10.0, 140, 4, 11, "NORMAL", ""),
    ("Shyam Wire", "North", "SAE 1008", 5.5, 160, 1, 5, "URGENT", ""),
    ("Tata Wiron", "West", "SAE 1008", 6.5, 140, 1, 6, "URGENT", ""),
    ("Usha Martin", "East", "SAE 1045", 10.0, 120, 3, 10, "HIGH", ""),
    ("Mesco Steel", "North", "SAE 1045", 8.0, 100, 3, 9, "HIGH", ""),
    ("JSW Steel", "South", "SAE 1065", 5.5, 120, 4, 7, "HIGH", ""),
    ("Bansal Wire", "West", "SAE 1065", 6.5, 120, 4, 8, "HIGH", ""),
    ("Rathi Bars", "Central", "SAE 1035", 8.0, 160, 5, 13, "NORMAL", ""),
    ("Sunflag Iron", "West", "SAE 1035", 6.5, 120, 5, 14, "NORMAL", ""),
    ("Kalyani Carpenter", "West", "SAE 1018", 8.0, 160, 6, 15, "NORMAL", "BIG-1018"),
    ("Tata Wiron", "West", "SAE 1018", 8.0, 120, 6, 16, "NORMAL", "BIG-1018"),
    ("Facor Steels", "Central", "SAE 1008", 8.0, 100, 4, 9, "HIGH", ""),
    ("Usha Martin", "East", "SAE 1018", 5.5, 140, 2, 6, "URGENT", ""),
]

next_so_num = len(so_rows) + 1
for cust, region, grade, sec, qty_mt, order_days_ago, due_in_days, priority, manual_campaign_id in stress_orders:
    grade_code = grade.replace(" ", "").replace("-", "")
    sec_code = str(sec).replace(".", "")
    sku = f"FG-WR-{grade_code}-{sec_code}"
    order_date = base - timedelta(days=order_days_ago)
    delivery_date = base + timedelta(days=due_in_days)
    camp_key = grade.replace(" ", "-")
    if camp_key not in campaign_map:
        campaign_map[camp_key] = f"CMP-{grade_code[:6]}"
    so_rows.append((
        f"SO-{next_so_num:03d}", cust, region, sku, grade, sec,
        qty_mt, int(qty_mt / 2.0),
        order_date.strftime("%Y-%m-%d"),
        delivery_date.strftime("%Y-%m-%d"),
        priority, campaign_map[camp_key], manual_campaign_id, "Open"
    ))
    next_so_num += 1

open_program_profiles = [
    ("Usha Martin", "East", "SAE 1008", 5.5, 140),
    ("JSW Steel", "South", "SAE 1008", 6.5, 160),
    ("Bansal Wire", "West", "SAE 1018", 5.5, 160),
    ("Sunflag Iron", "West", "SAE 1018", 8.0, 180),
    ("Mesco Steel", "North", "SAE 1035", 6.5, 150),
    ("Rathi Bars", "Central", "SAE 1035", 8.0, 170),
    ("Facor Steels", "Central", "SAE 1045", 8.0, 150),
    ("Kalyani Carpenter", "West", "SAE 1065", 6.5, 140),
    ("Shyam Wire", "North", "SAE 1080", 5.5, 120),
    ("Tata Wiron", "West", "CHQ 1006", 6.5, 120),
]

for wave in range(3):
    for cust, region, grade, sec, base_qty in open_program_profiles:
        grade_code = grade.replace(" ", "").replace("-", "")
        sec_code = str(sec).replace(".", "")
        sku = f"FG-WR-{grade_code}-{sec_code}"
        camp_key = grade.replace(" ", "-")
        if camp_key not in campaign_map:
            campaign_map[camp_key] = f"CMP-{grade_code[:6]}"
        qty_mt = max(60, base_qty + random.choice([-20, 0, 20, 40]))
        order_date = base - timedelta(days=random.randint(1, 6))
        due_in_days = 7 + wave * 7 + random.randint(0, 4)
        priority = "URGENT" if due_in_days <= 7 else ("HIGH" if due_in_days <= 14 else "NORMAL")
        so_rows.append((
            f"SO-{next_so_num:03d}", cust, region, sku, grade, sec,
            qty_mt, int(qty_mt / 2.0),
            order_date.strftime("%Y-%m-%d"),
            (base + timedelta(days=due_in_days)).strftime("%Y-%m-%d"),
            priority, campaign_map[camp_key], "", "Open"
        ))
        next_so_num += 1

long_horizon_contract_profiles = [
    ("Usha Martin", "East", "SAE 1008", 5.5, 380),
    ("JSW Steel", "South", "SAE 1008", 8.0, 420),
    ("Bansal Wire", "West", "SAE 1018", 5.5, 420),
    ("Sunflag Iron", "West", "SAE 1018", 8.0, 460),
    ("Tata Wiron", "West", "SAE 1018", 12.0, 520),
    ("Mesco Steel", "North", "SAE 1035", 6.5, 360),
    ("Rathi Bars", "Central", "SAE 1035", 8.0, 420),
    ("Facor Steels", "Central", "SAE 1045", 8.0, 380),
    ("Kalyani Carpenter", "West", "SAE 1045", 10.0, 420),
    ("Shyam Wire", "North", "SAE 1065", 5.5, 320),
    ("Usha Martin", "East", "SAE 1065", 6.5, 340),
    ("JSW Steel", "South", "SAE 1080", 5.5, 300),
    ("Bansal Wire", "West", "SAE 1080", 6.5, 320),
    ("Sunflag Iron", "West", "CHQ 1006", 5.5, 280),
    ("Mesco Steel", "North", "CHQ 1006", 6.5, 300),
    ("Facor Steels", "Central", "Cr-Mo 4140", 8.0, 240),
    ("Kalyani Carpenter", "West", "Cr-Mo 4140", 10.0, 260),
    ("Rathi Bars", "Central", "SAE 1008", 6.5, 360),
]

for week in range(12):
    for cust, region, grade, sec, base_qty in long_horizon_contract_profiles:
        grade_code = grade.replace(" ", "").replace("-", "")
        sec_code = str(sec).replace(".", "")
        sku = f"FG-WR-{grade_code}-{sec_code}"
        camp_key = grade.replace(" ", "-")
        if camp_key not in campaign_map:
            campaign_map[camp_key] = f"CMP-{grade_code[:6]}"
        qty_mt = max(120, base_qty + random.choice([-40, -20, 0, 20, 40, 60]))
        due_in_days = 21 + week * 7 + random.randint(0, 5)
        priority = "HIGH" if due_in_days <= 35 else "NORMAL"
        so_rows.append((
            f"SO-{next_so_num:03d}", cust, region, sku, grade, sec,
            qty_mt, int(qty_mt / 2.0),
            (base - timedelta(days=random.randint(1, 10))).strftime("%Y-%m-%d"),
            (base + timedelta(days=due_in_days)).strftime("%Y-%m-%d"),
            priority, campaign_map[camp_key], "", "Backlog"
        ))
        next_so_num += 1

monthly_block_orders = [
    ("OEM Wire Program", "West", "SAE 1018", 8.0, 1800),
    ("Auto Spring Contract", "North", "SAE 1080", 6.5, 1200),
    ("Fastener Quality Program", "Central", "CHQ 1006", 5.5, 1100),
    ("Alloy Engineering Contract", "South", "Cr-Mo 4140", 8.0, 900),
]

for month_idx, due_in_days in enumerate([30, 60, 90], start=1):
    for cust, region, grade, sec, qty_mt in monthly_block_orders:
        grade_code = grade.replace(" ", "").replace("-", "")
        sec_code = str(sec).replace(".", "")
        sku = f"FG-WR-{grade_code}-{sec_code}"
        camp_key = grade.replace(" ", "-")
        if camp_key not in campaign_map:
            campaign_map[camp_key] = f"CMP-{grade_code[:6]}"
        so_rows.append((
            f"SO-{next_so_num:03d}", cust, region, sku, grade, sec,
            qty_mt, int(qty_mt / 2.0),
            (base - timedelta(days=month_idx * 3)).strftime("%Y-%m-%d"),
            (base + timedelta(days=due_in_days)).strftime("%Y-%m-%d"),
            "NORMAL", campaign_map[camp_key], "", "Backlog"
        ))
        next_so_num += 1

write_table(ws, headers, so_rows, start_row=r,
            col_widths=[9, 22, 8, 26, 14, 12, 14, 12, 13, 13, 9, 16, 14, 8])

# Colour priority rows
priority_colors = {"URGENT": C_LIGHT_RED, "HIGH": C_LIGHT_AMBER, "NORMAL": "FFFFFF"}
for row_idx in range(r+1, r+1+len(so_rows)):
    pri = ws.cell(row_idx, 11).value
    color = priority_colors.get(pri, "FFFFFF")
    if color != "FFFFFF":
        for col in range(1, len(headers)+1):
            ws.cell(row_idx, col).fill = PatternFill("solid", fgColor=color)


# ════════════════════════════════════════════════════════════════════════════
# SHEET 7: Resource_Master
# ════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Resource_Master")
ws.sheet_properties.tabColor = "4472C4"
r = sheet_title(ws, "Resource Master", "All production units. Capacity = Avail_Hours × Max_Capacity_MT_Hr.")

headers = ["Resource_ID", "Resource_Name", "Plant", "Type",
           "Avail_Hours_Day", "Max_Capacity_MT_Hr", "Capacity_MT_Day",
           "Heat_Size_MT", "Efficiency_%", "Status",
           "Operation_Group", "Default_Cycle_Min", "Default_Setup_Min", "Operation_Color"]

res_rows = [
    # Blast Furnace
    ("BF-01",  "Blast Furnace 1",         "Blast Furnace", "Iron Making",    24, 10.0, 240,  "-",  92, "Active", "BF", 60, 15, "FFF2CC"),
    # SMS: EAF (2 units, 50T, 90 min tap-to-tap → 20 hrs × 60 / 90 heats × 50T)
    ("EAF-01", "Electric Arc Furnace 1",  "SMS",           "Melting",        20, 33.3, 667,  50,   90, "Active", "EAF", 90, 30, "DDEBF7"),
    ("EAF-02", "Electric Arc Furnace 2",  "SMS",           "Melting",        20, 33.3, 667,  50,   90, "Active", "EAF", 90, 30, "DDEBF7"),
    # LRF (3 units, 40 min per heat → 22 hrs × 60 / 40 = 33 heats/day × 50T)
    ("LRF-01", "Ladle Refining Furnace 1","SMS",           "Refining",       22, 75.0, 1650, 50,   95, "Active", "LRF", 40, 10, "FFCC99"),
    ("LRF-02", "Ladle Refining Furnace 2","SMS",           "Refining",       22, 75.0, 1650, 50,   95, "Active", "LRF", 40, 10, "FFCC99"),
    ("LRF-03", "Ladle Refining Furnace 3","SMS",           "Refining",       22, 75.0, 1650, 50,   95, "Active", "LRF", 40, 10, "FFCC99"),
    # VD (1 unit, 45 min per heat)
    ("VD-01",  "Vacuum Degasser 1",       "SMS",           "Degassing",      20, 66.7, 1333, 50,   93, "Active", "VD", 45, 15, "D9E1F2"),
    # CCM (2 units, 50 min per heat for 130mm; 60 min for 150mm)
    ("CCM-01", "Continuous Caster 1",     "SMS",           "Casting",        22, 60.0, 1320, 50,   93, "Active", "CCM", 60, 20, "E2EFDA"),
    ("CCM-02", "Continuous Caster 2",     "SMS",           "Casting",        22, 50.0, 1100, 50,   92, "Active", "CCM", 50, 20, "E2EFDA"),
    # Rolling Mills
    ("RM-01",  "Rolling Mill 1",          "Rolling Mill",  "Rolling",        20, 80.0, 1600, "-",  88, "Active", "RM", "", 40, "F8CBB5"),
    ("RM-02",  "Rolling Mill 2",          "Rolling Mill",  "Rolling",        20, 70.0, 1400, "-",  87, "Active", "RM", "", 40, "F8CBB5"),
]
write_table(ws, headers, res_rows, start_row=r,
            col_widths=[10, 28, 16, 14, 16, 20, 18, 14, 13, 9, 16, 18, 18, 16])

# Colour by plant
plant_colors = {"Blast Furnace": "FFF2CC", "SMS": "DDEBF7", "Rolling Mill": "E2EFDA"}
for row_idx in range(r+1, r+1+len(res_rows)):
    plant = ws.cell(row_idx, 3).value
    color = plant_colors.get(plant, "FFFFFF")
    for col in range(1, len(headers)+1):
        ws.cell(row_idx, col).fill = PatternFill("solid", fgColor=color)


# ════════════════════════════════════════════════════════════════════════════
# SHEET 8: Routing
# ════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Routing")
ws.sheet_properties.tabColor = "4472C4"
r = sheet_title(ws, "Routing", "Process route per grade. Seq 10=EAF, 20=LRF, 25=VD (if needed), 30=CCM, 40=RM.")

headers = ["SKU_ID", "Grade", "Needs_VD", "Op_Seq", "Operation",
           "Resource_Group", "Preferred_Resource", "Cycle_Time_Min_Heat",
           "Setup_Time_Min", "Min_Campaign_MT", "Max_Campaign_MT", "Note",
           "Sequence", "Is_Optional", "Optional_Condition", "Transfer_Time_Min"]

routing_rows = []
# Billet section → CCM time: 130mm = 50 min/heat, 150mm = 60 min/heat
billet_ccm_time = {
    "BIL-130-1008": 50, "BIL-130-1018": 50, "BIL-130-1035": 50,
    "BIL-150-1045": 60, "BIL-150-1065": 60, "BIL-150-1080": 60,
    "BIL-150-CHQ": 60,  "BIL-150-4140": 60,
}
# Wire rod section → RM cycle time per MT (min/MT)
sec_rm_time = {5.5: 0.6, 6.5: 0.5, 8.0: 0.4, 10.0: 0.35, 12.0: 0.30}
sec_setup   = {5.5: 45,  6.5: 40,  8.0: 35,  10.0: 30,   12.0: 25}

for grade, sec, vd in coils:
    grade_code = grade.replace(" ", "").replace("-", "")
    sec_code = str(sec).replace(".", "")
    sku = f"FG-WR-{grade_code}-{sec_code}"
    billet_sku = coil_to_billet[grade]
    ccm_t = billet_ccm_time[billet_sku]
    rm_ct = sec_rm_time.get(sec, 0.5)
    rm_su = sec_setup.get(sec, 40)

    # EAF → LRF → (VD) → CCM for billet; RM for coil
    # We route by Billet SKU for SMS, by Coil SKU for RM
    # RM route only
    routing_rows.append((sku, grade, vd, 40, "Rolling", "RM", "RM-01",
                         int(2.0 * rm_ct * 60), rm_su, 100, 400,
                         f"Campaign min 100 MT", 40, "N", "", 30))

# Billet routes (SMS: EAF → LRF → VD? → CCM)
for billet_sku, billet_name in billet_names.items():
    grade_part = billet_sku.split("-", 2)[2]
    grade = next((g for g, s, v in coils if coil_to_billet[g] == billet_sku), "-")
    vd = "Y" if billet_sku in ["BIL-150-1080", "BIL-150-CHQ", "BIL-150-4140"] else "N"
    ccm_t = billet_ccm_time[billet_sku]
    routing_rows.append((billet_sku, grade, vd, 10, "Melting",  "EAF", "EAF-01", 90, 30, 50,400,"Tap-to-tap 90 min", 10, "N", "", 0))
    routing_rows.append((billet_sku, grade, vd, 20, "Refining", "LRF", "LRF-01", 40, 10, 50,400,"LRF 40 min/heat", 20, "N", "", 5))
    if vd == "Y":
        routing_rows.append((billet_sku, grade, vd, 25, "Degassing","VD", "VD-01", 45, 15, 50,400,"VD for special grades", 25, "Y", "Needs_VD", 5))
    routing_rows.append((billet_sku, grade, vd, 30, "Casting",  "CCM", "CCM-01",ccm_t,20,50,400,f"CCM {ccm_t} min/heat", 30, "N", "", 5))

write_table(ws, headers, routing_rows, start_row=r,
            col_widths=[26, 14, 10, 8, 12, 14, 18, 20, 14, 18, 18, 28, 10, 12, 18, 18])


# ════════════════════════════════════════════════════════════════════════════
# SHEET 9: Campaign_Config
# ════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Campaign_Config")
ws.sheet_properties.tabColor = C_ORANGE_TAB
r = sheet_title(ws, "Campaign Configuration",
                "Rules for grouping SOs into campaigns. Grade sequencing on CCM/RM (low→high carbon).")

headers = ["Grade", "Carbon_Range", "Needs_VD", "Min_Campaign_Heats",
           "Max_Campaign_Heats", "Grade_Seq_Order", "CCM_Preferred",
           "RM_Changeover_Min", "Note"]
camp_rows = [
    ("SAE 1008",   "0.06-0.10%",  "N", 3, 14, 1, "CCM-01",  45, "High-volume mesh/wire grade. Preferred for long low-carbon caster sequences."),
    ("SAE 1018",   "0.15-0.20%",  "N", 3, 14, 2, "CCM-01",  50, "Backbone low-carbon grade. Large monthly contracts can be pooled aggressively."),
    ("SAE 1035",   "0.32-0.38%",  "N", 2, 10, 3, "CCM-01",  70, "Medium-carbon family. Keep after low-carbon campaigns on CCM/RM."),
    ("SAE 1045",   "0.43-0.50%",  "N", 2, 10, 4, "CCM-01",  80, "Medium-carbon structural grade. Use moderate campaign windows."),
    ("SAE 1065",   "0.60-0.70%",  "N", 2, 8,  5, "CCM-02",  95, "High-carbon spring-wire family. Prefer dedicated CCM-02 windows."),
    ("SAE 1080",   "0.75-0.88%",  "Y", 2, 8,  6, "CCM-02", 120, "High-carbon VD grade. Keep separate from low-carbon tails and plan shorter runs."),
    ("CHQ 1006",   "0.04-0.08%",  "Y", 2, 7,  7, "CCM-01", 135, "Cold-heading quality. Tight chemistry and VD; avoid mixing with standard commodity pools."),
    ("Cr-Mo 4140", "0.38-0.43%",  "Y", 2, 5,  8, "CCM-02", 165, "Alloy contract grade. FeCr-sensitive; keep smallest campaign window and longest RM cleanup."),
]
write_table(ws, headers, camp_rows, start_row=r,
            col_widths=[14, 15, 10, 20, 20, 17, 16, 20, 35])

ws.cell(r+len(camp_rows)+3, 1).value = "Grade Sequencing Rule:"
ws.cell(r+len(camp_rows)+3, 1).font = Font(bold=True, color=C_DARK_BLUE, name="Calibri")
ws.cell(r+len(camp_rows)+4, 1).value = "On CCM: sequence campaigns low→high carbon to avoid skull risk."
ws.cell(r+len(camp_rows)+5, 1).value = "On RM: sequence low→high carbon for guide wear minimization."
ws.cell(r+len(camp_rows)+6, 1).value = "VD grades always scheduled LAST in a CCM sequence before a relining stop."


# ════════════════════════════════════════════════════════════════════════════
# SHEET 10: Changeover_Matrix
# ════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Changeover_Matrix")
ws.sheet_properties.tabColor = C_ORANGE_TAB
r = sheet_title(ws, "Changeover Matrix (Rolling Mill)",
                "Time in minutes to switch from one grade+section to another on RM. Use by scheduler.")

grades = ["SAE 1008", "SAE 1018", "SAE 1035", "SAE 1045",
          "SAE 1065", "SAE 1080", "CHQ 1006", "Cr-Mo 4140"]

# Changeover matrix — diagonal = 0 (same grade);
# low carbon → next low carbon = 30-45 min; crossing carbon levels = 90-150 min
# Also: section change adds 15-30 min
matrix_base = [
    [0,   30,  60,  90,  120, 150, 60,  120],
    [30,   0,  45,  75,  105, 135, 60,  105],
    [60,  45,   0,  30,   75, 120, 90,   90],
    [90,  75,  30,   0,   45,  90, 90,   60],
    [120, 105,  75,  45,   0,  45, 120,  90],
    [150, 135, 120,  90,  45,   0, 150, 120],
    [60,  60,   90,  90, 120, 150,  0,  120],
    [120, 105,  90,  60,  90, 120, 120,   0],
]

ws.cell(r, 1).value = "From \\ To"
ws.cell(r, 1).font = Font(bold=True, color=C_WHITE, name="Calibri", size=10)
ws.cell(r, 1).fill = PatternFill("solid", fgColor=C_DARK_BLUE)
ws.cell(r, 1).alignment = Alignment(horizontal="center", vertical="center")
ws.column_dimensions["A"].width = 14

for c, g in enumerate(grades, 2):
    hdr_style(ws, r, c, g, bg=C_MID_BLUE)
    ws.column_dimensions[get_column_letter(c)].width = 14

for ri, grade in enumerate(grades):
    c_row = ws.cell(r+ri+1, 1, value=grade)
    c_row.font = Font(bold=True, color=C_DARK_BLUE, name="Calibri", size=10)
    c_row.fill = PatternFill("solid", fgColor=C_LIGHT_BLUE)
    c_row.border = thin_border()
    for ci, val in enumerate(matrix_base[ri]):
        c = ws.cell(r+ri+1, ci+2, value=val)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border()
        if val == 0:
            c.fill = PatternFill("solid", fgColor="D9E2F3")
        elif val <= 45:
            c.fill = PatternFill("solid", fgColor=C_LIGHT_GREEN)
        elif val <= 90:
            c.fill = PatternFill("solid", fgColor=C_LIGHT_AMBER)
        else:
            c.fill = PatternFill("solid", fgColor=C_LIGHT_RED)

ws.cell(r+len(grades)+3, 1).value = "Legend:"
ws.cell(r+len(grades)+3, 1).font = Font(bold=True)
ws.cell(r+len(grades)+4, 1).value = "Green = ≤45 min  |  Amber = 46-90 min  |  Red = >90 min"


# ════════════════════════════════════════════════════════════════════════════
# SHEET 11: Queue_Times
# ════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Queue_Times")
ws.sheet_properties.tabColor = "FF0000"
r = sheet_title(ws, "Queue Times", "Critical queue-time rules between consecutive operations. These limits protect steel temperature, transfer, and billet handling realism.")

headers = ["From_Operation", "To_Operation", "Min_Queue_Min", "Max_Queue_Min", "Enforcement", "Note"]
queue_rows = [
    ("EAF", "LRF", 0, 120, "Hard", "Liquid steel temperature loss limit before refining."),
    ("LRF", "VD", 0, 90, "Hard", "Steel temperature loss: VD must start promptly."),
    ("LRF", "CCM", 0, 90, "Hard", "Steel temperature loss: casting should begin quickly after LRF."),
    ("VD", "CCM", 0, 60, "Hard", "VD to CCM must be near-immediate."),
    ("CCM", "RM", 30, 480, "Soft", "Billet cooling plus transfer; 8-hour soft maximum."),
]
write_table(ws, headers, queue_rows, start_row=r, hdr_bg=C_RED, col_widths=[18, 18, 16, 16, 14, 52])
for row_idx in range(r + 1, r + 1 + len(queue_rows)):
    fill = C_LIGHT_RED if ws.cell(row_idx, 5).value == "Hard" else C_LIGHT_AMBER
    for col in range(1, len(headers) + 1):
        ws.cell(row_idx, col).fill = PatternFill("solid", fgColor=fill)


# ════════════════════════════════════════════════════════════════════════════
# SHEET 12: Scenarios
# ════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Scenarios")
ws.sheet_properties.tabColor = C_ORANGE_TAB
r = sheet_title(ws, "Scenario Parameters", "Adjust parameters and click 'Run Scenarios' to compare outcomes.")

ws.cell(r, 1).value = "Parameter"
ws.cell(r, 2).value = "Value"
ws.cell(r, 3).value = "Unit"
ws.cell(r, 4).value = "Description"
for c in range(1, 5):
    hdr_style(ws, r, c, ws.cell(r, c).value)

params = [
    ("Demand Spike (%)",        15.0,  "%",     "% increase in all order quantities for spike scenario"),
    ("Machine Down (Hrs)",      8.0,   "Hrs",   "Downtime hours for breakdown scenario"),
    ("Machine Down Resource",   "EAF-01", "",   "Which resource goes down in breakdown scenario"),
    ("Machine Down Start (Hr)", 0.0,   "Hr",    "Hour offset from plan start when downtime begins"),
    ("Solver Time Limit (sec)", 30.0,  "Sec",   "CP-SAT time limit for the finite scheduler"),
    ("Planning Horizon (Days)", 14,    "Days",  "Rolling planning window"),
    ("Yield Loss (%)",          0.0,   "%",     "Additional yield loss applied in campaign heat calculations"),
    ("Rush Order MT",           0.0,   "MT",    "Inject one urgent rush order using the earliest due product family"),
    ("Extra Shift Hours",       0.0,   "Hrs",   "Add overtime hours per day to all resources"),
    ("Min Campaign MT",         100.0, "MT",    "Minimum tonnage to form a rolling campaign"),
    ("Max Campaign MT",         500.0, "MT",    "Maximum campaign size on Rolling Mill"),
]
for ri, (param, val, unit, desc) in enumerate(params, r+1):
    data_cell(ws, ri, 1, param, bold=True, align="left")
    c_val = data_cell(ws, ri, 2, val)
    c_val.fill = PatternFill("solid", fgColor="FFF2CC")
    c_val.font = Font(bold=True, color=C_AMBER, name="Calibri", size=10)
    data_cell(ws, ri, 3, unit)
    data_cell(ws, ri, 4, desc, align="left")

for col, width in [(1, 28), (2, 12), (3, 8), (4, 50)]:
    ws.column_dimensions[get_column_letter(col)].width = width


# ════════════════════════════════════════════════════════════════════════════
# SHEET 13: CTP_Request
# ════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("CTP_Request")
ws.sheet_properties.tabColor = C_ORANGE_TAB
r = sheet_title(ws, "CTP Request", "Planner-entered capable-to-promise checks. Seeded sample requests are provided below and can be edited before running CTP.")

out_headers = ["Request_ID", "SKU_ID", "Qty_MT", "Requested_Date", "Notes"]
for c, h in enumerate(out_headers, 1):
    hdr_style(ws, r, c, h)
    ws.column_dimensions[get_column_letter(c)].width = 18 if h != "Notes" else 36
ctp_seed_rows = [
    ("CTP-001", "FG-WR-SAE1008-55", 20, (base + timedelta(days=2)).strftime("%Y-%m-%d"), "Small near-term stock or schedule check"),
    ("CTP-002", "FG-WR-SAE1018-80", 150, (base + timedelta(days=7)).strftime("%Y-%m-%d"), "Likely merge into an existing low-carbon release"),
    ("CTP-003", "FG-WR-SAE1080-55", 90, (base + timedelta(days=6)).strftime("%Y-%m-%d"), "VD grade with a tight due date"),
    ("CTP-004", "FG-WR-CHQ1006-65", 80, (base + timedelta(days=8)).strftime("%Y-%m-%d"), "Special quality product"),
    ("CTP-005", "FG-WR-CrMo4140-80", 60, (base + timedelta(days=10)).strftime("%Y-%m-%d"), "Alloy grade likely constrained by alloy inventory"),
]
for row_idx, row_vals in enumerate(ctp_seed_rows, start=r + 1):
    for col_idx, val in enumerate(row_vals, start=1):
        data_cell(ws, row_idx, col_idx, val, align="left" if col_idx in {1, 2, 5} else "center")
ws.freeze_panes = ws.cell(row=r+1, column=1)


# ════════════════════════════════════════════════════════════════════════════
# SHEET 14: BOM_Output (output)
# ════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("BOM_Output")
ws.sheet_properties.tabColor = C_GREEN_TAB
r = sheet_title(ws, "BOM Explosion Output",
                "Output: End-to-end material requirements grouped by plant, process stage, and material type. Updated by 'Run BOM Explosion'.")

out_headers = ["Plant", "Stage", "Material_Type", "Material_Category", "Parent_SKUs",
               "SKU_ID", "SKU_Name", "BOM_Level", "Gross_Req", "Available_Before",
               "Covered_By_Stock", "Produced_Qty", "Net_Req", "Status"]
for c, h in enumerate(out_headers, 1):
    hdr_style(ws, r, c, h)
    ws.column_dimensions[get_column_letter(c)].width = 18
ws.cell(r+1, 1).value = "Run 'Run BOM Explosion' to populate"
ws.cell(r+1, 1).font = Font(italic=True, color="AAAAAA", name="Calibri")
ws.freeze_panes = ws.cell(row=r+1, column=1)


# ════════════════════════════════════════════════════════════════════════════
# SHEET 15: Capacity_Map (output)
# ════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Capacity_Map")
ws.sheet_properties.tabColor = C_GREEN_TAB
r = sheet_title(ws, "Capacity Map", "Output: rough-cut heuristic demand vs available hours per machine. Not a finite-schedule mirror. Updated by 'Run Capacity Map'.")

out_headers = ["Resource_ID", "Resource_Name", "Plant", "Avail_Hrs_14d",
               "Demand_Hrs", "Idle_Hrs", "Overload_Hrs", "Utilisation_%", "Status", "Capacity_Basis"]
for c, h in enumerate(out_headers, 1):
    hdr_style(ws, r, c, h)
    ws.column_dimensions[get_column_letter(c)].width = 18

ws.cell(r+1, 1).value = "Run 'Run Capacity Map' to populate"
ws.cell(r+1, 1).font = Font(italic=True, color="AAAAAA", name="Calibri")
ws.freeze_panes = ws.cell(row=r+1, column=1)


# ════════════════════════════════════════════════════════════════════════════
# SHEET 16: Schedule_Output (output)
# ════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Schedule_Output")
ws.sheet_properties.tabColor = C_GREEN_TAB
r = sheet_title(ws, "Schedule Output",
                "Output: Master planner view across SMS and RM, with campaign sections, heat grouping, and RM dispatch rows.")

out_headers = ["Job_ID", "Campaign", "SO_ID", "Grade", "Section_mm", "SKU_ID",
               "Operation", "Resource_ID", "Planned_Start", "Planned_End",
               "Duration_Hrs", "Heat_No", "Qty_MT", "Status"]
for c, h in enumerate(out_headers, 1):
    hdr_style(ws, r, c, h)
    width_map = {
        "Job_ID": 36, "Campaign": 12, "SO_ID": 34, "Grade": 14, "Section_mm": 18,
        "SKU_ID": 24, "Operation": 12, "Resource_ID": 12, "Planned_Start": 18,
        "Planned_End": 18, "Duration_Hrs": 11, "Heat_No": 10, "Qty_MT": 10, "Status": 18,
    }
    ws.column_dimensions[get_column_letter(c)].width = width_map.get(h, 14)

ws.cell(r+1, 1).value = "Run 'Run Schedule' to populate"
ws.cell(r+1, 1).font = Font(italic=True, color="AAAAAA", name="Calibri")
ws.freeze_panes = ws["A4"]


# ════════════════════════════════════════════════════════════════════════════
# SHEET 17: Campaign_Schedule (output — campaign-level view)
# ════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Campaign_Schedule")
ws.sheet_properties.tabColor = C_GREEN_TAB
r = sheet_title(ws, "Campaign Schedule",
                "Output: One row per released or held campaign, including SO pool, release status, timing, and due-date margin.")

out_headers = ["Campaign_ID", "Campaign_Group", "Grade", "Section_mm", "Sections_Covered",
               "Total_MT", "Heats", "Heats_Calc_Method", "Heats_Calc_Warnings", "Order_Count", "Priority", "Release_Status",
               "Material_Issue", "EAF_Start", "CCM_Start", "RM_Start", "RM_End",
               "Duration_Hrs", "Due_Date", "Margin_Hrs", "Status", "SOs_Covered"]

# -- Differentiated column widths --
camp_col_widths = {
    "Campaign_ID": 16, "Campaign_Group": 16, "Grade": 14, "Section_mm": 12,
    "Sections_Covered": 20, "Total_MT": 11, "Heats": 8, "Heats_Calc_Method": 18,
    "Heats_Calc_Warnings": 34, "Order_Count": 11,
    "Priority": 10, "Release_Status": 15, "Material_Issue": 24,
    "EAF_Start": 18, "CCM_Start": 18, "RM_Start": 18, "RM_End": 18,
    "Duration_Hrs": 12, "Due_Date": 16, "Margin_Hrs": 12, "Status": 13,
    "SOs_Covered": 34,
}
# -- Headers that benefit from wrap-text --
wrap_headers = {"Sections_Covered", "Material_Issue", "Release_Status", "Heats_Calc_Warnings",
                "Duration_Hrs", "Margin_Hrs", "Order_Count"}

for c, h in enumerate(out_headers, 1):
    hdr_style(ws, r, c, h, wrap=(h in wrap_headers))
    ws.column_dimensions[get_column_letter(c)].width = camp_col_widths.get(h, 14)

# Taller header row for wrapped text
ws.row_dimensions[r].height = 28

# Prompt row
ws.cell(r+1, 1).value = "Run 'Run Schedule' to populate"
ws.cell(r+1, 1).font = Font(italic=True, color="AAAAAA", name="Calibri")

# Color legend note
legend_row = r + 2
ws.cell(legend_row, 1).value = "Color Key:"
ws.cell(legend_row, 1).font = Font(bold=True, size=9, color=C_DARK_BLUE, name="Calibri")
legend_items = [
    (2, "On Time", C_LIGHT_GREEN),
    (3, "Tight (<24h)", C_LIGHT_AMBER),
    (4, "Late", C_LIGHT_RED),
    (5, "Held", C_LIGHT_AMBER),
]
for col_off, label, color in legend_items:
    c = ws.cell(legend_row, col_off, value=label)
    c.font = Font(size=8, color="4F4F4F", name="Calibri")
    c.fill = PatternFill("solid", fgColor=color)
    c.alignment = Alignment(horizontal="center", vertical="center")

# Freeze panes: keep title/header visible + first 2 ID columns pinned
ws.freeze_panes = ws.cell(row=r+1, column=3)

# AutoFilter on the header row
ws.auto_filter.ref = f"A{r}:{get_column_letter(len(out_headers))}{r}"


# ════════════════════════════════════════════════════════════════════════════
# SHEET 18: Material_Plan (output)
# ════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Material_Plan")
ws.sheet_properties.tabColor = C_GREEN_TAB
r = sheet_title(ws, "Material Plan",
                "Output: Per-campaign material consumption and shortage trace. Updated by 'Run Schedule'.")

out_headers = ["Campaign_ID", "Plant", "Material_Type", "Material_SKU", "Material_Name",
               "Required_Qty", "Available_Before", "Consumed", "Remaining_After", "Status"]
for c, h in enumerate(out_headers, 1):
    hdr_style(ws, r, c, h)
    ws.column_dimensions[get_column_letter(c)].width = 18

ws.cell(r+1, 1).value = "Run 'Run Schedule' to populate campaign material usage"
ws.cell(r+1, 1).font = Font(italic=True, color="AAAAAA", name="Calibri")
ws.freeze_panes = ws.cell(row=r+1, column=1)


# ════════════════════════════════════════════════════════════════════════════
# SHEET 19: Equipment_Schedule (output)
# ════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Equipment_Schedule")
ws.sheet_properties.tabColor = C_GREEN_TAB
r = sheet_title(ws, "Equipment Schedule",
                "Output: Separate dispatch tables by plant and equipment. Updated by 'Run Schedule'.")

out_headers = ["Job_ID", "Campaign", "SO_ID", "Grade", "Section_mm", "SKU_ID", "Operation",
               "Planned_Start", "Planned_End", "Duration_Hrs", "Qty_MT", "Status"]
for c, h in enumerate(out_headers, 1):
    hdr_style(ws, r, c, h)
    ws.column_dimensions[get_column_letter(c)].width = 18

ws.cell(r+1, 1).value = "Run 'Run Schedule' to populate grouped equipment tables"
ws.cell(r+1, 1).font = Font(italic=True, color="AAAAAA", name="Calibri")
ws.freeze_panes = ws.cell(row=r+1, column=1)


# ════════════════════════════════════════════════════════════════════════════
# SHEET 20: Schedule_Gantt (output)
# ════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Schedule_Gantt")
ws.sheet_properties.tabColor = C_GREEN_TAB
r = sheet_title(ws, "Schedule Gantt",
                "Output: Resource swim-lane Gantt by campaign, with adaptive time buckets and plant-separated lanes.")

out_headers = ["Plant", "Resource_ID", "Utilisation_%", "Campaigns", "Ops"]
for c, h in enumerate(out_headers, 1):
    hdr_style(ws, r+2, c, h)
    ws.column_dimensions[get_column_letter(c)].width = 14 if h != "Resource_ID" else 16

ws.cell(r, 1).value = "Each row is one resource lane; colored bars show campaign occupancy windows across the horizon."
ws.cell(r, 1).font = Font(size=10, color="808080", italic=True, name="Calibri")
ws.freeze_panes = ws["F6"]


# ════════════════════════════════════════════════════════════════════════════
# SHEET 21: Scenario_Output (output)
# ════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Scenario_Output")
ws.sheet_properties.tabColor = C_GREEN_TAB
r = sheet_title(ws, "Scenario Output",
                "Output: Comparison of baseline and what-if scenarios. Updated by 'Run Scenarios'.")

out_headers = ["Scenario", "Heats", "Campaigns", "Released", "Held",
               "On_Time_%", "Weighted_Lateness_Hrs", "Bottleneck", "Throughput_MT_Day", "Avg_Margin_Hrs",
               "Solver", "Overloaded",
               "BF-01", "EAF-01", "EAF-02", "LRF-01", "LRF-02",
               "LRF-03", "VD-01", "CCM-01", "CCM-02", "RM-01", "RM-02"]
for c, h in enumerate(out_headers, 1):
    hdr_style(ws, r, c, h)
    ws.column_dimensions[get_column_letter(c)].width = 16

ws.cell(r+1, 1).value = "Run 'Run Scenarios' to populate"
ws.cell(r+1, 1).font = Font(italic=True, color="AAAAAA", name="Calibri")
ws.freeze_panes = ws.cell(row=r+1, column=1)


# ════════════════════════════════════════════════════════════════════════════
# SHEET 22: CTP_Output (output)
# ════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("CTP_Output")
ws.sheet_properties.tabColor = C_GREEN_TAB
r = sheet_title(ws, "CTP Output", "Output: capable-to-promise response per request, including delivery feasibility when modeled, plus blockers and completion surrogate.")

out_headers = ["Request_ID", "SKU_ID", "Qty_MT", "Requested_Date", "Earliest_Completion",
               "Plant_Completion_Feasible", "Earliest_Delivery", "Delivery_Feasible", "Lateness_Days",
               "Inventory_Lineage", "Material_Gaps", "Campaign_Action", "Merged_Campaigns", "New_Campaigns", "Solver_Status"]
for c, h in enumerate(out_headers, 1):
    hdr_style(ws, r, c, h)
    ws.column_dimensions[get_column_letter(c)].width = 18

ws.cell(r+1, 1).value = "Future Run CTP output will appear here"
ws.cell(r+1, 1).font = Font(italic=True, color="AAAAAA", name="Calibri")
ws.freeze_panes = ws.cell(row=r+1, column=1)


# ════════════════════════════════════════════════════════════════════════════
# SHEET 23: Theo_vs_Actual (output)
# ════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Theo_vs_Actual")
ws.sheet_properties.tabColor = C_GREEN_TAB
r = sheet_title(ws, "Theoretical vs Actual", "Feed actual times here. Deviations >10% flagged automatically.")

out_headers = ["Job_ID", "Resource_ID", "Operation", "Grade",
               "Planned_Start", "Planned_End", "Planned_Hrs",
               "Actual_Start", "Actual_End", "Actual_Hrs",
               "Deviation_Hrs", "Deviation_%", "Flag"]
for c, h in enumerate(out_headers, 1):
    hdr_style(ws, r, c, h)
    ws.column_dimensions[get_column_letter(c)].width = 17
ws.freeze_panes = ws.cell(row=r+1, column=1)


# ════════════════════════════════════════════════════════════════════════════
# SHEET 24: KPI_Dashboard (output)
# ════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("KPI_Dashboard")
ws.sheet_properties.tabColor = C_GREEN_TAB
ws.cell(1, 1).value = "KPI Dashboard"
ws.cell(1, 1).font = Font(bold=True, size=16, color=C_DARK_BLUE, name="Calibri")
ws.cell(2, 1).value = "Last run: refreshes automatically after Capacity, Schedule, or Scenario runs"
ws.cell(2, 1).font = Font(size=10, color="808080", italic=True, name="Calibri")
ws.cell(4, 1).value = "Charts and KPI tiles are generated by the Python runtime, including throughput and scenario comparisons."
ws.cell(4, 1).font = Font(size=10, color="808080", italic=True, name="Calibri")

kpi_headers = ["KPI", "Current", "Target", "Unit", "Status"]
kpis = [
    ("EAF Utilisation %",          "", "≥85",  "%",   ""),
    ("CCM Utilisation %",          "", "≥88",  "%",   ""),
    ("RM Utilisation %",           "", "≥80",  "%",   ""),
    ("Total Heats Scheduled",      "", "-",    "Heats",""),
    ("Campaigns Scheduled",        "", "-",    "No",  ""),
    ("On-Time Delivery %",         "", "≥95",  "%",   ""),
    ("Late Jobs",                  "", "0",    "No",  ""),
    ("EAF-01 Overload Hrs",        "", "0",    "Hrs", ""),
    ("EAF-02 Overload Hrs",        "", "0",    "Hrs", ""),
    ("Total Changeover Time (RM)", "", "-",    "Min", ""),
]
for c, h in enumerate(kpi_headers, 1):
    hdr_style(ws, 4, c, h)
    ws.column_dimensions[get_column_letter(c)].width = 28

for ri, (kpi, curr, tgt, unit, status) in enumerate(kpis, 5):
    data_cell(ws, ri, 1, kpi, align="left")
    data_cell(ws, ri, 2, curr)
    data_cell(ws, ri, 3, tgt)
    data_cell(ws, ri, 4, unit)
    data_cell(ws, ri, 5, status)


# ════════════════════════════════════════════════════════════════════════════
# Help Sheet Content
# ════════════════════════════════════════════════════════════════════════════
help_sections = {
    "Config": (5, [
        "Role: Workbook-level APS control sheet for defaults that should not live as Python constants.",
        "Use it to define batch size, campaign grouping fields, horizon, solver limit, and queue enforcement defaults.",
        "Scenarios can still override selected runtime parameters, but Config is the structural baseline.",
        "This sheet is Phase 1 schema groundwork; later reader and engine phases will consume it directly."
    ]),
    "Control_Panel": (10, [
        "Role: Central navigation sheet for planners and supervisors.",
        "Use this sheet to launch full APS actions without hunting for individual tabs.",
        "Read the plant summary first, then use the buttons to refresh BOM, capacity, schedule, and scenario outputs.",
        "Keep this sheet as the operational landing page when briefing or demoing the plan.",
        "Local actions also exist on the relevant input/output sheets for faster workflow."
    ]),
    "SKU_Master": (15, [
        "Role: Product and material master for finished goods, stage outputs, byproducts, and raw materials.",
        "The sheet now includes BF hot metal, EAF/LRF/VD outputs, CCM billets, RM outputs, slag, crop, and end-cut items.",
        "New schema columns Route_Variant, Product_Family, and Attribute_1 prepare the APS for data-driven campaign grouping.",
        "Update this sheet when a new grade, size, process intermediate, or tracked waste stream is introduced to the plant.",
        "Downstream impact: drives stagewise BOM explosion, material traceability, and material-plan labels."
    ]),
    "BOM": (12, [
        "Role: Stagewise input master for recipe, yield, and byproduct relationships.",
        "Each row says how a parent SKU consumes an input child SKU or generates a byproduct child SKU; Flow_Type separates the two.",
        "Yield_Pct now sits beside Scrap_% so the workbook can carry explicit yield assumptions instead of relying on code constants.",
        "Edit this sheet when BF burden, EAF charge mix, metallurgical losses, CCM crop, RM end-cut assumptions, or yield assumptions change.",
        "Results no longer write here; run BOM Explosion and review the separate BOM_Output sheet."
    ]),
    "Inventory": (10, [
        "Role: On-hand stock by SKU and location, including FG, billets, hot metal, and optional WIP or waste sinks.",
        "Available_Qty is what the APS engine can net against before asking the plant to make more material at an upstream stage.",
        "Update this sheet whenever warehouse balances, billet stocks, hot-metal availability, or raw material receipts change materially.",
        "Downstream impact: directly reduces net requirements shown in BOM_Output and campaign material holds."
    ]),
    "Sales_Orders": (14, [
        "Role: Demand input sheet for customer orders waiting to be planned.",
        "Each row is an order line with due date, quantity, and product attributes that feed campaign building.",
        "Update this sheet for new orders, cancellations, priority changes, or due-date moves.",
        "Downstream impact: this is the main driver of campaigns, capacity usage, and dispatch orders."
    ]),
    "Resource_Master": (16, [
        "Role: Plant resource master covering BF, EAF, LRF, VD, CCM, and RM assets.",
        "Use it to maintain available hours, nominal capacity, efficiency, and active status per resource.",
        "New schema columns Operation_Group, Default_Cycle_Min, Default_Setup_Min, and Operation_Color prepare runtime aliasing, fallback timing, and color rendering.",
        "Update when an asset is added, derated, idled, or operating hours change.",
        "Downstream impact: capacity calculations and schedule feasibility depend on this sheet."
    ]),
    "Routing": (18, [
        "Role: Standard process path and timing assumptions by product family or billet.",
        "Each row defines one operation step, preferred resource group, cycle time, and setup time.",
        "New schema columns Sequence, Is_Optional, Optional_Condition, and Transfer_Time_Min prepare data-driven op order, optional-step logic, and transfer timing.",
        "Update this sheet when process routing or standard times change on any unit.",
        "Downstream impact: controls machine-hour loading and the duration of schedule orders."
    ]),
    "Campaign_Config": (12, [
        "Role: Campaign-building rules for grouping orders into rolling runs.",
        "This sheet captures min/max campaign size, metallurgical sequencing order, VD need, and changeover assumptions.",
        "Use it when planners want to tune how aggressively orders are clustered before scheduling.",
        "Downstream impact: changes the number of campaigns, heats, and sequencing pressure on CCM/RM."
    ]),
    "Changeover_Matrix": (12, [
        "Role: Rolling-mill changeover reference between grade families.",
        "The scheduler already uses this matrix partially for same-machine RM changeover gaps.",
        "Full steel sequencing is still simplified, so maintain this data as the reference baseline for future upgrades.",
        "Use it when reviewing whether RM campaign sequence and setup assumptions are realistic."
    ]),
    "Queue_Times": (8, [
        "Role: Critical hold-time constraint master between consecutive operations.",
        "These rows define min and max queue gaps, such as liquid-steel transfer windows and billet cooling/transfer windows.",
        "Hard rows are intended to make infeasible plans fail; Soft rows are intended for penalised-but-allowed overruns.",
        "This sheet is part of the Phase 1 schema rollout and will be consumed in a later scheduler phase."
    ]),
    "Scenarios": (8, [
        "Role: Active planning-control sheet plus what-if scenario driver.",
        "Adjust demand spike, machine-down timing, solver time limit, yield loss, rush order, extra shift, or campaign-size assumptions here before comparing alternatives or rerunning the main plan.",
        "Run Schedule, Capacity Map, or BOM after changes if you want these settings to affect the live plan.",
        "Scenario comparison results are written to Scenario_Output; this sheet remains the editable control surface."
    ]),
    "CTP_Request": (8, [
        "Role: Planner input sheet for capable-to-promise checks.",
        "Each row is a requested SKU, quantity, and requested delivery date that the future CTP action will evaluate.",
        "Use this when sales or PPC asks whether extra demand can fit around the committed plan.",
        "This shell is part of the Phase 1 schema rollout; the runtime action arrives in a later phase."
    ]),
    "BOM_Output": (15, [
        "Role: Total requirement sheet across the whole demand pool, grouped by responsible plant and material type.",
        "Each section shows plant, process stage, material type, parent linkage, gross requirement, stock cover, and remaining net need.",
        "Use BOM_Output to decide what the network needs in total; it is not campaign-sequenced.",
        "Material_Plan is different: it shows campaign-by-campaign commitment after earlier campaigns consume stock."
    ]),
    "Capacity_Map": (11, [
        "Role: Output sheet for rough-cut load-versus-capacity by machine.",
        "Demand_Hrs is a heuristic allocation from campaign demand, routing, and standard times; it is not the finite scheduler result.",
        "Use Idle_Hrs, Overload_Hrs, and Utilisation_% for quick screening, then confirm constraints and sequencing in Schedule_Output.",
        "Queue limits, frozen jobs, downtime interactions, and detailed sequencing can make this diverge from the finite schedule."
    ]),
    "Schedule_Output": (20, [
        "Role: Master planning table across the whole plant.",
        "Campaign header rows split the plan into release buckets, then heat rows group the SMS chain as Heat n of total.",
        "Heat_No follows actual SMS production order within the campaign, starting from the first EAF heat as Heat 1.",
        "Campaigns are strict PPC release batches, so campaign n+1 should begin only after campaign n completes end-to-end.",
        "Detail rows are one planned operation on one resource for one heat or one rolling order, not one whole sales order.",
        "SMS rows are campaign heats; SO_ID shows the pooled order list and Section_mm shows the covered downstream size mix.",
        "Use this sheet when you want one sortable and filterable master view of the end-to-end plan across EAF, LRF, VD, CCM, and RM.",
        "Equipment_Schedule shows the same plan as machine-specific packets; this sheet stays flat on purpose so planners can analyze the whole network."
    ]),
    "Campaign_Schedule": (22, [
        "Role: APS campaign-release summary for the end-to-end plan.",
        "Each row is one PPC-released batch of sales orders, with the campaign group, covered sections, order count, release status, and any material blocker.",
        "A later campaign should not start until the earlier campaign is fully complete, so this sheet also reflects release order.",
        "Duration_Hrs shows planned end-to-end span and Margin_Hrs shows due-date slack, with margin coloring for quick review.",
        "Use it to review whether a campaign was released or held, then check stage timing only for the released campaigns.",
        "This is the best management-summary view; use Schedule_Output for detailed dispatch and Equipment_Schedule for machine packets."
    ]),
    "Material_Plan": (12, [
        "Role: Campaign-by-campaign material allocation and shortage trace.",
        "Each campaign is grouped separately, then split by plant so planners can see where material pressure sits.",
        "Rows show whether a line was drawn from stock, needs make/convert, is partially covered, or is short, and a summary block sits on the right.",
        "Use this when a planner needs to understand why a campaign was released or held."
    ]),
    "Equipment_Schedule": (16, [
        "Role: Equipment-by-equipment dispatch packet.",
        "Each machine gets its own table so supervisors can review only the work relevant to that asset.",
        "Section and SKU are shown so mixed-size campaigns remain readable at the equipment level.",
        "Blank rows separate equipment sections and make printouts or screenshots cleaner for shift handover.",
        "Use this when the master Schedule_Output feels too dense for operational release."
    ]),
    "Schedule_Gantt": (16, [
        "Role: Resource swim-lane view of the same dispatch schedule.",
        "Each row is one equipment lane and the colored buckets show which campaign occupies it over time.",
        "Campaign overlap follows Config > Campaign_Serialization_Mode, so later campaigns may wait for full completion or only for SMS completion.",
        "Use this to spot bunching, idle gaps, machine handoff timing, and campaign continuity at a glance.",
        "The sheet refreshes automatically whenever you rerun the schedule."
    ]),
    "Scenario_Output": (25, [
        "Role: Output comparison sheet for baseline and stress-test scenarios.",
        "Each row is one scenario and the columns show released/held campaigns, on-time %, weighted lateness, bottleneck, throughput, and utilisation snapshots.",
        "On-time and utilisation columns are color-coded, and the best scenario row is highlighted automatically.",
        "The sheet always includes the 4 base scenarios and adds extra rows when yield loss, rush order, or overtime parameters are populated.",
        "Use it to answer what breaks first when demand rises, downtime starts mid-plan, yield falls, or a rush order is injected.",
        "This sheet is for comparison and decision support; the editable assumptions stay on the Scenarios input sheet."
    ]),
    "CTP_Output": (13, [
        "Role: Future capable-to-promise result sheet.",
        "Each row will echo a request and show plant completion surrogate, modeled delivery feasibility when available, plus material gaps, campaign join behavior, and solver status.",
        "If delivery is not modeled yet, Feasible should show UNMODELED rather than pretending plant completion equals customer delivery.",
        "Workbook CTP should block if committed inventory lineage is non-authoritative instead of burying that risk in metadata.",
        "This shell is part of the Phase 1 schema rollout; the runtime action arrives in a later phase."
    ]),
    "Theo_vs_Actual": (15, [
        "Role: Future closed-loop tracking sheet for execution feedback.",
        "Planned columns come from APS, while Actual columns are meant to be fed back from operations.",
        "Use it to quantify deviation, identify unstable resources, and improve planning standards over time.",
        "This sheet is currently a prepared template and not yet populated automatically."
    ]),
    "KPI_Dashboard": (8, [
        "Role: Executive summary of the latest APS run.",
        "Use this sheet to see top-line KPI tiles plus chart views of utilisation, campaign outcome, operation mix, throughput, scenario volumes, and scenario KPI comparison.",
        "It is intended for quick review after every run before the planner dives into detailed tabs.",
        "The Python runtime redraws this sheet automatically, so its numbers and charts always reflect the latest APS outputs."
    ]),
}

help_ws = wb.create_sheet("Help", 2)
build_help_sheet(help_ws, help_sections)

wb.active = wb.sheetnames.index("Control_Panel")


# ════════════════════════════════════════════════════════════════════════════
# Save
# ════════════════════════════════════════════════════════════════════════════
wb.save(OUTPUT)
print(f"Saved: {OUTPUT}")
print(f"Sheets: {wb.sheetnames}")
print(f"Sales Orders: {len(so_rows)}")
print(f"SKUs: {len(rows)}")
print(f"BOM rows: {len(bom_rows)}")
print(f"Routing rows: {len(routing_rows)}")
