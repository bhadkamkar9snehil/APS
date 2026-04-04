"""Create Algorithm_Config sheet in APS workbook with all 47 parameters."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

# All 47 parameters in order
PARAMETERS = [
    # SCHEDULER - Cycle Times (5)
    ("CYCLE_TIME_EAF_MIN", "SCHEDULER", "EAF Cycle Time (min)", 90, "Duration", 60, 180, "min", "Time for EAF heat melting", "HIGH"),
    ("CYCLE_TIME_LRF_MIN", "SCHEDULER", "LRF Cycle Time (min)", 40, "Duration", 20, 120, "min", "Time for LRF refining (heating)", "HIGH"),
    ("CYCLE_TIME_VD_MIN", "SCHEDULER", "VD Cycle Time (min)", 45, "Duration", 20, 120, "min", "Time for VD vacuum degassing", "HIGH"),
    ("CYCLE_TIME_CCM_130_MIN", "SCHEDULER", "CCM-130 Cycle Time (min)", 50, "Duration", 30, 120, "min", "Time to cast 130mm billets on CCM", "HIGH"),
    ("CYCLE_TIME_CCM_150_MIN", "SCHEDULER", "CCM-150 Cycle Time (min)", 60, "Duration", 40, 150, "min", "Time to cast 150mm billets on CCM", "HIGH"),

    # SCHEDULER - Objective Weights (6)
    ("OBJECTIVE_QUEUE_VIOLATION_WEIGHT", "SCHEDULER", "Queue Violation Penalty", 500, "Weight", 1, 1000, "points", "Penalty per minute over queue max", "HIGH"),
    ("PRIORITY_WEIGHT_URGENT", "SCHEDULER", "Lateness Weight: URGENT", 4, "Weight", 1, 10, "mult", "Multiplier on lateness for URGENT orders", "HIGH"),
    ("PRIORITY_WEIGHT_HIGH", "SCHEDULER", "Lateness Weight: HIGH", 3, "Weight", 1, 10, "mult", "Multiplier on lateness for HIGH orders", "MEDIUM"),
    ("PRIORITY_WEIGHT_NORMAL", "SCHEDULER", "Lateness Weight: NORMAL", 2, "Weight", 1, 10, "mult", "Multiplier on lateness for NORMAL orders", "MEDIUM"),
    ("PRIORITY_WEIGHT_LOW", "SCHEDULER", "Lateness Weight: LOW", 1, "Weight", 1, 10, "mult", "Multiplier on lateness for LOW orders", "LOW"),
    ("OBJECTIVE_SMS_LATENESS_RATIO", "SCHEDULER", "SMS Lateness vs RM Lateness", 0.5, "Ratio", 0, 1, "ratio", "SMS lateness as % of RM lateness weight", "MEDIUM"),

    # SCHEDULER - Solver Parameters (5)
    ("PLANNING_HORIZON_DAYS", "SCHEDULER", "Planning Horizon (days)", 14, "Count", 1, 90, "days", "Days to look ahead in schedule", "MEDIUM"),
    ("PLANNING_HORIZON_EXTENSION_DAYS", "SCHEDULER", "Horizon Extension (days)", 7, "Count", 0, 30, "days", "Extra days beyond horizon for overflow", "LOW"),
    ("SOLVER_TIME_LIMIT_SECONDS", "SCHEDULER", "CP-SAT Solver Timeout", 30, "Count", 1, 300, "sec", "Maximum time solver searches", "MEDIUM"),
    ("SOLVER_NUM_SEARCH_WORKERS", "SCHEDULER", "Solver Search Workers", 4, "Count", 1, 16, "count", "Parallel search threads", "LOW"),
    ("SETUP_TIME_FIRST_HEAT_ONLY", "SCHEDULER", "Setup on First Heat Only", True, "Boolean", None, None, "flag", "Include setup only on first heat", "MEDIUM"),

    # CAMPAIGN - Batch Sizing (3)
    ("HEAT_SIZE_MT", "CAMPAIGN", "Standard Heat Size", 50.0, "Quantity", 10, 200, "MT", "Standard SMS batch size", "HIGH"),
    ("CAMPAIGN_MIN_SIZE_MT", "CAMPAIGN", "Minimum Campaign Size", 100.0, "Quantity", 50, 1000, "MT", "Smallest campaign released to SMS", "MEDIUM"),
    ("CAMPAIGN_MAX_SIZE_MT", "CAMPAIGN", "Maximum Campaign Size", 500.0, "Quantity", 100, 5000, "MT", "Largest campaign before auto-split", "MEDIUM"),

    # CAMPAIGN - Yield Factors (8)
    ("YIELD_CCM_PCT", "CAMPAIGN", "CCM Casting Yield", 95, "Percentage", 80, 100, "%", "Casting process yield factor", "HIGH"),
    ("YIELD_RM_DEFAULT_PCT", "CAMPAIGN", "RM Rolling Yield (default)", 89, "Percentage", 70, 98, "%", "Default rolling yield", "HIGH"),
    ("YIELD_RM_5_5MM_PCT", "CAMPAIGN", "RM Yield for 5.5mm Section", 88, "Percentage", 75, 98, "%", "Rolling mill yield for 5.5mm wire", "MEDIUM"),
    ("YIELD_RM_6_5MM_PCT", "CAMPAIGN", "RM Yield for 6.5mm Section", 89, "Percentage", 75, 98, "%", "Rolling mill yield for 6.5mm wire", "MEDIUM"),
    ("YIELD_RM_8_0MM_PCT", "CAMPAIGN", "RM Yield for 8.0mm Section", 90, "Percentage", 75, 98, "%", "Rolling mill yield for 8.0mm wire", "MEDIUM"),
    ("YIELD_RM_10_0MM_PCT", "CAMPAIGN", "RM Yield for 10.0mm Section", 91, "Percentage", 75, 98, "%", "Rolling mill yield for 10.0mm wire", "MEDIUM"),
    ("YIELD_RM_12_0MM_PCT", "CAMPAIGN", "RM Yield for 12.0mm Section", 92, "Percentage", 75, 98, "%", "Rolling mill yield for 12.0mm wire", "MEDIUM"),
    ("YIELD_LOSS_DEFAULT_PCT", "CAMPAIGN", "Default Yield Loss", 0, "Percentage", 0, 10, "%", "Default % loss applied to BOM calcs", "LOW"),

    # CAMPAIGN - Material Rules (3)
    ("LOW_CARBON_BILLET_GRADES", "CAMPAIGN", "Low Carbon Grades (comma-sep)", "1008,1018,1035", "List", None, None, "", "Grades that use BIL-130", "HIGH"),
    ("VD_REQUIRED_GRADES", "CAMPAIGN", "VD Required Grades (comma-sep)", "1080,CHQ1006,CrMo4140", "List", None, None, "", "Grades requiring VD degassing", "HIGH"),
    ("BOM_MAX_DEPTH", "CAMPAIGN", "BOM Explosion Max Depth", 12, "Count", 5, 20, "levels", "Maximum nesting levels in BOM", "LOW"),

    # BOM - Rules (8)
    ("YIELD_MIN_BOUND_PCT", "BOM", "Minimum Yield Bound", 1, "Percentage", 0, 50, "%", "Floor on yield to prevent div-by-zero", "LOW"),
    ("YIELD_MAX_BOUND_PCT", "BOM", "Maximum Yield Bound", 100, "Percentage", 50, 100, "%", "Ceiling on yield (safety check)", "LOW"),
    ("YIELD_COLUMN_PREFERENCE", "BOM", "Yield Column Priority", "Yield_Pct,Scrap_%", "List", None, None, "order", "Which column to prefer: Yield_Pct or Scrap_%", "MEDIUM"),
    ("BYPRODUCT_INVENTORY_MODE", "BOM", "Byproduct Availability Mode", "deferred", "Choice", None, None, "mode", "When byproducts become available", "LOW"),
    ("INPUT_FLOW_TYPES", "BOM", "Input Flow Types (comma-sep)", ",INPUT,CONSUME,CONSUMED,REQUIRED", "List", None, None, "", "Which Flow_Type values count as inputs", "LOW"),
    ("BYPRODUCT_FLOW_TYPES", "BOM", "Byproduct Flow Types (comma-sep)", "BYPRODUCT,OUTPUT,CO_PRODUCT,COPRODUCT,WASTE", "List", None, None, "", "Which Flow_Type values are byproducts", "LOW"),
    ("ZERO_TOLERANCE_THRESHOLD", "BOM", "Quantity Zero Tolerance", 0.000001, "Threshold", 0, 0.01, "MT", "Below this, treat qty as zero", "LOW"),

    # CTP - Rules (6)
    ("CTP_SCORE_STOCK_ONLY", "CTP", "Score: Stock-Only Promise", 60, "Points", 0, 100, "pts", "Points for fulfilling from stock", "MEDIUM"),
    ("CTP_SCORE_MERGE_CAMPAIGN", "CTP", "Score: Merge Existing Campaign", 10, "Points", 0, 100, "pts", "Points for merging with existing campaign", "MEDIUM"),
    ("CTP_SCORE_NEW_CAMPAIGN", "CTP", "Score: New Campaign", 4, "Points", 0, 100, "pts", "Points for creating new campaign", "MEDIUM"),
    ("CTP_MERGEABLE_SCORE_THRESHOLD", "CTP", "Mergeable Score Threshold", 55, "Points", 0, 100, "pts", "Min score to consider merge viable", "MEDIUM"),
    ("CTP_INVENTORY_ZERO_TOLERANCE", "CTP", "Inventory Zero Tolerance", 0.000000001, "Threshold", 0, 0.01, "MT", "Below this, inventory = zero", "LOW"),
    ("CTP_MERGE_PENALTY", "CTP", "Merge Non-Selection Penalty", 1, "Cost", 0, 10, "cost", "Penalty if merge not selected", "LOW"),

    # CAPACITY - Rules (3)
    ("CAPACITY_HORIZON_DAYS", "CAPACITY", "Capacity Planning Horizon", 14, "Count", 1, 90, "days", "Days to analyze capacity utilization", "MEDIUM"),
    ("CAPACITY_SETUP_HOURS_DEFAULT", "CAPACITY", "Setup Hours Default", 0.0, "Duration", 0, 24, "hrs", "Initial setup hours before calculation", "LOW"),
    ("CAPACITY_CHANGEOVER_HOURS_DEFAULT", "CAPACITY", "Changeover Hours Default", 0.0, "Duration", 0, 24, "hrs", "Initial changeover hours before calculation", "LOW"),
]

def create_algorithm_config_sheet():
    """Create Algorithm_Config sheet with all 47 parameters."""

    wb = openpyxl.load_workbook("c:/Users/bhadk/Documents/APS/APS_BF_SMS_RM.xlsx")

    # Remove existing Algorithm_Config sheet if it exists
    if "Algorithm_Config" in wb.sheetnames:
        del wb["Algorithm_Config"]

    ws = wb.create_sheet("Algorithm_Config", 0)  # Insert at position 0

    # Header row
    headers = [
        "Config_Key", "Category", "Parameter_Name", "Current_Value", "Data_Type",
        "Min_Value", "Max_Value", "Unit", "Description", "Impact_Level",
        "Valid_Options", "Notes", "Last_Updated", "Updated_By", "Change_Reason"
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data rows
    for row_idx, param in enumerate(PARAMETERS, 2):
        config_key, category, param_name, current_val, data_type, min_val, max_val, unit, desc, impact = param

        ws.cell(row=row_idx, column=1, value=config_key)
        ws.cell(row=row_idx, column=2, value=category)
        ws.cell(row=row_idx, column=3, value=param_name)
        ws.cell(row=row_idx, column=4, value=current_val)
        ws.cell(row=row_idx, column=5, value=data_type)
        ws.cell(row=row_idx, column=6, value=min_val if min_val is not None else "")
        ws.cell(row=row_idx, column=7, value=max_val if max_val is not None else "")
        ws.cell(row=row_idx, column=8, value=unit)
        ws.cell(row=row_idx, column=9, value=desc)
        ws.cell(row=row_idx, column=10, value=impact)
        ws.cell(row=row_idx, column=11, value="")  # Valid_Options
        ws.cell(row=row_idx, column=12, value="")  # Notes
        ws.cell(row=row_idx, column=13, value=datetime.now().strftime("%Y-%m-%d"))
        ws.cell(row=row_idx, column=14, value="SYSTEM")
        ws.cell(row=row_idx, column=15, value="Initial configuration")

    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 10
    ws.column_dimensions['I'].width = 40
    ws.column_dimensions['J'].width = 10
    ws.column_dimensions['K'].width = 15
    ws.column_dimensions['L'].width = 20
    ws.column_dimensions['M'].width = 12
    ws.column_dimensions['N'].width = 12
    ws.column_dimensions['O'].width = 25

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save("c:/Users/bhadk/Documents/APS/APS_BF_SMS_RM.xlsx")
    print(f"[OK] Created Algorithm_Config sheet with {len(PARAMETERS)} parameters")

if __name__ == "__main__":
    create_algorithm_config_sheet()
