import shutil
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from data.loader import load_all
from engine.config import (
    load_workbook_config_snapshot,
    update_algorithm_config_in_workbook,
    upgrade_workbook_config,
)
from engine.ctp import _scenario_rank_key
from engine.masterdata_audit import audit_workbook_masterdata


def _copy_workbook(tmp_path: Path) -> Path:
    workbook_copy = tmp_path / "APS_config_runtime_test.xlsx"
    shutil.copy2(ROOT / "APS_BF_SMS_RM.xlsx", workbook_copy)
    return workbook_copy


def _copy_and_upgrade_workbook(tmp_path: Path) -> Path:
    workbook = _copy_workbook(tmp_path)
    upgrade_workbook_config(workbook)
    return workbook


def _sheet_value(ws, key_column: str, value_column: str, key: str, *, header_row: int) -> object:
    headers = {
        str(ws.cell(header_row, col).value).strip(): col
        for col in range(1, ws.max_column + 1)
        if ws.cell(header_row, col).value is not None
    }
    for row in range(header_row + 1, ws.max_row + 1):
        if str(ws.cell(row, headers[key_column]).value or "").strip() == key:
            return ws.cell(row, headers[value_column]).value
    raise KeyError(key)


def _set_algorithm_value(workbook_path: Path, key: str, value: object) -> None:
    wb = openpyxl.load_workbook(workbook_path)
    try:
        ws = wb["Algorithm_Config"]
        headers = {
            str(ws.cell(1, col).value).strip(): col
            for col in range(1, ws.max_column + 1)
            if ws.cell(1, col).value is not None
        }
        for row in range(2, ws.max_row + 1):
            if str(ws.cell(row, headers["Config_Key"]).value or "").strip() == key:
                ws.cell(row, headers["Current_Value"]).value = value
                wb.save(workbook_path)
                return
        raise KeyError(key)
    finally:
        wb.close()


def _set_config_value(workbook_path: Path, key: str, value: object) -> None:
    wb = openpyxl.load_workbook(workbook_path)
    try:
        ws = wb["Config"]
        headers = {
            str(ws.cell(3, col).value).strip(): col
            for col in range(1, ws.max_column + 1)
            if ws.cell(3, col).value is not None
        }
        for row in range(4, ws.max_row + 1):
            if str(ws.cell(row, headers["Key"]).value or "").strip() == key:
                ws.cell(row, headers["Value"]).value = value
                wb.save(workbook_path)
                return
        raise KeyError(key)
    finally:
        wb.close()


def test_workbook_upgrade_adds_new_config_rows_and_syncs_legacy_conflict(tmp_path):
    workbook = _copy_workbook(tmp_path)
    result = upgrade_workbook_config(workbook)

    assert "DEFAULT_HEAT_DURATION_HOURS" in result["inserted_keys"]
    assert "PRIORITY_SEQUENCE" in result["inserted_keys"]
    assert "CTP_DECISION_PRECEDENCE_SEQUENCE" in result["inserted_keys"]
    assert "Planning_Horizon_Days" in result["updated_system_config_keys"]

    snapshot = load_workbook_config_snapshot(workbook)

    assert snapshot.runtime_config["HEAT_SIZE_MT"] == 50.0
    assert snapshot.runtime_config["Default_Batch_Size_MT"] == 50.0
    assert snapshot.runtime_config["Planning_Horizon_Days"] == 14
    assert snapshot.runtime_config["Allow_Scheduler_Default_Masters"] is False
    assert snapshot.runtime_config["PRIORITY_SEQUENCE"] == ["URGENT", "HIGH", "NORMAL", "LOW"]
    assert not snapshot.conflicts


def test_runtime_config_reflects_algorithm_config_excel_edits(tmp_path):
    workbook = _copy_and_upgrade_workbook(tmp_path)
    _set_algorithm_value(workbook, "HEAT_SIZE_MT", 64)

    snapshot = load_workbook_config_snapshot(workbook)

    assert snapshot.runtime_config["HEAT_SIZE_MT"] == 64.0
    assert snapshot.runtime_config["Default_Batch_Size_MT"] == 64.0

    data = load_all(workbook)
    assert data["config"]["HEAT_SIZE_MT"] == 64.0
    assert data["config"]["Default_Batch_Size_MT"] == 64.0


def test_algorithm_config_update_persists_and_syncs_legacy_alias(tmp_path):
    workbook = _copy_and_upgrade_workbook(tmp_path)

    result = update_algorithm_config_in_workbook(
        workbook,
        "CAMPAIGN_MIN_SIZE_MT",
        180,
        user="pytest",
        reason="sync legacy alias",
    )

    assert result["key"] == "CAMPAIGN_MIN_SIZE_MT"
    assert result["new_value"] == 180.0

    wb = openpyxl.load_workbook(workbook, data_only=True)
    try:
        algorithm_value = _sheet_value(
            wb["Algorithm_Config"],
            "Config_Key",
            "Current_Value",
            "CAMPAIGN_MIN_SIZE_MT",
            header_row=1,
        )
        config_value = _sheet_value(
            wb["Config"],
            "Key",
            "Value",
            "Min_Campaign_MT",
            header_row=3,
        )
    finally:
        wb.close()

    assert algorithm_value == 180.0
    assert config_value == 180.0


def test_runtime_config_prefers_algorithm_sheet_for_exact_duplicate_keys(tmp_path):
    workbook = _copy_and_upgrade_workbook(tmp_path)
    _set_algorithm_value(workbook, "Allow_Scheduler_Default_Masters", True)
    _set_config_value(workbook, "Allow_Scheduler_Default_Masters", "N")

    snapshot = load_workbook_config_snapshot(workbook)

    assert snapshot.runtime_config["Allow_Scheduler_Default_Masters"] is True
    assert any(item["key"] == "Allow_Scheduler_Default_Masters" for item in snapshot.conflicts)


def test_algorithm_config_update_syncs_exact_duplicate_config_key(tmp_path):
    workbook = _copy_and_upgrade_workbook(tmp_path)

    result = update_algorithm_config_in_workbook(
        workbook,
        "Allow_Scheduler_Default_Masters",
        True,
        user="pytest",
        reason="sync direct duplicate",
    )

    assert result["key"] == "Allow_Scheduler_Default_Masters"
    assert result["new_value"] is True

    wb = openpyxl.load_workbook(workbook, data_only=True)
    try:
        algorithm_value = _sheet_value(
            wb["Algorithm_Config"],
            "Config_Key",
            "Current_Value",
            "Allow_Scheduler_Default_Masters",
            header_row=1,
        )
        config_value = _sheet_value(
            wb["Config"],
            "Key",
            "Value",
            "Allow_Scheduler_Default_Masters",
            header_row=3,
        )
    finally:
        wb.close()

    assert str(algorithm_value).strip().upper() == "TRUE"
    assert str(config_value).strip().upper() == "TRUE"


def test_masterdata_audit_reports_required_input_sheets_and_config_conflicts(tmp_path):
    workbook = _copy_workbook(tmp_path)
    report = audit_workbook_masterdata(workbook)

    for sheet_name in [
        "SKU_Master",
        "BOM",
        "Inventory",
        "Sales_Orders",
        "Resource_Master",
        "Routing",
        "Campaign_Config",
        "Changeover_Matrix",
        "Queue_Times",
        "Scenarios",
        "CTP_Request",
    ]:
        sheet = report["master_sheets"][sheet_name]
        assert sheet["exists"] is True
        assert not sheet["used_columns_missing"]
        assert sheet["deprecated"] is False

    assert report["master_sheets"]["Config"]["deprecated"] is True
    assert report["config_duplicates"]["primary_runtime_sheet"] == "Algorithm_Config"
    assert report["config_duplicates"]["compatibility_sheet"] == "Config"
    assert report["config_duplicates"]["conflict_count"] >= 1
    assert any(
        conflict["key"] == "Planning_Horizon_Days"
        for conflict in report["config_duplicates"]["conflicts"]
    )


def test_ctp_decision_precedence_sequence_is_configurable():
    stock_only = {
        "decision_class": "PROMISE_CONFIRMED_STOCK_ONLY",
        "promised_qty_mt": 100.0,
        "promised_completion_date": "2026-01-01T00:00:00",
        "promise_confidence": "HIGH",
        "merged_campaign_ids": [],
    }
    later_date = {
        "decision_class": "PROMISE_LATER_DATE",
        "promised_qty_mt": 100.0,
        "promised_completion_date": "2026-01-01T00:00:00",
        "promise_confidence": "HIGH",
        "merged_campaign_ids": [],
    }

    assert _scenario_rank_key(stock_only) < _scenario_rank_key(later_date)

    custom_config = {
        "CTP_DECISION_PRECEDENCE_SEQUENCE": [
            "PROMISE_LATER_DATE",
            "PROMISE_CONFIRMED_STOCK_ONLY",
        ]
    }
    assert _scenario_rank_key(later_date, config=custom_config) < _scenario_rank_key(stock_only, config=custom_config)
