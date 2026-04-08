"""
Master Data Fixer for APS Steel Plant

Implements phase 1 (cleanup) + phase 2 (routing) optimizations
for proper SO->Material->Production mapping.

Steel industry assumptions:
  - Heat size: 50 MT (EAF capacity)
  - Billet families: BIL-130 (low-C), BIL-150 (medium-high C)
  - SMS sequence: EAF -> LRF -> {VD} -> CCM
  - RM: CCM billets -> wire rod coils
"""

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
import shutil
import os
from datetime import datetime

EXCEL_PATH = r'c:\Users\bhadk\Documents\APS\APS_BF_SMS_RM.xlsx'

def backup_workbook():
    """Create timestamped backup."""
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    backup_path = f"{EXCEL_PATH}.{timestamp}.backup"
    if not os.path.exists(backup_path):
        shutil.copy(EXCEL_PATH, backup_path)
        print(f"[BACKUP] Created: {backup_path}")
    return backup_path

def phase1_cleanup(wb):
    """Phase 1: Remove orphaned rows, add Product_Type, set defaults."""
    print("\n" + "="*70)
    print("PHASE 1: DATA CLEANUP & NORMALIZATION")
    print("="*70)

    # --- FIX 1.1: Delete orphaned SO rows ---
    print("\n[FIX 1.1] Removing orphaned Sales Orders...")
    so_ws = wb['Sales_Orders']

    # Find and mark rows with null SO_ID and null SKU_ID
    rows_to_delete = []
    for row_idx, row in enumerate(so_ws.iter_rows(min_row=2, max_row=so_ws.max_row), start=2):
        so_id = row[0].value  # Column A: SO_ID
        sku_id = row[3].value  # Column D: SKU_ID

        if so_id is None and sku_id is None:
            rows_to_delete.append(row_idx)
            print(f"  [DEL] Row {row_idx} (orphaned)")

    # Delete in reverse order to maintain row numbers
    for row_idx in reversed(rows_to_delete):
        so_ws.delete_rows(row_idx, 1)
        print(f"      Deleted row {row_idx}")

    print(f"  [DONE] Removed {len(rows_to_delete)} orphaned rows")

    # --- FIX 1.2: Add Product_Type to SKU_Master ---
    print("\n[FIX 1.2] Adding Product_Type column to SKU_Master...")
    sku_ws = wb['SKU_Master']

    # Find or create Product_Type column (after Product_Family)
    header_row = 3  # Row 3 is the actual header
    col_idx = None

    # Find existing Product_Type or add it
    for cell in sku_ws[header_row]:
        if cell.value == 'Product_Type':
            col_idx = cell.column
            print(f"  -> Found Product_Type at column {col_idx}")
            break

    if col_idx is None:
        # Add new column after Product_Family
        col_idx = 13  # Column M (assuming Product_Family is L)
        sku_ws.cell(row=header_row, column=col_idx).value = 'Product_Type'
        print(f"  -> Created Product_Type column at {col_idx}")

    # Assign Product_Type based on SKU_ID prefix
    product_type_map = {
        'FG-': 'FINISHED_GOODS',
        'BIL-': 'PRODUCTION_INTERMEDIATE',
        'RM-OUT-': 'PRODUCTION_INTERMEDIATE',
        'EAF-OUT-': 'PROCESS_INTERMEDIATE',
        'LRF-OUT-': 'PROCESS_INTERMEDIATE',
        'VD-OUT-': 'PROCESS_INTERMEDIATE',
        'BF-': 'RAW_MATERIAL',
        'RM-': 'RAW_MATERIAL',  # RM-FECR (raw scrap)
    }

    assigned = 0
    for row_idx, row in enumerate(sku_ws.iter_rows(min_row=2, max_row=sku_ws.max_row), start=2):
        sku_id = row[0].value  # Column A
        current_type = sku_ws.cell(row=row_idx, column=col_idx).value

        if sku_id and (current_type is None or str(current_type).strip() == ''):
            # Determine product type
            product_type = 'UNKNOWN'
            for prefix, ptype in product_type_map.items():
                if str(sku_id).startswith(prefix):
                    product_type = ptype
                    break

            sku_ws.cell(row=row_idx, column=col_idx).value = product_type
            assigned += 1

    print(f"  [DONE] Assigned Product_Type to {assigned} SKUs")

    # --- FIX 1.3: Set default Lead_Time_Days ---
    print("\n[FIX 1.3] Setting Lead_Time_Days (steel industry defaults)...")
    lt_col = None
    for cell in sku_ws[header_row]:
        if cell.value == 'Lead_Time_Days':
            lt_col = cell.column
            break

    if lt_col is None:
        lt_col = 14  # Column N
        sku_ws.cell(row=header_row, column=lt_col).value = 'Lead_Time_Days'

    lt_defaults = {
        'RAW_MATERIAL': 7,  # External supply
        'PROCESS_INTERMEDIATE': 0,  # Internal, transient
        'PRODUCTION_INTERMEDIATE': 1,  # Internal, 1 day CCM->RM
        'FINISHED_GOODS': 0,  # RM bottleneck drives
    }

    lt_assigned = 0
    for row_idx, row in enumerate(sku_ws.iter_rows(min_row=2, max_row=sku_ws.max_row), start=2):
        sku_id = row[0].value
        current_lt = sku_ws.cell(row=row_idx, column=lt_col).value
        product_type = sku_ws.cell(row=row_idx, column=col_idx).value

        if sku_id and (current_lt is None or current_lt == 0):
            default_lt = lt_defaults.get(str(product_type or '').strip(), 0)
            sku_ws.cell(row=row_idx, column=lt_col).value = default_lt
            lt_assigned += 1

    print(f"  [DONE] Set Lead_Time_Days for {lt_assigned} SKUs")

    # --- FIX 1.4: Set default Safety_Stock_MT ---
    print("\n[FIX 1.4] Setting Safety_Stock_MT (steel industry defaults)...")
    ss_col = None
    for cell in sku_ws[header_row]:
        if cell.value == 'Safety_Stock_MT':
            ss_col = cell.column
            break

    if ss_col is None:
        ss_col = 15  # Column O
        sku_ws.cell(row=header_row, column=ss_col).value = 'Safety_Stock_MT'

    ss_defaults = {
        'RAW_MATERIAL': 100.0,  # 2 heats buffer
        'PROCESS_INTERMEDIATE': 0.0,  # Transient, no storage
        'PRODUCTION_INTERMEDIATE': 50.0,  # 1 heat buffer
        'FINISHED_GOODS': 30.0,  # 0.5-1 heat
    }

    ss_assigned = 0
    for row_idx, row in enumerate(sku_ws.iter_rows(min_row=2, max_row=sku_ws.max_row), start=2):
        sku_id = row[0].value
        current_ss = sku_ws.cell(row=row_idx, column=ss_col).value
        product_type = sku_ws.cell(row=row_idx, column=col_idx).value

        if sku_id and (current_ss is None or current_ss == 0):
            default_ss = ss_defaults.get(str(product_type or '').strip(), 0.0)
            sku_ws.cell(row=row_idx, column=ss_col).value = default_ss
            ss_assigned += 1

    print(f"  [DONE] Set Safety_Stock_MT for {ss_assigned} SKUs")

def phase2_routing(wb):
    """Phase 2: Add routing for missing intermediate SKUs."""
    print("\n" + "="*70)
    print("PHASE 2: ROUTING FOR INTERMEDIATE MATERIALS")
    print("="*70)

    routing_ws = wb['Routing']
    sku_master_ws = wb['SKU_Master']

    # Get existing routing SKUs
    existing_routes = set()
    for row in routing_ws.iter_rows(min_row=2, max_row=routing_ws.max_row):
        sku_id = row[0].value
        if sku_id:
            existing_routes.add(str(sku_id).strip())

    print(f"\nExisting routes: {len(existing_routes)} SKUs")

    # Define routing for intermediate outputs
    # Format: (SKU_ID, Operation, Primary_Resource, Duration_Min, Secondary_Resource, Next_Op)
    new_routes = []

    # EAF outputs (liquid steel) -> LRF
    eaf_outputs = [
        ('EAF-OUT-SAE1008', 'REFINING', 'LRF-01', 40, 'LRF-02;LRF-03', 'REFINING'),
        ('EAF-OUT-SAE1018', 'REFINING', 'LRF-01', 40, 'LRF-02;LRF-03', 'REFINING'),
        ('EAF-OUT-SAE1035', 'REFINING', 'LRF-01', 40, 'LRF-02;LRF-03', 'REFINING'),
        ('EAF-OUT-SAE1045', 'REFINING', 'LRF-01', 40, 'LRF-02;LRF-03', 'REFINING'),
        ('EAF-OUT-SAE1065', 'REFINING', 'LRF-01', 40, 'LRF-02;LRF-03', 'REFINING'),
        ('EAF-OUT-SAE1080', 'REFINING', 'LRF-01', 40, 'LRF-02;LRF-03', 'REFINING'),
        ('EAF-OUT-CHQ1006', 'REFINING', 'LRF-01', 40, 'LRF-02;LRF-03', 'REFINING'),
        ('EAF-OUT-CrMo4140', 'REFINING', 'LRF-01', 40, 'LRF-02;LRF-03', 'REFINING'),
    ]

    # LRF outputs: grades requiring VD -> VD, others -> CCM
    vd_required = {'1080', 'CHQ1006', 'CrMo4140'}
    lrf_outputs_vd = [
        ('LRF-OUT-SAE1080', 'DEGASSING', 'VD-01', 45, '', 'CASTING'),
        ('LRF-OUT-CHQ1006', 'DEGASSING', 'VD-01', 45, '', 'CASTING'),
        ('LRF-OUT-CrMo4140', 'DEGASSING', 'VD-01', 45, '', 'CASTING'),
    ]

    lrf_outputs_ccm = [
        ('LRF-OUT-SAE1008', 'CASTING', 'CCM-01', 50, 'CCM-02', 'CASTING'),
        ('LRF-OUT-SAE1018', 'CASTING', 'CCM-01', 50, 'CCM-02', 'CASTING'),
        ('LRF-OUT-SAE1035', 'CASTING', 'CCM-01', 50, 'CCM-02', 'CASTING'),
        ('LRF-OUT-SAE1045', 'CASTING', 'CCM-01', 50, 'CCM-02', 'CASTING'),
        ('LRF-OUT-SAE1065', 'CASTING', 'CCM-01', 50, 'CCM-02', 'CASTING'),
    ]

    # VD outputs -> CCM
    vd_outputs = [
        ('VD-OUT-SAE1080', 'CASTING', 'CCM-01', 50, 'CCM-02', 'CASTING'),
        ('VD-OUT-CHQ1006', 'CASTING', 'CCM-01', 50, 'CCM-02', 'CASTING'),
        ('VD-OUT-CrMo4140', 'CASTING', 'CCM-01', 50, 'CCM-02', 'CASTING'),
    ]

    all_new = eaf_outputs + lrf_outputs_vd + lrf_outputs_ccm + vd_outputs

    # Check which ones actually need to be added
    to_add = []
    for sku_id, operation, resource, duration, secondary, next_op in all_new:
        if sku_id not in existing_routes:
            to_add.append((sku_id, operation, resource, duration, secondary, next_op))

    if not to_add:
        print("\n  All intermediate routes already exist. [OK]")
        return

    # Add new routes
    print(f"\n[ACTION] Adding {len(to_add)} new routes for intermediate materials:")

    # Find the routing sheet header columns
    header_row = 3
    col_map = {}
    for cell in routing_ws[header_row]:
        col_map[cell.value] = cell.column

    print(f"  Routing columns: {col_map}")

    # Add rows
    next_row = routing_ws.max_row + 1
    for sku_id, operation, resource, duration, secondary, next_op in to_add:
        # SKU_ID (col A)
        routing_ws.cell(row=next_row, column=col_map.get('SKU_ID', 1)).value = sku_id
        # Operation (col B)
        routing_ws.cell(row=next_row, column=col_map.get('Operation', 2)).value = operation
        # Primary_Resource (col C - may be named 'Primary Resource')
        if 'Primary_Resource' in col_map:
            routing_ws.cell(row=next_row, column=col_map['Primary_Resource']).value = resource
        elif 'Primary Resource' in col_map:
            routing_ws.cell(row=next_row, column=col_map['Primary Resource']).value = resource

        # Duration if column exists
        if 'Duration_Min' in col_map:
            routing_ws.cell(row=next_row, column=col_map['Duration_Min']).value = duration

        print(f"  -> {sku_id:20s} : {operation:12s} on {resource:10s} ({duration} min)")
        next_row += 1

    print(f"\n  [DONE] Added {len(to_add)} routing records")

def phase3_bom_completeness(wb):
    """Phase 3: Verify BOM completeness for demanded SKUs."""
    print("\n" + "="*70)
    print("PHASE 3: BOM COMPLETENESS CHECK")
    print("="*70)

    so_ws = wb['Sales_Orders']
    bom_ws = wb['BOM']
    sku_ws = wb['SKU_Master']

    # Get demanded SKUs (from active SOs)
    demanded = set()
    for row in so_ws.iter_rows(min_row=2, max_row=so_ws.max_row):
        sku_id = row[3].value  # Column D: SKU_ID
        if sku_id:
            demanded.add(str(sku_id).strip())

    # Get BOM parents
    bom_parents = set()
    for row in bom_ws.iter_rows(min_row=2, max_row=bom_ws.max_row):
        parent = row[1].value  # Column B: Parent_SKU
        if parent:
            bom_parents.add(str(parent).strip())

    missing_bom = demanded - bom_parents

    print(f"\nDemanded SKUs: {len(demanded)}")
    print(f"SKUs with BOM: {len(bom_parents & demanded)}")

    if missing_bom:
        print(f"\nWARNING: {len(missing_bom)} demanded SKUs have no BOM:")
        for sku in sorted(missing_bom):
            print(f"  - {sku}")
        print("\n  These SKUs cannot be produced. Add BOM entries manually.")
    else:
        print("\n[OK] All demanded SKUs have BOM defined")

def save_workbook(wb):
    """Save workbook."""
    print("\n" + "="*70)
    print("SAVING WORKBOOK")
    print("="*70)
    wb.save(EXCEL_PATH)
    print(f"\n[SAVED] {EXCEL_PATH}")

def main():
    print("\n")
    print("=" * 70)
    print("MASTER DATA OPTIMIZATION - APS STEEL PLANT")
    print("Implementing data quality and structure improvements")
    print("=" * 70)

    # Backup
    backup_path = backup_workbook()

    # Load workbook
    print(f"\n[LOAD] Opening workbook: {EXCEL_PATH}")
    wb = load_workbook(EXCEL_PATH)

    # Run phases
    try:
        phase1_cleanup(wb)
        phase2_routing(wb)
        phase3_bom_completeness(wb)
        save_workbook(wb)

        print("\n" + "="*70)
        print("SUCCESS")
        print("="*70)
        print("\nOptimization complete. Backup saved: " + backup_path)
        print("\nNext steps:")
        print("  1. Run tests to verify data integrity")
        print("  2. Update BOM for any missing demanded SKUs")
        print("  3. Run planning workflow")

    except Exception as e:
        print(f"\n[ERROR] {e}")
        print(f"\nRestore from backup: {backup_path}")
        raise

if __name__ == '__main__':
    main()
