"""
APS Flask API Server
--------------------
Reads APS_BF_SMS_RM.xlsx directly (no xlwings/Excel COM needed).
Calls engine functions, returns JSON to the web UI.

Start:  python api_server.py
Serves: http://localhost:5000
"""
from __future__ import annotations

import json
import os
import sys
import traceback
from datetime import date, datetime
from pathlib import Path

import numpy as np
import pandas as pd
from flask import Flask, jsonify, request
from flask_cors import CORS

sys.path.insert(0, str(Path(__file__).parent))

from engine.bom_explosion import consolidate_demand, explode_bom_details, net_requirements
from engine.campaign import build_campaigns
from engine.capacity import capacity_map, compute_demand_hours
from engine.ctp import capable_to_promise
from engine.scheduler import schedule

# ── Config ──────────────────────────────────────────────────────────────────
WORKBOOK = Path(__file__).parent / "APS_BF_SMS_RM.xlsx"
HEADER_ROW = 2   # rows 0–1 = title/subtitle, row 2 = column headers (0-indexed)

app = Flask(__name__)
CORS(app, origins=["http://localhost:3131", "http://127.0.0.1:3131", "*"])


# ── JSON helpers ─────────────────────────────────────────────────────────────
class _Enc(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, (np.integer,)):      return int(obj)
        if isinstance(obj, (np.floating,)):     return None if np.isnan(obj) else float(obj)
        if isinstance(obj, np.ndarray):         return obj.tolist()
        if isinstance(obj, (datetime, date, pd.Timestamp)):
            try:    return obj.isoformat()
            except: return str(obj)
        try:
            if pd.isna(obj): return None
        except Exception:
            pass
        return super().default(obj)


def _safe(v):
    """Make a single value JSON-safe."""
    if v is None: return None
    if isinstance(v, float) and (np.isnan(v) or np.isinf(v)): return None
    if isinstance(v, (np.integer,)):  return int(v)
    if isinstance(v, (np.floating,)): return float(v)
    if isinstance(v, (datetime, date, pd.Timestamp)):
        try:    return v.isoformat()
        except: return str(v)
    try:
        if pd.isna(v): return None
    except Exception:
        pass
    return v


def _df_to_records(df: pd.DataFrame) -> list:
    if df is None or df.empty:
        return []
    records = []
    for _, row in df.iterrows():
        records.append({k: _safe(v) for k, v in row.items()})
    return records


def _jsonify(data):
    return app.response_class(
        json.dumps(data, cls=_Enc, ensure_ascii=False),
        mimetype="application/json",
    )


# ── Excel readers ────────────────────────────────────────────────────────────
def _read_sheet(sheet: str, required: list | None = None) -> pd.DataFrame:
    df = pd.read_excel(WORKBOOK, sheet_name=sheet, header=HEADER_ROW, dtype=str)
    df = df.dropna(how="all").reset_index(drop=True)
    if required:
        df = df.dropna(subset=[c for c in required if c in df.columns], how="all")
    return df


def _read_config() -> dict:
    try:
        df = _read_sheet("Config", ["Key"])
        return {str(r["Key"]).strip(): r.get("Value") for _, r in df.iterrows() if str(r.get("Key", "")).strip()}
    except Exception:
        return {}


def _read_queue_times() -> dict:
    try:
        df = _read_sheet("Queue_Times", ["From_Operation"])
        out = {}
        for _, r in df.iterrows():
            k = (str(r.get("From_Operation", "") or "").strip().upper(),
                 str(r.get("To_Operation", "") or "").strip().upper())
            try:
                out[k] = {
                    "min": float(r.get("Min_Queue_Min") or 0),
                    "max": float(r.get("Max_Queue_Min") or 120),
                    "enforcement": str(r.get("Enforcement") or "Hard").strip(),
                }
            except Exception:
                pass
        return out
    except Exception:
        return {}


def _load_all() -> dict:
    """Load all input sheets → engine-ready DataFrames."""
    config  = _read_config()
    so_raw  = _read_sheet("Sales_Orders",   ["SO_ID", "SKU_ID"])
    res     = _read_sheet("Resource_Master", ["Resource_ID"])
    skus    = _read_sheet("SKU_Master",      ["SKU_ID"])
    routing = _read_sheet("Routing",         ["SKU_ID", "Operation"])
    bom     = _read_sheet("BOM",             ["Parent_SKU", "Child_SKU"])
    inv     = _read_sheet("Inventory",       ["SKU_ID"])
    queue   = _read_queue_times()

    # ── numeric coercion ────────────────────────────────────────────────────
    def _num(df, col, default=0.0):
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(default)
        return df

    # Inventory
    if "Available_Qty" not in inv.columns and "Stock_Qty" in inv.columns:
        inv = inv.rename(columns={"Stock_Qty": "Available_Qty"})
    inv = _num(inv, "Available_Qty", 0.0)

    # Resources
    res = _num(res, "Avail_Hours_Day", 20.0)
    res = _num(res, "Max_Capacity_MT_Hr", 33.3)
    if "Plant" not in res.columns:
        res["Plant"] = "Plant"

    # BOM
    bom = _num(bom, "Qty_Per", 1.0)
    bom = _num(bom, "Yield_Pct", 100.0)
    bom = _num(bom, "Level", 1.0)

    # Sales Orders
    so_raw["Delivery_Date"] = pd.to_datetime(so_raw.get("Delivery_Date"), errors="coerce")
    so_raw["Order_Date"]    = pd.to_datetime(so_raw.get("Order_Date"),    errors="coerce")
    so_raw["Status"]        = so_raw.get("Status", "Open").fillna("Open")
    so_raw = _num(so_raw, "Order_Qty_MT", 0.0)
    so_raw = _num(so_raw, "Section_mm",  6.5)
    if "Order_Qty_MT" not in so_raw.columns and "Order_Qty" in so_raw.columns:
        so_raw["Order_Qty_MT"] = pd.to_numeric(so_raw["Order_Qty"], errors="coerce").fillna(0)
    if "Order_Qty" not in so_raw.columns:
        so_raw["Order_Qty"] = so_raw.get("Order_Qty_MT", 0)

    open_mask = so_raw["Status"].astype(str).str.strip().str.upper().isin({"OPEN", "CONFIRMED", "PLANNED", ""})
    open_so   = so_raw[open_mask].copy()
    if open_so.empty:
        open_so = so_raw.copy()

    # Routing numeric
    routing = _num(routing, "Cycle_Time_Min_Heat", 60.0)
    routing = _num(routing, "Setup_Time_Min", 0.0)
    routing = _num(routing, "Transfer_Time_Min", 0.0)
    routing = _num(routing, "Op_Seq", 10.0)
    if "Sequence" not in routing.columns and "Op_Seq" in routing.columns:
        routing["Sequence"] = routing["Op_Seq"]

    return {
        "config":        config,
        "sales_orders":  open_so,
        "all_orders":    so_raw,
        "resources":     res,
        "skus":          skus,
        "routing":       routing,
        "bom":           bom,
        "inventory":     inv,
        "queue_times":   queue,
    }


# ── Shared run-state ─────────────────────────────────────────────────────────
_state: dict = {
    "last_run":        None,
    "campaigns":       [],
    "heat_schedule":   pd.DataFrame(),
    "camp_schedule":   pd.DataFrame(),
    "capacity":        pd.DataFrame(),
    "solver_status":   "NOT RUN",
    "error":           None,
}


# ── Campaign serialiser ───────────────────────────────────────────────────────
def _camps_to_json(campaigns: list, camp_df: pd.DataFrame | None) -> list:
    result = []
    for c in campaigns:
        cid = str(c.get("campaign_id", ""))
        sched = {}
        if camp_df is not None and not camp_df.empty and "Campaign_ID" in camp_df.columns:
            match = camp_df[camp_df["Campaign_ID"].astype(str) == cid]
            if not match.empty:
                sched = {k: _safe(v) for k, v in match.iloc[0].to_dict().items()}

        shortages = [
            {"sku_id": k, "qty": round(float(v or 0), 2)}
            for k, v in (c.get("material_shortages") or {}).items()
        ]

        # Compute Margin_Hrs = Due_Date - last_operation_end
        margin_hrs = None
        try:
            due = pd.to_datetime(sched.get("Due_Date") or c.get("due_date"), errors="coerce")
            # Use the last relevant end time: RM_End (if RM campaign) else CCM_End else EAF_End
            end_key = next((k for k in ("RM_End", "CCM_End", "EAF_End") if sched.get(k) not in (None, "")), None)
            if end_key and pd.notna(due):
                end_dt = pd.to_datetime(sched[end_key], errors="coerce")
                if pd.notna(end_dt):
                    margin_hrs = round((due - end_dt).total_seconds() / 3600.0, 1)
        except Exception:
            pass

        result.append({
            "campaign_id":    cid,
            "campaign_group": str(c.get("campaign_group", "")),
            "grade":          str(c.get("grade", "")),
            "heats":          int(c.get("heats", 0) or 0),
            "total_mt":       round(float(c.get("total_coil_mt", 0) or 0), 1),
            "release_status": str(c.get("release_status", "HELD")),
            "needs_vd":       bool(c.get("needs_vd", False)),
            "billet_family":  str(c.get("billet_family", "")),
            "shortages":      shortages,
            "so_list":        [str(s) for s in (c.get("so_list") or [])],
            "Margin_Hrs":     margin_hrs,
            **sched,
        })
    return result


def _kpis_from_state() -> dict:
    campaigns  = _state["campaigns"]
    camp_df    = _state["camp_schedule"]
    cap_df     = _state["capacity"]

    released   = [c for c in campaigns if str(c.get("release_status", "")).upper() == "RELEASED"]
    held       = [c for c in campaigns if str(c.get("release_status", "")).upper() == "MATERIAL HOLD"]
    total_mt   = sum(float(c.get("total_coil_mt", 0) or 0) for c in campaigns)
    total_heats= sum(int(c.get("heats", 0) or 0) for c in campaigns)

    # Count late from camp_schedule Status column (covers all scheduled campaigns)
    late_count = 0
    if camp_df is not None and not camp_df.empty and "Status" in camp_df.columns:
        late_count = int((camp_df["Status"].astype(str).str.upper() == "LATE").sum())

    # On-time % = non-late campaigns out of all campaigns (not just released)
    total_camps = len(campaigns)
    on_time_count = max(0, total_camps - late_count)
    on_time_pct = round(100 * on_time_count / max(total_camps, 1), 1) if campaigns else 0.0

    # Throughput: total MT / horizon days
    horizon = 14
    throughput = round(total_mt / horizon, 1) if total_mt else 0.0

    # Bottleneck
    bottleneck = None
    max_util   = 0.0
    if cap_df is not None and not cap_df.empty and "Utilisation_%" in cap_df.columns:
        non_bf = cap_df[~cap_df.get("Resource_ID", pd.Series("")).astype(str).str.startswith("BF")]
        if not non_bf.empty:
            idx = non_bf["Utilisation_%"].idxmax()
            bottleneck = str(non_bf.loc[idx, "Resource_ID"])
            max_util   = round(float(non_bf.loc[idx, "Utilisation_%"]), 1)

    # Utilisation list per resource
    util_list = []
    if cap_df is not None and not cap_df.empty:
        for _, r in cap_df.iterrows():
            util_list.append({
                "resource_id":   _safe(r.get("Resource_ID")),
                "resource_name": _safe(r.get("Resource_Name", r.get("Resource_ID"))),
                "utilisation":   _safe(r.get("Utilisation_%")),
                "demand_hrs":    _safe(r.get("Demand_Hrs")),
                "avail_hrs":     _safe(r.get("Avail_Hrs_14d", r.get("Avail_Hours_Day", 20) or 20)),
                "status":        _safe(r.get("Status")),
                "operation":     _safe(r.get("Operation_Group", "")),
            })

    # Material shortages summary
    shortage_alerts = []
    for c in held:
        for sku, qty in (c.get("material_shortages") or {}).items():
            shortage_alerts.append({
                "campaign_id": str(c.get("campaign_id", "")),
                "sku_id":      str(sku),
                "shortage_qty": round(float(qty or 0), 2),
                "severity":    "HIGH" if float(qty or 0) > 20 else "MEDIUM",
            })

    return {
        "solver_status":       _state["solver_status"],
        "last_run":            _state["last_run"],
        "campaigns_total":     len(campaigns),
        "campaigns_released":  len(released),
        "campaigns_held":      len(held),
        "campaigns_late":      late_count,
        "total_heats":         total_heats,
        "total_mt":            round(total_mt, 1),
        "on_time_pct":         on_time_pct,
        "throughput_mt_day":   throughput,
        "bottleneck":          bottleneck,
        "max_utilisation":     max_util,
        "utilisation":         util_list,
        "shortage_alerts":     shortage_alerts,
    }


# ── Routes ────────────────────────────────────────────────────────────────────

@app.route("/api/health")
def health():
    exists = WORKBOOK.exists()
    mtime  = datetime.fromtimestamp(WORKBOOK.stat().st_mtime).isoformat() if exists else None
    return _jsonify({
        "ok":             exists,
        "workbook":       str(WORKBOOK),
        "workbook_mtime": mtime,
        "last_run":       _state["last_run"],
        "solver_status":  _state["solver_status"],
        "solver_detail":  _state.get("solver_detail", ""),
    })

@app.route("/api/data/config")
def get_config():
    try:
        # Load sheets directly for raw view
        df_cfg = _read_sheet("Config", ["Key"])
        conf_list = [{"Key": r.get("Key",""), "Value": _safe(r.get("Value","")), "Description": r.get("Description","")} for _, r in df_cfg.iterrows()]

        res_df   = _read_sheet("Resource_Master", ["Resource_ID"])
        skus_df  = _read_sheet("SKU_Master", ["SKU_ID"])
        rout_df  = _read_sheet("Routing", ["SKU_ID", "Operation"])
        queue_df = _read_sheet("Queue_Times", ["From_Operation", "To_Operation"])

        return _jsonify({
            "config": conf_list,
            "resources": _df_to_records(res_df),
            "skus": _df_to_records(skus_df),
            "routing": _df_to_records(rout_df),
            "queue": _df_to_records(queue_df),
        })
    except Exception as e:
        return _jsonify({"error": str(e)}), 500

@app.route("/api/data/dashboard")
def dashboard():
    return _jsonify(_kpis_from_state())


@app.route("/api/data/campaigns")
def get_campaigns():
    if not _state["campaigns"]:
        return _jsonify({"campaigns": [], "note": "Run schedule first"})
    return _jsonify({"campaigns": _camps_to_json(_state["campaigns"], _state["camp_schedule"])})


@app.route("/api/data/gantt")
def get_gantt():
    return _jsonify({"jobs": _df_to_records(_state["heat_schedule"])})


@app.route("/api/data/capacity")
def get_capacity():
    return _jsonify({"capacity": _df_to_records(_state["capacity"])})


@app.route("/api/data/orders")
def get_orders():
    try:
        d = _load_all()
        return _jsonify({"orders": _df_to_records(d["all_orders"])})
    except Exception as e:
        return _jsonify({"error": str(e), "trace": traceback.format_exc()}), 500


@app.route("/api/run/schedule", methods=["POST"])
def run_schedule_api():
    try:
        d       = _load_all()
        config  = d["config"]
        min_cmt = float(config.get("Min_Campaign_MT", 100.0) or 100.0)
        max_cmt = float(config.get("Max_Campaign_MT", 500.0) or 500.0)
        horizon = int(float(config.get("Planning_Horizon_Days", 14) or 14))
        sec_lim = float(config.get("Default_Solver_Limit_Sec", 30.0) or 30.0)

        # Override from request body
        body = request.get_json(silent=True) or {}
        horizon = int(body.get("horizon", horizon))
        sec_lim = float(body.get("solver_sec", body.get("time_limit", sec_lim)))

        campaigns = build_campaigns(
            d["sales_orders"],
            min_campaign_mt=min_cmt,
            max_campaign_mt=max_cmt,
            inventory=d["inventory"],
            bom=d["bom"],
            config=config,
            skus=d["skus"],
        )

        released = [c for c in campaigns if str(c.get("release_status", "")).upper() == "RELEASED"]

        result = schedule(
            campaigns,
            d["resources"],
            planning_horizon_days=horizon,
            planning_start=datetime.now(),
            routing=d["routing"],
            queue_times=d["queue_times"],
            config=config,
            solver_time_limit_sec=sec_lim,
        )

        heat_df  = result.get("heat_schedule", pd.DataFrame())
        camp_df  = result.get("campaign_schedule", pd.DataFrame())
        solver   = result.get("solver_status", "UNKNOWN")
        s_detail = result.get("solver_detail", "CP_SAT_NA")

        demand_hrs = compute_demand_hours(
            released, d["resources"], routing=d["routing"],
        )
        cap_df = capacity_map(demand_hrs, d["resources"], horizon_days=horizon)

        # Store in state
        _state["last_run"]       = datetime.now().isoformat()
        _state["campaigns"]      = campaigns
        _state["heat_schedule"]  = heat_df
        _state["camp_schedule"]  = camp_df
        _state["capacity"]       = cap_df
        _state["solver_status"]  = solver
        _state["solver_detail"]  = s_detail
        _state["error"]          = None

        kpis = _kpis_from_state()
        return _jsonify({
            **kpis,
            "campaigns": _camps_to_json(campaigns, camp_df),
            "gantt":     _df_to_records(heat_df),
            "capacity":  _df_to_records(cap_df),
            "solver_detail": s_detail,
        })

    except Exception as e:
        _state["error"] = str(e)
        return _jsonify({"error": str(e), "trace": traceback.format_exc()}), 500


@app.route("/api/run/bom", methods=["POST"])
def run_bom_api():
    try:
        d      = _load_all()
        demand = consolidate_demand(d["sales_orders"])
        din    = demand[["SKU_ID", "Total_Qty"]].rename(columns={"Total_Qty": "Required_Qty"})
        gross  = explode_bom_details(din, d["bom"])
        netted = net_requirements(gross, d["inventory"])
        return _jsonify({"bom": _df_to_records(netted), "rows": len(netted)})
    except Exception as e:
        return _jsonify({"error": str(e), "trace": traceback.format_exc()}), 500


@app.route("/api/run/ctp", methods=["POST"])
def run_ctp_api():
    try:
        body           = request.get_json(silent=True) or {}
        sku_id         = str(body.get("sku_id", ""))
        qty_mt         = float(body.get("qty_mt", 100.0))
        requested_date = body.get("requested_date", (datetime.now() + pd.Timedelta(days=14)).isoformat())

        d        = _load_all()
        config   = d["config"]
        min_cmt  = float(config.get("Min_Campaign_MT", 100.0) or 100.0)
        max_cmt  = float(config.get("Max_Campaign_MT", 500.0) or 500.0)
        campaigns = _state["campaigns"] or build_campaigns(
            d["sales_orders"], min_cmt, max_cmt,
            inventory=d["inventory"], bom=d["bom"], config=config, skus=d["skus"],
        )

        res = capable_to_promise(
            sku_id=sku_id,
            qty_mt=qty_mt,
            requested_date=requested_date,
            campaigns=campaigns,
            resources=d["resources"],
            bom=d["bom"],
            inventory=d["inventory"],
            routing=d["routing"],
            skus=d["skus"],
            planning_start=datetime.now(),
            config=config,
            queue_times=d["queue_times"],
        )
        return _jsonify({k: _safe(v) for k, v in res.items()})

    except Exception as e:
        return _jsonify({"error": str(e), "trace": traceback.format_exc()}), 500


@app.route("/api/orders/assign", methods=["POST"])
def assign_orders():
    """Write Campaign_ID back to Sales_Orders in Excel."""
    try:
        import openpyxl
        body        = request.get_json(silent=True) or {}
        assignments = body.get("assignments", [])   # [{so_id, campaign_id}]

        if not assignments:
            return _jsonify({"ok": True, "updated": 0})

        wb = openpyxl.load_workbook(WORKBOOK)
        ws = wb["Sales_Orders"]

        header_row = None
        so_col = cid_col = None
        for r in ws.iter_rows(min_row=1, max_row=6):
            for cell in r:
                if str(cell.value or "").strip() == "SO_ID":
                    header_row = cell.row
                    break
            if header_row:
                break

        if header_row:
            for cell in ws[header_row]:
                val = str(cell.value or "").strip()
                if val == "SO_ID":      so_col  = cell.column
                if val == "Campaign_ID": cid_col = cell.column

        if not so_col or not cid_col:
            return _jsonify({"error": "Cannot find SO_ID / Campaign_ID columns", "ok": False}), 400

        assign_map = {str(a["so_id"]).strip(): str(a["campaign_id"]).strip() for a in assignments}
        updated = 0
        for row in ws.iter_rows(min_row=header_row + 1):
            so_val = str(row[so_col - 1].value or "").strip()
            if so_val in assign_map:
                row[cid_col - 1].value = assign_map[so_val]
                updated += 1

        wb.save(WORKBOOK)
        return _jsonify({"ok": True, "updated": updated})

    except Exception as e:
        return _jsonify({"error": str(e), "trace": traceback.format_exc()}), 500


# ── Sales orders CRUD ───────────────────────────────────────────────────────

def _open_workbook_and_sheet(sheet_name="Sales_Orders"):
    import openpyxl
    wb = openpyxl.load_workbook(WORKBOOK)
    return wb, wb[sheet_name]


def _find_header_row(ws):
    for r in ws.iter_rows(min_row=1, max_row=10):
        if any(str(c.value or "").strip().upper() == "SO_ID" for c in r):
            return r[0].row, [str(c.value or "").strip() for c in r]
    return None, []


def _normalize_order_obj(o):
    if not isinstance(o, dict):
        return {}
    return {str(k).strip(): v for k, v in o.items()}


@app.route("/api/orders", methods=["GET", "POST"])
def orders_collection():
    if request.method == "GET":
        try:
            d = _load_all()
            return _jsonify({"orders": _df_to_records(d["all_orders"])})
        except Exception as e:
            return _jsonify({"error": str(e), "trace": traceback.format_exc()}), 500

    if request.method == "POST":
        body = _normalize_order_obj(request.get_json(silent=True) or {})
        so_id = str(body.get("SO_ID") or body.get("SO_ID", "")).strip()
        if not so_id:
            return _jsonify({"error": "SO_ID is required"}), 400

        d = _load_all()
        existing = d["all_orders"]
        if so_id in existing.get("SO_ID", pd.Series([], dtype=str)).astype(str).tolist():
            return _jsonify({"error": f"SO_ID {so_id} already exists"}), 409

        try:
            wb, ws = _open_workbook_and_sheet("Sales_Orders")
            header_row, header_cols = _find_header_row(ws)
            if not header_row:
                return _jsonify({"error": "Cannot find SO_ID header row"}), 500

            data_row = [body.get(col, "") for col in header_cols]
            ws.append(data_row)
            wb.save(WORKBOOK)
            return _jsonify({"ok": True, "so_id": so_id})
        except Exception as e:
            return _jsonify({"error": str(e), "trace": traceback.format_exc()}), 500


@app.route("/api/orders/<so_id>", methods=["GET", "PUT", "DELETE"])
def order_item(so_id):
    so_id = str(so_id or "").strip()
    if not so_id:
        return _jsonify({"error": "Invalid SO_ID"}), 400

    if request.method == "GET":
        try:
            d = _load_all()
            row = d["all_orders"][d["all_orders"]["SO_ID"].astype(str).str.strip() == so_id]
            if row.empty:
                return _jsonify({"error": "Not found"}), 404
            return _jsonify({"order": _df_to_records(row)[0]})
        except Exception as e:
            return _jsonify({"error": str(e), "trace": traceback.format_exc()}), 500

    try:
        wb, ws = _open_workbook_and_sheet("Sales_Orders")
        header_row, header_cols = _find_header_row(ws)
        if not header_row:
            return _jsonify({"error": "Cannot find SO_ID header row"}), 500

        # Map column names to openpyxl index
        col_index = {name: idx + 1 for idx, name in enumerate(header_cols)}

        target_row_idx = None
        for row in ws.iter_rows(min_row=header_row + 1):
            cell_so_id = str(row[col_index.get("SO_ID", 1) - 1].value or "").strip()
            if cell_so_id == so_id:
                target_row_idx = row[0].row
                break

        if target_row_idx is None:
            return _jsonify({"error": "Not found"}), 404

        if request.method == "PUT":
            body = _normalize_order_obj(request.get_json(silent=True) or {})
            for k, v in body.items():
                if k in col_index and k != "SO_ID":
                    ws.cell(row=target_row_idx, column=col_index[k], value=v)
            wb.save(WORKBOOK)
            return _jsonify({"ok": True, "so_id": so_id})

        if request.method == "DELETE":
            ws.delete_rows(target_row_idx, 1)
            wb.save(WORKBOOK)
            return _jsonify({"ok": True, "so_id": so_id})

    except Exception as e:
        return _jsonify({"error": str(e), "trace": traceback.format_exc()}), 500


@app.route("/api/data/skus")
def get_skus():
    try:
        d = _load_all()
        skus = d["skus"]
        fg = skus[skus.get("Category", pd.Series("")).astype(str).str.contains("Finished", case=False, na=False)]
        records = [{"sku_id": r.get("SKU_ID", ""), "sku_name": r.get("SKU_Name", "")}
                   for _, r in fg.iterrows()]
        return _jsonify({"skus": records})
    except Exception as e:
        return _jsonify({"error": str(e)}), 500


if __name__ == "__main__":
    print(f"\n  APS API  ->  http://localhost:5000")
    print(f"  Workbook ->  {WORKBOOK}\n")
    app.run(host="0.0.0.0", port=5000, debug=False, use_reloader=False)
