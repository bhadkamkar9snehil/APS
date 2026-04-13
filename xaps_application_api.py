"""
X-APS Application API
---------------------
Concept-oriented Flask API for the X-APS HTML application.

This server treats the Excel workbook as the persistence layer while exposing
APS application concepts as the API contract:
- dashboard
- orders
- campaigns
- schedule
- dispatch
- capacity
- material
- CTP
- scenarios
- master data

It also preserves the legacy route names already used by the existing UI.

Run:
    python xaps_application_api.py
"""
from __future__ import annotations

import json
import os
import sys
import time
import traceback
import uuid
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional
from zipfile import BadZipFile
import zipfile

import numpy as np
import pandas as pd
from flask import Flask, request
from flask_cors import CORS

sys.path.insert(0, str(Path(__file__).parent))
from engine.aps_planner import APSPlanner, SalesOrder, PlanningHorizon, HeatBatch
from engine.bom_explosion import consolidate_demand, explode_bom_details, net_requirements
from engine.campaign import build_campaigns
from engine.capacity import capacity_map, compute_demand_hours
from engine.config import (
    canonicalize_config_key,
    get_config,
    load_workbook_config_snapshot,
    update_algorithm_config_in_workbook,
)
from engine.ctp import capable_to_promise, _frozen_jobs_from_schedule_dataframe
from engine.excel_store import ExcelStore
from engine.masterdata_audit import audit_workbook_masterdata
from engine.scheduler import schedule
from engine.workbook_schema import SHEETS

raw_workbook_path = os.getenv("WORKBOOK_PATH", str(Path(__file__).parent / "APS_BF_SMS_RM.xlsx"))
raw_workbook_path = raw_workbook_path.strip().strip('"').strip("'")
WORKBOOK = Path(raw_workbook_path)
HEADER_ROW = 2
PORT = int(os.getenv("PORT", "5000"))

app = Flask(__name__)
CORS(app, origins=["http://localhost:3131", "http://127.0.0.1:3131", "*"])
store = ExcelStore(WORKBOOK)

MASTERDATA_SECTION_TO_SHEET = {
    "config": "config",
    "resources": "resource-master",
    "routing": "routing",
    "queue": "queue-times",
    "skus": "sku-master",
    "bom": "bom",
    "inventory": "inventory",
    "campaign-config": "campaign-config",
    "changeover": "changeover-matrix",
    "scenarios": "scenarios",
}

OUTPUT_SECTION_TO_SHEET = {
    "capacity": "capacity-map",
    "campaigns": "campaign-schedule",
    "material": "material-plan",
    "dispatch": "equipment-schedule",
    "schedule": "schedule-output",
    "gantt": "schedule-gantt",
    "kpis": "kpi-dashboard",
    "ctp-output": "ctp-output",
    "scenario-output": "scenario-output",
}


class _Enc(json.JSONEncoder):
    def default(self, obj: Any):
        if isinstance(obj, (np.integer,)):
            return int(obj)
        if isinstance(obj, (np.floating,)):
            return None if np.isnan(obj) else float(obj)
        if isinstance(obj, np.ndarray):
            return obj.tolist()
        if isinstance(obj, (datetime, date, pd.Timestamp)):
            try:
                return obj.isoformat()
            except Exception:
                return str(obj)
        try:
            if pd.isna(obj):
                return None
        except Exception:
            pass
        return super().default(obj)


def _safe(v: Any):
    if v is None:
        return None
    if isinstance(v, float) and (np.isnan(v) or np.isinf(v)):
        return None
    if isinstance(v, (np.integer,)):
        return int(v)
    if isinstance(v, (np.floating,)):
        return float(v)
    if isinstance(v, (datetime, date, pd.Timestamp)):
        try:
            return v.isoformat()
        except Exception:
            return str(v)
    try:
        if pd.isna(v):
            return None
    except Exception:
        pass
    return v


def _jsonify(data, status: int = 200):
    return app.response_class(json.dumps(data, cls=_Enc, ensure_ascii=False), status=status, mimetype="application/json")


def _clean_mode(val, default: str) -> str:
    """Safely convert a pandas cell to a clean uppercase string, returning default for NaN/None/blank."""
    if val is None:
        return default
    try:
        if pd.isna(val):
            return default
    except (TypeError, ValueError):
        pass
    s = str(val).strip().upper()
    return default if s in ("", "NAN", "NONE", "NA", "NULL", "N/A") else s or default


# Compatibility cache only. Authoritative reads must come from the active run artifact first.
_state: dict = {
    "last_run": None,
    "run_id": None,
    "campaigns": [],
    "heat_schedule": pd.DataFrame(),
    "camp_schedule": pd.DataFrame(),
    "capacity": pd.DataFrame(),
    "material_plan_data": None,
    "bom_explosion": None,
    "solver_status": "NOT RUN",
    "solver_detail": "",
    "error": None,
}

_run_artifacts: Dict[str, Dict[str, Any]] = {}
_active_run_id: Optional[str] = None

# BOM grouping sort order constants (mirror of aps_functions.py)
_PLANT_SORT_ORDER = {
    "Rolling Mill": 0,
    "SMS": 1,
    "Blast Furnace": 2,
    "Shared Stores": 3,
    "Other": 9,
}

_MAT_TYPE_SORT_ORDER = {
    "Finished Good Demand": 0,
    "RM Output": 1,
    "CCM Output": 2,
    "VD Output": 3,
    "LRF Output": 4,
    "EAF Output": 5,
    "EAF Raw Charge": 6,
    "Hot Metal": 7,
    "BF Raw Mix": 8,
    "BF Raw Material": 9,
    "SMS Raw Material": 10,
    "Shared Flux": 11,
    "BF Byproduct": 12,
    "EAF Byproduct": 13,
    "LRF Byproduct": 14,
    "VD Byproduct": 15,
    "CCM Byproduct": 16,
    "RM Byproduct": 17,
    "Byproduct/Waste": 18,
    "Raw Material": 19,
    "Material": 20,
}


def _current_trace_id() -> str:
    artifact = _get_active_run_artifact()
    if artifact:
        return str(artifact.get("run_id") or "unknown")
    return str(_state.get("run_id") or "unknown")


def _error_response(
    error_message: str,
    error_code: str = "INTERNAL_ERROR",
    error_domain: str = "API",
    trace_id: str | None = None,
    details: str | None = None,
    degraded_mode: bool = False,
    status_code: int = 500,
):
    response = {
        "error": True,
        "error_code": error_code,
        "error_message": error_message,
        "error_domain": error_domain,
        "trace_id": trace_id or _current_trace_id(),
        "degraded_mode": degraded_mode,
        "details": details,
        "timestamp": datetime.now().isoformat(),
    }
    return _jsonify(response, status=status_code), status_code


def _mat_type_for_sku(sku_id: str, category: str = "", sku_name: str = "") -> str:
    """Derive material type from SKU ID, category, or name. Mirrors aps_functions._material_type_for_sku."""
    sku = str(sku_id or "").strip().upper()
    category_text = str(category or "").strip()
    if sku.startswith("FG-WR-"):
        return "Finished Good Demand"
    if sku.startswith("RM-OUT-"):
        return "RM Output"
    if sku.startswith("BIL-"):
        return "CCM Output"
    if sku.startswith("VD-OUT-"):
        return "VD Output"
    if sku.startswith("LRF-OUT-"):
        return "LRF Output"
    if sku.startswith("EAF-OUT-"):
        return "EAF Output"
    if sku.startswith("EAF-RAW-"):
        return "EAF Raw Charge"
    if sku == "BF-HM":
        return "Hot Metal"
    if sku == "BF-RAW-MIX":
        return "BF Raw Mix"
    if sku in {"RM-IRON", "RM-COAL"}:
        return "BF Raw Material"
    if sku in {"RM-SCRAP", "RM-FESI", "RM-FEMN", "RM-FECR", "RM-ELEC"}:
        return "SMS Raw Material"
    if sku in {"RM-LIME", "RM-DOLO"}:
        return "Shared Flux"
    if sku == "BF-SLAG":
        return "BF Byproduct"
    if sku == "EAF-SLAG":
        return "EAF Byproduct"
    if sku == "LRF-WASTE":
        return "LRF Byproduct"
    if sku == "VD-WASTE":
        return "VD Byproduct"
    if sku == "CCM-CROP":
        return "CCM Byproduct"
    if sku in {"RM-ENDCUT", "RM-SCALE"}:
        return "RM Byproduct"
    if category_text:
        return category_text
    if sku_name:
        return str(sku_name).strip()
    return "Material"


def _plant_for_sku(sku_id: str, category: str = "", material_type: str = "") -> str:
    """Derive plant location from SKU ID, category, or material type. Mirrors aps_functions._plant_for_material."""
    sku = str(sku_id or "").strip().upper()
    material_type = str(material_type or "").strip()
    category_text = str(category or "").strip()
    if sku.startswith("FG-WR-") or sku.startswith("RM-OUT-") or sku in {"RM-ENDCUT", "RM-SCALE"}:
        return "Rolling Mill"
    if (
        sku.startswith("BIL-")
        or sku.startswith("VD-OUT-")
        or sku.startswith("LRF-OUT-")
        or sku.startswith("EAF-OUT-")
        or sku.startswith("EAF-RAW-")
        or sku in {"EAF-SLAG", "LRF-WASTE", "VD-WASTE", "CCM-CROP", "RM-SCRAP", "RM-FESI", "RM-FEMN", "RM-FECR", "RM-ELEC"}
    ):
        return "SMS"
    if sku in {"BF-HM", "BF-RAW-MIX", "BF-SLAG", "RM-IRON", "RM-COAL"}:
        return "Blast Furnace"
    if sku in {"RM-LIME", "RM-DOLO"}:
        return "Shared Stores"
    if material_type == "Raw Material":
        return "Shared Stores"
    if category_text == "Raw Material":
        return "Shared Stores"
    return "Other"


def _stage_for_sku(sku_id: str, material_type: str = "") -> str:
    """Derive production stage from SKU ID or material type. Mirrors aps_functions._stage_for_material."""
    sku = str(sku_id or "").strip().upper()
    material_type = str(material_type or "").strip()
    if sku.startswith("FG-WR-") or sku.startswith("RM-OUT-") or sku in {"RM-ENDCUT", "RM-SCALE"}:
        return "Rolling"
    if sku.startswith("BIL-") or sku == "CCM-CROP":
        return "CCM"
    if sku.startswith("VD-OUT-") or sku == "VD-WASTE":
        return "VD"
    if sku.startswith("LRF-OUT-") or sku == "LRF-WASTE":
        return "LRF"
    if sku.startswith("EAF-OUT-") or sku.startswith("EAF-RAW-") or sku == "EAF-SLAG":
        return "EAF"
    if sku in {"BF-HM", "BF-RAW-MIX", "BF-SLAG", "RM-IRON", "RM-COAL"}:
        return "BF"
    if sku in {"RM-LIME", "RM-DOLO"}:
        return "Shared Stores"
    if material_type == "Raw Material":
        return "Stores"
    return "Other"


def _df_to_records(df: pd.DataFrame | None) -> list:
    if df is None or df.empty:
        return []
    return [{k: _safe(v) for k, v in row.items()} for _, row in df.iterrows()]


def _sheet_items(api_name: str, **kwargs) -> List[Dict[str, Any]]:
    if api_name not in SHEETS:
        return []
    return store.list_rows(api_name, **kwargs)["items"]


def _read_sheet(sheet: str, required: list | None = None) -> pd.DataFrame:
    max_retries = 3
    retry_delay = 0.5
    last_error: Exception | None = None
    for attempt in range(max_retries):
        try:
            df = pd.read_excel(WORKBOOK, sheet_name=sheet, header=HEADER_ROW, dtype=str)
            df = df.dropna(how="all").reset_index(drop=True)
            if required:
                df = df.dropna(subset=[c for c in required if c in df.columns], how="all")
            return df
        except Exception as e:
            last_error = e
            if attempt < max_retries - 1:
                time.sleep(retry_delay)
    if last_error:
        raise last_error
    raise RuntimeError("Failed to read Excel sheet after multiple retries")


def _read_config() -> dict:
    try:
        df = _read_sheet("Config", ["Key"])
        return {str(r["Key"]).strip(): r.get("Value") for _, r in df.iterrows() if str(r.get("Key", "")).strip()}
    except Exception:
        return {}


def _read_queue_times(default_enforcement: str = "Hard") -> dict:
    try:
        df = _read_sheet("Queue_Times", ["From_Operation"])
        out = {}
        for _, r in df.iterrows():
            key = (
                str(r.get("From_Operation", "") or "").strip().upper(),
                str(r.get("To_Operation", "") or "").strip().upper(),
            )
            try:
                out[key] = {
                    "min": float(r.get("Min_Queue_Min") or 0),
                    "max": float(r.get("Max_Queue_Min") or 120),
                    "enforcement": str(r.get("Enforcement") or default_enforcement or "Hard").strip(),
                }
            except Exception:
                pass
        return out
    except Exception:
        return {}


def _load_all() -> dict:
    config_snapshot = load_workbook_config_snapshot(WORKBOOK, reload_algorithm=True)
    config = config_snapshot.runtime_config

    so_raw = _read_sheet("Sales_Orders", ["SO_ID", "SKU_ID"])
    so_raw = so_raw[so_raw["SO_ID"].notna()].copy()
    so_raw["SO_ID"] = so_raw["SO_ID"].astype(str).str.strip()
    so_raw = so_raw[so_raw["SO_ID"] != ""].copy()
    so_raw = so_raw.drop_duplicates(subset=["SO_ID"], keep="first").reset_index(drop=True)

    res = _read_sheet("Resource_Master", ["Resource_ID"])
    skus = _read_sheet("SKU_Master", ["SKU_ID"])
    routing = _read_sheet("Routing", ["SKU_ID", "Operation"])
    bom = _read_sheet("BOM", ["Parent_SKU", "Child_SKU"])
    inv = _read_sheet("Inventory", ["SKU_ID"])
    campaign_config = _read_sheet("Campaign_Config", ["Grade"])
    queue = _read_queue_times(str(config.get("Queue_Enforcement", "Hard") or "Hard"))
    changeover = _read_sheet("Changeover_Matrix", ["From \\ To"])

    def _num(df, col, default=0.0):
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(default)
        return df

    # Inventory
    if "Available_Qty" not in inv.columns and "Stock_Qty" in inv.columns:
        inv = inv.rename(columns={"Stock_Qty": "Available_Qty"})
    inv = _num(inv, "Available_Qty", 0.0)
    inv = _num(inv, "Reserved_Qty", 0.0)

    # Resources
    res = _num(res, "Avail_Hours_Day", 20.0)
    res = _num(res, "Max_Capacity_MT_Hr", 33.3)
    res = _num(res, "Default_Cycle_Min", 0.0)
    res = _num(res, "Default_Setup_Min", 0.0)
    res = _num(res, "Efficiency_%", 100.0)
    if "Plant" not in res.columns:
        res["Plant"] = "Plant"

    # BOM
    bom = _num(bom, "Qty_Per", 1.0)
    bom = _num(bom, "Yield_Pct", 100.0)
    bom = _num(bom, "Level", 1.0)

    # Sales orders
    so_raw["Delivery_Date"] = pd.to_datetime(so_raw.get("Delivery_Date"), format="mixed", errors="coerce")
    so_raw["Order_Date"] = pd.to_datetime(so_raw.get("Order_Date"), format="mixed", errors="coerce")
    so_raw["Status"] = so_raw.get("Status", "Open").fillna("Open")
    so_raw = _num(so_raw, "Order_Qty_MT", 0.0)
    so_raw = _num(so_raw, "Section_mm", 6.5)
    if "Order_Qty_MT" not in so_raw.columns and "Order_Qty" in so_raw.columns:
        so_raw["Order_Qty_MT"] = pd.to_numeric(so_raw["Order_Qty"], errors="coerce").fillna(0.0)
    if "Order_Qty" not in so_raw.columns:
        so_raw["Order_Qty"] = so_raw.get("Order_Qty_MT", 0.0)

    so_raw = so_raw[
        so_raw["SO_ID"].astype(str).str.strip().ne("")
        & so_raw["SKU_ID"].notna()
        & (so_raw["Order_Qty_MT"] > 0)
        & so_raw["Delivery_Date"].notna()
    ].copy()

    campaign_group = so_raw.get("Campaign_Group", "").fillna("").astype(str).str.strip().str.upper()
    aps_released = campaign_group.eq("APS_RELEASED")
    open_mask = so_raw["Status"].astype(str).str.strip().str.upper().isin({"OPEN", "CONFIRMED", "PLANNED", ""}) & ~aps_released
    open_so = so_raw[open_mask].copy()
    if open_so.empty:
        open_so = so_raw.copy()

    # Routing
    routing = _num(routing, "Cycle_Time_Min_Heat", 60.0)
    routing = _num(routing, "Setup_Time_Min", 0.0)
    routing = _num(routing, "Transfer_Time_Min", 0.0)
    routing = _num(routing, "Op_Seq", 10.0)
    if "Sequence" not in routing.columns and "Op_Seq" in routing.columns:
        routing["Sequence"] = routing["Op_Seq"]

    # Changeover matrix numeric coercion
    if not changeover.empty:
        first_col = changeover.columns[0]
        for col in changeover.columns[1:]:
            changeover[col] = pd.to_numeric(changeover[col], errors="coerce").fillna(0.0)
        changeover[first_col] = changeover[first_col].astype(str).str.strip()

    return {
        "config": config,
        "config_conflicts": config_snapshot.conflicts,
        "algorithm_config": config_snapshot.algorithm_config.all_params(),
        "system_config": config_snapshot.system_config,
        "sales_orders": open_so,
        "all_orders": so_raw,
        "resources": res,
        "skus": skus,
        "routing": routing,
        "bom": bom,
        "inventory": inv,
        "campaign_config": campaign_config,
        "queue_times": queue,
        "changeover_matrix": changeover,
    }

def _config_flag(config: dict | None, key: str, default: str = "N") -> bool:
    value = str((config or {}).get(key, default) or default).strip().upper()
    return value in {"Y", "YES", "TRUE", "1", "ON"}


# ----- Run artifact model -----

def _create_run_artifact(
    *,
    config: Dict[str, Any],
    campaigns: List[Dict[str, Any]],
    heat_schedule: pd.DataFrame,
    campaign_schedule: pd.DataFrame,
    capacity_map_df: pd.DataFrame,
    solver_status: str,
    solver_detail: str,
    material_plan: Dict[str, Any] | None = None,
    warnings: List[str] | None = None,
    degraded_flags: Dict[str, bool] | None = None,
) -> str:
    run_id = str(uuid.uuid4())
    created_at = datetime.now().isoformat()
    config_snapshot = {k: v for k, v in (config or {}).items()}
    allow_default_masters = _config_flag(config_snapshot, "Allow_Scheduler_Default_Masters", "N")
    campaign_serialization_mode = str(config_snapshot.get("Campaign_Serialization_Mode", "STANDARD") or "STANDARD")

    # Auto-compute material plan if not provided
    if material_plan is None:
        material_plan = _calculate_material_plan(campaigns, detail_level="campaign", run_id=run_id)

    artifact = {
        "run_id": run_id,
        "created_at": created_at,
        "workbook_path": str(WORKBOOK),
        "input_snapshot": {
            "campaign_count": len(campaigns),
            "solver_config": {
                "solver_status": solver_status,
                "solver_detail": solver_detail,
                "allow_default_masters": allow_default_masters,
                "campaign_serialization_mode": campaign_serialization_mode,
            },
        },
        "config_snapshot": config_snapshot,
        "results": {
            "campaigns": campaigns,
            "heat_schedule": _df_to_records(heat_schedule),
            "campaign_schedule": _df_to_records(campaign_schedule),
            "capacity": _df_to_records(capacity_map_df),
            "material_plan": material_plan,
        },
        "solver_metadata": {
            "status": solver_status,
            "detail": solver_detail,
            "allow_default_masters": allow_default_masters,
            "campaign_serialization_mode": campaign_serialization_mode,
        },
        "warnings": warnings or [],
        "degraded_flags": degraded_flags or {
            "default_masters_used": allow_default_masters,
            "greedy_fallback": solver_status in {"GREEDY", "GREEDY_FALLBACK", "UNKNOWN"},
            "material_incomplete": False,
            "inventory_lineage_degraded": False,
            "capacity_unavailable": False,
        },
    }
    _run_artifacts[run_id] = artifact
    return run_id


def _get_active_run_artifact() -> Dict[str, Any] | None:
    if _active_run_id and _active_run_id in _run_artifacts:
        return _run_artifacts[_active_run_id]
    return None


def _set_active_run_artifact(run_id: str) -> bool:
    global _active_run_id
    if run_id in _run_artifacts:
        _active_run_id = run_id
        return True
    return False


def _invalidate_active_run_cache(*, clear_simulation: bool = False) -> None:
    """Clear run artifact/cache snapshots after planning-state mutations."""
    global _active_run_id
    _active_run_id = None
    _run_artifacts.clear()
    _state["run_id"] = None
    _state["campaigns"] = []
    _state["heat_schedule"] = pd.DataFrame()
    _state["camp_schedule"] = pd.DataFrame()
    _state["capacity"] = pd.DataFrame()
    _state["material_plan_data"] = None
    _state["bom_explosion"] = None
    _state["solver_status"] = "NOT RUN"
    _state["solver_detail"] = "Planning state changed. Re-run simulation/schedule."

    if clear_simulation:
        aps_planning_simulate._last_result = None
        aps_planning_simulate._scheduler_inputs = []
        aps_planning_simulate._heat_batches = []


# ----- Active run reader utilities -----

def _active_run_campaigns() -> List[Dict[str, Any]]:
    artifact = _get_active_run_artifact()
    if artifact:
        return list(artifact.get("results", {}).get("campaigns", []))
    if _state.get("campaigns"):
        return list(_state["campaigns"])
    try:
        if OUTPUT_SECTION_TO_SHEET.get("campaigns", "campaign-schedule") in SHEETS:
            return _sheet_items(OUTPUT_SECTION_TO_SHEET.get("campaigns", "campaign-schedule"))
    except Exception:
        pass
    return []


def _active_run_heat_schedule() -> pd.DataFrame:
    artifact = _get_active_run_artifact()
    if artifact:
        return pd.DataFrame(artifact.get("results", {}).get("heat_schedule", []))
    if isinstance(_state.get("heat_schedule"), pd.DataFrame) and not _state["heat_schedule"].empty:
        return _state["heat_schedule"]
    try:
        if OUTPUT_SECTION_TO_SHEET.get("schedule", "schedule-output") in SHEETS:
            return _read_sheet(OUTPUT_SECTION_TO_SHEET.get("schedule", "schedule-output"))
    except Exception:
        pass
    return pd.DataFrame()


def _active_run_campaign_schedule() -> pd.DataFrame:
    artifact = _get_active_run_artifact()
    if artifact:
        return pd.DataFrame(artifact.get("results", {}).get("campaign_schedule", []))
    if isinstance(_state.get("camp_schedule"), pd.DataFrame) and not _state["camp_schedule"].empty:
        return _state["camp_schedule"]
    return pd.DataFrame()


def _active_run_capacity() -> pd.DataFrame:
    artifact = _get_active_run_artifact()
    if artifact:
        return pd.DataFrame(artifact.get("results", {}).get("capacity", []))
    if isinstance(_state.get("capacity"), pd.DataFrame) and not _state["capacity"].empty:
        return _state["capacity"]
    try:
        if OUTPUT_SECTION_TO_SHEET.get("capacity", "capacity-map") in SHEETS:
            return _read_sheet(OUTPUT_SECTION_TO_SHEET.get("capacity", "capacity-map"))
    except Exception:
        pass
    return pd.DataFrame()


def _workbook_material_plan_payload() -> Dict[str, Any]:
    if OUTPUT_SECTION_TO_SHEET.get("material", "material-plan") not in SHEETS:
        return {"summary": {}, "campaigns": [], "detail_level": "campaign"}

    try:
        rows = _sheet_items(OUTPUT_SECTION_TO_SHEET.get("material", "material-plan"))
        if not rows:
            return {"summary": {}, "campaigns": [], "detail_level": "campaign"}

        campaigns_dict: Dict[str, Dict[str, Any]] = {}
        for row in rows:
            camp_id = str(row.get("Campaign_ID") or row.get("campaign_id") or "").strip()
            if not camp_id or camp_id.startswith("Run"):
                continue

            camp = campaigns_dict.setdefault(
                camp_id,
                {
                    "campaign_id": camp_id,
                    "grade": str(row.get("Grade") or row.get("grade") or ""),
                    "release_status": str(row.get("Release_Status") or row.get("release_status") or "OPEN"),
                    "required_qty": 0.0,
                    "shortage_qty": 0.0,
                    "material_status": "WORKBOOK_FALLBACK",
                    "feasible": True,
                    "structure_errors": [],
                    "material_shortages": {},
                    "material_consumed": {},
                    "material_gross_requirements": {},
                    "detail_rows": [],
                },
            )

            sku = str(row.get("Material_SKU") or row.get("material_sku") or "").strip()
            req_qty = float(pd.to_numeric(pd.Series([row.get("Required_Qty") or row.get("required_qty")]), errors="coerce").fillna(0).iloc[0])
            consumed = float(pd.to_numeric(pd.Series([row.get("Consumed") or row.get("consumed")]), errors="coerce").fillna(0).iloc[0])
            available_before = float(pd.to_numeric(pd.Series([row.get("Available_Before") or row.get("available_before")]), errors="coerce").fillna(0).iloc[0])
            status = str(row.get("Status") or row.get("status") or "OK").strip()

            camp["required_qty"] += req_qty
            if sku and consumed > 0:
                camp["material_consumed"][sku] = round(float(camp["material_consumed"].get(sku, 0.0)) + consumed, 3)
            if sku and req_qty > 0:
                camp["material_gross_requirements"][sku] = round(float(camp["material_gross_requirements"].get(sku, 0.0)) + req_qty, 3)

            shortage_qty = max(req_qty - available_before, 0.0) if "SHORT" in status.upper() else 0.0
            if sku and shortage_qty > 0:
                camp["material_shortages"][sku] = round(float(camp["material_shortages"].get(sku, 0.0)) + shortage_qty, 3)
                camp["shortage_qty"] += shortage_qty
                camp["feasible"] = False
                camp["material_status"] = "SHORTAGE"

            camp["detail_rows"].append(
                {
                    "sku_id": sku,
                    "required_qty": round(req_qty, 3),
                    "consumed_qty": round(consumed, 3),
                    "available_before": round(available_before, 3),
                    "type": status or "WORKBOOK_ROW",
                }
            )

        campaigns = list(campaigns_dict.values())
        summary = {
            "Campaigns": len(campaigns),
            "Released": sum(1 for c in campaigns if "RELEASED" in str(c.get("release_status", "")).upper()),
            "Held": sum(1 for c in campaigns if "HOLD" in str(c.get("release_status", "")).upper()),
            "Shortage Lines": sum(1 for c in campaigns if float(c.get("shortage_qty", 0.0) or 0.0) > 1e-9),
            "BOM Structure Errors": sum(1 for c in campaigns if c.get("structure_errors")),
            "Total Required Qty": round(sum(float(c.get("required_qty", 0.0) or 0.0) for c in campaigns), 3),
            "Inventory Covered Qty": round(
                sum(max(float(c.get("required_qty", 0.0) or 0.0) - float(c.get("shortage_qty", 0.0) or 0.0), 0.0) for c in campaigns),
                3,
            ),
            "Shortage Qty": round(sum(float(c.get("shortage_qty", 0.0) or 0.0) for c in campaigns), 3),
        }
        return {"summary": summary, "campaigns": campaigns, "detail_level": "campaign"}
    except Exception:
        return {"summary": {}, "campaigns": [], "detail_level": "campaign"}


def _active_run_material() -> Dict[str, Any]:
    artifact = _get_active_run_artifact()
    if artifact:
        material_plan = artifact.get("results", {}).get("material_plan")
        if material_plan is not None:
            return material_plan
    if _state.get("material_plan_data"):
        return _state["material_plan_data"]
    return _workbook_material_plan_payload()


def _active_run_bom() -> Optional[Dict[str, Any]]:
    """Read BOM explosion results from artifact first, then compatibility cache."""
    artifact = _get_active_run_artifact()
    if artifact:
        bom = artifact.get("results", {}).get("bom_explosion")
        if bom is not None:
            return bom
    return _state.get("bom_explosion")


def _active_run_solver_metadata() -> Dict[str, Any]:
    artifact = _get_active_run_artifact()
    if artifact:
        meta = artifact.get("solver_metadata", {})
        return {
            "solver_status": meta.get("status", "UNKNOWN"),
            "solver_detail": meta.get("detail", ""),
            "allow_default_masters": bool(meta.get("allow_default_masters", False)),
            "campaign_serialization_mode": meta.get("campaign_serialization_mode", "STANDARD"),
            "degraded_flags": artifact.get("degraded_flags", {}),
            "run_id": artifact.get("run_id"),
            "created_at": artifact.get("created_at"),
            "warnings": artifact.get("warnings", []),
        }
    return {
        "solver_status": _state.get("solver_status", "UNKNOWN"),
        "solver_detail": _state.get("solver_detail", ""),
        "allow_default_masters": False,
        "campaign_serialization_mode": "STANDARD",
        "degraded_flags": {},
        "run_id": _state.get("run_id"),
        "created_at": _state.get("last_run"),
        "warnings": [],
    }


def _campaigns_to_view(campaigns: list, camp_df: pd.DataFrame | None) -> list:
    result = []
    for c in campaigns:
        cid = str(c.get("campaign_id") or c.get("Campaign_ID") or "")
        sched = {}
        if camp_df is not None and not camp_df.empty:
            key_col = "Campaign_ID" if "Campaign_ID" in camp_df.columns else "campaign_id" if "campaign_id" in camp_df.columns else None
            if key_col:
                match = camp_df[camp_df[key_col].astype(str) == cid]
                if not match.empty:
                    sched = {k: _safe(v) for k, v in match.iloc[0].to_dict().items()}
        result.append(
            {
                "campaign_id": cid,
                "campaign_group": str(c.get("campaign_group", "")),
                "grade": str(c.get("grade", "")),
                "heats": int(c.get("heats", 0) or 0),
                "total_mt": round(float(c.get("total_coil_mt", c.get("total_mt", 0)) or 0), 1),
                "release_status": str(c.get("release_status", "HELD")),
                "material_status": str(c.get("material_status", "")),
                "material_issue": str(c.get("material_issue", "")),
                "needs_vd": bool(c.get("needs_vd", False)),
                "billet_family": str(c.get("billet_family", "")),
                "shortages": [{"sku_id": k, "qty": round(float(v or 0), 2)} for k, v in (c.get("material_shortages") or {}).items()],
                "so_list": [str(s) for s in (c.get("so_ids") or c.get("so_list") or [])],
                **sched,
            }
        )
    return result


def _campaign_rows() -> List[Dict[str, Any]]:
    campaigns = _active_run_campaigns()
    if campaigns:
        return _campaigns_to_view(campaigns, _active_run_campaign_schedule())
    return []


def _schedule_rows() -> List[Dict[str, Any]]:
    return _df_to_records(_active_run_heat_schedule())


def _infer_operation_group(resource_id: str, resources_df: pd.DataFrame | None = None) -> str:
    rid = str(resource_id or "").strip()
    if not rid:
        return ""
    rid_upper = rid.upper()
    if resources_df is not None and not getattr(resources_df, "empty", True):
        if "Resource_ID" in resources_df.columns and "Operation_Group" in resources_df.columns:
            rows = resources_df.copy()
            rows["Resource_ID"] = rows["Resource_ID"].astype(str).str.strip().str.upper()
            rows["Operation_Group"] = rows["Operation_Group"].astype(str).str.strip().str.upper()
            match = rows.loc[rows["Resource_ID"] == rid_upper, "Operation_Group"]
            if not match.empty and str(match.iloc[0]).strip():
                return str(match.iloc[0]).strip().upper()

    if rid_upper.startswith("EAF"):
        return "EAF"
    if rid_upper.startswith("LRF"):
        return "LRF"
    if rid_upper.startswith("VD"):
        return "VD"
    if rid_upper.startswith("CCM"):
        return "CCM"
    if rid_upper.startswith("RM"):
        return "RM"
    return ""


def _derive_capacity_status(util_pct: float) -> str:
    if util_pct > 100.0:
        return "OVERLOADED"
    if util_pct < 60.0:
        return "UNDERUTILISED"
    if util_pct > 85.0:
        return "HIGH"
    return "OK"


def _capacity_rows() -> List[Dict[str, Any]]:
    rows = _df_to_records(_active_run_capacity())
    if not rows:
        return []

    resources_df = None
    try:
        resources_df = _load_all().get("resources")
    except Exception:
        resources_df = None

    enriched: List[Dict[str, Any]] = []
    for row in rows:
        normalized = dict(row or {})
        rid = str(
            normalized.get("Resource_ID")
            or normalized.get("resource_id")
            or ""
        ).strip()

        demand_hrs = float(normalized.get("Demand_Hrs") or normalized.get("demand_hrs") or 0.0)
        process_hrs = float(normalized.get("Process_Hrs") or normalized.get("process_hrs") or 0.0)
        setup_hrs = float(normalized.get("Setup_Hrs") or normalized.get("setup_hrs") or 0.0)
        changeover_hrs = float(normalized.get("Changeover_Hrs") or normalized.get("changeover_hrs") or 0.0)
        task_count = int(float(normalized.get("Task_Count") or normalized.get("task_count") or 0))

        avail_hrs = float(
            normalized.get("Avail_Hrs_14d")
            or normalized.get("avail_hrs")
            or normalized.get("Avail_Hours")
            or 0.0
        )
        if avail_hrs <= 0:
            avail_day = float(normalized.get("Avail_Hours_Day") or normalized.get("avail_hours_day") or 0.0)
            if avail_day > 0:
                avail_hrs = avail_day * 14.0

        util_pct = float(normalized.get("Utilisation_%") or normalized.get("utilisation") or 0.0)
        if avail_hrs > 0 and util_pct <= 0:
            util_pct = (demand_hrs / avail_hrs) * 100.0

        idle_hrs = float(normalized.get("Idle_Hrs") or max(avail_hrs - demand_hrs, 0.0))
        overload_hrs = float(normalized.get("Overload_Hrs") or max(demand_hrs - avail_hrs, 0.0))
        operation_group = str(
            normalized.get("Operation_Group")
            or normalized.get("Operation")
            or _infer_operation_group(rid, resources_df=resources_df)
        ).strip()
        status = str(normalized.get("Status") or _derive_capacity_status(util_pct)).strip().upper()

        normalized.update(
            {
                "Resource_ID": rid,
                "Operation_Group": operation_group,
                "Demand_Hrs": round(demand_hrs, 2),
                "Process_Hrs": round(process_hrs, 2),
                "Setup_Hrs": round(setup_hrs, 2),
                "Changeover_Hrs": round(changeover_hrs, 2),
                "Task_Count": task_count,
                "Avail_Hrs_14d": round(avail_hrs, 2),
                "Idle_Hrs": round(idle_hrs, 2),
                "Overload_Hrs": round(overload_hrs, 2),
                "Utilisation_%": round(util_pct, 1),
                "Status": status,
            }
        )
        enriched.append(normalized)

    return enriched


def _masterdata_health_summary(audit_payload: Dict[str, Any]) -> Dict[str, Any]:
    checks: List[Dict[str, Any]] = []

    def add_check(level: str, code: str, title: str, count: int, detail: str) -> None:
        checks.append(
            {
                "level": str(level).lower(),
                "code": str(code).upper(),
                "title": title,
                "count": int(max(count, 0)),
                "detail": detail,
            }
        )

    master_sheets = audit_payload.get("master_sheets", {}) if isinstance(audit_payload, dict) else {}
    missing_sheets = [name for name, report in master_sheets.items() if not bool((report or {}).get("exists"))]
    if missing_sheets:
        add_check(
            "critical",
            "MISSING_MASTER_SHEETS",
            "Missing master-data sheets",
            len(missing_sheets),
            f"Missing: {', '.join(sorted(missing_sheets)[:8])}",
        )
    else:
        add_check("ok", "MASTER_SHEETS_PRESENT", "All expected master-data sheets present", 0, "No missing sheets.")

    missing_columns: List[str] = []
    for sheet_name, report in (master_sheets or {}).items():
        for col in (report or {}).get("used_columns_missing", []) or []:
            missing_columns.append(f"{sheet_name}.{col}")
    if missing_columns:
        add_check(
            "critical",
            "MISSING_USED_COLUMNS",
            "Required columns missing",
            len(missing_columns),
            ", ".join(missing_columns[:8]),
        )
    else:
        add_check("ok", "MASTER_COLUMNS_OK", "Required master columns are present", 0, "No required-column gaps.")

    config_dupes = (audit_payload or {}).get("config_duplicates", {}) if isinstance(audit_payload, dict) else {}
    conflict_count = int(len((config_dupes or {}).get("conflicts", []) or []))
    if conflict_count > 0:
        add_check(
            "warning",
            "CONFIG_CONFLICTS",
            "Config value conflicts",
            conflict_count,
            "Algorithm_Config and Config contain conflicting keys. Review merged runtime config.",
        )
    else:
        add_check("ok", "CONFIG_CONFLICT_FREE", "No config conflicts detected", 0, "Config sheets are aligned.")

    try:
        data = _load_all()
        resources = data.get("resources", pd.DataFrame())
        routing = data.get("routing", pd.DataFrame())
        queue_times = data.get("queue_times", {}) or {}
        all_orders = data.get("all_orders", pd.DataFrame())

        inactive_count = 0
        stale_capacity_count = 0
        resource_ops: set[str] = set()
        if resources is not None and not getattr(resources, "empty", True):
            if "Status" in resources.columns:
                status = resources["Status"].fillna("").astype(str).str.strip().str.upper()
                inactive_count = int(status.isin({"INACTIVE", "DOWN", "DISABLED"}).sum())
            if "Avail_Hours_Day" in resources.columns:
                avail = pd.to_numeric(resources["Avail_Hours_Day"], errors="coerce").fillna(0.0)
                stale_capacity_count = int((avail <= 0).sum())
            if "Operation_Group" in resources.columns:
                resource_ops = {
                    str(op).strip().upper()
                    for op in resources["Operation_Group"].fillna("").astype(str)
                    if str(op).strip()
                }
        if inactive_count > 0 or stale_capacity_count > 0:
            add_check(
                "warning",
                "RESOURCE_HEALTH_WARN",
                "Inactive or stale resources detected",
                inactive_count + stale_capacity_count,
                f"Inactive/Down={inactive_count}, non-positive available hours={stale_capacity_count}.",
            )
        else:
            add_check("ok", "RESOURCE_HEALTH_OK", "Resource master health is clean", 0, "No stale/inactive resource issues.")

        routing_ops: set[str] = set()
        setup_missing_ops: List[str] = []
        if routing is not None and not getattr(routing, "empty", True):
            if "Operation" in routing.columns:
                routing_ops = {
                    str(op).strip().upper()
                    for op in routing["Operation"].fillna("").astype(str)
                    if str(op).strip()
                }
            if {"Operation", "Setup_Time_Min"}.issubset(set(routing.columns)):
                rt = routing.copy()
                rt["Operation"] = rt["Operation"].fillna("").astype(str).str.strip().str.upper()
                rt["Setup_Time_Min"] = pd.to_numeric(rt["Setup_Time_Min"], errors="coerce").fillna(0.0)
                for op, grp in rt.groupby("Operation", sort=False):
                    if op and float(grp["Setup_Time_Min"].max()) <= 0.0:
                        setup_missing_ops.append(op)

        required_ops = {"EAF", "LRF", "CCM", "RM"}
        if all_orders is not None and not getattr(all_orders, "empty", True):
            so = all_orders.copy()
            route_variant = so.get("Route_Variant")
            needs_vd = so.get("Needs_VD")
            needs_vd_flag = False
            if route_variant is not None:
                rv = route_variant.fillna("").astype(str).str.strip().str.upper()
                needs_vd_flag = bool(rv.eq("Y").any())
            if not needs_vd_flag and needs_vd is not None:
                nv = needs_vd.fillna("").astype(str).str.strip().str.upper()
                needs_vd_flag = bool(nv.isin({"Y", "YES", "TRUE", "1", "ON"}).any())
            if needs_vd_flag:
                required_ops.add("VD")

        missing_in_routing = sorted(op for op in required_ops if op not in routing_ops)
        missing_in_resources = sorted(op for op in required_ops if op not in resource_ops)
        if missing_in_routing or missing_in_resources:
            add_check(
                "critical",
                "ROUTING_RESOURCE_GAPS",
                "Routing/resource operation coverage gaps",
                len(missing_in_routing) + len(missing_in_resources),
                f"Missing in Routing={missing_in_routing or ['None']}; missing in Resource_Master={missing_in_resources or ['None']}.",
            )
        else:
            add_check(
                "ok",
                "ROUTING_RESOURCE_COVERAGE_OK",
                "Routing and resources cover required operations",
                0,
                "Required operation coverage is complete.",
            )

        if setup_missing_ops:
            add_check(
                "warning",
                "SETUP_TIME_MISSING",
                "Required setup-time warnings",
                len(setup_missing_ops),
                f"Operations with zero setup defaults: {', '.join(setup_missing_ops[:8])}",
            )
        else:
            add_check("ok", "SETUP_TIME_OK", "Setup-time defaults are present", 0, "No setup-time gaps found.")

        expected_pairs = [("EAF", "LRF"), ("CCM", "RM")]
        if "VD" in required_ops:
            expected_pairs.extend([("LRF", "VD"), ("VD", "CCM")])
        else:
            expected_pairs.append(("LRF", "CCM"))

        missing_queue_pairs = [pair for pair in expected_pairs if pair not in queue_times]
        invalid_queue_rules = []
        for pair, rule in queue_times.items():
            try:
                min_q = float((rule or {}).get("min", 0) or 0)
                max_q = float((rule or {}).get("max", 9999) or 9999)
                if max_q < min_q:
                    invalid_queue_rules.append((pair, f"min {min_q:g} > max {max_q:g}"))
            except Exception:
                invalid_queue_rules.append((pair, "rule parse error"))

        if missing_queue_pairs or invalid_queue_rules:
            add_check(
                "warning",
                "QUEUE_TIME_WARN",
                "Queue-time coverage/validity warnings",
                len(missing_queue_pairs) + len(invalid_queue_rules),
                f"Missing pairs={missing_queue_pairs[:4]}; invalid rules={invalid_queue_rules[:4]}",
            )
        else:
            add_check("ok", "QUEUE_TIME_OK", "Queue-time rules look valid", 0, "No queue-time gaps detected.")

    except Exception as exc:
        add_check(
            "warning",
            "HEALTH_CHECK_PARTIAL",
            "Master-data runtime health checks partially unavailable",
            1,
            f"Runtime checks could not complete: {exc}",
        )

    critical_count = sum(1 for item in checks if item.get("level") == "critical")
    warning_count = sum(1 for item in checks if item.get("level") == "warning")
    ok_count = sum(1 for item in checks if item.get("level") == "ok")
    return {
        "critical_count": critical_count,
        "warning_count": warning_count,
        "ok_count": ok_count,
        "checks": checks,
        "generated_at": datetime.now().isoformat(),
    }


def _material_plan_payload(detail_level: str = "campaign") -> Dict[str, Any]:
    requested_level = _normalize_material_detail_level(detail_level)
    payload = _active_run_material()
    if not payload or not (payload.get("campaigns") or payload.get("summary")):
        payload = _state.get("material_plan_data") or _workbook_material_plan_payload()
    return _material_plan_for_detail_level(payload or {}, requested_level)


def _material_holds_payload() -> List[Dict[str, Any]]:
    payload = _material_plan_payload(detail_level="campaign")
    items: List[Dict[str, Any]] = []
    for camp in payload.get("campaigns", []):
        shortage_qty = float(camp.get("shortage_qty", 0.0) or 0.0)
        material_status = str(camp.get("material_status", "") or "").upper()
        if shortage_qty > 1e-9 or material_status not in {"", "OK", "READY"}:
            items.append(camp)
    return items


def _planning_orders_release_signature(planning_orders: List[Dict[str, Any]]) -> str:
    """Build a stable signature for schedule-affecting planning-order content."""
    canonical: List[Dict[str, Any]] = []
    for po in (planning_orders or []):
        po_id = str(po.get("po_id", "")).strip()
        if not po_id:
            continue

        raw_so_ids = po.get("selected_so_ids") or []
        if isinstance(raw_so_ids, str):
            raw_so_ids = [x.strip() for x in raw_so_ids.split(",") if str(x).strip()]
        so_ids = sorted({
            str(so_id).strip()
            for so_id in raw_so_ids
            if str(so_id).strip()
        })
        due_window = po.get("due_window") or ("", "")
        if isinstance(due_window, list):
            due_window = tuple(due_window)
        if not isinstance(due_window, tuple) or len(due_window) < 2:
            due_window = ("", "")

        canonical.append(
            {
                "po_id": po_id,
                "selected_so_ids": so_ids,
                "total_qty_mt": round(float(po.get("total_qty_mt", 0) or 0), 3),
                "grade_family": str(po.get("grade_family", "")).strip(),
                "size_family": str(po.get("size_family", "")).strip(),
                "due_start": str(due_window[0] or "").strip(),
                "due_end": str(due_window[1] or "").strip(),
                "route_family": str(po.get("route_family", "SMS→RM")).strip() or "SMS→RM",
                "heats_required": int(po.get("heats_required", 0) or 0),
                "frozen_flag": bool(po.get("frozen_flag", False)),
                "order_type": str(po.get("order_type", "MTO")).strip().upper() or "MTO",
                "rolling_mode": str(po.get("rolling_mode", "HOT")).strip().upper() or "HOT",
                "priority": str(po.get("priority", "NORMAL")).strip().upper() or "NORMAL",
            }
        )

    canonical.sort(key=lambda row: row["po_id"])
    return json.dumps(canonical, sort_keys=True, separators=(",", ":"))


def _dispatch_rows() -> List[Dict[str, Any]]:
    if "dispatch" in OUTPUT_SECTION_TO_SHEET and OUTPUT_SECTION_TO_SHEET["dispatch"] in SHEETS:
        artifact = _get_active_run_artifact()
        if artifact:
            jobs = _schedule_rows()
        else:
            try:
                return _sheet_items(OUTPUT_SECTION_TO_SHEET["dispatch"])
            except Exception:
                jobs = _schedule_rows()
        if not artifact:
            jobs = _schedule_rows()
    else:
        jobs = _schedule_rows()

    grouped: Dict[str, Dict[str, Any]] = {}
    for job in jobs:
        rid = str(job.get("Resource_ID") or job.get("resource_id") or "UNKNOWN")
        bucket = grouped.setdefault(rid, {"resource_id": rid, "jobs": [], "job_count": 0, "total_mt": 0.0})
        bucket["jobs"].append(job)
        bucket["job_count"] += 1
        try:
            bucket["total_mt"] += float(job.get("Qty_MT") or 0)
        except Exception:
            pass
    return list(grouped.values())


def _dashboard_payload() -> Dict[str, Any]:
    campaigns = _campaign_rows()
    capacity = _capacity_rows()
    material_data = _material_plan_payload()
    solver_meta = _active_run_solver_metadata()

    total_mt = sum(float(x.get("total_mt") or 0) for x in campaigns)
    total_heats = sum(float(x.get("heats") or 0) for x in campaigns)
    released = sum(1 for x in campaigns if str(x.get("release_status") or "").upper() == "RELEASED")
    held = sum(1 for x in campaigns if "HOLD" in str(x.get("release_status") or "").upper())
    late = sum(1 for x in campaigns if str(x.get("Status") or "").upper() == "LATE")
    released_campaigns = [c for c in campaigns if str(c.get("release_status") or "").upper() == "RELEASED"]
    released_heats = sum(float(x.get("heats") or 0) for x in released_campaigns)
    released_mt = round(sum(float(x.get("total_mt") or 0) for x in released_campaigns), 1)

    # Fall back to in-memory planning orders when workbook campaigns are absent
    _in_mem_pos = list(getattr(aps_planning_orders_propose, "_planning_orders", []) or [])
    if not campaigns and _in_mem_pos:
        total_mt = round(sum(float(po.get("total_qty_mt") or 0) for po in _in_mem_pos), 1)
        total_heats = sum(int(po.get("heats_required") or 0) for po in _in_mem_pos)
        released = sum(1 for po in _in_mem_pos if str(po.get("planner_status") or "").upper() == "RELEASED")
        released_mt = round(sum(float(po.get("total_qty_mt") or 0) for po in _in_mem_pos if str(po.get("planner_status") or "").upper() == "RELEASED"), 1)
        released_heats = sum(int(po.get("heats_required") or 0) for po in _in_mem_pos if str(po.get("planner_status") or "").upper() == "RELEASED")

    bottleneck = None
    if capacity:
        bottleneck = max(capacity, key=lambda x: float(x.get("Utilisation_%") or 0))

    alerts = []
    for camp in material_data.get("campaigns", []):
        shortage_qty = float(camp.get("shortage_qty", 0.0) or 0.0)
        if shortage_qty > 1e-9:
            alerts.append({
                "campaign_id": camp.get("campaign_id"),
                "shortage_qty": round(shortage_qty, 2),
                "severity": "HIGH",
                "material_status": camp.get("material_status"),
            })
        elif str(camp.get("material_status", "")).upper() not in {"", "OK", "READY"}:
            alerts.append({
                "campaign_id": camp.get("campaign_id"),
                "shortage_qty": 0.0,
                "severity": "MEDIUM",
                "material_status": camp.get("material_status"),
            })

    return {
        "solver_status": solver_meta.get("solver_status", "WORKBOOK"),
        "solver_detail": solver_meta.get("solver_detail", ""),
        "degraded_flags": solver_meta.get("degraded_flags", {}),
        "run_id": solver_meta.get("run_id"),
        "active_run_created_at": solver_meta.get("created_at"),
        "warnings": solver_meta.get("warnings", []),
        "last_run": _state.get("last_run"),
        "campaigns_total": len(campaigns),
        "campaigns_released": released,
        "campaigns_held": held,
        "campaigns_late": late,
        "total_heats": total_heats,
        "total_mt": round(total_mt, 1),
        "released_heats": released_heats,
        "released_mt": released_mt,
        "on_time_pct": round(100 * max(0, len(campaigns) - late) / max(len(campaigns), 1), 1) if campaigns else 0.0,
        "throughput_mt_day": round(total_mt / 14, 1) if total_mt else 0.0,
        "bottleneck": bottleneck.get("Resource_ID") if bottleneck else None,
        "max_utilisation": bottleneck.get("Utilisation_%") if bottleneck else None,
        "shortage_alerts": alerts,
        "utilisation": [
            {
                "resource_id": r.get("Resource_ID"),
                "resource_name": r.get("Resource_Name", r.get("Resource_ID")),
                "utilisation": r.get("Utilisation_%"),
                "demand_hrs": r.get("Demand_Hrs"),
                "avail_hrs": r.get("Avail_Hrs_14d"),
                "status": r.get("Status"),
                "operation": r.get("Operation_Group", ""),
            }
            for r in capacity
        ],
        "material": material_data.get("summary", {}),
    }


def _masterdata_payload() -> Dict[str, Any]:
    return {section: _sheet_items(api_name) for section, api_name in MASTERDATA_SECTION_TO_SHEET.items() if api_name in SHEETS}


def _section_sheet(section: str) -> Optional[str]:
    return MASTERDATA_SECTION_TO_SHEET.get(section)


def _output_sheet(section: str) -> Optional[str]:
    return OUTPUT_SECTION_TO_SHEET.get(section)


def _validate_workbook_schema() -> Dict[str, Any]:
    result = {
        "valid": True,
        "missing_sheets": [],
        "missing_columns": {},
        "sheet_details": {},
        "required_sheets": list(MASTERDATA_SECTION_TO_SHEET.values()),
    }

    required_columns = {
        "Config": ["Key"],
        "Sales_Orders": ["SO_ID", "SKU_ID"],
        "BOM": ["Parent_SKU", "Child_SKU"],
        "Inventory": ["SKU_ID"],
        "Resource_Master": ["Resource_ID"],
        "Routing": ["SKU_ID", "Operation"],
        "SKU_Master": ["SKU_ID"],
    }

    try:
        import openpyxl
        wb = openpyxl.load_workbook(WORKBOOK, read_only=True, data_only=False)
        sheet_names = set(wb.sheetnames)
        wb.close()
    except Exception as e:
        result["valid"] = False
        result["schema_error"] = f"Failed to read workbook sheets: {str(e)}"
        return result

    for sheet_name in required_columns.keys():
        if sheet_name not in sheet_names:
            result["missing_sheets"].append(sheet_name)
            result["valid"] = False

    for sheet_name, cols in required_columns.items():
        if sheet_name not in sheet_names:
            continue
        try:
            df = _read_sheet(sheet_name)
            result["sheet_details"][sheet_name] = {
                "exists": True,
                "row_count": len(df),
                "columns": list(df.columns),
            }
            missing_cols = [col for col in cols if col not in df.columns]
            if missing_cols:
                result["missing_columns"][sheet_name] = missing_cols
                result["valid"] = False
        except Exception as e:
            result["sheet_details"][sheet_name] = {"exists": False, "error": str(e)}
            result["valid"] = False
    return result


@app.route('/api/health')
def health():
    quick = str(request.args.get("quick") or "").strip().lower() in {"1", "true", "yes", "y"}
    exists = WORKBOOK.exists()
    mtime = datetime.fromtimestamp(WORKBOOK.stat().st_mtime).isoformat() if exists else None

    workbook_ok = False
    workbook_error = None
    schema_validation = {}
    if exists:
        try:
            if quick:
                workbook_ok = zipfile.is_zipfile(WORKBOOK)
                schema_validation = {
                    "valid": workbook_ok,
                    "mode": "quick",
                    "zip_container_ok": workbook_ok,
                }
                if not workbook_ok:
                    workbook_error = "Workbook is not a valid .xlsx zip container"
            else:
                import openpyxl
                wb = openpyxl.load_workbook(WORKBOOK, read_only=True, data_only=True)
                wb.close()
                workbook_ok = True
                schema_validation = _validate_workbook_schema()
        except Exception as e:
            workbook_error = str(e)

    active_run = _get_active_run_artifact()
    active_run_campaign_count = 0
    active_run_schedule_row_count = 0
    active_run_capacity_row_count = 0
    artifact_backed_bom_exists = False
    artifact_backed_material_exists = False
    if active_run:
        active_run_campaign_count = len(active_run.get("results", {}).get("campaigns", []))
        active_run_schedule_row_count = len(active_run.get("results", {}).get("heat_schedule", []))
        active_run_capacity_row_count = len(active_run.get("results", {}).get("capacity", []))
        artifact_backed_bom_exists = bool(active_run.get("results", {}).get("bom_explosion") is not None)
        artifact_backed_material_exists = bool(active_run.get("results", {}).get("material_plan") is not None)

    return _jsonify({
        "ok": exists and workbook_ok and schema_validation.get("valid", True),
        "api": "up",
        "workbook": str(WORKBOOK),
        "workbook_exists": exists,
        "workbook_ok": workbook_ok,
        "workbook_mtime": mtime,
        "workbook_error": workbook_error,
        "schema": schema_validation,
        "active_run_exists": bool(active_run),
        "active_run_id": active_run.get("run_id") if active_run else None,
        "active_run_created_at": active_run.get("created_at") if active_run else None,
        "active_run_campaign_count": active_run_campaign_count,
        "active_run_schedule_row_count": active_run_schedule_row_count,
        "active_run_capacity_row_count": active_run_capacity_row_count,
        "artifact_backed_reads": bool(active_run),
        "artifact_backed_bom_exists": artifact_backed_bom_exists,
        "artifact_backed_material_exists": artifact_backed_material_exists,
        "active_run_degraded_flags": active_run.get("degraded_flags") if active_run else None,
        "last_run": _state.get("last_run"),
        "run_id": _current_trace_id(),
        "solver_status": _active_run_solver_metadata().get("solver_status", "NOT RUN"),
        "solver_detail": _active_run_solver_metadata().get("solver_detail", ""),
    }, 200 if exists and workbook_ok else 500)


# ============================================================================
# Configuration Management API
# ============================================================================

@app.route('/api/config/algorithm', methods=['GET'])
def get_algorithm_config():
    """Get all algorithm configuration parameters."""
    config = get_config(WORKBOOK, reload=True)
    params = config.all_params()
    metadata = config.metadata

    result = []
    for key, value in params.items():
        meta = metadata.get(key, {})
        result.append({
            'key': key,
            'value': value,
            'category': meta.get('category'),
            'data_type': meta.get('data_type'),
            'description': meta.get('description'),
            'min': meta.get('min'),
            'max': meta.get('max'),
        })

    return _jsonify({
        'status': 'success',
        'parameters': sorted(result, key=lambda x: x['category']),
        'total': len(result),
        'by_category': {
            cat: len([p for p in result if p['category'] == cat])
            for cat in set(p['category'] for p in result)
        }
    })


@app.route('/api/config/algorithm/<key>', methods=['GET'])
def get_algorithm_config_param(key):
    """Get single algorithm parameter."""
    config = get_config(WORKBOOK, reload=True)
    canonical_key = canonicalize_config_key(key)
    value = config.get(canonical_key)

    if value is None:
        return _jsonify({
            'status': 'error',
            'error': f'Parameter {key} not found'
        }, 404)
    meta = config.metadata.get(canonical_key, {})
    return _jsonify({
        'status': 'success',
        'key': canonical_key,
        'value': value,
        'category': meta.get('category'),
        'data_type': meta.get('data_type'),
        'description': meta.get('description'),
        'min': meta.get('min'),
        'max': meta.get('max'),
    })


@app.route('/api/config/algorithm/category/<category>', methods=['GET'])
def get_algorithm_config_by_category(category):
    """Get all parameters in a category."""
    config = get_config(WORKBOOK, reload=True)
    params = config.params_by_category(category.upper())

    if not params:
        return _jsonify({
            'status': 'error',
            'error': f'Category {category} not found or has no parameters'
        }, 404)

    metadata = config.metadata
    result = []
    for key, value in params.items():
        meta = metadata.get(key, {})
        result.append({
            'key': key,
            'value': value,
            'data_type': meta.get('data_type'),
            'description': meta.get('description'),
            'min': meta.get('min'),
            'max': meta.get('max'),
        })

    return _jsonify({
        'status': 'success',
        'category': category.upper(),
        'parameters': result,
        'total': len(result),
    })


@app.route('/api/config/algorithm/<key>', methods=['PUT'])
def update_algorithm_config_param(key):
    """Update a single algorithm parameter."""
    try:
        data = request.get_json() or {}
        new_value = data.get('value')
        user = data.get('user', 'API')
        reason = data.get('reason', '')

        if new_value is None:
            return _jsonify({
                'status': 'error',
                'error': 'Missing required field: value'
            }, 400)

        updated = update_algorithm_config_in_workbook(
            WORKBOOK,
            key,
            new_value,
            user=user,
            reason=reason,
        )
        return _jsonify({
            'status': 'success',
            'key': updated['key'],
            'old_value': updated['old_value'],
            'new_value': updated['new_value'],
            'user': user,
            'reason': reason,
            'timestamp': datetime.now().isoformat(),
        })
    except KeyError as e:
        return _jsonify({
            'status': 'error',
            'error': str(e)
        }, 404)
    except PermissionError as e:
        return _jsonify({
            'status': 'error',
            'error': str(e)
        }, 409)
    except ValueError as e:
        return _jsonify({
            'status': 'error',
            'error': str(e)
        }, 400)
    except Exception as e:
        return _jsonify({
            'status': 'error',
            'error': str(e)
        }, 500)


@app.route('/api/config/algorithm/validate', methods=['POST'])
def validate_algorithm_config():
    """Pre-validate configuration changes before committing."""
    try:
        data = request.get_json() or {}
        changes = data.get('changes', {})

        if not changes:
            return _jsonify({
                'status': 'error',
                'error': 'Missing required field: changes'
            }, 400)

        config = get_config(WORKBOOK, reload=True)
        metadata = config.metadata

        validations = {}
        all_valid = True

        for key, new_value in changes.items():
            canonical_key = canonicalize_config_key(key)
            current = config.get(canonical_key)
            meta = metadata.get(canonical_key, {})

            if current is None:
                validations[key] = {
                    'valid': False,
                    'error': f'Parameter {key} not found'
                }
                all_valid = False
                continue

            is_valid = config._validate_value(
                canonical_key, config._convert_value(new_value, meta.get('data_type')), meta.get('data_type'),
                meta.get('min'), meta.get('max')
            )

            validations[key] = {
                'valid': is_valid,
                'current_value': current,
                'proposed_value': new_value,
                'data_type': meta.get('data_type'),
                'min': meta.get('min'),
                'max': meta.get('max'),
            }

            if not is_valid:
                all_valid = False

        return _jsonify({
            'status': 'success',
            'all_valid': all_valid,
            'validations': validations,
        })

    except Exception as e:
        return _jsonify({
            'status': 'error',
            'error': str(e)
        }, 500)


@app.route('/api/config/algorithm/export', methods=['POST'])
def export_algorithm_config():
    """Export current configuration to CSV."""
    try:
        config = get_config(WORKBOOK, reload=True)
        params = config.all_params()
        metadata = config.metadata

        # Build CSV rows
        csv_lines = ['Config_Key,Category,Current_Value,Data_Type,Description']
        for key, value in params.items():
            meta = metadata.get(key, {})
            cat = meta.get('category', '')
            dtype = meta.get('data_type', '')
            desc = meta.get('description', '').replace(',', ';')  # Escape commas in description

            csv_lines.append(f'{key},{cat},{value},{dtype},"{desc}"')

        csv_content = '\n'.join(csv_lines)

        return _jsonify({
            'status': 'success',
            'timestamp': datetime.now().isoformat(),
            'parameters': len(params),
            'csv': csv_content,
        })

    except Exception as e:
        return _jsonify({
            'status': 'error',
            'error': str(e)
        }, 500)


@app.route('/api/data/dashboard')
def dashboard():
    return _jsonify(_dashboard_payload())


@app.route('/api/data/config')
def data_config():
    md = _masterdata_payload()
    return _jsonify({
        "config": md.get("config", []),
        "resources": md.get("resources", []),
        "routing": md.get("routing", []),
        "queue": md.get("queue", []),
        "skus": md.get("skus", []),
        "bom": md.get("bom", []),
        "inventory": md.get("inventory", []),
        "campaign_config": md.get("campaign-config", []),
        "changeover": md.get("changeover", []),
        "scenarios": md.get("scenarios", []),
    })


@app.route('/api/data/orders')
def data_orders():
    return _jsonify({"orders": _sheet_items('sales-orders')})


@app.route('/api/data/skus')
def data_skus():
    d = _load_all()
    skus = d['skus']
    fg = skus[skus.get('Category', pd.Series('')).astype(str).str.contains('Finished', case=False, na=False)]
    return _jsonify({"skus": [{"sku_id": r.get('SKU_ID', ''), "sku_name": r.get('SKU_Name', '')} for _, r in fg.iterrows()]})


@app.route('/api/data/campaigns')
def data_campaigns():
    return _jsonify({"run_id": _current_trace_id(), "campaigns": _campaign_rows()})


@app.route('/api/data/gantt')
def data_gantt():
    return _jsonify({"run_id": _current_trace_id(), "jobs": _schedule_rows()})


@app.route('/api/data/capacity')
def data_capacity():
    return _jsonify({"run_id": _current_trace_id(), "capacity": _capacity_rows()})


def _build_bom_grouped_view(
    netted_df: pd.DataFrame | None,
    skus_df: pd.DataFrame | None = None,
    bom_df: pd.DataFrame | None = None,
) -> tuple[list, dict]:
    """Build grouped BOM view (Plant → Material_Type hierarchy) from netted DataFrame.

    Returns:
        (grouped_bom: list of plant dicts, summary: dict with overall stats)
    """
    if netted_df is None or netted_df.empty:
        return [], {
            "total_sku_lines": 0,
            "short_lines": 0,
            "partial_lines": 0,
            "covered_lines": 0,
            "byproduct_lines": 0,
            "total_gross_req": 0.0,
            "total_net_req": 0.0,
            "total_covered_by_stock": 0.0,
            "total_produced_qty": 0.0,
        }

    # Build lookup maps for enrichment
    sku_name_lookup = {}
    sku_category_lookup = {}
    if skus_df is not None and not skus_df.empty:
        for _, row in skus_df.iterrows():
            sku_id = str(row.get("SKU_ID", "")).strip()
            if sku_id:
                sku_name_lookup[sku_id] = str(row.get("SKU_Name", sku_id)).strip()
                sku_category_lookup[sku_id] = str(row.get("Category", "")).strip()

    parent_sku_lookup = {}
    if bom_df is not None and not bom_df.empty:
        for _, row in bom_df.iterrows():
            child = str(row.get("Child_SKU", "")).strip()
            parent = str(row.get("Parent_SKU", "")).strip()
            if child and parent:
                if child in parent_sku_lookup:
                    if parent not in parent_sku_lookup[child]:
                        parent_sku_lookup[child] += f", {parent}"
                else:
                    parent_sku_lookup[child] = parent

    # Enrich each row
    enriched_rows = []
    for _, row in netted_df.iterrows():
        sku_id = str(row.get("SKU_ID", "")).strip()
        category = sku_category_lookup.get(sku_id, "")
        material_type = _mat_type_for_sku(sku_id, category)
        stage = _stage_for_sku(sku_id, material_type)
        plant = _plant_for_sku(sku_id, category, material_type)

        gross_req = float(row.get("Gross_Req", row.get("Required_Qty", 0)) or 0)
        net_req = float(row.get("Net_Req", 0) or 0)
        flow_type = str(row.get("Flow_Type", "INPUT")).strip().upper()
        available_before = float(row.get("Available", row.get("Available_Before", 0)) or 0)
        produced_qty = float(row.get("Produced_Qty", 0) or 0)

        # Determine if byproduct
        is_byproduct = flow_type in {"BYPRODUCT", "OUTPUT", "CO_PRODUCT", "COPRODUCT", "WASTE"}
        covered_by_stock = 0.0 if is_byproduct else max(gross_req - net_req, 0.0)

        # Derive status
        if is_byproduct:
            status = "BYPRODUCT"
        elif net_req <= 1e-9:
            status = "COVERED"
        elif available_before > 1e-9:
            status = "PARTIAL SHORT"
        else:
            status = "SHORT"

        enriched_rows.append({
            "plant": plant,
            "stage": stage,
            "material_type": material_type,
            "material_category": category,
            "parent_skus": parent_sku_lookup.get(sku_id, ""),
            "sku_id": sku_id,
            "sku_name": sku_name_lookup.get(sku_id, sku_id),
            "bom_level": int(row.get("BOM_Level", 0) or 0),
            "gross_req": round(gross_req, 3),
            "available_before": round(available_before, 3),
            "covered_by_stock": round(covered_by_stock, 3),
            "produced_qty": round(produced_qty, 3),
            "net_req": round(net_req, 3),
            "status": status,
            "flow_type": flow_type,
        })

    # Sort by plant, material_type, bom_level, sku_id
    enriched_rows.sort(
        key=lambda r: (
            _PLANT_SORT_ORDER.get(r["plant"], 99),
            _MAT_TYPE_SORT_ORDER.get(r["material_type"], 99),
            r["bom_level"],
            r["sku_id"],
        )
    )

    # Group into plant -> material_type -> rows
    grouped_bom = []
    plant_groups = {}
    for row in enriched_rows:
        plant = row["plant"]
        mat_type = row["material_type"]
        if plant not in plant_groups:
            plant_groups[plant] = {}
        if mat_type not in plant_groups[plant]:
            plant_groups[plant][mat_type] = []
        plant_groups[plant][mat_type].append(row)

    # Build grouped output
    for plant in sorted(plant_groups.keys(), key=lambda p: _PLANT_SORT_ORDER.get(p, 99)):
        plant_rows_all = []
        material_types = []

        for mat_type in sorted(
            plant_groups[plant].keys(), key=lambda m: _MAT_TYPE_SORT_ORDER.get(m, 99)
        ):
            rows = plant_groups[plant][mat_type]
            gross = sum(r["gross_req"] for r in rows)
            produced = sum(r["produced_qty"] for r in rows)
            net = sum(r["net_req"] for r in rows)

            material_types.append({
                "material_type": mat_type,
                "gross_req": round(gross, 3),
                "produced_qty": round(produced, 3),
                "net_req": round(net, 3),
                "row_count": len(rows),
                "rows": rows,
            })
            plant_rows_all.extend(rows)

        plant_gross = sum(r["gross_req"] for r in plant_rows_all)
        plant_produced = sum(r["produced_qty"] for r in plant_rows_all)
        plant_net = sum(r["net_req"] for r in plant_rows_all)

        grouped_bom.append({
            "plant": plant,
            "gross_req": round(plant_gross, 3),
            "produced_qty": round(plant_produced, 3),
            "net_req": round(plant_net, 3),
            "row_count": len(plant_rows_all),
            "material_types": material_types,
        })

    # Compute summary
    total_short = sum(1 for r in enriched_rows if r["status"] == "SHORT")
    total_partial = sum(1 for r in enriched_rows if r["status"] == "PARTIAL SHORT")
    total_covered = sum(1 for r in enriched_rows if r["status"] == "COVERED")
    total_byproduct = sum(1 for r in enriched_rows if r["status"] == "BYPRODUCT")
    total_gross = sum(r["gross_req"] for r in enriched_rows)
    total_net = sum(r["net_req"] for r in enriched_rows)
    total_covered_by_stock = sum(r["covered_by_stock"] for r in enriched_rows)
    total_produced = sum(r["produced_qty"] for r in enriched_rows)

    summary = {
        "total_sku_lines": len(enriched_rows),
        "short_lines": total_short,
        "partial_lines": total_partial,
        "covered_lines": total_covered,
        "byproduct_lines": total_byproduct,
        "total_gross_req": round(total_gross, 3),
        "total_net_req": round(total_net, 3),
        "total_covered_by_stock": round(total_covered_by_stock, 3),
        "total_produced_qty": round(total_produced, 3),
    }

    return grouped_bom, summary


@app.route('/api/run/bom', methods=['POST'])
def run_bom_api():
    try:
        d = _load_all()
        demand = consolidate_demand(d['sales_orders'])
        din = demand[['SKU_ID', 'Total_Qty']].rename(columns={'Total_Qty': 'Required_Qty'})
        explosion_result = explode_bom_details(din, d['bom'])
        gross_bom_df = explosion_result.get('exploded', pd.DataFrame())
        structure_errors = explosion_result.get('structure_errors', [])
        feasible = explosion_result.get('feasible', True)

        if not feasible and structure_errors:
            error_summary = "; ".join(f"{e.get('type')}: {e.get('path')}" for e in structure_errors[:3])
            return _error_response(
                error_message="BOM structure validation failed",
                error_code="BOM_STRUCTURE_ERROR",
                error_domain="BOM",
                details=error_summary,
                degraded_mode=False,
                status_code=400,
            )[0]

        netted = net_requirements(gross_bom_df, d['inventory'])

        # Build grouped view
        grouped_bom_data, bom_summary = _build_bom_grouped_view(
            netted, d.get("skus"), d.get("bom")
        )

        # Build BOM payload (additive: includes grouped_bom + summary)
        bom_payload = {
            "gross_bom": _df_to_records(gross_bom_df),
            "net_bom": _df_to_records(netted),
            "grouped_bom": grouped_bom_data,
            "summary": bom_summary,
            "structure_errors": structure_errors,
            "feasible": feasible,
            "rows": len(netted),
            "run_id": _current_trace_id(),
        }

        # Persist to active run artifact
        artifact = _get_active_run_artifact()
        if artifact is not None:
            artifact.setdefault("results", {})["bom_explosion"] = bom_payload

        # Mirror to compatibility cache
        _state["bom_explosion"] = bom_payload

        return _jsonify(bom_payload)
    except Exception as e:
        return _error_response(
            error_message=f"BOM explosion failed: {str(e)}",
            error_code="BOM_EXPLOSION_ERROR",
            error_domain="BOM",
            details=traceback.format_exc(),
            status_code=500,
        )[0]


def _calculate_material_plan(
    campaigns: List[Dict],
    detail_level: str = "campaign",
    run_id: str | None = None,
    skus: list | None = None,
) -> Dict[str, Any]:
    """
    Build a planner-friendly material summary per campaign.

    Important design rules:
    1. One material row per SKU per campaign.
    2. Do not emit duplicate semantic rows for the same SKU.
    3. Keep quantities explicit:
       - gross_required_qty
       - available_before
       - inventory_covered_qty
       - make_convert_qty
       - consumed_qty
       - shortage_qty
       - remaining_after
    4. Keep grouped plant buckets for UI summary.
    5. Keep backward compatibility fields where practical, but make them
       deterministic and non-duplicative.
    """
    try:
        sku_name_lookup: Dict[str, str] = {}
        sku_category_lookup: Dict[str, str] = {}

        if skus:
            try:
                if isinstance(skus, pd.DataFrame):
                    skus_iter = skus.iterrows()
                else:
                    skus_iter = enumerate(skus)

                for _, row in skus_iter:
                    if isinstance(row, pd.Series):
                        sku_id = str(row.get("SKU_ID", "")).strip()
                        sku_name = str(row.get("SKU_Name", sku_id)).strip()
                        sku_cat = str(row.get("Category", "")).strip()
                    else:
                        sku_id = str(row.get("SKU_ID", "")).strip()
                        sku_name = str(row.get("SKU_Name", sku_id)).strip()
                        sku_cat = str(row.get("Category", "")).strip()

                    if sku_id:
                        sku_name_lookup[sku_id] = sku_name or sku_id
                        sku_category_lookup[sku_id] = sku_cat
            except Exception:
                pass

        campaigns_data: List[Dict[str, Any]] = []

        for camp in campaigns:
            camp_id = str(camp.get("campaign_id") or camp.get("Campaign_ID") or "").strip()
            if not camp_id:
                continue

            grade = str(camp.get("grade") or camp.get("Grade") or "").strip()
            release_status = str(camp.get("release_status") or camp.get("Release_Status") or "OPEN").strip()
            req_qty = float(camp.get("total_coil_mt") or camp.get("total_mt") or camp.get("Total_MT") or 0.0)
            material_status = str(camp.get("material_status") or "UNREVIEWED").strip()

            shortages = dict(camp.get("material_shortages") or {})
            material_consumed = dict(camp.get("material_consumed") or {})
            material_gross_requirements = dict(camp.get("material_gross_requirements") or {})
            structure_errors = list(camp.get("material_structure_errors") or [])
            inventory_before = dict(camp.get("inventory_before") or {})
            inventory_after = dict(camp.get("inventory_after") or {})

            shortage_qty = round(sum(float(v or 0.0) for v in shortages.values()), 3)
            feasible = not structure_errors and shortage_qty <= 1e-9

            # One canonical row per SKU
            all_skus = sorted(
                set(material_gross_requirements.keys())
                | set(material_consumed.keys())
                | set(shortages.keys())
                | set(inventory_before.keys())
                | set(inventory_after.keys())
            )

            canonical_rows: List[Dict[str, Any]] = []
            material_rows_by_plant: Dict[str, List[Dict[str, Any]]] = {}

            for sku_id in all_skus:
                gross_required_qty = float(material_gross_requirements.get(sku_id, 0.0) or 0.0)
                consumed_qty = float(material_consumed.get(sku_id, 0.0) or 0.0)
                shortage_qty_sku = float(shortages.get(sku_id, 0.0) or 0.0)
                available_before = float(inventory_before.get(sku_id, 0.0) or 0.0)
                remaining_after = float(inventory_after.get(sku_id, 0.0) or 0.0)

                category = sku_category_lookup.get(sku_id, "")
                material_type = _mat_type_for_sku(
                    sku_id=sku_id,
                    category=category,
                    sku_name=sku_name_lookup.get(sku_id, sku_id),
                )
                plant = _plant_for_sku(
                    sku_id=sku_id,
                    category=category,
                    material_type=material_type,
                )
                stage = _stage_for_sku(sku_id, material_type)

                inventory_covered_qty = max(min(gross_required_qty, available_before), 0.0)
                make_convert_qty = max(gross_required_qty - inventory_covered_qty - shortage_qty_sku, 0.0)

                # Status precedence
                if shortage_qty_sku > 1e-6:
                    status = "SHORTAGE"
                elif make_convert_qty > 1e-6 and inventory_covered_qty > 1e-6:
                    status = "PARTIAL COVER"
                elif inventory_covered_qty > 1e-6 and make_convert_qty <= 1e-6:
                    status = "DRAWN FROM STOCK"
                elif make_convert_qty > 1e-6:
                    status = "MAKE / CONVERT"
                else:
                    status = "COVERED"

                row = {
                    "material_sku": sku_id,
                    "material_name": sku_name_lookup.get(sku_id, sku_id),
                    "material_category": category,
                    "material_type": material_type,
                    "plant": plant,
                    "stage": stage,
                    "gross_required_qty": round(gross_required_qty, 3),
                    "required_qty": round(gross_required_qty, 3),  # backward-compatible alias
                    "available_before": round(available_before, 3),
                    "inventory_covered_qty": round(inventory_covered_qty, 3),
                    "make_convert_qty": round(make_convert_qty, 3),
                    "consumed_qty": round(consumed_qty, 3),
                    "consumed": round(consumed_qty, 3),  # backward-compatible alias
                    "shortage_qty": round(shortage_qty_sku, 3),
                    "remaining_after": round(remaining_after, 3),
                    "status": status,
                }
                canonical_rows.append(row)

                if plant not in material_rows_by_plant:
                    material_rows_by_plant[plant] = []
                material_rows_by_plant[plant].append(row)

            # Sort plant buckets for stable UI
            plants: List[Dict[str, Any]] = []
            for plant_name in sorted(material_rows_by_plant.keys(), key=lambda p: _PLANT_SORT_ORDER.get(p, 99)):
                plant_rows = sorted(
                    material_rows_by_plant[plant_name],
                    key=lambda r: (
                        _MAT_TYPE_SORT_ORDER.get(r.get("material_type", "Material"), 99),
                        str(r.get("material_sku", "")),
                    ),
                )

                plant_required = sum(float(r.get("gross_required_qty", 0.0) or 0.0) for r in plant_rows)
                plant_inventory_covered = sum(float(r.get("inventory_covered_qty", 0.0) or 0.0) for r in plant_rows)
                plant_make_convert = sum(float(r.get("make_convert_qty", 0.0) or 0.0) for r in plant_rows)
                plant_shortage = sum(float(r.get("shortage_qty", 0.0) or 0.0) for r in plant_rows)

                plants.append({
                    "plant": plant_name,
                    "required_qty": round(plant_required, 3),
                    "inventory_covered_qty": round(plant_inventory_covered, 3),
                    "make_convert_qty": round(plant_make_convert, 3),
                    "shortage_qty": round(plant_shortage, 3),
                    "rows": plant_rows,
                })

            total_inventory_covered_qty = sum(float(r.get("inventory_covered_qty", 0.0) or 0.0) for r in canonical_rows)
            total_make_convert_qty = sum(float(r.get("make_convert_qty", 0.0) or 0.0) for r in canonical_rows)

            # Keep backward compatibility, but do not duplicate the same SKU into multiple semantic rows
            detail_rows = [
                {
                    "sku_id": r["material_sku"],
                    "material_name": r["material_name"],
                    "material_type": r["material_type"],
                    "plant": r["plant"],
                    "required_qty": r["gross_required_qty"],
                    "available_before": r["available_before"],
                    "inventory_covered_qty": r["inventory_covered_qty"],
                    "make_convert_qty": r["make_convert_qty"],
                    "consumed_qty": r["consumed_qty"],
                    "shortage_qty": r["shortage_qty"],
                    "remaining_after": r["remaining_after"],
                    "status": r["status"],
                    "type": r["status"],  # backward-compatible loose field
                }
                for r in canonical_rows
            ]

            camp_data = {
                "campaign_id": camp_id,
                "grade": grade,
                "release_status": release_status,
                "required_qty": round(req_qty, 3),
                "shortage_qty": round(shortage_qty, 3),
                "inventory_covered_qty": round(total_inventory_covered_qty, 3),
                "make_convert_qty": round(total_make_convert_qty, 3),
                "material_status": material_status,
                "feasible": feasible,
                "material_structure_errors": structure_errors,
                "material_shortages": {k: round(float(v or 0.0), 3) for k, v in shortages.items()},
                "material_consumed": {k: round(float(v or 0.0), 3) for k, v in material_consumed.items()},
                "material_gross_requirements": {k: round(float(v or 0.0), 3) for k, v in material_gross_requirements.items()},
                "inventory_before": {k: round(float(v or 0.0), 3) for k, v in inventory_before.items()},
                "inventory_after": {k: round(float(v or 0.0), 3) for k, v in inventory_after.items()},
                "materials": canonical_rows,
                "plants": plants,
                "detail_rows": detail_rows,
            }
            campaigns_data.append(camp_data)

        total_required = sum(float(c.get("required_qty", 0.0) or 0.0) for c in campaigns_data)
        total_shortages = sum(float(c.get("shortage_qty", 0.0) or 0.0) for c in campaigns_data)
        total_inventory_covered = sum(float(c.get("inventory_covered_qty", 0.0) or 0.0) for c in campaigns_data)
        total_make_convert = sum(float(c.get("make_convert_qty", 0.0) or 0.0) for c in campaigns_data)

        released_count = sum(1 for c in campaigns_data if "RELEASED" in str(c.get("release_status", "")).upper())
        held_count = sum(1 for c in campaigns_data if "HOLD" in str(c.get("release_status", "")).upper())
        shortage_count = sum(1 for c in campaigns_data if float(c.get("shortage_qty", 0.0) or 0.0) > 1e-6)
        error_count = sum(1 for c in campaigns_data if not bool(c.get("feasible", True)))

        summary = {
            "Campaigns": len(campaigns_data),
            "Released": released_count,
            "Held": held_count,
            "Shortage Lines": shortage_count,
            "BOM Structure Errors": error_count,
            "Total Required Qty": round(total_required, 3),
            "Inventory Covered Qty": round(total_inventory_covered, 3),
            "Shortage Qty": round(total_shortages, 3),
            "Make / Convert Qty": round(total_make_convert, 3),
        }

        return {
            "campaigns": campaigns_data,
            "detail_level": detail_level,
            "summary": summary,
            "run_id": run_id,
        }

    except Exception:
        return {
            "campaigns": [],
            "detail_level": detail_level,
            "summary": {},
            "run_id": run_id,
        }


def _normalize_material_detail_level(value: Any) -> str:
    level = str(value or "campaign").strip().lower()
    if level in {"campaign", "campaigns"}:
        return "campaign"
    if level in {"po", "planning_order", "planning-order", "planning_orders", "planning-orders"}:
        return "po"
    if level in {"heat", "heats"}:
        return "heat"
    return "campaign"


def _material_float(value: Any, default: float = 0.0) -> float:
    try:
        parsed = pd.to_numeric(pd.Series([value]), errors="coerce").fillna(default).iloc[0]
        return float(parsed)
    except Exception:
        return float(default)


def _material_rows_from_entity(entity: Dict[str, Any]) -> List[Dict[str, Any]]:
    rows = list(entity.get("materials") or [])
    if rows:
        normalized = []
        for row in rows:
            sku_id = str(row.get("material_sku") or row.get("sku_id") or "").strip()
            if not sku_id:
                continue
            gross_required_qty = _material_float(row.get("gross_required_qty") or row.get("required_qty") or 0.0)
            available_before = _material_float(row.get("available_before") or 0.0)
            inventory_covered_qty = _material_float(
                row.get("inventory_covered_qty")
                if row.get("inventory_covered_qty") is not None
                else max(min(gross_required_qty, available_before), 0.0)
            )
            shortage_qty = _material_float(row.get("shortage_qty") or 0.0)
            make_convert_qty = _material_float(
                row.get("make_convert_qty")
                if row.get("make_convert_qty") is not None
                else max(gross_required_qty - inventory_covered_qty - shortage_qty, 0.0)
            )
            consumed_qty = _material_float(row.get("consumed_qty") or row.get("consumed") or 0.0)
            remaining_after = _material_float(row.get("remaining_after") or 0.0)
            material_type = str(row.get("material_type") or "Material")
            plant = str(row.get("plant") or "Unassigned")
            status = str(row.get("status") or row.get("type") or "COVERED")

            normalized.append(
                {
                    "material_sku": sku_id,
                    "material_name": str(row.get("material_name") or sku_id),
                    "material_category": str(row.get("material_category") or ""),
                    "material_type": material_type,
                    "plant": plant,
                    "stage": str(row.get("stage") or _stage_for_sku(sku_id, material_type)),
                    "gross_required_qty": round(gross_required_qty, 3),
                    "required_qty": round(gross_required_qty, 3),
                    "available_before": round(available_before, 3),
                    "inventory_covered_qty": round(inventory_covered_qty, 3),
                    "make_convert_qty": round(make_convert_qty, 3),
                    "consumed_qty": round(consumed_qty, 3),
                    "consumed": round(consumed_qty, 3),
                    "shortage_qty": round(shortage_qty, 3),
                    "remaining_after": round(remaining_after, 3),
                    "status": status,
                }
            )
        return normalized

    detail_rows = list(entity.get("detail_rows") or [])
    normalized = []
    for row in detail_rows:
        sku_id = str(row.get("sku_id") or row.get("material_sku") or "").strip()
        if not sku_id:
            continue
        gross_required_qty = _material_float(row.get("required_qty") or row.get("gross_required_qty") or 0.0)
        available_before = _material_float(row.get("available_before") or 0.0)
        inventory_covered_qty = _material_float(
            row.get("inventory_covered_qty")
            if row.get("inventory_covered_qty") is not None
            else max(min(gross_required_qty, available_before), 0.0)
        )
        shortage_qty = _material_float(row.get("shortage_qty") or 0.0)
        make_convert_qty = _material_float(
            row.get("make_convert_qty")
            if row.get("make_convert_qty") is not None
            else max(gross_required_qty - inventory_covered_qty - shortage_qty, 0.0)
        )
        consumed_qty = _material_float(row.get("consumed_qty") or row.get("consumed") or 0.0)
        remaining_after = _material_float(row.get("remaining_after") or 0.0)
        material_type = str(row.get("material_type") or "Material")
        plant = str(row.get("plant") or "Unassigned")
        status = str(row.get("status") or row.get("type") or "COVERED")

        normalized.append(
            {
                "material_sku": sku_id,
                "material_name": str(row.get("material_name") or sku_id),
                "material_category": str(row.get("material_category") or ""),
                "material_type": material_type,
                "plant": plant,
                "stage": str(row.get("stage") or _stage_for_sku(sku_id, material_type)),
                "gross_required_qty": round(gross_required_qty, 3),
                "required_qty": round(gross_required_qty, 3),
                "available_before": round(available_before, 3),
                "inventory_covered_qty": round(inventory_covered_qty, 3),
                "make_convert_qty": round(make_convert_qty, 3),
                "consumed_qty": round(consumed_qty, 3),
                "consumed": round(consumed_qty, 3),
                "shortage_qty": round(shortage_qty, 3),
                "remaining_after": round(remaining_after, 3),
                "status": status,
            }
        )
    return normalized


def _material_rows_to_plants(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    by_plant: Dict[str, Dict[str, Any]] = {}
    for row in rows:
        plant = str(row.get("plant") or "Unassigned")
        bucket = by_plant.setdefault(
            plant,
            {
                "plant": plant,
                "required_qty": 0.0,
                "inventory_covered_qty": 0.0,
                "make_convert_qty": 0.0,
                "shortage_qty": 0.0,
                "rows": [],
            },
        )
        bucket["required_qty"] += _material_float(row.get("gross_required_qty", 0.0) or 0.0)
        bucket["inventory_covered_qty"] += _material_float(row.get("inventory_covered_qty", 0.0) or 0.0)
        bucket["make_convert_qty"] += _material_float(row.get("make_convert_qty", 0.0) or 0.0)
        bucket["shortage_qty"] += _material_float(row.get("shortage_qty", 0.0) or 0.0)
        bucket["rows"].append(row)

    plants: List[Dict[str, Any]] = []
    for plant_name in sorted(by_plant.keys(), key=lambda p: _PLANT_SORT_ORDER.get(p, 99)):
        bucket = by_plant[plant_name]
        plants.append(
            {
                "plant": plant_name,
                "required_qty": round(_material_float(bucket["required_qty"]), 3),
                "inventory_covered_qty": round(_material_float(bucket["inventory_covered_qty"]), 3),
                "make_convert_qty": round(_material_float(bucket["make_convert_qty"]), 3),
                "shortage_qty": round(_material_float(bucket["shortage_qty"]), 3),
                "rows": sorted(
                    bucket["rows"],
                    key=lambda r: (
                        _MAT_TYPE_SORT_ORDER.get(str(r.get("material_type", "Material")), 99),
                        str(r.get("material_sku", "")),
                    ),
                ),
            }
        )
    return plants


def _material_rows_to_detail_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    return [
        {
            "sku_id": str(r.get("material_sku") or ""),
            "material_name": str(r.get("material_name") or r.get("material_sku") or ""),
            "material_type": str(r.get("material_type") or "Material"),
            "plant": str(r.get("plant") or "Unassigned"),
            "required_qty": round(_material_float(r.get("gross_required_qty", 0.0) or 0.0), 3),
            "available_before": round(_material_float(r.get("available_before", 0.0) or 0.0), 3),
            "inventory_covered_qty": round(_material_float(r.get("inventory_covered_qty", 0.0) or 0.0), 3),
            "make_convert_qty": round(_material_float(r.get("make_convert_qty", 0.0) or 0.0), 3),
            "consumed_qty": round(_material_float(r.get("consumed_qty", 0.0) or 0.0), 3),
            "shortage_qty": round(_material_float(r.get("shortage_qty", 0.0) or 0.0), 3),
            "remaining_after": round(_material_float(r.get("remaining_after", 0.0) or 0.0), 3),
            "status": str(r.get("status") or "COVERED"),
            "type": str(r.get("status") or "COVERED"),
        }
        for r in rows
    ]


def _scaled_material_rows(rows: List[Dict[str, Any]], ratio: float) -> List[Dict[str, Any]]:
    ratio = max(_material_float(ratio, 0.0), 0.0)
    scaled: List[Dict[str, Any]] = []
    for row in rows:
        gross_required_qty = _material_float(row.get("gross_required_qty", 0.0) or 0.0) * ratio
        available_before = _material_float(row.get("available_before", 0.0) or 0.0) * ratio
        inventory_covered_qty = _material_float(row.get("inventory_covered_qty", 0.0) or 0.0) * ratio
        make_convert_qty = _material_float(row.get("make_convert_qty", 0.0) or 0.0) * ratio
        consumed_qty = _material_float(row.get("consumed_qty", 0.0) or 0.0) * ratio
        shortage_qty = _material_float(row.get("shortage_qty", 0.0) or 0.0) * ratio
        remaining_after = _material_float(row.get("remaining_after", 0.0) or 0.0) * ratio

        if shortage_qty > 1e-6:
            status = "SHORTAGE"
        elif make_convert_qty > 1e-6 and inventory_covered_qty > 1e-6:
            status = "PARTIAL COVER"
        elif inventory_covered_qty > 1e-6 and make_convert_qty <= 1e-6:
            status = "DRAWN FROM STOCK"
        elif make_convert_qty > 1e-6:
            status = "MAKE / CONVERT"
        else:
            status = "COVERED"

        scaled.append(
            {
                **row,
                "gross_required_qty": round(gross_required_qty, 3),
                "required_qty": round(gross_required_qty, 3),
                "available_before": round(available_before, 3),
                "inventory_covered_qty": round(inventory_covered_qty, 3),
                "make_convert_qty": round(make_convert_qty, 3),
                "consumed_qty": round(consumed_qty, 3),
                "consumed": round(consumed_qty, 3),
                "shortage_qty": round(shortage_qty, 3),
                "remaining_after": round(remaining_after, 3),
                "status": status,
            }
        )
    return scaled


def _material_summary_from_entities(entities: List[Dict[str, Any]], detail_level: str) -> Dict[str, Any]:
    entity_label = {
        "campaign": "Campaigns",
        "po": "Planning Orders",
        "heat": "Heats",
    }.get(detail_level, "Campaigns")

    total_required = sum(_material_float(x.get("required_qty", 0.0) or 0.0) for x in entities)
    total_shortages = sum(_material_float(x.get("shortage_qty", 0.0) or 0.0) for x in entities)
    total_inventory_covered = sum(_material_float(x.get("inventory_covered_qty", 0.0) or 0.0) for x in entities)
    total_make_convert = sum(_material_float(x.get("make_convert_qty", 0.0) or 0.0) for x in entities)
    released_count = sum(1 for x in entities if "RELEASED" in str(x.get("release_status", "")).upper())
    held_count = sum(1 for x in entities if "HOLD" in str(x.get("release_status", "")).upper())
    shortage_count = sum(1 for x in entities if _material_float(x.get("shortage_qty", 0.0) or 0.0) > 1e-6)
    error_count = sum(1 for x in entities if not bool(x.get("feasible", True)))

    return {
        entity_label: len(entities),
        "Released": released_count,
        "Held": held_count,
        "Shortage Lines": shortage_count,
        "BOM Structure Errors": error_count,
        "Total Required Qty": round(total_required, 3),
        "Inventory Covered Qty": round(total_inventory_covered, 3),
        "Shortage Qty": round(total_shortages, 3),
        "Make / Convert Qty": round(total_make_convert, 3),
    }


def _material_plan_for_detail_level(base_payload: Dict[str, Any], detail_level: str) -> Dict[str, Any]:
    requested_level = _normalize_material_detail_level(detail_level)
    base_campaigns = list(base_payload.get("campaigns") or [])
    run_id = base_payload.get("run_id")

    if requested_level == "campaign":
        payload = dict(base_payload)
        payload["detail_level"] = "campaign"
        payload["entity_type"] = "campaign"
        payload["entities"] = list(payload.get("campaigns") or [])
        return payload

    campaign_map = {
        str(c.get("campaign_id") or "").strip(): dict(c)
        for c in base_campaigns
        if str(c.get("campaign_id") or "").strip()
    }
    planning_orders = list(getattr(aps_planning_orders_propose, "_planning_orders", []) or [])
    po_map = {
        str(po.get("po_id") or "").strip(): dict(po)
        for po in planning_orders
        if str(po.get("po_id") or "").strip()
    }

    def _material_due_fields(source: Dict[str, Any]) -> tuple[str, str, List[str]]:
        raw_window = source.get("due_window") or ("", "")
        if isinstance(raw_window, list):
            raw_window = tuple(raw_window)
        if not isinstance(raw_window, tuple):
            raw_window = ("", "")
        due_start = str(raw_window[0] if len(raw_window) >= 1 else "").strip()
        due_end = str(raw_window[1] if len(raw_window) >= 2 else "").strip()
        fallback_due = str(source.get("due_end") or source.get("due_date") or source.get("Due_Date") or "").strip()
        if not due_end and fallback_due:
            due_end = fallback_due
        if not due_start and due_end:
            due_start = due_end
        return due_start, due_end, [due_start, due_end]

    po_entities: List[Dict[str, Any]] = []
    if planning_orders:
        for po in planning_orders:
            po_id = str(po.get("po_id") or "").strip()
            if not po_id:
                continue
            base_entity = dict(campaign_map.get(po_id) or {})
            planner_status = str(po.get("planner_status") or "PROPOSED").strip().upper()
            release_status = base_entity.get("release_status")
            if not release_status:
                release_status = "RELEASED" if planner_status == "RELEASED" else "PLANNED"

            required_qty = _material_float(base_entity.get("required_qty") or po.get("total_qty_mt") or 0.0)
            materials = _material_rows_from_entity(base_entity)
            plants = list(base_entity.get("plants") or _material_rows_to_plants(materials))
            detail_rows = list(base_entity.get("detail_rows") or _material_rows_to_detail_rows(materials))
            due_start, due_end, due_window = _material_due_fields(po)
            if not due_start and not due_end:
                due_start, due_end, due_window = _material_due_fields(base_entity)
            priority = str(
                po.get("priority")
                or po.get("Priority")
                or base_entity.get("priority")
                or base_entity.get("Priority")
                or "NORMAL"
            ).strip().upper() or "NORMAL"

            entity = {
                "entity_id": po_id,
                "planning_order_id": po_id,
                "campaign_id": str(base_entity.get("campaign_id") or po_id),
                "grade": str(base_entity.get("grade") or po.get("grade_family") or ""),
                "release_status": str(release_status),
                "required_qty": round(required_qty, 3),
                "shortage_qty": round(_material_float(base_entity.get("shortage_qty") or 0.0), 3),
                "inventory_covered_qty": round(_material_float(base_entity.get("inventory_covered_qty") or 0.0), 3),
                "make_convert_qty": round(_material_float(base_entity.get("make_convert_qty") or 0.0), 3),
                "material_status": str(base_entity.get("material_status") or "UNREVIEWED"),
                "feasible": bool(base_entity.get("feasible", True)),
                "material_structure_errors": list(base_entity.get("material_structure_errors") or []),
                "material_shortages": dict(base_entity.get("material_shortages") or {}),
                "material_consumed": dict(base_entity.get("material_consumed") or {}),
                "material_gross_requirements": dict(base_entity.get("material_gross_requirements") or {}),
                "inventory_before": dict(base_entity.get("inventory_before") or {}),
                "inventory_after": dict(base_entity.get("inventory_after") or {}),
                "materials": materials,
                "plants": plants,
                "detail_rows": detail_rows,
                "selected_so_ids": list(po.get("selected_so_ids") or []),
                "heats_required": int(po.get("heats_required") or 0),
                "planner_status": planner_status,
                "priority": priority,
                "due_start": due_start,
                "due_end": due_end,
                "due_window": due_window,
            }
            po_entities.append(entity)
    else:
        for camp in base_campaigns:
            cid = str(camp.get("campaign_id") or "").strip()
            if not cid:
                continue
            entity = dict(camp)
            due_start, due_end, due_window = _material_due_fields(entity)
            entity["priority"] = str(entity.get("priority") or entity.get("Priority") or "NORMAL").strip().upper() or "NORMAL"
            entity["due_start"] = due_start
            entity["due_end"] = due_end
            entity["due_window"] = due_window
            entity["entity_id"] = cid
            entity["planning_order_id"] = cid
            po_entities.append(entity)

    if requested_level == "po":
        return {
            "campaigns": po_entities,
            "entities": po_entities,
            "detail_level": "po",
            "entity_type": "planning_order",
            "summary": _material_summary_from_entities(po_entities, "po"),
            "run_id": run_id,
            "source_detail_level": "campaign",
        }

    # requested_level == "heat"
    heat_batches = list(getattr(aps_planning_simulate, "_heat_batches", []) or [])
    normalized_heats: List[Dict[str, Any]] = []
    for heat in heat_batches:
        heat_id = str(heat.get("heat_id") or heat.get("Heat_ID") or "").strip()
        po_id = str(heat.get("planning_order_id") or heat.get("po_id") or "").strip()
        if not heat_id or not po_id:
            continue
        normalized_heats.append(
            {
                "heat_id": heat_id,
                "planning_order_id": po_id,
                "qty_mt": _material_float(heat.get("qty_mt") or 0.0),
                "grade": str(heat.get("grade") or ""),
                "heat_number_seq": int(heat.get("heat_number_seq") or 0),
                "upstream_route": str(heat.get("upstream_route") or ""),
                "expected_duration_hours": _material_float(heat.get("expected_duration_hours") or 0.0),
            }
        )

    if not normalized_heats and planning_orders:
        for po in planning_orders:
            po_id = str(po.get("po_id") or "").strip()
            if not po_id:
                continue
            heats_required = max(int(po.get("heats_required") or 0), 1)
            total_qty = _material_float(po.get("total_qty_mt") or 0.0)
            per_heat_qty = total_qty / heats_required if heats_required > 0 else total_qty
            for idx in range(heats_required):
                normalized_heats.append(
                    {
                        "heat_id": f"{po_id}-H{idx + 1:02d}",
                        "planning_order_id": po_id,
                        "qty_mt": round(per_heat_qty, 3),
                        "grade": str(po.get("grade_family") or ""),
                        "heat_number_seq": idx + 1,
                        "upstream_route": str(po.get("route_family") or "SMS→RM"),
                        "expected_duration_hours": 2.0,
                    }
                )

    if not normalized_heats:
        for camp in base_campaigns:
            cid = str(camp.get("campaign_id") or "").strip()
            if not cid:
                continue
            normalized_heats.append(
                {
                    "heat_id": f"{cid}-H01",
                    "planning_order_id": cid,
                    "qty_mt": _material_float(camp.get("required_qty") or 0.0),
                    "grade": str(camp.get("grade") or ""),
                    "heat_number_seq": 1,
                    "upstream_route": "SMS→RM",
                    "expected_duration_hours": 2.0,
                }
            )

    po_entity_map = {
        str(entity.get("planning_order_id") or entity.get("entity_id") or "").strip(): entity
        for entity in po_entities
        if str(entity.get("planning_order_id") or entity.get("entity_id") or "").strip()
    }
    heat_count_by_po: Dict[str, int] = {}
    for heat in normalized_heats:
        po_id = str(heat.get("planning_order_id") or "").strip()
        if not po_id:
            continue
        heat_count_by_po[po_id] = heat_count_by_po.get(po_id, 0) + 1

    heat_entities: List[Dict[str, Any]] = []
    for heat in normalized_heats:
        heat_id = str(heat.get("heat_id") or "").strip()
        po_id = str(heat.get("planning_order_id") or "").strip()
        if not heat_id or not po_id:
            continue

        base_entity = dict(po_entity_map.get(po_id) or campaign_map.get(po_id) or {})
        po_source = dict(po_map.get(po_id) or {})
        base_rows = _material_rows_from_entity(base_entity)
        heat_qty = _material_float(heat.get("qty_mt") or 0.0)
        po_total = _material_float(base_entity.get("required_qty") or po_map.get(po_id, {}).get("total_qty_mt") or 0.0)
        if po_total > 1e-9 and heat_qty > 1e-9:
            ratio = heat_qty / po_total
        else:
            ratio = 1.0 / max(heat_count_by_po.get(po_id, 1), 1)

        rows = _scaled_material_rows(base_rows, ratio)
        plants = _material_rows_to_plants(rows)
        detail_rows = _material_rows_to_detail_rows(rows)

        required_qty = round(sum(_material_float(r.get("gross_required_qty", 0.0) or 0.0) for r in rows), 3)
        inventory_covered_qty = round(sum(_material_float(r.get("inventory_covered_qty", 0.0) or 0.0) for r in rows), 3)
        make_convert_qty = round(sum(_material_float(r.get("make_convert_qty", 0.0) or 0.0) for r in rows), 3)
        shortage_qty = round(sum(_material_float(r.get("shortage_qty", 0.0) or 0.0) for r in rows), 3)
        material_status = "SHORTAGE" if shortage_qty > 1e-6 else ("MAKE / CONVERT" if make_convert_qty > 1e-6 else "READY")
        due_start, due_end, due_window = _material_due_fields(po_source)
        if not due_start and not due_end:
            due_start, due_end, due_window = _material_due_fields(base_entity)
        priority = str(
            po_source.get("priority")
            or po_source.get("Priority")
            or base_entity.get("priority")
            or base_entity.get("Priority")
            or "NORMAL"
        ).strip().upper() or "NORMAL"

        heat_entities.append(
            {
                "entity_id": heat_id,
                "heat_id": heat_id,
                "planning_order_id": po_id,
                "campaign_id": str(base_entity.get("campaign_id") or po_id),
                "grade": str(heat.get("grade") or base_entity.get("grade") or ""),
                "release_status": str(base_entity.get("release_status") or "PLANNED"),
                "required_qty": required_qty if required_qty > 0 else round(heat_qty, 3),
                "shortage_qty": shortage_qty,
                "inventory_covered_qty": inventory_covered_qty,
                "make_convert_qty": make_convert_qty,
                "material_status": material_status,
                "feasible": shortage_qty <= 1e-6,
                "material_structure_errors": list(base_entity.get("material_structure_errors") or []),
                "material_shortages": {},
                "material_consumed": {},
                "material_gross_requirements": {},
                "inventory_before": {},
                "inventory_after": {},
                "materials": rows,
                "plants": plants,
                "detail_rows": detail_rows,
                "selected_so_ids": list(base_entity.get("selected_so_ids") or []),
                "heats_required": int(base_entity.get("heats_required") or heat_count_by_po.get(po_id, 1)),
                "planner_status": str(base_entity.get("planner_status") or ""),
                "heat_number_seq": int(heat.get("heat_number_seq") or 0),
                "upstream_route": str(heat.get("upstream_route") or ""),
                "expected_duration_hours": _material_float(heat.get("expected_duration_hours") or 0.0),
                "qty_mt": round(heat_qty, 3),
                "priority": priority,
                "due_start": due_start,
                "due_end": due_end,
                "due_window": due_window,
            }
        )

    return {
        "campaigns": heat_entities,
        "entities": heat_entities,
        "detail_level": "heat",
        "entity_type": "heat",
        "summary": _material_summary_from_entities(heat_entities, "heat"),
        "run_id": run_id,
        "source_detail_level": "campaign",
    }

def _store_compatibility_cache(
    *,
    run_id: str,
    campaigns: List[Dict[str, Any]],
    heat_df: pd.DataFrame,
    camp_df: pd.DataFrame,
    cap_df: pd.DataFrame,
    material_payload: Dict[str, Any],
    solver: str,
    solver_detail: str,
) -> None:
    _state['last_run'] = datetime.now().isoformat()
    _state['run_id'] = run_id
    _state['campaigns'] = campaigns
    _state['heat_schedule'] = heat_df
    _state['camp_schedule'] = camp_df
    _state['capacity'] = cap_df
    _state['material_plan_data'] = material_payload
    _state['solver_status'] = solver
    _state['solver_detail'] = solver_detail
    _state['error'] = None


def _publish_planning_snapshot(
    *,
    campaigns: List[Dict[str, Any]],
    skus: list | pd.DataFrame | None,
    heat_df: pd.DataFrame | None = None,
    camp_df: pd.DataFrame | None = None,
    cap_df: pd.DataFrame | None = None,
    solver_status: str | None = None,
    solver_detail: str | None = None,
) -> Dict[str, Any]:
    """Publish planning-scoped state for dashboard/campaign/material readers."""
    material_payload = _calculate_material_plan(
        campaigns,
        detail_level="campaign",
        run_id=None,
        skus=skus,
    )

    _state["last_run"] = datetime.now().isoformat()
    _state["run_id"] = None
    _state["campaigns"] = campaigns
    _state["heat_schedule"] = (
        heat_df.copy()
        if isinstance(heat_df, pd.DataFrame)
        else pd.DataFrame()
    )
    _state["camp_schedule"] = (
        camp_df.copy()
        if isinstance(camp_df, pd.DataFrame)
        else pd.DataFrame()
    )
    _state["capacity"] = (
        cap_df.copy()
        if isinstance(cap_df, pd.DataFrame)
        else pd.DataFrame()
    )
    _state["material_plan_data"] = material_payload
    if solver_status is not None:
        _state["solver_status"] = str(solver_status)
    if solver_detail is not None:
        _state["solver_detail"] = str(solver_detail)
    _state["error"] = None
    return material_payload


def _publish_planning_orders_snapshot(
    planning_orders: List[Dict[str, Any]],
    *,
    workbook_data: Dict[str, Any] | None = None,
    solver_detail: str = "Planning state updated. Derive heats and run Feasibility Check.",
) -> Dict[str, Any]:
    """Publish planning-order scoped campaigns/material for cross-tab freshness."""
    d = workbook_data or _load_all()
    campaigns = _planning_orders_to_scheduler_campaigns(
        planning_orders=list(planning_orders or []),
        all_orders_df=d.get("all_orders", pd.DataFrame()),
        config=d.get("config", {}),
    )
    return _publish_planning_snapshot(
        campaigns=campaigns,
        skus=d.get("skus"),
        solver_status="PLANNING READY",
        solver_detail=solver_detail,
    )


@app.route('/api/run/schedule', methods=['POST'])
def run_schedule():
    return run_schedule_api()


def run_schedule_api():
    try:
        d = _load_all()
        config = d["config"]
        horizon = int(float(config.get("Planning_Horizon_Days", 14) or 14))
        sec_lim = float(config.get("Default_Solver_Limit_Sec", 30.0) or 30.0)

        body = request.get_json(silent=True) or {}
        horizon = int(body.get("horizon", body.get("horizon_days", horizon)))
        sec_lim = float(body.get("solver_sec", body.get("time_limit", sec_lim)))

        released_pos = _released_sales_orders_to_planning_orders(d)

        if released_pos:
            campaigns = _planning_orders_to_scheduler_campaigns(
                planning_orders=released_pos,
                all_orders_df=d["all_orders"],
                config=config,
            )
        else:
            min_cmt = float(config.get("Min_Campaign_MT", 100.0) or 100.0)
            max_cmt = float(config.get("Max_Campaign_MT", 500.0) or 500.0)
            campaigns = build_campaigns(
                d["sales_orders"],
                min_campaign_mt=min_cmt,
                max_campaign_mt=max_cmt,
                inventory=d["inventory"],
                bom=d["bom"],
                routing=d["routing"],
                config=config,
                skus=d["skus"],
                campaign_config=d["campaign_config"],
            )

        released = [c for c in campaigns if str(c.get("release_status", "")).upper() == "RELEASED"]

        result = schedule(
            campaigns,
            d["resources"],
            planning_horizon_days=horizon,
            planning_start=datetime.now(),
            routing=d["routing"],
            queue_times=d["queue_times"],
            changeover_matrix=d["changeover_matrix"],
            config=config,
            solver_time_limit_sec=sec_lim,
        )

        heat_df = result.get("heat_schedule", pd.DataFrame())
        camp_df = result.get("campaign_schedule", pd.DataFrame())
        solver = result.get("solver_status", "UNKNOWN")
        s_detail = result.get("solver_detail", "CP_SAT_NA")

        allow_defaults = bool(result.get("allow_default_masters", False)) or _config_flag(config, "Allow_Scheduler_Default_Masters", "N")
        warnings: List[str] = []
        degraded_flags = {
            "default_masters_used": allow_defaults,
            "greedy_fallback": solver in {"GREEDY", "GREEDY_FALLBACK", "UNKNOWN"},
            "material_incomplete": False,
            "inventory_lineage_degraded": False,
            "capacity_unavailable": False,
        }

        try:
            demand_hrs = compute_demand_hours(
                released,
                d["resources"],
                routing=d["routing"],
                changeover_matrix=d["changeover_matrix"],
                allow_defaults=allow_defaults,
            )
            cap_df = capacity_map(demand_hrs, d["resources"], horizon_days=horizon)
        except Exception as cap_err:
            cap_df = pd.DataFrame()
            degraded_flags["capacity_unavailable"] = True
            warnings.append(f"Capacity map failed: {cap_err}")

        material_payload = _calculate_material_plan(campaigns, detail_level="campaign", run_id=None,skus=d["skus"],)
        run_id = _create_run_artifact(
            config=config,
            campaigns=campaigns,
            heat_schedule=heat_df,
            campaign_schedule=camp_df,
            capacity_map_df=cap_df,
            material_plan={**material_payload, "run_id": None},
            solver_status=solver,
            solver_detail=s_detail,
            warnings=warnings,
            degraded_flags=degraded_flags,
        )
        _set_active_run_artifact(run_id)

        material_payload = _calculate_material_plan(campaigns, detail_level="campaign", run_id=run_id,skus=d["skus"],)
        _run_artifacts[run_id]["results"]["material_plan"] = material_payload
        _store_compatibility_cache(
            run_id=run_id,
            campaigns=campaigns,
            heat_df=heat_df,
            camp_df=camp_df,
            cap_df=cap_df,
            material_payload=material_payload,
            solver=solver,
            solver_detail=s_detail,
        )

        response_payload = {
            **_dashboard_payload(),
            "campaigns": _campaigns_to_view(campaigns, camp_df),
            "gantt": _df_to_records(heat_df),
            "capacity": _df_to_records(cap_df),
            "material": material_payload,
            "run_id": run_id,
        }
        return _jsonify(response_payload)
    except Exception as e:
        _state["error"] = str(e)
        return _error_response(
            error_message=f"Scheduling failed: {str(e)}",
            error_code="SCHEDULE_ERROR",
            error_domain="SCHEDULER",
            details=traceback.format_exc(),
            degraded_mode=False,
            status_code=500,
        )[0]

@app.route('/api/run/ctp', methods=['POST'])
def run_ctp_api():
    try:
        body = request.get_json(silent=True) or {}
        sku_id = str(body.get("sku_id", "")).strip()
        qty_mt = float(body.get("qty_mt", 100.0))
        requested_date = body.get("requested_date", (datetime.now() + pd.Timedelta(days=14)).isoformat())

        d = _load_all()
        config = d["config"]

        released_pos = _released_sales_orders_to_planning_orders(d)
        if released_pos:
            campaigns = _planning_orders_to_scheduler_campaigns(
                planning_orders=released_pos,
                all_orders_df=d["all_orders"],
                config=config,
            )
        else:
            min_cmt = float(config.get("Min_Campaign_MT", 100.0) or 100.0)
            max_cmt = float(config.get("Max_Campaign_MT", 500.0) or 500.0)
            campaigns = _active_run_campaigns() or build_campaigns(
                d["sales_orders"],
                min_cmt,
                max_cmt,
                inventory=d["inventory"],
                bom=d["bom"],
                routing=d["routing"],
                config=config,
                skus=d["skus"],
                campaign_config=d["campaign_config"],
            )

        schedule_df = _active_run_heat_schedule()
        frozen_jobs = _frozen_jobs_from_schedule_dataframe(schedule_df)

        res = capable_to_promise(
            sku_id=sku_id,
            qty_mt=qty_mt,
            requested_date=requested_date,
            campaigns=campaigns,
            resources=d["resources"],
            bom=d["bom"],
            inventory=d["inventory"],
            routing=d["routing"],
            skus=d["skus"],
            planning_start=datetime.now(),
            config=config,
            campaign_config=d["campaign_config"],
            queue_times=d["queue_times"],
            changeover_matrix=d["changeover_matrix"],
            frozen_jobs=frozen_jobs,
        )
        decision_class = str(res.get("decision_class") or "").strip().upper()
        lineage_status = str(res.get("inventory_lineage_status") or "").strip().upper()
        lineage_label_map = {
            "AUTHORITATIVE_SNAPSHOT_CHAIN": "Authoritative",
            "NO_COMMITTED_CAMPAIGNS": "Authoritative",
            "RECOMPUTED_FROM_CONSUMPTION": "Recomputed",
            "CONSERVATIVE_BLEND": "Conservative Blend",
        }
        res["lineage_trust_level"] = lineage_label_map.get(lineage_status, "Unknown")
        res["decision_family"] = (
            "Confirmed"
            if decision_class.startswith("PROMISE_CONFIRMED")
            else "Split"
            if decision_class == "PROMISE_SPLIT_REQUIRED"
            else "Later Date"
            if decision_class == "PROMISE_LATER_DATE"
            else "Material Block"
            if decision_class == "CANNOT_PROMISE_MATERIAL"
            else "Capacity Block"
            if decision_class == "CANNOT_PROMISE_CAPACITY"
            else "Policy Block"
            if decision_class == "CANNOT_PROMISE_POLICY_ONLY"
            else "Inventory Trust Block"
            if decision_class == "CANNOT_PROMISE_INVENTORY_TRUST"
            else "Master Data Block"
            if decision_class == "CANNOT_PROMISE_MASTER_DATA"
            else "Mixed Blocker"
            if decision_class == "CANNOT_PROMISE_MIXED_BLOCKERS"
            else "Unclassified"
        )
        return _jsonify({k: _safe(v) for k, v in res.items()})
    except Exception as e:
        return _error_response(
            error_message=f"CTP evaluation failed: {str(e)}",
            error_code="CTP_ERROR",
            error_domain="CTP",
            details=traceback.format_exc(),
            status_code=500,
        )[0]


@app.route('/api/orders/assign', methods=['POST'])
def orders_assign():
    payload = request.get_json(silent=True) or {}
    assignments = payload.get('assignments', []) if isinstance(payload, dict) else []
    updated = 0
    for assignment in assignments:
        so_id = str(assignment.get('so_id', '')).strip()
        campaign_id = assignment.get('campaign_id')
        if not so_id:
            continue
        try:
            store.update_row('sales-orders', so_id, {'Campaign_ID': campaign_id}, partial=True)
            updated += 1
        except Exception:
            continue
    return _jsonify({"ok": True, "updated": updated})


@app.route('/api/orders', methods=['GET', 'POST'])
def orders_collection():
    if request.method == 'GET':
        return _jsonify({"orders": _sheet_items('sales-orders')})
    payload = request.get_json(silent=True) or {}
    data = payload if isinstance(payload, dict) and 'data' not in payload else payload.get('data', {})
    try:
        return _jsonify({"ok": True, "item": store.create_row('sales-orders', data)})
    except ValueError as e:
        return _jsonify({"error": str(e)}, 400)


@app.route('/api/orders/<so_id>', methods=['GET', 'PUT', 'DELETE'])
def order_item(so_id):
    if request.method == 'GET':
        row = store.get_row('sales-orders', so_id)
        return _jsonify({"order": row} if row else {"error": 'Not found'}, 200 if row else 404)
    if request.method == 'DELETE':
        try:
            store.delete_row('sales-orders', so_id)
            return _jsonify({"ok": True, 'so_id': so_id})
        except KeyError as e:
            return _jsonify({"error": str(e)}, 404)
    payload = request.get_json(silent=True) or {}
    data = payload if isinstance(payload, dict) and 'data' not in payload else payload.get('data', {})
    try:
        return _jsonify({"ok": True, 'item': store.update_row('sales-orders', so_id, data, partial=False)})
    except KeyError as e:
        return _jsonify({"error": str(e)}, 404)


@app.route('/api/aps/dashboard/overview')
def aps_dashboard_overview():
    dashboard = _dashboard_payload()
    campaigns = _campaign_rows()
    last_simulation = dict(getattr(aps_planning_simulate, "_last_result", None) or {}) or None
    return _jsonify({
        "summary": dashboard,
        "campaigns": campaigns[:8],
        "alerts": dashboard.get('shortage_alerts', [])[:10],
        "utilisation": dashboard.get('utilisation', [])[:8],
        "release_queue": sorted(campaigns, key=lambda x: str(x.get('Due_Date') or x.get('due_date') or '9999-12-31'))[:8],
        "last_simulation": last_simulation,
    })


@app.route('/api/masterdata/audit')
def masterdata_audit():
    payload = audit_workbook_masterdata(WORKBOOK)
    payload["health_summary"] = _masterdata_health_summary(payload)
    return _jsonify(payload)


@app.route('/api/aps/orders/list')
def aps_orders_list():
    items = _sheet_items('sales-orders', search=request.args.get('search'))
    # Filter out rows with null/empty SO_ID
    items = [x for x in items if x.get('SO_ID')]
    priority = request.args.get('priority')
    grade = request.args.get('grade')
    if priority:
        items = [x for x in items if str(x.get('Priority', '')).upper() == priority.upper()]
    if grade:
        items = [x for x in items if str(x.get('Grade', '')) == grade]
    return _jsonify({"items": items, "total": len(items)})


@app.route('/api/aps/orders/<so_id>', methods=['GET', 'PUT', 'DELETE'])
def aps_order_item(so_id):
    return order_item(so_id)


@app.route('/api/aps/orders', methods=['POST'])
def aps_order_create():
    return orders_collection()


@app.route('/api/aps/orders/assign', methods=['POST'])
def aps_orders_assign():
    return orders_assign()


@app.route('/api/aps/campaigns/list')
def aps_campaigns_list():
    items = _campaign_rows()
    status = request.args.get('status')
    if status:
        items = [x for x in items if status.upper() in str(x.get('release_status') or x.get('Release_Status') or x.get('Status') or '').upper()]
    return _jsonify({"run_id": _current_trace_id(), "items": items, "total": len(items)})


@app.route('/api/aps/campaigns/release-queue')
def aps_campaign_release_queue():
    items = sorted(_campaign_rows(), key=lambda x: str(x.get('Due_Date') or x.get('due_date') or '9999-12-31'))
    return _jsonify({"run_id": _current_trace_id(), "items": items})


@app.route('/api/aps/campaigns/<campaign_id>')
def aps_campaign_item(campaign_id):
    for item in _campaign_rows():
        cid = str(item.get('campaign_id') or item.get('Campaign_ID') or '')
        if cid == campaign_id:
            return _jsonify({"run_id": _current_trace_id(), "item": item})
    return _jsonify({"error": 'Campaign not found'}, 404)


@app.route('/api/aps/campaigns/<campaign_id>/status', methods=['PATCH'])
def aps_campaign_status_update(campaign_id):
    if 'campaign-schedule' not in SHEETS:
        return _jsonify({"error": 'Campaign_Schedule sheet not available for updates'}, 400)
    payload = request.get_json(silent=True) or {}
    data = payload if isinstance(payload, dict) and 'data' not in payload else payload.get('data', {})
    try:
        item = store.update_row('campaign-schedule', campaign_id, data, partial=True)
        return _jsonify({"item": item})
    except KeyError as e:
        return _jsonify({"error": str(e)}, 404)
    except ValueError as e:
        return _jsonify({"error": str(e)}, 400)


@app.route('/api/aps/schedule/gantt')
def aps_schedule_gantt():
    return _jsonify({"run_id": _current_trace_id(), "jobs": _schedule_rows()})


@app.route('/api/aps/schedule/run', methods=['POST'])
def aps_schedule_run():
    return run_schedule_api()


@app.route('/api/aps/schedule/jobs/<job_id>')
def aps_schedule_job_item(job_id):
    for item in _schedule_rows():
        jid = str(item.get('Job_ID') or item.get('job_id') or '')
        if jid == job_id:
            return _jsonify({"run_id": _current_trace_id(), "item": item})
    return _jsonify({"error": 'Job not found'}, 404)


@app.route('/api/aps/schedule/jobs/<job_id>/reschedule', methods=['PATCH'])
def aps_schedule_job_reschedule(job_id):
    if 'schedule-output' not in SHEETS:
        return _jsonify({"error": 'Schedule_Output sheet not available for updates'}, 400)
    payload = request.get_json(silent=True) or {}
    data = payload if isinstance(payload, dict) and 'data' not in payload else payload.get('data', {})
    try:
        item = store.update_row('schedule-output', job_id, data, partial=True)
        return _jsonify({"item": item})
    except KeyError as e:
        return _jsonify({"error": str(e)}, 404)


@app.route('/api/aps/dispatch/board')
def aps_dispatch_board():
    return _jsonify({"run_id": _current_trace_id(), "resources": _dispatch_rows()})


@app.route('/api/aps/dispatch/resources/<resource_id>')
def aps_dispatch_resource_item(resource_id):
    items = _dispatch_rows()
    for item in items:
        rid = str(item.get('resource_id') or item.get('Resource_ID') or '')
        if rid == resource_id:
            return _jsonify({"run_id": _current_trace_id(), "item": item})
    return _jsonify({"error": 'Resource not found'}, 404)


@app.route('/api/aps/capacity/map')
def aps_capacity_map():
    return _jsonify({"run_id": _current_trace_id(), "items": _capacity_rows()})


@app.route('/api/aps/capacity/bottlenecks')
def aps_capacity_bottlenecks():
    items = sorted(_capacity_rows(), key=lambda x: float(x.get('Utilisation_%') or 0), reverse=True)
    return _jsonify({"run_id": _current_trace_id(), "items": items[:10]})


@app.route('/api/aps/material/plan')
def aps_material_plan():
    detail_level = request.args.get("detail_level", "campaign")
    return _jsonify(_material_plan_payload(detail_level=detail_level))


@app.route('/api/aps/material/holds')
def aps_material_holds():
    items = _material_holds_payload()
    return _jsonify({"run_id": _current_trace_id(), "items": items, "total": len(items)})


@app.route('/api/aps/bom/explosion')
def aps_bom_explosion():
    payload = _active_run_bom()
    if payload is None:
        return _error_response(
            error_message="No BOM explosion data available. Run POST /api/run/bom first.",
            error_code="BOM_NOT_RUN",
            error_domain="BOM",
            status_code=404,
        )[0]
    return _jsonify(payload)



@app.route('/api/aps/bom/tree', methods=['POST'])
def aps_bom_tree():
    """Return a netted BOM payload for an explicit selected root demand set."""
    try:
        body = request.get_json(silent=True) or {}
        items = body.get('items') or []

        empty_payload = {
            'roots': [],
            'gross_bom': [],
            'net_bom': [],
            'summary': {},
            'structure_errors': [],
            'feasible': True,
            'rows': 0,
        }
        if not isinstance(items, list) or not items:
            return _jsonify(empty_payload)

        root_qty_map: dict[str, float] = {}
        for item in items:
            sku_id = str(item.get('sku_id') or '').strip()
            try:
                qty_mt = float(item.get('qty_mt') or 0)
            except Exception:
                qty_mt = 0.0
            if not sku_id or qty_mt <= 0:
                continue
            root_qty_map[sku_id] = round(root_qty_map.get(sku_id, 0.0) + qty_mt, 6)

        if not root_qty_map:
            return _jsonify(empty_payload)

        d = _load_all()
        gross_frames = []
        structure_errors = []
        feasible = True

        for sku_id, qty_mt in root_qty_map.items():
            demand_df = pd.DataFrame([{'SKU_ID': sku_id, 'Required_Qty': qty_mt}])
            explosion_result = explode_bom_details(
                demand_df,
                d.get('bom'),
                max_levels=10,
                on_structure_error='record',
            )

            exploded_df = explosion_result.get('exploded', pd.DataFrame())
            if exploded_df is not None and not exploded_df.empty:
                exploded_df = exploded_df.copy()
                exploded_df['Root_SKU'] = sku_id
                gross_frames.append(exploded_df)

            for error in explosion_result.get('structure_errors', []):
                structure_errors.append({
                    **error,
                    'root_sku': sku_id,
                })
            feasible = feasible and explosion_result.get('feasible', True)

        gross_bom_df = (
            pd.concat(gross_frames, ignore_index=True)
            if gross_frames
            else pd.DataFrame(columns=['SKU_ID', 'Required_Qty', 'Produced_Qty', 'BOM_Level', 'Parent_SKU', 'Flow_Type', 'Root_SKU'])
        )

        if not feasible and structure_errors:
            error_summary = "; ".join(f"{e.get('type')}: {e.get('path')}" for e in structure_errors[:3])
            return _error_response(
                error_message="BOM structure validation failed",
                error_code="BOM_STRUCTURE_ERROR",
                error_domain="BOM",
                details=error_summary,
                degraded_mode=False,
                status_code=400,
            )[0]

        netted = net_requirements(gross_bom_df, d.get('inventory'))
        grouped_bom_data, bom_summary = _build_bom_grouped_view(netted, d.get('skus'), d.get('bom'))

        return _jsonify({
            'roots': [
                {
                    'sku_id': sku_id,
                    'required_qty': round(qty_mt, 3),
                }
                for sku_id, qty_mt in root_qty_map.items()
            ],
            'gross_bom': _df_to_records(gross_bom_df),
            'net_bom': _df_to_records(netted),
            'grouped_bom': grouped_bom_data,
            'summary': bom_summary,
            'structure_errors': structure_errors,
            'feasible': feasible,
            'rows': len(netted),
        })
    except Exception as e:
        return _error_response(
            error_message=f"BOM tree failed: {str(e)}",
            error_code="BOM_TREE_ERROR",
            error_domain="BOM",
            details=traceback.format_exc(),
            status_code=500,
        )[0]


@app.route('/api/aps/bom/for-skus', methods=['POST'])
def aps_bom_for_skus():
    """Return exploded input materials for each requested FG SKU.

    Request:
      { "items": [ { "sku_id": "...", "qty_mt": 123.4 }, ... ] }

    Response:
      {
        "materials": {
          "FG-...": [
            {"sku_id": "RM-...", "label": "RM-...", "qty_mt": 10.5},
            ...
          ]
        }
      }
    }
    """
    try:
        body = request.get_json(silent=True) or {}
        items = body.get('items') or []

        if not isinstance(items, list) or not items:
            return _jsonify({'materials': {}})

        # Load workbook-backed master data
        d = _load_all()
        bom_df = d.get('bom') if d else None

        if bom_df is None or bom_df.empty:
            return _jsonify({'materials': {}})

        from engine.bom_explosion import explode_bom_details

        result_materials = {}

        for item in items:
            sku_id = str(item.get('sku_id') or '').strip()

            try:
                qty_mt = float(item.get('qty_mt') or 0)
            except Exception:
                qty_mt = 0.0

            if not sku_id or qty_mt <= 0:
                if sku_id:
                    result_materials[sku_id] = []
                continue

            try:
                # Single-row demand for one FG SKU
                demand_df = pd.DataFrame([
                    {
                        'SKU_ID': sku_id,
                        'Required_Qty': qty_mt
                    }
                ])

                explosion = explode_bom_details(
                    demand=demand_df,
                    bom=bom_df,
                    max_levels=10,
                    on_structure_error='record'
                )

                exp_df = explosion.get('exploded')
                if exp_df is None or exp_df.empty:
                    result_materials[sku_id] = []
                    continue

                # Keep only input / consumed materials, not byproducts
                if 'Flow_Type' in exp_df.columns:
                    input_materials = exp_df[
                        exp_df['Flow_Type']
                        .fillna('')
                        .astype(str)
                        .str.strip()
                        .str.upper()
                        .eq('INPUT')
                    ].copy()
                else:
                    input_materials = exp_df.copy()

                if input_materials.empty:
                    result_materials[sku_id] = []
                    continue

                # Aggregate duplicate material rows by SKU_ID
                agg_df = (
                    input_materials
                    .groupby('SKU_ID', as_index=False)['Required_Qty']
                    .sum()
                )

                materials = []
                for _, row in agg_df.iterrows():
                    child_sku = str(row.get('SKU_ID') or '').strip()
                    req_qty = float(row.get('Required_Qty') or 0)

                    if not child_sku or req_qty <= 0:
                        continue

                    materials.append({
                        'sku_id': child_sku,
                        'label': child_sku,
                        'qty_mt': round(req_qty, 3)
                    })

                result_materials[sku_id] = materials

            except Exception as e:
                print(f"Error expanding BOM for {sku_id}: {e}")
                result_materials[sku_id] = []

        return _jsonify({'materials': result_materials})

    except Exception as e:
        print(f"BOM for-skus error: {e}")
        return _jsonify({'materials': {}})


@app.route('/api/aps/clear-outputs', methods=['POST'])
def aps_clear_outputs():
    try:
        _state['campaigns'] = []
        _state['heat_schedule'] = pd.DataFrame()
        _state['camp_schedule'] = pd.DataFrame()
        _state['capacity'] = pd.DataFrame()
        _state['material_plan_data'] = None
        _state['bom_explosion'] = None
        _state['last_run'] = None
        _state['run_id'] = None
        _state['solver_status'] = 'CLEARED'
        _state['solver_detail'] = ''
        _state['error'] = None

        global _active_run_id
        _active_run_id = None
        _run_artifacts.clear()

        return _jsonify({"ok": True, "message": "All outputs cleared"})
    except Exception as e:
        _state['error'] = str(e)
        return _jsonify({"error": str(e), "trace": traceback.format_exc()}, 500)


@app.route('/api/aps/ctp/check', methods=['POST'])
def aps_ctp_check():
    return run_ctp_api()


@app.route('/api/aps/ctp/requests', methods=['GET', 'POST'])
def aps_ctp_requests():
    if request.method == 'GET':
        return _jsonify({"items": _sheet_items('ctp-request') if 'ctp-request' in SHEETS else []})
    if 'ctp-request' not in SHEETS:
        return _jsonify({"error": 'CTP_Request sheet not available'}, 400)
    payload = request.get_json(silent=True) or {}
    data = payload if isinstance(payload, dict) and 'data' not in payload else payload.get('data', {})
    try:
        return _jsonify({"item": store.create_row('ctp-request', data)})
    except ValueError as e:
        return _jsonify({"error": str(e)}, 400)


@app.route('/api/aps/ctp/output')
def aps_ctp_output():
    return _jsonify({"items": _sheet_items('ctp-output') if 'ctp-output' in SHEETS else []})


@app.route('/api/aps/scenarios/list')
def aps_scenarios_list():
    return _jsonify({"items": _sheet_items('scenarios') if 'scenarios' in SHEETS else []})


@app.route('/api/aps/scenarios', methods=['POST'])
def aps_scenario_create():
    if 'scenarios' not in SHEETS:
        return _jsonify({"error": 'Scenarios sheet not available'}, 400)
    payload = request.get_json(silent=True) or {}
    data = payload if isinstance(payload, dict) and 'data' not in payload else payload.get('data', {})
    try:
        return _jsonify({"item": store.create_row('scenarios', data)})
    except ValueError as e:
        return _jsonify({"error": str(e)}, 400)


@app.route('/api/aps/scenarios/<key_value>', methods=['PUT', 'PATCH', 'DELETE'])
def aps_scenario_item(key_value):
    if 'scenarios' not in SHEETS:
        return _jsonify({"error": 'Scenarios sheet not available'}, 400)
    if request.method == 'DELETE':
        try:
            store.delete_row('scenarios', key_value)
            return _jsonify({"deleted": True, 'key': key_value})
        except KeyError as e:
            return _jsonify({"error": str(e)}, 404)
    payload = request.get_json(silent=True) or {}
    data = payload if isinstance(payload, dict) and 'data' not in payload else payload.get('data', {})
    partial = request.method == 'PATCH'
    try:
        return _jsonify({"item": store.update_row('scenarios', key_value, data, partial=partial)})
    except KeyError as e:
        return _jsonify({"error": str(e)}, 404)


@app.route('/api/aps/scenarios/output')
def aps_scenarios_output():
    return _jsonify({"items": _sheet_items('scenario-output') if 'scenario-output' in SHEETS else []})


@app.route('/api/aps/scenarios/apply', methods=['POST'])
def aps_scenarios_apply():
    payload = request.get_json(silent=True) or {}
    scenario = payload.get('scenario')
    return _jsonify({
        "ok": True,
        'applied_scenario': scenario,
        'note': 'Workbook-backed scenario selection acknowledged. Use /api/aps/schedule/run to recompute the live plan.',
    })


@app.route('/api/aps/masterdata')
def aps_masterdata():
    return _jsonify(_masterdata_payload())


@app.route('/api/aps/masterdata/<section>')
def aps_masterdata_section(section):
    api_name = _section_sheet(section)
    if not api_name:
        return _jsonify({"error": f'Unknown master data section: {section}'}, 404)
    return _jsonify({"items": _sheet_items(api_name)})


@app.route('/api/aps/masterdata/<section>', methods=['POST'])
def aps_masterdata_section_create(section):
    api_name = _section_sheet(section)
    if not api_name:
        return _jsonify({"error": f"Unknown master data section: {section}"}, 404)

    payload = request.get_json(silent=True) or {}
    data = payload if isinstance(payload, dict) and "data" not in payload else payload.get("data", {})

    try:
        clean = _validate_masterdata_payload(section, data)
        return _jsonify({"item": store.create_row(api_name, clean)})
    except ValueError as e:
        return _jsonify({"error": str(e)}, 400)


@app.route('/api/aps/masterdata/<section>/<key_value>', methods=['GET', 'PUT', 'PATCH', 'DELETE'])
def aps_masterdata_section_item(section, key_value):
    api_name = _section_sheet(section)
    if not api_name:
        return _jsonify({"error": f"Unknown master data section: {section}"}, 404)

    if request.method == "GET":
        row = store.get_row(api_name, key_value)
        return _jsonify({"item": row} if row else {"error": "Not found"}, 200 if row else 404)

    if request.method == "DELETE":
        try:
            store.delete_row(api_name, key_value)
            return _jsonify({"deleted": True, "key": key_value})
        except KeyError as e:
            return _jsonify({"error": str(e)}, 404)
        except ValueError as e:
            return _jsonify({"error": str(e)}, 400)

    payload = request.get_json(silent=True) or {}
    data = payload if isinstance(payload, dict) and "data" not in payload else payload.get("data", {})
    partial = request.method == "PATCH"

    try:
        current = store.get_row(api_name, key_value) or {}
        merged = current.copy()
        merged.update(data if partial else data)
        clean = _validate_masterdata_payload(section, merged, existing=current)
        return _jsonify({"item": store.update_row(api_name, key_value, clean, partial=False)})
    except KeyError as e:
        return _jsonify({"error": str(e)}, 404)
    except ValueError as e:
        return _jsonify({"error": str(e)}, 400)


@app.route('/api/aps/masterdata/<section>/bulk-replace', methods=['PUT'])
def aps_masterdata_bulk_replace(section):
    api_name = _section_sheet(section)
    if not api_name:
        return _jsonify({"error": f"Unknown master data section: {section}"}, 404)

    payload = request.get_json(silent=True) or {}
    items = payload.get("items", []) if isinstance(payload, dict) else []

    try:
        cleaned = [_validate_masterdata_payload(section, item) for item in items]
        return _jsonify(store.bulk_replace(api_name, cleaned))
    except ValueError as e:
        return _jsonify({"error": str(e)}, 400)


# ===== APS PLANNING WORKFLOW ENDPOINTS =====

@app.route('/api/aps/planning/orders/pool', methods=['GET'])
def aps_planning_orders_pool():
    """List all valid open sales orders for planning."""
    try:
        d = _load_all()
        so_df = d["sales_orders"].copy()
        rolling_mode_default = str(d["config"].get("ROLLING_MODE_DEFAULT", "HOT") or "HOT").strip().upper() or "HOT"

        orders = []
        for _, so in so_df.iterrows():
            orders.append({
                "so_id": str(so.get("SO_ID", "")).strip(),
                "customer_id": so.get("Customer"),
                "grade": so.get("Grade"),
                "section_mm": float(so.get("Section_mm", 0) or 0),
                "qty_mt": float(so.get("Order_Qty_MT", 0) or 0),
                "due_date": _safe(so.get("Delivery_Date")),
                "priority": so.get("Priority", "NORMAL"),
                "route_family": "SMS→RM",
                "status": so.get("Status", "Open"),
                "sku_id": so.get("SKU_ID"),
                "order_type": _clean_mode(so.get("Order_Type"), "MTO"),
                "rolling_mode": _clean_mode(so.get("Rolling_Mode"), rolling_mode_default),
            })

        return _jsonify({
            "total_orders": len(orders),
            "orders": orders,
        })
    except Exception as e:
        return _jsonify({"error": str(e), "traceback": traceback.format_exc()}, 500)


@app.route('/api/aps/planning/orders', methods=['GET'])
def aps_planning_orders_list():
    """Return the current in-memory planning orders (proposed/frozen/released)."""
    pos = list(getattr(aps_planning_orders_propose, "_planning_orders", []) or [])
    return _jsonify({"planning_orders": pos, "count": len(pos)})


@app.route('/api/aps/planning/window/select', methods=['POST'])
def aps_planning_window_select():
    """Select planning window and return candidate SOs."""
    try:
        payload = request.get_json(silent=True) or {}
        window_days = int(payload.get("days", 7) or 7)

        d = _load_all()
        all_sos = []
        rolling_mode_default = str(d["config"].get("ROLLING_MODE_DEFAULT", "HOT") or "HOT").strip().upper() or "HOT"
        for _, so in d["sales_orders"].iterrows():
            all_sos.append(
                SalesOrder(
                    so_id=str(so.get("SO_ID", "")).strip(),
                    customer_id=so.get("Customer"),
                    grade=so.get("Grade"),
                    section_mm=float(so.get("Section_mm", 0) or 0),
                    qty_mt=float(so.get("Order_Qty_MT", 0) or 0),
                    due_date=_safe(so.get("Delivery_Date")) or "2099-12-31",
                    priority=so.get("Priority", "NORMAL"),
                    route_family="SMS→RM",
                    status=so.get("Status", "Open"),
                    order_date=_safe(so.get("Order_Date")),
                    order_type=_clean_mode(so.get("Order_Type"), "MTO"),
                    rolling_mode=_clean_mode(so.get("Rolling_Mode"), rolling_mode_default),
                )
            )

        planner = APSPlanner(d["config"])
        window_map = {
            3: PlanningHorizon.NEXT_3_DAYS,
            7: PlanningHorizon.NEXT_7_DAYS,
            10: PlanningHorizon.NEXT_10_DAYS,
            14: PlanningHorizon.NEXT_14_DAYS,
        }
        window = window_map.get(window_days, PlanningHorizon.NEXT_7_DAYS)
        selected = planner.select_planning_window(all_sos, window)

        return _jsonify({
            "window_days": window_days,
            "candidate_count": len(selected),
            "candidates": [
                {
                    "so_id": so.so_id,
                    "customer_id": so.customer_id,
                    "grade": so.grade,
                    "section_mm": so.section_mm,
                    "qty_mt": so.qty_mt,
                    "due_date": so.due_date,
                    "priority": so.priority,
                    "hours_until_due": round(max(0, so.hours_until_due()), 2),
                    "order_type": so.order_type,
                    "rolling_mode": so.rolling_mode,
                }
                for so in selected
            ],
        })
    except Exception as e:
        return _jsonify({"error": str(e), "traceback": traceback.format_exc()}, 500)


@app.route('/api/aps/planning/orders/propose', methods=['POST'])
def aps_planning_orders_propose():
    """Propose Planning Orders from valid SOs in the selected window."""
    try:
        payload = request.get_json(silent=True) or {}
        window_days = int(payload.get("days", 7) or 7)
        selected_so_ids_raw = payload.get("so_ids", []) if isinstance(payload, dict) else []
        selected_so_ids = selected_so_ids_raw or []
        if selected_so_ids and not isinstance(selected_so_ids, list):
            selected_so_ids = [selected_so_ids]
        selected_so_ids = [str(x).strip() for x in selected_so_ids if x]
        if isinstance(selected_so_ids_raw, list) and len(selected_so_ids_raw) == 0:
            return _jsonify({"error": "Select at least one Sales Order before proposing Planning Orders."}, 400)

        d = _load_all()
        all_sos = []
        rolling_mode_default = str(d["config"].get("ROLLING_MODE_DEFAULT", "HOT") or "HOT").strip().upper() or "HOT"
        for _, so in d["sales_orders"].iterrows():
            all_sos.append(
                SalesOrder(
                    so_id=str(so.get("SO_ID", "")).strip(),
                    customer_id=so.get("Customer"),
                    grade=so.get("Grade"),
                    section_mm=float(so.get("Section_mm", 0) or 0),
                    qty_mt=float(so.get("Order_Qty_MT", 0) or 0),
                    due_date=_safe(so.get("Delivery_Date")) or "2099-12-31",
                    priority=so.get("Priority", "NORMAL"),
                    route_family="SMS→RM",
                    status=so.get("Status", "Open"),
                    order_date=_safe(so.get("Order_Date")),
                    order_type=_clean_mode(so.get("Order_Type"), "MTO"),
                    rolling_mode=_clean_mode(so.get("Rolling_Mode"), rolling_mode_default),
                )
            )

        # Apply planner rolling_mode overrides (set in UI per SO)
        rolling_overrides = {str(k).strip(): _clean_mode(v, rolling_mode_default)
                             for k, v in (payload.get("rolling_mode_overrides") or {}).items()}
        if rolling_overrides:
            for so in all_sos:
                if so.so_id in rolling_overrides:
                    so.rolling_mode = rolling_overrides[so.so_id]

        planner = APSPlanner(d["config"])
        window_map = {
            3: PlanningHorizon.NEXT_3_DAYS,
            7: PlanningHorizon.NEXT_7_DAYS,
            10: PlanningHorizon.NEXT_10_DAYS,
            14: PlanningHorizon.NEXT_14_DAYS,
        }
        window = window_map.get(window_days, PlanningHorizon.NEXT_7_DAYS)
        window_sos = planner.select_planning_window(all_sos, window)

        # If specific SOs were selected, filter to only those
        if selected_so_ids:
            window_sos = [so for so in window_sos if so.so_id in selected_so_ids]

        pos = planner.propose_planning_orders(window_sos)
        validation = planner.validate_planning_orders(pos)

        aps_planning_orders_propose._planning_orders = [po.to_dict() for po in pos]
        _invalidate_active_run_cache(clear_simulation=True)
        _publish_planning_orders_snapshot(
            aps_planning_orders_propose._planning_orders,
            workbook_data=d,
            solver_detail=(
                f"{len(aps_planning_orders_propose._planning_orders)} planning orders proposed. "
                "Derive heats and run Feasibility Check."
            ),
        )

        return _jsonify({
            "window_days": window_days,
            "po_count": len(pos),
            "validation": validation,
            "planning_orders": aps_planning_orders_propose._planning_orders,
        })
    except Exception as e:
        return _jsonify({"error": str(e), "traceback": traceback.format_exc()}, 500)

@app.route('/api/aps/planning/orders/update', methods=['POST'])
def aps_planning_orders_update():
    """Planner adjustments to Planning Orders (replace / merge / split / freeze / status)."""
    try:
        payload = request.get_json(silent=True) or {}

        current_orders = list(getattr(aps_planning_orders_propose, "_planning_orders", []) or [])
        if not current_orders:
            return _jsonify({
                "error": "No proposed planning orders found. Run /api/aps/planning/orders/propose first."
            }, 400)

        def _norm_po(po: dict) -> dict:
            po = dict(po or {})
            po_id = str(po.get("po_id", "")).strip()
            selected_so_ids = [str(x).strip() for x in (po.get("selected_so_ids") or []) if str(x).strip()]
            total_qty_mt = float(pd.to_numeric(pd.Series([po.get("total_qty_mt", 0)]), errors="coerce").fillna(0).iloc[0])
            heats_required = int(pd.to_numeric(pd.Series([po.get("heats_required", 0)]), errors="coerce").fillna(0).iloc[0])
            due_window = po.get("due_window", ("", ""))
            if isinstance(due_window, list):
                due_window = tuple(due_window)
            if not isinstance(due_window, tuple):
                due_window = ("", "")
            return {
                "po_id": po_id,
                "selected_so_ids": selected_so_ids,
                "total_qty_mt": total_qty_mt,
                "grade_family": str(po.get("grade_family", "")).strip(),
                "size_family": str(po.get("size_family", "")).strip(),
                "due_window": due_window,
                "route_family": str(po.get("route_family", "SMS→RM")).strip() or "SMS→RM",
                "heats_required": heats_required,
                "planner_status": str(po.get("planner_status", "PROPOSED")).strip() or "PROPOSED",
                "frozen_flag": bool(po.get("frozen_flag", False)),
                "order_type": str(po.get("order_type", "MTO")).strip().upper() or "MTO",
                "rolling_mode": str(po.get("rolling_mode", "HOT")).strip().upper() or "HOT",
            }

        def _validate_orders(orders: list[dict]) -> list[str]:
            issues = []
            seen_po_ids = set()
            seen_so_ids = set()

            for po in orders:
                po_id = str(po.get("po_id", "")).strip()
                if not po_id:
                    issues.append("Blank po_id found.")
                elif po_id in seen_po_ids:
                    issues.append(f"Duplicate po_id: {po_id}")
                seen_po_ids.add(po_id)

                if float(po.get("total_qty_mt", 0) or 0) <= 0:
                    issues.append(f"{po_id}: total_qty_mt must be > 0")

                if int(po.get("heats_required", 0) or 0) <= 0:
                    issues.append(f"{po_id}: heats_required must be > 0")

                so_ids = [str(x).strip() for x in (po.get("selected_so_ids") or []) if str(x).strip()]
                if not so_ids:
                    issues.append(f"{po_id}: selected_so_ids cannot be empty")

                for so_id in so_ids:
                    if so_id in seen_so_ids:
                        issues.append(f"SO {so_id} exists in more than one Planning Order")
                    seen_so_ids.add(so_id)

            return issues

        # Normalize current state first
        current_orders = [_norm_po(po) for po in current_orders]

        # Mode 1: full replacement from UI
        if isinstance(payload.get("planning_orders"), list):
            updated_orders = [_norm_po(po) for po in payload.get("planning_orders", [])]
            issues = _validate_orders(updated_orders)
            if issues:
                return _jsonify({
                    "updated": False,
                    "error": "Planning order validation failed",
                    "issues": issues,
                }, 400)

            aps_planning_orders_propose._planning_orders = updated_orders
            _invalidate_active_run_cache(clear_simulation=True)
            _publish_planning_orders_snapshot(
                updated_orders,
                solver_detail=(
                    f"Planning orders replaced ({len(updated_orders)} rows). "
                    "Re-run Derive and Feasibility Check."
                ),
            )
            return _jsonify({
                "updated": True,
                "mode": "replace",
                "po_count": len(updated_orders),
                "planning_orders": updated_orders,
                "issues": [],
            })

        # Mode 2: action-based incremental update
        action = str(payload.get("action", "")).strip().lower()
        if not action:
            return _jsonify({
                "error": "No update action provided. Use planning_orders=[...] for full replace or action=merge/split/freeze/status."
            }, 400)

        po_map = {po["po_id"]: dict(po) for po in current_orders}

        if action == "merge":
            source_po_ids = [str(x).strip() for x in (payload.get("source_po_ids") or []) if str(x).strip()]
            target_po_id = str(payload.get("target_po_id") or "").strip()

            if len(source_po_ids) < 2:
                return _jsonify({"error": "merge requires at least two source_po_ids"}, 400)
            if not target_po_id:
                return _jsonify({"error": "merge requires target_po_id"}, 400)
            if any(pid not in po_map for pid in source_po_ids):
                missing = [pid for pid in source_po_ids if pid not in po_map]
                return _jsonify({"error": f"Unknown planning orders for merge: {missing}"}, 404)

            source_orders = [po_map[pid] for pid in source_po_ids]
            grades = {po["grade_family"] for po in source_orders}
            if len(grades) > 1:
                return _jsonify({"error": "Cannot merge planning orders with different grade_family values"}, 400)

            merged_so_ids = []
            merged_qty = 0.0
            merged_heats = 0
            due_starts = []
            due_ends = []
            size_families = set()
            route_family = source_orders[0]["route_family"]
            frozen_flag = any(po.get("frozen_flag") for po in source_orders)

            for po in source_orders:
                merged_so_ids.extend(po["selected_so_ids"])
                merged_qty += float(po["total_qty_mt"] or 0)
                merged_heats += int(po["heats_required"] or 0)
                if po.get("size_family"):
                    size_families.update([x.strip() for x in str(po["size_family"]).split(",") if x.strip()])
                dw = po.get("due_window") or ("", "")
                if len(dw) >= 2:
                    if dw[0]:
                        due_starts.append(dw[0])
                    if dw[1]:
                        due_ends.append(dw[1])

            for pid in source_po_ids:
                po_map.pop(pid, None)

            po_map[target_po_id] = {
                "po_id": target_po_id,
                "selected_so_ids": sorted(set(merged_so_ids)),
                "total_qty_mt": round(merged_qty, 3),
                "grade_family": source_orders[0]["grade_family"],
                "size_family": ",".join(sorted(size_families)),
                "due_window": (
                    min(due_starts) if due_starts else "",
                    max(due_ends) if due_ends else "",
                ),
                "route_family": route_family,
                "heats_required": max(1, merged_heats),
                "planner_status": "MERGED",
                "frozen_flag": frozen_flag,
            }

        elif action == "split":
            source_po_id = str(payload.get("source_po_id") or "").strip()
            split_map = payload.get("split_map") or {}

            if not source_po_id:
                return _jsonify({"error": "split requires source_po_id"}, 400)
            if source_po_id not in po_map:
                return _jsonify({"error": f"Unknown source planning order: {source_po_id}"}, 404)
            if not isinstance(split_map, dict) or not split_map:
                return _jsonify({"error": "split requires split_map of {new_po_id: [so_ids...]}"}, 400)

            source = po_map[source_po_id]
            source_so_set = set(source["selected_so_ids"])
            assigned_so_set = set()

            for new_po_id, so_ids in split_map.items():
                clean_ids = {str(x).strip() for x in (so_ids or []) if str(x).strip()}
                if not clean_ids:
                    return _jsonify({"error": f"Split target {new_po_id} has no SOs"}, 400)
                if not clean_ids.issubset(source_so_set):
                    unknown = sorted(clean_ids - source_so_set)
                    return _jsonify({"error": f"Split target {new_po_id} contains SOs not in source PO: {unknown}"}, 400)
                if assigned_so_set.intersection(clean_ids):
                    overlap = sorted(assigned_so_set.intersection(clean_ids))
                    return _jsonify({"error": f"Split targets overlap on SOs: {overlap}"}, 400)
                assigned_so_set.update(clean_ids)

            if assigned_so_set != source_so_set:
                missing = sorted(source_so_set - assigned_so_set)
                extra = sorted(assigned_so_set - source_so_set)
                return _jsonify({
                    "error": "split_map must fully partition the source selected_so_ids",
                    "missing_from_split": missing,
                    "extra_in_split": extra,
                }, 400)

            # need SO master rows to recalc qty and due windows
            d = _load_all()
            so_df = d["all_orders"].copy()
            so_df["SO_ID"] = so_df["SO_ID"].astype(str).str.strip()
            so_map = {str(row["SO_ID"]).strip(): row for _, row in so_df.iterrows()}

            po_map.pop(source_po_id, None)

            for new_po_id, so_ids in split_map.items():
                clean_so_ids = [str(x).strip() for x in (so_ids or []) if str(x).strip()]
                rows = [so_map[so_id] for so_id in clean_so_ids if so_id in so_map]

                total_qty = float(sum(float(pd.to_numeric(pd.Series([r.get("Order_Qty_MT", 0)]), errors="coerce").fillna(0).iloc[0]) for r in rows))
                due_dates = [pd.to_datetime(r.get("Delivery_Date"), errors="coerce") for r in rows]
                due_dates = [d for d in due_dates if pd.notna(d)]
                section_vals = sorted({
                    float(pd.to_numeric(pd.Series([r.get("Section_mm", 0)]), errors="coerce").fillna(0).iloc[0])
                    for r in rows
                })

                heat_size = float(pd.to_numeric(pd.Series([d["config"].get("HEAT_SIZE_MT", 50)]), errors="coerce").fillna(50).iloc[0])
                heats_required = max(1, int(np.ceil(total_qty / heat_size))) if heat_size > 0 else 1

                po_map[str(new_po_id).strip()] = {
                    "po_id": str(new_po_id).strip(),
                    "selected_so_ids": clean_so_ids,
                    "total_qty_mt": round(total_qty, 3),
                    "grade_family": source["grade_family"],
                    "size_family": ",".join(f"{x:g}mm" for x in section_vals if x > 0),
                    "due_window": (
                        min(due_dates).date().isoformat() if due_dates else source["due_window"][0],
                        max(due_dates).date().isoformat() if due_dates else source["due_window"][1],
                    ),
                    "route_family": source["route_family"],
                    "heats_required": heats_required,
                    "planner_status": "SPLIT",
                    "frozen_flag": bool(source.get("frozen_flag", False)),
                }

        elif action in {"freeze", "unfreeze"}:
            po_id = str(payload.get("po_id") or "").strip()
            if not po_id or po_id not in po_map:
                return _jsonify({"error": f"Unknown planning order: {po_id}"}, 404)
            po_map[po_id]["frozen_flag"] = (action == "freeze")
            po_map[po_id]["planner_status"] = "FROZEN" if action == "freeze" else "PROPOSED"

        elif action == "status":
            po_id = str(payload.get("po_id") or "").strip()
            new_status = str(payload.get("planner_status") or "").strip().upper()
            if not po_id or po_id not in po_map:
                return _jsonify({"error": f"Unknown planning order: {po_id}"}, 404)
            if not new_status:
                return _jsonify({"error": "status action requires planner_status"}, 400)
            po_map[po_id]["planner_status"] = new_status

        else:
            return _jsonify({"error": f"Unsupported planning update action: {action}"}, 400)

        updated_orders = list(po_map.values())
        issues = _validate_orders(updated_orders)
        if issues:
            return _jsonify({
                "updated": False,
                "error": "Planning order validation failed after update",
                "issues": issues,
            }, 400)

        aps_planning_orders_propose._planning_orders = updated_orders
        _invalidate_active_run_cache(clear_simulation=True)
        _publish_planning_orders_snapshot(
            updated_orders,
            solver_detail=(
                f"Planning orders updated via '{action}' ({len(updated_orders)} rows). "
                "Re-run Derive and Feasibility Check."
            ),
        )

        return _jsonify({
            "updated": True,
            "mode": "action",
            "action": action,
            "po_count": len(updated_orders),
            "planning_orders": updated_orders,
            "issues": [],
        })
    except Exception as e:
        return _jsonify({"error": str(e), "traceback": traceback.format_exc()}, 500)


@app.route('/api/aps/planning/heats/derive', methods=['POST'])
def aps_planning_heats_derive():
    """Derive heats from proposed Planning Orders."""
    try:
        payload = request.get_json(silent=True) or {}
        planning_orders_data = payload.get("planning_orders") or getattr(aps_planning_orders_propose, "_planning_orders", [])

        from engine.aps_planner import PlanningOrder

        pos = []
        for po_data in planning_orders_data:
            pos.append(
                PlanningOrder(
                    po_id=po_data.get("po_id"),
                    selected_so_ids=po_data.get("selected_so_ids", []),
                    total_qty_mt=float(po_data.get("total_qty_mt", 0) or 0),
                    grade_family=po_data.get("grade_family"),
                    size_family=po_data.get("size_family", ""),
                    due_window=tuple(po_data.get("due_window", ("", ""))),
                    route_family=po_data.get("route_family", "SMS→RM"),
                    heats_required=int(po_data.get("heats_required", 1) or 1),
                    planner_status=po_data.get("planner_status", "PROPOSED"),
                    frozen_flag=bool(po_data.get("frozen_flag", False)),
                )
            )

        runtime_config = load_workbook_config_snapshot(WORKBOOK, reload_algorithm=True).runtime_config
        planner = APSPlanner(runtime_config)
        heat_size_mt = payload.get("heat_size_mt")
        if heat_size_mt is None:
            heat_size_mt = get_config(WORKBOOK, reload=True).get("HEAT_SIZE_MT", 50)
        heat_size_mt = float(heat_size_mt or 50)
        if heat_size_mt <= 0:
            heat_size_mt = 50.0
        heats = planner.derive_heat_batches(
            pos,
            heat_size_mt=heat_size_mt,
        )

        aps_planning_orders_propose._planning_orders = [po.to_dict() for po in pos]
        aps_planning_simulate._heat_batches = [h.to_dict() for h in heats]
        _invalidate_active_run_cache(clear_simulation=False)
        aps_planning_simulate._last_result = None
        aps_planning_simulate._scheduler_inputs = []
        _publish_planning_orders_snapshot(
            aps_planning_orders_propose._planning_orders,
            solver_detail=(
                f"Derived {len(heats)} heat batches from {len(aps_planning_orders_propose._planning_orders)} planning orders. "
                "Run Feasibility Check to validate finite capacity."
            ),
        )

        return _jsonify({
            "total_heats": len(heats),
            "total_mt": round(sum(float(h.qty_mt or 0) for h in heats), 3),
            "heats": [h.to_dict() for h in heats],
        })
    except Exception as e:
        return _jsonify({"error": str(e), "traceback": traceback.format_exc()}, 500)

@app.route('/api/aps/planning/simulate', methods=['POST'])
def aps_planning_simulate():
    """Authoritative planning simulation using Planning Orders bridged into the real scheduler."""
    try:
        payload = request.get_json(silent=True) or {}
        planning_orders_data = payload.get("planning_orders") or getattr(aps_planning_orders_propose, "_planning_orders", [])
        solver_sec = float(payload.get("solver_sec", 30.0) or 30.0)
        horizon_days = int(payload.get("horizon_days", 7) or 7)

        # Read remediation filters from payload
        priority_filter = str(payload.get("priority_filter", "") or "").strip().upper()
        rolling_mode_filter = str(payload.get("rolling_mode_filter", "") or "").strip().upper()
        grade_filter = str(payload.get("grade_filter", "") or "").strip().upper()
        num_sms = int(payload.get("num_sms", 1) or 1)
        num_rm = int(payload.get("num_rm", 1) or 1)
        horizon_hours = payload.get("horizon_hours")
        if horizon_hours:
            horizon_days = int(float(horizon_hours) / 24)

        if not planning_orders_data:
            return _jsonify({
                "authoritative": True,
                "feasible": False,
                "solver_status": "NO_DATA",
                "message": "No planning orders available for simulation",
                "total_duration_hours": 0.0,
                "sms_hours": 0.0,
                "rm_hours": 0.0,
                "schedule_rows": [],
            })

        filtered_orders = planning_orders_data
        if priority_filter:
            if priority_filter == "URGENT":
                filtered_orders = [po for po in planning_orders_data if str(po.get("priority", "")).upper() == "URGENT"]
            elif priority_filter == "URGENT+HIGH":
                filtered_orders = [po for po in planning_orders_data if str(po.get("priority", "")).upper() in ("URGENT", "HIGH")]

        if rolling_mode_filter in ("HOT", "COLD"):
            filtered_orders = [
                po for po in filtered_orders
                if str(po.get("rolling_mode", "HOT")).strip().upper() == rolling_mode_filter
            ]

        if grade_filter:
            filtered_orders = [
                po for po in filtered_orders
                if str(po.get("grade", po.get("grade_family", ""))).strip().upper() == grade_filter
            ]

        if not filtered_orders:
            return _jsonify({
                "authoritative": True,
                "feasible": False,
                "solver_status": "NO_MATCHING_ORDERS",
                "message": "No planning orders matched the selected simulation filters",
                "total_duration_hours": 0.0,
                "sms_hours": 0.0,
                "rm_hours": 0.0,
                "schedule_rows": [],
                "applied_filters": {
                    "priority_filter": priority_filter,
                    "rolling_mode_filter": rolling_mode_filter,
                    "grade_filter": grade_filter,
                    "num_sms": num_sms,
                    "num_rm": num_rm,
                },
            })


        d = _load_all()
        scheduler_inputs = _planning_orders_to_scheduler_campaigns(
            planning_orders=filtered_orders,
            all_orders_df=d["all_orders"],
            config=d["config"],
        )

        resources_df = d.get("resources", pd.DataFrame()).copy() if d.get("resources") is not None else pd.DataFrame()
        applied_resource_ids = []
        if not resources_df.empty and "Resource_ID" in resources_df.columns:
            resources_df = resources_df.copy()
            resources_df["Resource_ID"] = resources_df["Resource_ID"].astype(str).str.strip()
            resources_df["_rid_upper"] = resources_df["Resource_ID"].str.upper()

            def _limit_family(df: pd.DataFrame, pattern: str, keep_count: int) -> pd.DataFrame:
                keep_count = max(int(keep_count or 1), 1)
                mask = df["_rid_upper"].str.contains(pattern, case=False, na=False)
                family = df[mask].sort_values("Resource_ID", kind="stable")
                if family.empty:
                    return df
                keep_ids = set(family.head(keep_count)["Resource_ID"].tolist())
                return df[(~mask) | (df["Resource_ID"].isin(keep_ids))].copy()

            # Treat "SMS lines" as parallel lanes through the SMS families.
            for pattern in ["^EAF", "^LRF", "^VD", "^CCM"]:
                resources_df = _limit_family(resources_df, pattern, num_sms)

            # RM lines are independent parallel rolling mills.
            resources_df = _limit_family(resources_df, "^RM", num_rm)
            applied_resource_ids = resources_df["Resource_ID"].tolist()
            resources_df = resources_df.drop(columns=["_rid_upper"], errors="ignore")

        # Suppress stdout during scheduling to avoid Unicode encoding errors on Windows
        import io
        old_stdout = sys.stdout
        try:
            sys.stdout = io.StringIO()
            result = schedule(
                scheduler_inputs,
                resources_df,
                planning_start=datetime.now(),
                planning_horizon_days=horizon_days,
                routing=d["routing"],
                queue_times=d["queue_times"],
                changeover_matrix=d["changeover_matrix"],
                config=d["config"],
                solver_time_limit_sec=solver_sec,
            )
        finally:
            sys.stdout = old_stdout

        heat_df = result.get("heat_schedule", pd.DataFrame())
        camp_df = result.get("campaign_schedule", pd.DataFrame())
        solver_status = result.get("solver_status", "UNKNOWN")
        solver_detail = result.get("solver_detail", "")

        if heat_df is not None and not heat_df.empty and "Planned_Start" in heat_df.columns and "Planned_End" in heat_df.columns:
            starts = pd.to_datetime(heat_df["Planned_Start"], errors="coerce")
            ends = pd.to_datetime(heat_df["Planned_End"], errors="coerce")
            total_duration_hours = round(((ends.max() - starts.min()).total_seconds() / 3600.0), 2)
        else:
            total_duration_hours = 0.0

        sms_hours = 0.0
        rm_hours = 0.0
        sms_span_hours = 0.0
        rm_span_hours = 0.0
        if heat_df is not None and not heat_df.empty and {"Operation", "Duration_Hrs"}.issubset(set(heat_df.columns)):
            dur = pd.to_numeric(heat_df["Duration_Hrs"], errors="coerce").fillna(0.0)
            ops = heat_df["Operation"].astype(str).str.upper()
            sms_mask = ops.isin(["EAF", "LRF", "VD", "CCM"])
            rm_mask = ops.eq("RM")
            sms_hours = round(dur[sms_mask].sum(), 2)
            rm_hours = round(dur[rm_mask].sum(), 2)

            if {"Planned_Start", "Planned_End"}.issubset(set(heat_df.columns)):
                def _family_span(mask) -> float:
                    subset = heat_df.loc[mask]
                    if subset.empty:
                        return 0.0
                    starts = pd.to_datetime(subset["Planned_Start"], errors="coerce").dropna()
                    ends = pd.to_datetime(subset["Planned_End"], errors="coerce").dropna()
                    if starts.empty or ends.empty:
                        return 0.0
                    return round(((ends.max() - starts.min()).total_seconds() / 3600.0), 2)

                sms_span_hours = _family_span(sms_mask)
                rm_span_hours = _family_span(rm_mask)

        horizon_hours = round(float(max(horizon_days * 24, 1)), 2)
        has_schedule = heat_df is not None and not heat_df.empty
        solver_found_schedule = solver_status not in {"INFEASIBLE", "MODEL_INVALID"} and has_schedule
        overflow_hours = round(max(total_duration_hours - horizon_hours, 0.0), 2)
        horizon_exceeded = solver_found_schedule and overflow_hours > 0.01
        feasible = solver_found_schedule and not horizon_exceeded

        if feasible:
            message = "Finite schedule generated within selected horizon"
        elif horizon_exceeded:
            message = (
                f"Finite schedule needs {total_duration_hours:.1f}h, exceeding the "
                f"{horizon_hours:.1f}h horizon by {overflow_hours:.1f}h"
            )
        elif not has_schedule:
            message = "No finite schedule rows were generated for the selected planning orders"
        else:
            message = "Finite schedule is infeasible with current APS planning orders / master data"

        # Count HOT vs COLD planning orders
        hot_count = sum(1 for campaign in scheduler_inputs if campaign.get("hot_charging", True))
        cold_count = len(scheduler_inputs) - hot_count

        allow_defaults = bool(result.get("allow_default_masters", False)) or _config_flag(
            d.get("config"), "Allow_Scheduler_Default_Masters", "N"
        )
        try:
            demand_hrs = compute_demand_hours(
                scheduler_inputs,
                resources_df,
                routing=d["routing"],
                changeover_matrix=d["changeover_matrix"],
                allow_defaults=allow_defaults,
            )
            cap_df = capacity_map(demand_hrs, resources_df, horizon_days=horizon_days)
        except Exception:
            cap_df = pd.DataFrame()

        _publish_planning_snapshot(
            campaigns=scheduler_inputs,
            skus=d.get("skus"),
            heat_df=heat_df,
            camp_df=camp_df,
            cap_df=cap_df,
            solver_status="HORIZON EXCEEDED" if horizon_exceeded else solver_status,
            solver_detail=message,
        )

        planning_signature = _planning_orders_release_signature(planning_orders_data)
        aps_planning_simulate._last_result = {
            "authoritative": True,
            "feasible": feasible,
            "solver_status": solver_status,
            "solver_detail": solver_detail,
            "horizon_hours": horizon_hours,
            "overflow_hours": overflow_hours,
            "horizon_exceeded": horizon_exceeded,
            "total_duration_hours": total_duration_hours,
            "message": message,
            "planning_signature": planning_signature,
        }
        _state["solver_status"] = "HORIZON EXCEEDED" if horizon_exceeded else solver_status
        _state["solver_detail"] = message
        aps_planning_simulate._scheduler_inputs = scheduler_inputs
        aps_planning_orders_propose._planning_orders = planning_orders_data

        return _jsonify({
            "authoritative": True,
            "feasible": feasible,
            "solver_status": solver_status,
            "solver_detail": solver_detail,
            "total_duration_hours": total_duration_hours,
            "horizon_hours": horizon_hours,
            "overflow_hours": overflow_hours,
            "horizon_exceeded": horizon_exceeded,
            "sms_span_hours": sms_span_hours,
            "rm_span_hours": rm_span_hours,
            "sms_hours": sms_hours,
            "rm_hours": rm_hours,
            "hot_planning_orders": hot_count,
            "cold_planning_orders": cold_count,
            "schedule_rows": _df_to_records(heat_df),
            "load_factor": f"{round((total_duration_hours / max(horizon_hours, 1)) * 100, 1)}%",
            "applied_filters": {
                "priority_filter": priority_filter,
                "rolling_mode_filter": rolling_mode_filter,
                "grade_filter": grade_filter,
                "num_sms": num_sms,
                "num_rm": num_rm,
                "resource_count": len(applied_resource_ids),
            },
            "message": message,
        })
    except Exception as e:
        return _jsonify({"error": str(e), "traceback": traceback.format_exc()}, 500)

@app.route('/api/aps/planning/release', methods=['POST'])
def aps_planning_release():
    """Release approved Planning Orders to execution and persist APS release markers to Sales_Orders."""
    try:
        payload = request.get_json(silent=True) or {}
        po_ids = [str(x).strip() for x in (payload.get("po_ids") or []) if str(x).strip()]

        if not po_ids:
            return _jsonify({"error": "po_ids is required and cannot be empty"}, 400)

        planning_orders = list(getattr(aps_planning_orders_propose, "_planning_orders", []) or [])
        if not planning_orders:
            return _jsonify({
                "error": "No planning orders available for release. Run /api/aps/planning/orders/propose first."
            }, 400)

        po_map = {
            str(po.get("po_id", "")).strip(): dict(po)
            for po in planning_orders
            if str(po.get("po_id", "")).strip()
        }

        missing_po_ids = [po_id for po_id in po_ids if po_id not in po_map]
        if missing_po_ids:
            return _jsonify({
                "error": "Some requested planning orders do not exist",
                "missing_po_ids": missing_po_ids,
            }, 404)

        sim_result = getattr(aps_planning_simulate, "_last_result", None)
        if sim_result is None:
            return _jsonify({"error": "Run Planning Simulation before release."}, 400)
        if not bool(sim_result.get("authoritative", False)):
            return _jsonify({"error": "Planning simulation is not authoritative. Re-run simulation before release."}, 400)
        if not bool(sim_result.get("feasible", False)):
            return _jsonify({"error": "Cannot release planning orders because the current simulation is infeasible."}, 400)

        current_signature = _planning_orders_release_signature(planning_orders)
        simulated_signature = str(sim_result.get("planning_signature") or "").strip()
        if not simulated_signature:
            return _jsonify({"error": "Planning simulation is stale. Re-run simulation before release."}, 400)
        if simulated_signature != current_signature:
            return _jsonify({
                "error": "Planning orders changed after simulation. Re-run simulation before release.",
            }, 409)

        released_orders = [po_map[po_id] for po_id in po_ids]

        so_to_po = {}
        for po in released_orders:
            po_id_for_map = str(po.get("po_id") or po.get("PO_ID") or "").strip()
            if not po_id_for_map:
                return _jsonify({
                    "error": "Invalid planning order payload: missing po_id"
                }, 400)
            raw_so_ids = po.get("selected_so_ids") or po.get("Selected_SO_IDs") or []
            if isinstance(raw_so_ids, str):
                raw_so_ids = [x.strip() for x in raw_so_ids.split(",") if str(x).strip()]
            for so_id in raw_so_ids:
                clean_so = str(so_id).strip()
                if clean_so:
                    if clean_so in so_to_po and so_to_po[clean_so] != po_id_for_map:
                        return _jsonify({
                            "error": f"SO {clean_so} is mapped to multiple release planning orders",
                            "existing_po": so_to_po[clean_so],
                            "new_po": po_id_for_map,
                        }, 400)
                    so_to_po[clean_so] = po_id_for_map

        if not so_to_po:
            return _jsonify({"error": "Selected planning orders have no sales orders to release"}, 400)

        result = store.list_rows("sales-orders", filters=[], limit=5000)
        sos = result.get("items", []) or []
        existing_so_ids = {str(so.get("SO_ID", "")).strip() for so in sos if str(so.get("SO_ID", "")).strip()}

        missing_sos = sorted([so_id for so_id in so_to_po if so_id not in existing_so_ids])
        if missing_sos:
            return _jsonify({
                "error": "Some Sales Orders linked to planning orders were not found in workbook",
                "missing_so_ids": missing_sos,
            }, 404)

        updated_sales_orders = []
        for so_id, po_id in so_to_po.items():
            update_data = {
                "Status": "Planned",
                "Campaign_ID": po_id,
                "Campaign_Group": "APS_RELEASED",
            }
            try:
                item = store.update_row("sales-orders", so_id, update_data, partial=True)
                updated_sales_orders.append(item)
            except PermissionError as pe:
                return _jsonify({"error": "Cannot write to Excel file - it may be open in another application. Please close the file and try again.", "details": str(pe)}, 409)
            except BadZipFile as ze:
                return _jsonify({
                    "error": "Workbook is not readable as a valid .xlsx file (zip container). Close Excel/OneDrive sync and restore a valid workbook, then retry unrelease.",
                    "details": str(ze),
                    "workbook": str(WORKBOOK),
                }, 409)
            except Exception as e:
                return _jsonify({"error": f"Failed to update sales order {so_id}: {str(e)}"}, 400)

        refreshed_orders = []
        for po in planning_orders:
            po_id = str(po.get("po_id", "")).strip()
            if po_id in po_ids:
                po = dict(po)
                po["planner_status"] = "RELEASED"
                po["released_at"] = datetime.now().isoformat()
            refreshed_orders.append(po)

        existing_heat_df = (
            _state.get("heat_schedule").copy()
            if isinstance(_state.get("heat_schedule"), pd.DataFrame)
            else pd.DataFrame()
        )
        existing_camp_df = (
            _state.get("camp_schedule").copy()
            if isinstance(_state.get("camp_schedule"), pd.DataFrame)
            else pd.DataFrame()
        )
        existing_cap_df = (
            _state.get("capacity").copy()
            if isinstance(_state.get("capacity"), pd.DataFrame)
            else pd.DataFrame()
        )
        previous_solver_status = str(_state.get("solver_status", "NOT RUN"))
        previous_solver_detail = str(_state.get("solver_detail", ""))

        aps_planning_orders_propose._planning_orders = refreshed_orders
        aps_planning_release._released_po_ids = po_ids
        aps_planning_release._released_sales_orders = list(so_to_po.keys())
        _invalidate_active_run_cache(clear_simulation=False)
        d = _load_all()
        planning_campaigns = _planning_orders_to_scheduler_campaigns(
            planning_orders=refreshed_orders,
            all_orders_df=d["all_orders"],
            config=d["config"],
        )
        _publish_planning_snapshot(
            campaigns=planning_campaigns,
            skus=d.get("skus"),
            heat_df=existing_heat_df,
            camp_df=existing_camp_df,
            cap_df=existing_cap_df,
            solver_status=previous_solver_status,
            solver_detail=previous_solver_detail,
        )

        total_mt = round(sum(float(po.get("total_qty_mt", 0) or 0) for po in released_orders), 3)
        total_heats = int(sum(int(po.get("heats_required", 0) or 0) for po in released_orders))

        return _jsonify({
            "released": True,
            "po_count": len(released_orders),
            "released_po_ids": po_ids,
            "sales_order_count": len(so_to_po),
            "released_so_ids": sorted(so_to_po.keys()),
            "total_mt": total_mt,
            "total_heats": total_heats,
            "message": "Planning orders released to execution",
        })
    except BadZipFile as ze:
        return _jsonify({
            "error": "Workbook is not readable as a valid .xlsx file (zip container). Close Excel/OneDrive sync and restore a valid workbook, then retry release.",
            "details": str(ze),
            "workbook": str(WORKBOOK),
        }, 409)
    except Exception as e:
        return _jsonify({"error": str(e), "traceback": traceback.format_exc()}, 500)


@app.route('/api/aps/planning/unrelease', methods=['POST'])
def aps_planning_unrelease():
    """Return released Planning Orders from execution back to planning and clear APS release markers."""
    try:
        payload = request.get_json(silent=True) or {}
        po_ids = [str(x).strip() for x in (payload.get("po_ids") or []) if str(x).strip()]

        if not po_ids:
            return _jsonify({"error": "po_ids is required and cannot be empty"}, 400)

        planning_orders = list(getattr(aps_planning_orders_propose, "_planning_orders", []) or [])
        d = _load_all()
        reconstructed = _released_sales_orders_to_planning_orders(d)
        po_map = {}
        for po in planning_orders + reconstructed:
            po_id = str(po.get("po_id", "")).strip()
            if po_id and po_id not in po_map:
                po_map[po_id] = dict(po)

        missing_po_ids = [po_id for po_id in po_ids if po_id not in po_map]
        if missing_po_ids:
            return _jsonify({
                "error": "Some requested planning orders do not exist",
                "missing_po_ids": missing_po_ids,
            }, 404)

        non_released_po_ids = [
            po_id for po_id in po_ids
            if str(po_map[po_id].get("planner_status", "")).strip().upper() != "RELEASED"
        ]
        if non_released_po_ids:
            return _jsonify({
                "error": "Only released planning orders can be unreleased",
                "non_released_po_ids": non_released_po_ids,
            }, 400)

        result = store.list_rows("sales-orders", filters=[], limit=5000)
        sos = result.get("items", []) or []
        so_lookup = {
            str(so.get("SO_ID", "")).strip(): so
            for so in sos
            if str(so.get("SO_ID", "")).strip()
        }

        so_to_po: Dict[str, str] = {}
        for po_id in po_ids:
            po = po_map[po_id]
            for so_id in (po.get("selected_so_ids") or []):
                clean_so = str(so_id).strip()
                if clean_so:
                    so_to_po[clean_so] = po_id

        for so in sos:
            so_id = str(so.get("SO_ID", "")).strip()
            so_po_id = str(so.get("Campaign_ID", "")).strip()
            so_group = str(so.get("Campaign_Group", "")).strip().upper()
            if so_id and so_po_id in po_ids and so_group == "APS_RELEASED":
                so_to_po[so_id] = so_po_id

        if not so_to_po:
            return _jsonify({"error": "Selected planning orders have no released sales orders to unrelease"}, 400)

        missing_sos = sorted([so_id for so_id in so_to_po if so_id not in so_lookup])
        if missing_sos:
            return _jsonify({
                "error": "Some Sales Orders linked to planning orders were not found in workbook",
                "missing_so_ids": missing_sos,
            }, 404)

        updated_sales_orders = []
        for so_id, po_id in so_to_po.items():
            row = so_lookup.get(so_id, {})
            row_po_id = str(row.get("Campaign_ID", "")).strip()
            row_group = str(row.get("Campaign_Group", "")).strip().upper()
            if row_po_id and row_po_id != po_id and row_group == "APS_RELEASED":
                return _jsonify({
                    "error": f"Sales order {so_id} is released under a different planning order",
                    "existing_po": row_po_id,
                    "requested_po": po_id,
                }, 409)

            update_data = {
                "Status": "OPEN",
                "Campaign_ID": "",
                "Campaign_Group": "",
            }
            try:
                item = store.update_row("sales-orders", so_id, update_data, partial=True)
                updated_sales_orders.append(item)
            except PermissionError as pe:
                return _jsonify({"error": "Cannot write to Excel file - it may be open in another application. Please close the file and try again.", "details": str(pe)}, 409)
            except BadZipFile as ze:
                return _jsonify({
                    "error": "Workbook is not readable as a valid .xlsx file (zip container). Close Excel/OneDrive sync and restore a valid workbook, then retry release.",
                    "details": str(ze),
                    "workbook": str(WORKBOOK),
                }, 409)
            except Exception as e:
                return _jsonify({"error": f"Failed to update sales order {so_id}: {str(e)}"}, 400)

        refreshed_orders = []
        seen_po_ids = set()
        for po in planning_orders + reconstructed:
            po_id = str(po.get("po_id", "")).strip()
            if not po_id or po_id in seen_po_ids:
                continue
            seen_po_ids.add(po_id)
            po = dict(po)
            if po_id in po_ids:
                po["planner_status"] = "PROPOSED"
                po.pop("released_at", None)
                po["unreleased_at"] = datetime.now().isoformat()
            refreshed_orders.append(po)

        existing_heat_df = (
            _state.get("heat_schedule").copy()
            if isinstance(_state.get("heat_schedule"), pd.DataFrame)
            else pd.DataFrame()
        )
        existing_camp_df = (
            _state.get("camp_schedule").copy()
            if isinstance(_state.get("camp_schedule"), pd.DataFrame)
            else pd.DataFrame()
        )
        existing_cap_df = (
            _state.get("capacity").copy()
            if isinstance(_state.get("capacity"), pd.DataFrame)
            else pd.DataFrame()
        )
        previous_solver_status = str(_state.get("solver_status", "NOT RUN"))
        previous_solver_detail = str(_state.get("solver_detail", ""))

        aps_planning_orders_propose._planning_orders = refreshed_orders
        aps_planning_release._released_po_ids = [
            str(po.get("po_id", "")).strip()
            for po in refreshed_orders
            if str(po.get("planner_status", "")).strip().upper() == "RELEASED"
        ]
        aps_planning_release._released_sales_orders = [
            so_id
            for po in refreshed_orders
            if str(po.get("planner_status", "")).strip().upper() == "RELEASED"
            for so_id in (po.get("selected_so_ids") or [])
            if str(so_id).strip()
        ]
        _invalidate_active_run_cache(clear_simulation=False)
        d = _load_all()
        planning_campaigns = _planning_orders_to_scheduler_campaigns(
            planning_orders=refreshed_orders,
            all_orders_df=d["all_orders"],
            config=d["config"],
        )
        _publish_planning_snapshot(
            campaigns=planning_campaigns,
            skus=d.get("skus"),
            heat_df=existing_heat_df,
            camp_df=existing_camp_df,
            cap_df=existing_cap_df,
            solver_status=previous_solver_status,
            solver_detail=previous_solver_detail,
        )

        total_mt = round(sum(float(po_map[po_id].get("total_qty_mt", 0) or 0) for po_id in po_ids), 3)
        total_heats = int(sum(int(po_map[po_id].get("heats_required", 0) or 0) for po_id in po_ids))

        return _jsonify({
            "unreleased": True,
            "po_count": len(po_ids),
            "unreleased_po_ids": po_ids,
            "sales_order_count": len(so_to_po),
            "unreleased_so_ids": sorted(so_to_po.keys()),
            "total_mt": total_mt,
            "total_heats": total_heats,
            "message": "Planning orders returned from execution to planning",
        })
    except BadZipFile as ze:
        return _jsonify({
            "error": "Workbook is not readable as a valid .xlsx file (zip container). Close Excel/OneDrive sync and restore a valid workbook, then retry release.",
            "details": str(ze),
            "workbook": str(WORKBOOK),
        }, 409)
    except Exception as e:
        return _jsonify({"error": str(e), "traceback": traceback.format_exc()}, 500)

@app.route('/api/meta/xaps/routes')
def xaps_route_manifest():
    return _jsonify({
        'legacy': [
            '/api/health',
            '/api/data/dashboard',
            '/api/data/config',
            '/api/data/orders',
            '/api/data/skus',
            '/api/data/campaigns',
            '/api/data/gantt',
            '/api/data/capacity',
            '/api/run/bom',
            '/api/run/schedule',
            '/api/run/ctp',
            '/api/orders',
            '/api/orders/<so_id>',
            '/api/orders/assign',
        ],
        'application': [
            # Dashboard
            '/api/aps/dashboard/overview',

            # Orders
            '/api/aps/orders/list',
            '/api/aps/orders',
            '/api/aps/orders/<so_id>',
            '/api/aps/orders/assign',

            # APS planning workflow
            '/api/aps/planning/orders/pool',
            '/api/aps/planning/window/select',
            '/api/aps/planning/orders/propose',
            '/api/aps/planning/orders/update',
            '/api/aps/planning/heats/derive',
            '/api/aps/planning/simulate',
            '/api/aps/planning/release',

            # Campaigns
            '/api/aps/campaigns/list',
            '/api/aps/campaigns/release-queue',
            '/api/aps/campaigns/<campaign_id>',
            '/api/aps/campaigns/<campaign_id>/status',

            # Schedule
            '/api/aps/schedule/gantt',
            '/api/aps/schedule/run',
            '/api/aps/schedule/jobs/<job_id>',
            '/api/aps/schedule/jobs/<job_id>/reschedule',

            # Dispatch
            '/api/aps/dispatch/board',
            '/api/aps/dispatch/resources/<resource_id>',

            # Capacity
            '/api/aps/capacity/map',
            '/api/aps/capacity/bottlenecks',

            # Material
            '/api/aps/material/plan',
            '/api/aps/material/holds',

            # CTP
            '/api/aps/ctp/check',
            '/api/aps/ctp/requests',
            '/api/aps/ctp/output',

            # Scenarios
            '/api/aps/scenarios/list',
            '/api/aps/scenarios',
            '/api/aps/scenarios/<key_value>',
            '/api/aps/scenarios/output',
            '/api/aps/scenarios/apply',

            # Master data
            '/api/aps/masterdata',
            '/api/aps/masterdata/<section>',
            '/api/aps/masterdata/<section>/<key_value>',
            '/api/aps/masterdata/<section>/bulk-replace',
        ],
        'notes': {
            'planning_model': 'SO → PlanningOrder → HeatBatch → finite schedule',
            'authoritative_scheduler': '/api/aps/planning/simulate and /api/run/schedule',
            'legacy_routes_retained': True,
        }
    })


def _planning_orders_to_scheduler_campaigns(
    planning_orders: List[Dict[str, Any]],
    all_orders_df: pd.DataFrame,
    config: Dict[str, Any],
) -> List[Dict[str, Any]]:
    """
    Bridge APS PlanningOrder objects into the legacy scheduler's expected campaign shape.

    This lets the real scheduler stay unchanged while APS becomes the primary planning model.
    """
    from engine.campaign import billet_family_for_grade, needs_vd_for_grade, priority_rank

    if all_orders_df is None or all_orders_df.empty:
        all_orders_df = pd.DataFrame()

    so_df = all_orders_df.copy()
    if "SO_ID" in so_df.columns:
        so_df["SO_ID"] = so_df["SO_ID"].astype(str).str.strip()
    so_map = {
        str(row.get("SO_ID", "")).strip(): row
        for _, row in so_df.iterrows()
        if str(row.get("SO_ID", "")).strip()
    }

    campaigns: List[Dict[str, Any]] = []
    heat_size = float(pd.to_numeric(pd.Series([config.get("HEAT_SIZE_MT", 50)]), errors="coerce").fillna(50).iloc[0])

    for po in planning_orders:
        po_id = str(po.get("po_id", "")).strip()
        if not po_id:
            continue

        so_ids = [str(x).strip() for x in (po.get("selected_so_ids") or []) if str(x).strip()]
        source_rows = [so_map[x] for x in so_ids if x in so_map]

        priorities = [str(r.get("Priority", "NORMAL")).strip().upper() for r in source_rows] or ["NORMAL"]
        if "URGENT" in priorities:
            priority = "URGENT"
        elif "HIGH" in priorities:
            priority = "HIGH"
        elif "LOW" in priorities:
            priority = "LOW"
        else:
            priority = "NORMAL"

        sections = []
        for r in source_rows:
            try:
                sections.append(float(pd.to_numeric(pd.Series([r.get("Section_mm", 0)]), errors="coerce").fillna(0).iloc[0]))
            except Exception:
                pass

        if not sections:
            for x in str(po.get("size_family", "")).split(","):
                x = x.strip().replace("mm", "")
                try:
                    sections.append(float(x))
                except Exception:
                    pass

        section_mm = min(sections) if sections else float(pd.to_numeric(pd.Series([config.get("Default_Section_Fallback", 6.5)]), errors="coerce").fillna(6.5).iloc[0])

        due_window = po.get("due_window", ("", ""))
        due_value = None
        if isinstance(due_window, (list, tuple)) and len(due_window) >= 2:
            due_value = due_window[1]
        due_ts = pd.to_datetime(due_value, errors="coerce")
        if pd.isna(due_ts):
            due_ts = pd.Timestamp.now() + pd.Timedelta(days=int(config.get("Planning_Horizon_Days", 7) or 7))

        total_mt = float(pd.to_numeric(pd.Series([po.get("total_qty_mt", 0)]), errors="coerce").fillna(0).iloc[0])
        heats_required = int(pd.to_numeric(pd.Series([po.get("heats_required", 0)]), errors="coerce").fillna(0).iloc[0])
        if heats_required <= 0:
            heats_required = max(1, int(np.ceil(total_mt / heat_size))) if heat_size > 0 else 1

        grade = str(po.get("grade_family", "")).strip()
        route_family = str(po.get("route_family", "SMS→RM")).strip() or "SMS→RM"
        planner_status = str(po.get("planner_status", "PROPOSED")).strip().upper()

        production_orders = []
        if source_rows:
            for idx, row in enumerate(source_rows, start=1):
                qty_mt = float(pd.to_numeric(pd.Series([row.get("Order_Qty_MT", row.get("Order_Qty", 0))]), errors="coerce").fillna(0).iloc[0])
                po_due = pd.to_datetime(row.get("Delivery_Date"), errors="coerce")
                production_orders.append({
                    "production_order_id": f"{po_id}-SO{idx:02d}",
                    "so_id": str(row.get("SO_ID", "")).strip(),
                    "sku_id": str(row.get("SKU_ID", "")).strip(),
                    "section_mm": float(pd.to_numeric(pd.Series([row.get("Section_mm", section_mm)]), errors="coerce").fillna(section_mm).iloc[0]),
                    "qty_mt": qty_mt,
                    "due_date": po_due if pd.notna(po_due) else due_ts,
                    "priority_rank": priority_rank(row.get("Priority", priority), config=config),
                    "priority": str(row.get("Priority", priority)).strip().upper(),
                })
        else:
            production_orders.append({
                "production_order_id": f"{po_id}-PO01",
                "so_id": ", ".join(so_ids),
                "sku_id": "",
                "section_mm": section_mm,
                "qty_mt": total_mt,
                "due_date": due_ts,
                "priority_rank": priority_rank(priority, config=config),
                "priority": priority,
            })

        order_type = str(po.get("order_type", "MTO")).strip().upper() or "MTO"
        rolling_mode = str(po.get("rolling_mode", "HOT")).strip().upper() or "HOT"
        hot_charging = rolling_mode == "HOT"

        campaigns.append({
            "campaign_id": po_id,
            "campaign_group": "APS",
            "grade": grade,
            "heats": heats_required,
            "needs_vd": bool(needs_vd_for_grade(grade)),
            "billet_family": billet_family_for_grade(grade),
            "due_date": due_ts,
            "priority": priority,
            "priority_rank": priority_rank(priority, config=config),
            "total_coil_mt": total_mt,
            "total_mt": total_mt,
            "section_mm": section_mm,
            "so_ids": so_ids,
            "release_status": "RELEASED" if planner_status == "RELEASED" else "PLANNED",
            "planner_status": planner_status,
            "sku_attributes": {
                "sections_covered": str(po.get("size_family", "")),
                "route_family": route_family,
            },
            "production_orders": production_orders,
            "order_type": order_type,
            "rolling_mode": rolling_mode,
            "hot_charging": hot_charging,
        })

    return campaigns


def _mode_or_default(series, default: str) -> str:
    """Extract modal non-null value from Series or return default."""
    if series is None or series.empty:
        return default
    mode_vals = series.dropna().astype(str).str.strip()
    if mode_vals.empty:
        return default
    mode_result = mode_vals.mode()
    if mode_result.empty:
        return default
    return str(mode_result.iloc[0]).strip().upper() or default


def _released_sales_orders_to_planning_orders(d: Dict[str, Any]) -> List[Dict[str, Any]]:
    """
    Reconstruct released APS planning orders from Sales_Orders workbook rows.
    Uses Campaign_ID = PO_ID and Campaign_Group = APS_RELEASED written by release step.
    """
    so_df = d["all_orders"].copy()
    if so_df.empty:
        return []

    for col in ["SO_ID", "Campaign_ID", "Campaign_Group", "Grade", "SKU_ID", "Priority", "Status"]:
        if col in so_df.columns:
            so_df[col] = so_df[col].astype(str).str.strip()

    released_mask = (
        so_df.get("Campaign_Group", pd.Series([""] * len(so_df))).astype(str).str.strip().str.upper().eq("APS_RELEASED")
        & so_df.get("Campaign_ID", pd.Series([""] * len(so_df))).astype(str).str.strip().ne("")
    )
    aps_rows = so_df[released_mask].copy()
    if aps_rows.empty:
        return []

    aps_rows["Order_Qty_MT"] = pd.to_numeric(aps_rows.get("Order_Qty_MT"), errors="coerce").fillna(
        pd.to_numeric(aps_rows.get("Order_Qty"), errors="coerce").fillna(0.0)
    )
    aps_rows["Section_mm"] = pd.to_numeric(aps_rows.get("Section_mm"), errors="coerce").fillna(
        float(pd.to_numeric(pd.Series([d["config"].get("Default_Section_Fallback", 6.5)]), errors="coerce").fillna(6.5).iloc[0])
    )
    aps_rows["Delivery_Date"] = pd.to_datetime(aps_rows.get("Delivery_Date"), errors="coerce")

    heat_size = float(pd.to_numeric(pd.Series([d["config"].get("HEAT_SIZE_MT", 50)]), errors="coerce").fillna(50).iloc[0])
    rolling_mode_default = str(d["config"].get("ROLLING_MODE_DEFAULT", "HOT") or "HOT").strip().upper() or "HOT"

    planning_orders: List[Dict[str, Any]] = []
    for po_id, grp in aps_rows.groupby("Campaign_ID", dropna=False):
        po_id = str(po_id).strip()
        if not po_id:
            continue

        grade_mode = grp["Grade"].mode()
        grade = str(grade_mode.iloc[0]).strip() if not grade_mode.empty else ""

        due_dates = grp["Delivery_Date"].dropna()
        sections = sorted({float(x) for x in grp["Section_mm"].dropna().tolist()})
        total_qty = float(grp["Order_Qty_MT"].sum())
        heats_required = max(1, int(np.ceil(total_qty / heat_size))) if heat_size > 0 else 1

        # Reconstruct order_type and rolling_mode
        order_type = _mode_or_default(grp.get("Order_Type"), "MTO")
        rolling_mode = _mode_or_default(grp.get("Rolling_Mode"), rolling_mode_default)

        planning_orders.append({
            "po_id": po_id,
            "selected_so_ids": [str(x).strip() for x in grp["SO_ID"].tolist() if str(x).strip()],
            "total_qty_mt": round(total_qty, 3),
            "grade_family": grade,
            "size_family": ",".join(f"{x:g}mm" for x in sections),
            "due_window": (
                due_dates.min().date().isoformat() if not due_dates.empty else "",
                due_dates.max().date().isoformat() if not due_dates.empty else "",
            ),
            "route_family": "SMS→RM",
            "heats_required": heats_required,
            "planner_status": "RELEASED",
            "frozen_flag": False,
            "order_type": order_type,
            "rolling_mode": rolling_mode,
        })

    return planning_orders


if __name__ == '__main__':
    print(f'\n  X-APS Application API -> http://localhost:{PORT}')
    print(f'  Workbook              -> {WORKBOOK}\n')
    app.run(host='0.0.0.0', port=PORT, debug=False, use_reloader=False)
