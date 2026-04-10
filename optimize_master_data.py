"""
APS Master Data Optimization Script

Applies 10 targeted fixes to APS_BF_SMS_RM.xlsx to enable realistic multi-resource
scheduling and scenario quality data.
"""

from openpyxl import load_workbook
import shutil
import os
import time

def optimize_master_data():
    filepath = r'c:\Users\bhadk\Documents\APS\APS_BF_SMS_RM.xlsx'

    # Create backup
    backup_path = filepath + '.backup'
    if not os.path.exists(backup_path):
        shutil.copy(filepath, backup_path)
        print(f"[*] Backup created: {backup_path}")

    # Load workbook
    print("[*] Loading workbook...")
    wb = load_workbook(filepath)

    # FIX 1: CCM Routing Assignment (HIGH severity)
    print("\n[Fix 1] CCM Routing Assignment...")
    routing_ws = wb['Routing']
    ccm_updates = 0
    
    for row in routing_ws.iter_rows(min_row=2, max_row=routing_ws.max_row):
        sku_id = row[0].value
        operation = row[1].value if len(row) > 1 else None
        
        if sku_id in ['BIL-150-1065', 'BIL-150-1080', 'BIL-150-4140']:
            if operation and 'CCM' in str(operation):
                for col_idx in range(len(row)):
                    if row[col_idx].value == 'CCM-01':
                        row[col_idx].value = 'CCM-02'
                        print(f"  - {sku_id} CCM route: CCM-01 -> CCM-02")
                        ccm_updates += 1
                        break
    
    print(f"  ({ccm_updates} rows updated)")

    # FIX 2: EAF Load Balancing (MODERATE severity)
    print("[Fix 2] EAF Load Balancing...")
    eaf_targets = ['BIL-150-1045', 'BIL-150-1065', 'BIL-150-1080', 'BIL-150-CHQ', 'BIL-150-4140']
    eaf_updates = 0
    
    for row in routing_ws.iter_rows(min_row=2, max_row=routing_ws.max_row):
        sku_id = row[0].value
        operation = row[1].value if len(row) > 1 else None
        
        if sku_id in eaf_targets:
            if operation and 'EAF' in str(operation):
                for col_idx in range(len(row)):
                    if row[col_idx].value == 'EAF-01':
                        row[col_idx].value = 'EAF-02'
                        print(f"  - {sku_id} EAF route: EAF-01 -> EAF-02")
                        eaf_updates += 1
                        break
    
    print(f"  ({eaf_updates} rows updated)")

    # FIX 3: Rolling Mill Load Balancing
    print("[Fix 3] Rolling Mill Load Balancing...")
    rm_targets = ['FG-WR-SAE1035-30', 'FG-WR-SAE1035-40', 'FG-WR-SAE1045-30', 
                  'FG-WR-SAE1045-40', 'FG-WR-SAE1045-55']
    rm_updates = 0
    
    for row in routing_ws.iter_rows(min_row=2, max_row=routing_ws.max_row):
        sku_id = row[0].value
        operation = row[1].value if len(row) > 1 else None
        
        if sku_id in rm_targets:
            if operation and 'RM' in str(operation):
                for col_idx in range(len(row)):
                    if row[col_idx].value == 'RM-01':
                        row[col_idx].value = 'RM-02'
                        print(f"  - {sku_id} RM route: RM-01 -> RM-02")
                        rm_updates += 1
                        break
    
    print(f"  ({rm_updates} rows updated)")

    # FIX 4: CrMo RM_Changeover_Min
    print("[Fix 4] CrMo RM_Changeover_Min Alignment...")
    campaign_ws = wb['Campaign_Config']
    cmo_updates = 0
    
    for row in campaign_ws.iter_rows(min_row=2, max_row=campaign_ws.max_row):
        grade = row[0].value
        if grade and ('CrMo' in str(grade) or 'Cr-Mo' in str(grade)):
            for col_idx in range(len(row)):
                if row[col_idx].value == 165:
                    row[col_idx].value = 120
                    print(f"  - Cr-Mo 4140 RM_Changeover_Min: 165 -> 120")
                    cmo_updates += 1
                    break
    
    if cmo_updates == 0:
        print("  - No matching rows found (may already be updated)")

    # FIX 5: Grade-Differentiated Safety Stock
    print("[Fix 5] Grade-Differentiated Safety Stocks...")
    sku_ws = wb['SKU_Master']
    
    fg_safety_stocks = {
        'FG-WR-SAE1008-30': 80, 'FG-WR-SAE1008-40': 80,
        'FG-WR-SAE1018-30': 60, 'FG-WR-SAE1018-40': 60,
        'FG-WR-SAE1035-30': 40, 'FG-WR-SAE1035-40': 40,
        'FG-WR-SAE1045-30': 30, 'FG-WR-SAE1045-40': 30, 'FG-WR-SAE1045-55': 30,
        'FG-WR-SAE1065-40': 20, 'FG-WR-SAE1065-50': 20,
        'FG-WR-SAE1080-40': 15, 'FG-WR-SAE1080-50': 15,
        'FG-WR-CHQ1006-55': 10,
        'FG-WR-CrMo4140-30': 10, 'FG-WR-CrMo4140-40': 10
    }
    
    billet_safety_stocks = {
        'BIL-130-1008': 200, 'BIL-130-1018': 150, 'BIL-130-1035': 100,
        'BIL-150-1045': 100, 'BIL-150-1065': 50, 'BIL-150-1080': 50,
        'BIL-150-CHQ': 50, 'BIL-150-4140': 30
    }
    
    sku_updates = 0
    for row in sku_ws.iter_rows(min_row=2, max_row=sku_ws.max_row):
        sku_id = row[0].value
        all_targets = {**fg_safety_stocks, **billet_safety_stocks}
        
        if sku_id in all_targets:
            for col_idx in range(len(row)):
                if col_idx == 7:  # Safety_Stock_MT column
                    old_val = row[col_idx].value
                    row[col_idx].value = all_targets[sku_id]
                    print(f"  - {sku_id}: {old_val} -> {all_targets[sku_id]} MT")
                    sku_updates += 1
                    break
    
    print(f"  ({sku_updates} rows updated)")

    # FIX 6: RM-FECR Replenishment
    print("[Fix 6] RM-FECR Replenishment...")
    inventory_ws = wb['Inventory']
    inv_updates = 0
    
    for row in inventory_ws.iter_rows(min_row=2, max_row=inventory_ws.max_row):
        sku_id = row[0].value
        if sku_id == 'RM-FECR':
            for col_idx in range(len(row)):
                if col_idx == 3:  # Available_Qty
                    old_val = row[col_idx].value
                    row[col_idx].value = 30
                    print(f"  - RM-FECR: {old_val} -> 30 MT")
                    inv_updates += 1
                    break
    
    if inv_updates == 0:
        print("  - RM-FECR not found in Inventory sheet")

    # FIX 7: BIL-150-1080 WIP
    print("[Fix 7] BIL-150-1080 WIP Inventory...")
    for row in inventory_ws.iter_rows(min_row=2, max_row=inventory_ws.max_row):
        sku_id = row[0].value
        if sku_id == 'BIL-150-1080':
            for col_idx in range(len(row)):
                if col_idx == 3:  # Available_Qty
                    old_val = row[col_idx].value
                    row[col_idx].value = 100
                    print(f"  - BIL-150-1080: {old_val} -> 100 MT")
                    break

    # FIX 8: BIL-150-CHQ WIP
    print("[Fix 8] BIL-150-CHQ WIP Inventory...")
    for row in inventory_ws.iter_rows(min_row=2, max_row=inventory_ws.max_row):
        sku_id = row[0].value
        if sku_id == 'BIL-150-CHQ':
            for col_idx in range(len(row)):
                if col_idx == 3:  # Available_Qty
                    old_val = row[col_idx].value
                    row[col_idx].value = 50
                    print(f"  - BIL-150-CHQ: {old_val} -> 50 MT")
                    break

    # FIX 9: Add Maintenance Scenario
    print("[Fix 9] Add Maintenance Scenario...")
    scenarios_ws = wb['Scenarios']
    last_row = scenarios_ws.max_row + 1
    
    scenarios_ws[f'A{last_row}'] = 'EAF-01'
    scenarios_ws[f'B{last_row}'] = 'Maintenance'
    scenarios_ws[f'C{last_row}'] = 'Scheduled downtime'
    scenarios_ws[f'D{last_row}'] = '2026-04-14'
    scenarios_ws[f'E{last_row}
