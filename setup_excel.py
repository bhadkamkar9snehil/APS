"""
Setup script: Embeds xlwings VBA module + macro buttons into the Excel workbook.
Run once:  python setup_excel.py

Prerequisites:
  - Excel must have "Trust access to the VBA project object model" enabled
    (File > Options > Trust Center > Trust Center Settings > Macro Settings)
  - A supported APS workbook should exist in this folder
"""
import sys
import time
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import xlwings as xw
import win32com.client
from pywintypes import com_error

XLWINGS_BAS = Path(xw.__file__).parent / "xlwings.bas"
THIS_DIR = Path(__file__).resolve().parent
ARCHIVE_DIR = THIS_DIR / "archive" / "rebuild_snapshots"
WORKBOOK_CANDIDATES = [
    "APS_BF_SMS_RM.xlsm",
    "APS_BF_SMS_RM.xlsx",
]
XL_OPEN_XML_MACRO_ENABLED = 52
PHASE1_REQUIRED_SHEETS = ("Config", "Queue_Times", "CTP_Request", "CTP_Output")
PHASE1_HEADER_PREFIXES = {
    "Sales_Orders": [
        "SO_ID", "Customer", "Region", "SKU_ID", "Grade", "Section_mm",
        "Order_Qty_MT", "Coils_Count", "Order_Date", "Delivery_Date",
        "Priority", "Campaign_Group", "Campaign_ID", "Status",
    ],
    "Resource_Master": [
        "Resource_ID", "Resource_Name", "Plant", "Type",
        "Avail_Hours_Day", "Max_Capacity_MT_Hr", "Capacity_MT_Day",
        "Heat_Size_MT", "Efficiency_%", "Status",
        "Operation_Group", "Default_Cycle_Min", "Default_Setup_Min", "Operation_Color",
    ],
    "Routing": [
        "SKU_ID", "Grade", "Needs_VD", "Op_Seq", "Operation",
        "Resource_Group", "Preferred_Resource", "Cycle_Time_Min_Heat",
        "Setup_Time_Min", "Min_Campaign_MT", "Max_Campaign_MT", "Note",
        "Sequence", "Is_Optional", "Optional_Condition", "Transfer_Time_Min",
    ],
    "SKU_Master": [
        "SKU_ID", "SKU_Name", "Category", "Grade", "Section_mm",
        "Coil_Wt_MT", "UOM", "Needs_VD", "Lead_Time_Days", "Safety_Stock_MT",
        "Route_Variant", "Product_Family", "Attribute_1",
    ],
    "BOM": [
        "BOM_ID", "Parent_SKU", "Child_SKU", "Flow_Type", "Qty_Per",
        "Scrap_%", "Yield_Pct", "Level", "UOM", "Note",
    ],
}

BTN_PRIMARY = (31, 56, 100)
BTN_SECONDARY = (68, 114, 196)
BTN_CLEAR = (192, 80, 77)
BTN_NAV = (233, 241, 250)
BTN_HELP = (226, 239, 218)
TXT_DARK = (31, 56, 100)
TXT_MUTED = (128, 128, 128)
TXT_LIGHT = (255, 255, 255)
CARD_BORDER = (191, 200, 214)
CARD_FILL = (255, 255, 255)
CARD_FILL_ALT = (247, 249, 252)
CARD_FILL_SOFT = (234, 242, 248)
CARD_FILL_WARN = (255, 244, 204)
TXT_WARN = (127, 96, 0)
BUTTON_COL_GAP = 10
BUTTON_WIDTH_COLS = 4
EXCEL_BUSY_HRESULT = -2146777998


def _excel_rgb(color):
    if isinstance(color, tuple):
        red, green, blue = color
        return red + (green << 8) + (blue << 16)
    return color


def _is_excel_busy_error(exc):
    try:
        hresult = getattr(exc, "hresult", None)
        if hresult is None and exc.args:
            hresult = exc.args[0]
        return int(hresult) == EXCEL_BUSY_HRESULT
    except Exception:
        return False


def _retry_excel_busy(action, fn, *, attempts=12, delay_sec=0.2):
    last_exc = None
    for attempt in range(attempts):
        try:
            return fn()
        except com_error as exc:
            if not _is_excel_busy_error(exc):
                raise
            last_exc = exc
            time.sleep(delay_sec * (attempt + 1))
    raise RuntimeError(
        f"Excel stayed busy while trying to {action}. Close any active cell edit, dialog, or modal prompt and re-run setup."
    ) from last_exc


CONTROL_PANEL_BUTTON_SPECS = [
    ("G8:L8", "Run Schedule", "RunSchedule", BTN_PRIMARY),
    ("G11:I11", "Run BOM Explosion", "RunBOMExplosion", BTN_SECONDARY),
    ("J11:L11", "Run Capacity Map", "RunCapacityMap", BTN_SECONDARY),
    ("G14:I14", "Run Scenarios", "RunScenarios", BTN_SECONDARY),
    ("J14:L14", "Run CTP", "RunCTP", BTN_SECONDARY),
    ("G17:I17", "Open Help", "GoToHelpSheet", BTN_HELP),
    ("J17:L17", "Clear All Outputs", "ClearOutputs", BTN_CLEAR),
]


BUTTON_LAYOUTS = {
    "BOM": {
        "actions": [(1, "Run BOM Explosion", "RunBOMExplosion", BTN_SECONDARY)],
    },
    "BOM_Output": {
        "actions": [
            (1, "Run BOM Explosion", "RunBOMExplosion", BTN_SECONDARY),
            (4, "Clear All Outputs", "ClearOutputs", BTN_CLEAR),
        ],
    },
    "Capacity_Map": {
        "actions": [
            (1, "Run Capacity Map", "RunCapacityMap", BTN_SECONDARY),
            (4, "Clear All Outputs", "ClearOutputs", BTN_CLEAR),
        ],
    },
    "Scenarios": {
        "actions": [(1, "Run Scenarios", "RunScenarios", BTN_SECONDARY)],
    },
    "Scenario_Output": {
        "actions": [
            (1, "Run Scenarios", "RunScenarios", BTN_SECONDARY),
            (4, "Clear All Outputs", "ClearOutputs", BTN_CLEAR),
        ],
    },
    "Schedule_Output": {
        "actions": [
            (1, "Run Schedule", "RunSchedule", BTN_SECONDARY),
            (4, "Clear All Outputs", "ClearOutputs", BTN_CLEAR),
        ],
    },
    "Campaign_Schedule": {
        "actions": [
            (1, "Run Schedule", "RunSchedule", BTN_SECONDARY),
            (4, "Clear All Outputs", "ClearOutputs", BTN_CLEAR),
        ],
    },
    "Material_Plan": {
        "actions": [
            (1, "Run Schedule", "RunSchedule", BTN_SECONDARY),
            (4, "Clear All Outputs", "ClearOutputs", BTN_CLEAR),
        ],
    },
    "Equipment_Schedule": {
        "actions": [
            (1, "Run Schedule", "RunSchedule", BTN_SECONDARY),
            (4, "Clear All Outputs", "ClearOutputs", BTN_CLEAR),
        ],
    },
    "Schedule_Gantt": {
        "actions": [
            (1, "Run Schedule", "RunSchedule", BTN_SECONDARY),
            (4, "Clear All Outputs", "ClearOutputs", BTN_CLEAR),
        ],
    },
    "CTP_Request": {
        "actions": [
            (1, "Run CTP", "RunCTP", BTN_SECONDARY),
            (4, "Clear All Outputs", "ClearOutputs", BTN_CLEAR),
        ],
    },
    "CTP_Output": {
        "actions": [
            (1, "Run CTP", "RunCTP", BTN_SECONDARY),
            (4, "Clear All Outputs", "ClearOutputs", BTN_CLEAR),
        ],
    },
    "Help": {},
}


def resolve_workbook_path():
    """Choose the workbook to set up, preferring the canonical BF/SMS/RM template."""
    if len(sys.argv) > 1:
        requested = Path(sys.argv[1])
        return requested if requested.is_absolute() else (THIS_DIR / requested)

    for candidate in WORKBOOK_CANDIDATES:
        path = THIS_DIR / candidate
        if path.exists():
            return path

    raise FileNotFoundError(
        "Could not find a supported APS workbook. Expected one of: "
        + ", ".join(WORKBOOK_CANDIDATES)
    )


def _row3_prefix(ws, count):
    return [ws.cell(3, col).value for col in range(1, count + 1)]


def _style_com_range(rng, *, fill=None, font_color=TXT_DARK, bold=False, size=10, italic=False, align=-4108, border_color=CARD_BORDER):
    if fill is None:
        rng.Interior.Pattern = -4142
    else:
        rng.Interior.Color = _excel_rgb(fill)
    rng.Font.Name = "Calibri"
    rng.Font.Size = size
    rng.Font.Bold = bold
    rng.Font.Italic = italic
    rng.Font.Color = _excel_rgb(font_color)
    rng.HorizontalAlignment = align
    rng.VerticalAlignment = -4108
    rng.WrapText = True
    rng.Borders.LineStyle = 1
    rng.Borders.Color = _excel_rgb(border_color)
    rng.Borders.Weight = 2 if bold else 1


def _set_panel_row(ws, range_ref, value, *, fill=CARD_FILL, font_color=TXT_DARK, bold=False, size=10, italic=False, align=-4108):
    rng = ws.Range(range_ref)
    try:
        rng.UnMerge()
    except Exception:
        pass
    rng.Merge()
    rng.Value = value
    _style_com_range(rng, fill=fill, font_color=font_color, bold=bold, size=size, italic=italic, align=align)


def _col_span(start_col, width_cols=BUTTON_WIDTH_COLS):
    return f"{get_column_letter(start_col)}:{get_column_letter(start_col + width_cols - 1)}"


def _anchor_cell(start_col, row):
    return f"{get_column_letter(start_col)}{row}"


def _last_used_col(ws):
    try:
        used = ws.UsedRange
        if used is None:
            return 1
        return int(used.Column) + int(used.Columns.Count) - 1
    except Exception:
        return 1


def workbook_button_start_col(wb_com):
    existing_sheets = [wb_com.Worksheets(i) for i in range(1, wb_com.Worksheets.Count + 1)]
    widest = max((_last_used_col(ws) for ws in existing_sheets), default=1)
    return widest + BUTTON_COL_GAP


def format_control_panel_sheet(wb_com, button_start_col):
    if "Control_Panel" not in {wb_com.Worksheets(i).Name for i in range(1, wb_com.Worksheets.Count + 1)}:
        return

    ws = wb_com.Worksheets("Control_Panel")
    top_area = ws.Range(f"A1:{get_column_letter(max(button_start_col + BUTTON_WIDTH_COLS + 1, 17))}28")
    try:
        top_area.UnMerge()
    except Exception:
        pass
    top_area.Clear()
    top_area.Interior.Pattern = -4142
    top_area.Font.Name = "Calibri"
    top_area.Borders.LineStyle = 0

    column_widths = {
        "A": 2.5, "B": 16, "C": 16, "D": 16, "E": 16,
        "F": 3.5, "G": 13, "H": 13, "I": 13, "J": 13, "K": 13, "L": 13,
        "M": 3.5, "N": 13, "O": 13, "P": 13, "Q": 13,
    }
    for col, width in column_widths.items():
        ws.Columns(col).ColumnWidth = width

    for row in range(1, 29):
        ws.Rows(row).RowHeight = 20
    ws.Rows(1).RowHeight = 24
    ws.Rows(2).RowHeight = 20
    for row in (8, 11, 14, 17):
        ws.Rows(row).RowHeight = 28

    _set_panel_row(
        ws,
        "B1:Q1",
        "APS - Steel Plant (BF | SMS | Rolling Mill)",
        fill=None,
        font_color=TXT_DARK,
        bold=True,
        size=18,
        align=-4131,
    )
    _set_panel_row(
        ws,
        "B2:Q2",
        "Advanced Planning & Scheduling | Primary batch = Heat | Standard heat size = 50 MT",
        fill=None,
        font_color=TXT_MUTED,
        size=11,
        italic=True,
        align=-4131,
    )
    _set_panel_row(
        ws,
        "B3:F3",
        f"Workbook refreshed: {datetime.now().strftime('%d-%b-%Y %H:%M')}",
        fill=None,
        font_color=(170, 170, 170),
        size=9,
        align=-4131,
    )

    _set_panel_row(ws, "B5:E5", "Plant Snapshot", fill=BTN_PRIMARY, font_color=TXT_LIGHT, bold=True, size=11)
    plant_rows = [
        ("B6:E6", "1 Blast Furnace", CARD_FILL),
        ("B7:E7", "2 EAF | 3 LRF | 1 VD", CARD_FILL_ALT),
        ("B8:E8", "2 Continuous Casters", CARD_FILL),
        ("B9:E9", "2 Rolling Mills", CARD_FILL_ALT),
        ("B10:E10", "Planning unit: Heat | 50 MT", CARD_FILL_WARN),
    ]
    for range_ref, text, fill in plant_rows:
        _set_panel_row(
            ws,
            range_ref,
            text,
            fill=fill,
            font_color=TXT_WARN if fill == CARD_FILL_WARN else TXT_DARK,
            size=10,
        )

    _set_panel_row(ws, "G5:L5", "APS Actions", fill=BTN_PRIMARY, font_color=TXT_LIGHT, bold=True, size=11)
    _set_panel_row(
        ws,
        "G6:L6",
        "Run the full plan here. Other sheets keep their local buttons in one shared strip further to the right.",
        fill=CARD_FILL_SOFT,
        font_color=TXT_MUTED,
        size=10,
        italic=True,
    )
    for range_ref, text, fill in [
        ("G20:L20", "1. Refresh inputs only when demand, inventory, or scenario settings change.", CARD_FILL),
        ("G21:L21", "2. Run Schedule for the operational plan. Use BOM/Capacity only when investigating.", CARD_FILL_ALT),
        ("G22:L22", "3. Review Campaign_Schedule, Equipment_Schedule, and Schedule_Gantt together.", CARD_FILL),
    ]:
        _set_panel_row(ws, range_ref, text, fill=fill, font_color=TXT_DARK, size=10, align=-4131)

    _set_panel_row(ws, "N5:Q5", "Recommended Flow", fill=BTN_PRIMARY, font_color=TXT_LIGHT, bold=True, size=11)
    for range_ref, text, fill in [
        ("N6:Q6", "1. Update Sales_Orders, Inventory, or Scenarios.", CARD_FILL),
        ("N7:Q7", "2. Run Schedule for the live plan.", CARD_FILL_ALT),
        ("N8:Q8", "3. Review Campaign_Schedule and Equipment_Schedule.", CARD_FILL),
        ("N9:Q9", "4. Compare cases in Scenario_Output only when needed.", CARD_FILL_ALT),
        ("N10:Q10", "5. Use Help for detailed sheet explanations.", CARD_FILL),
    ]:
        _set_panel_row(ws, range_ref, text, fill=fill, font_color=TXT_DARK, size=10, align=-4131)

    _set_panel_row(ws, "N12:Q12", "Review Outputs", fill=BTN_PRIMARY, font_color=TXT_LIGHT, bold=True, size=11)
    for range_ref, text, fill in [
        ("N13:Q13", "Campaign_Schedule: release order and due-date margin.", CARD_FILL),
        ("N14:Q14", "Schedule_Output: master row-level plan across SMS and RM.", CARD_FILL_ALT),
        ("N15:Q15", "Equipment_Schedule: machine-wise dispatch packets.", CARD_FILL),
        ("N16:Q16", "Material_Plan: campaign material commitment and shortages.", CARD_FILL_ALT),
        ("N17:Q17", "Schedule_Gantt: visual resource timeline.", CARD_FILL),
    ]:
        _set_panel_row(ws, range_ref, text, fill=fill, font_color=TXT_DARK, size=10, align=-4131)

    _set_panel_row(
        ws,
        "N19:Q19",
        f"Sheet buttons on working tabs start at {get_column_letter(button_start_col)}.",
        fill=CARD_FILL_SOFT,
        font_color=TXT_MUTED,
        size=10,
        italic=True,
        align=-4131,
    )


def phase1_schema_issues(workbook_path):
    """Return a list of Phase 1 schema gaps for the workbook path."""
    workbook_path = Path(workbook_path)
    if not workbook_path.exists():
        return [f"Workbook not found: {workbook_path.name}"]

    wb = load_workbook(workbook_path, keep_vba=workbook_path.suffix.lower() == ".xlsm")
    issues = []

    for sheet_name in PHASE1_REQUIRED_SHEETS:
        if sheet_name not in wb.sheetnames:
            issues.append(f"Missing sheet: {sheet_name}")

    for sheet_name, expected_headers in PHASE1_HEADER_PREFIXES.items():
        if sheet_name not in wb.sheetnames:
            issues.append(f"Missing sheet: {sheet_name}")
            continue
        actual_headers = _row3_prefix(wb[sheet_name], len(expected_headers))
        if actual_headers != expected_headers:
            issues.append(
                f"{sheet_name} row 3 header mismatch: expected {expected_headers}, got {actual_headers}"
            )

    return issues


def backup_and_retire_workbook(workbook_path, *, reason="pre_phase1_schema"):
    """Move the existing workbook aside before rebuilding it from the refreshed .xlsx."""
    workbook_path = Path(workbook_path)
    ARCHIVE_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = ARCHIVE_DIR / f"{workbook_path.stem}.{reason}_{timestamp}{workbook_path.suffix}"
    workbook_path.replace(backup_path)
    return backup_path


def _template_newer_than_workbook(workbook_path, template_path) -> bool:
    workbook_path = Path(workbook_path)
    template_path = Path(template_path)
    if not workbook_path.exists() or not template_path.exists():
        return False
    try:
        # Small tolerance avoids needless churn from filesystem timestamp rounding.
        return template_path.stat().st_mtime > workbook_path.stat().st_mtime + 1.0
    except Exception:
        return False


def ensure_phase1_schema_rollout(workbook_path):
    """
    Ensure the macro workbook is on the Phase 1 schema.

    If the live .xlsm is stale but the regenerated .xlsx exists, move the stale workbook
    into archive/rebuild_snapshots so setup can recreate a fresh .xlsm from the current template.
    """
    workbook_path = Path(workbook_path).with_suffix(".xlsm")
    template_path = workbook_path.with_suffix(".xlsx")

    if not workbook_path.exists():
        return

    if template_path.exists() and _template_newer_than_workbook(workbook_path, template_path):
        try:
            template_issues = phase1_schema_issues(template_path)
        except Exception as exc:
            raise RuntimeError(
                f"Could not inspect newer template {template_path.name} before rebuild."
            ) from exc

        if template_issues:
            formatted = "\n  - ".join(template_issues)
            raise RuntimeError(
                f"{template_path.name} is newer than {workbook_path.name} but is not Phase 1 compliant.\n"
                f"  - {formatted}"
            )

        try:
            backup_path = backup_and_retire_workbook(workbook_path, reason="template_refresh")
        except PermissionError as exc:
            raise RuntimeError(
                f"{template_path.name} is newer than {workbook_path.name}, but the old macro workbook "
                f"could not be archived. Close Excel for this workbook, then re-run setup_excel.py."
            ) from exc

        print(f"  Archived older workbook -> {backup_path.name}")
        print(f"  {template_path.name} is newer and will now be used to recreate {workbook_path.name}")
        return

    issues = phase1_schema_issues(workbook_path)
    if not issues:
        return

    if not template_path.exists():
        formatted = "\n  - ".join(issues)
        raise RuntimeError(
            f"{workbook_path.name} is missing Phase 1 schema elements, and {template_path.name} "
            f"is not available to rebuild it.\n  - {formatted}"
        )

    try:
        template_issues = phase1_schema_issues(template_path)
    except Exception as exc:
        raise RuntimeError(
            f"Could not inspect {template_path.name} before rebuild."
        ) from exc

    if template_issues:
        formatted = "\n  - ".join(template_issues)
        raise RuntimeError(
            f"{template_path.name} is not Phase 1 compliant, so setup will not rebuild from it.\n"
            f"  - {formatted}"
        )

    try:
        backup_path = backup_and_retire_workbook(workbook_path, reason="pre_phase1_schema")
    except PermissionError as exc:
        formatted = "\n  - ".join(issues)
        raise RuntimeError(
            f"{workbook_path.name} is still on the pre-Phase-1 schema and could not be archived.\n"
            f"Close Excel for this workbook, then re-run setup_excel.py.\n"
            f"  - {formatted}"
        ) from exc

    print(f"  Archived stale workbook -> {backup_path.name}")
    for issue in issues:
        print(f"    - {issue}")
    print(f"  {template_path.name} will now be used to recreate {workbook_path.name}")


def get_excel_app():
    """Attach to a running Excel instance or start one."""
    try:
        xl = win32com.client.GetActiveObject("Excel.Application")
        print("  Using running Excel instance")
    except com_error:
        print("  Excel is not running; starting a new instance...")
        xl = win32com.client.gencache.EnsureDispatch("Excel.Application")
        xl.Visible = True
    return xl


def find_open_workbook(xl, *candidate_paths):
    """Return the first open workbook matching any provided path."""
    wanted = {str(path.resolve()).lower() for path in candidate_paths}

    for wb in xl.Workbooks:
        try:
            if str(Path(wb.FullName).resolve()).lower() in wanted:
                return wb
        except Exception:
            continue

    return None


def open_workbook(xl, workbook_path, *, prefer_template_rebuild=False):
    """Find or open the macro-enabled workbook, converting from .xlsx if needed."""
    xlsx_fallback_path = workbook_path.with_suffix(".xlsx")

    if prefer_template_rebuild and xlsx_fallback_path.exists():
        wb_com = find_open_workbook(xl, workbook_path)
        if wb_com is not None:
            print(f"  Closing open workbook before rebuild: {wb_com.Name}")
            _retry_excel_busy(f"save workbook {wb_com.Name}", wb_com.Save)
            _retry_excel_busy(
                f"close workbook {wb_com.Name}",
                lambda: wb_com.Close(SaveChanges=False),
            )

        if workbook_path.exists():
            backup_path = backup_and_retire_workbook(workbook_path, reason="template_refresh")
            print(f"  Archived existing workbook -> {backup_path.name}")

        wb_xlsx = find_open_workbook(xl, xlsx_fallback_path)
        if wb_xlsx is not None:
            print(f"  Converting open workbook {wb_xlsx.Name} -> {workbook_path.name}")
        else:
            print(f"  Opening {xlsx_fallback_path.name} and saving as {workbook_path.name}")
            wb_xlsx = xl.Workbooks.Open(str(xlsx_fallback_path))
        xl.DisplayAlerts = False
        try:
            wb_xlsx.SaveAs(str(workbook_path), FileFormat=XL_OPEN_XML_MACRO_ENABLED)
        finally:
            xl.DisplayAlerts = True
        return wb_xlsx

    wb_com = find_open_workbook(xl, workbook_path)
    if wb_com is not None:
        print(f"  Found open workbook: {wb_com.Name}")
        return wb_com

    if workbook_path.exists():
        print(f"  Opening workbook: {workbook_path.name}")
        return xl.Workbooks.Open(str(workbook_path))

    wb_xlsx = find_open_workbook(xl, xlsx_fallback_path)
    if wb_xlsx is not None:
        print(f"  Converting open workbook {wb_xlsx.Name} -> {workbook_path.name}")
        xl.DisplayAlerts = False
        try:
            wb_xlsx.SaveAs(str(workbook_path), FileFormat=XL_OPEN_XML_MACRO_ENABLED)
        finally:
            xl.DisplayAlerts = True
        return wb_xlsx

    if xlsx_fallback_path.exists():
        print(f"  Opening {xlsx_fallback_path.name} and saving as {workbook_path.name}")
        wb_xlsx = xl.Workbooks.Open(str(xlsx_fallback_path))
        xl.DisplayAlerts = False
        try:
            wb_xlsx.SaveAs(str(workbook_path), FileFormat=XL_OPEN_XML_MACRO_ENABLED)
        finally:
            xl.DisplayAlerts = True
        return wb_xlsx

    raise FileNotFoundError(
        f"Could not find {workbook_path.name} or {xlsx_fallback_path.name} in {THIS_DIR}"
    )


def get_vba_project(workbook_path, *, prefer_template_rebuild=False):
    """Get VBA project via win32com for the target workbook."""
    xl = get_excel_app()
    wb_com = open_workbook(xl, workbook_path, prefer_template_rebuild=prefer_template_rebuild)
    try:
        vbp = wb_com.VBProject
    except com_error as exc:
        raise RuntimeError(
            'Excel blocked VBA project access. Enable "Trust access to the VBA project object model" '
            "in Excel Trust Center settings, then re-run this script."
        ) from exc
    return xl, wb_com, vbp


def remove_module(vbproject, name):
    """Remove a VBA module by name if it exists."""
    for comp in vbproject.VBComponents:
        if comp.Name == name:
            vbproject.VBComponents.Remove(comp)
            print(f"  Removed old module: {name}")
            return


def import_bas_file(vbproject, bas_path, module_name=None):
    """Import a .bas file into VBA project."""
    vbproject.VBComponents.Import(str(bas_path))
    print(f"  Imported: {bas_path.name}")


def add_macro_module(vbproject):
    """Add our APS macro module with RunPython calls."""
    remove_module(vbproject, "APS_Macros")

    mod = vbproject.VBComponents.Add(1)  # 1 = vbext_ct_StdModule
    mod.Name = "APS_Macros"
    mod.CodeModule.AddFromString("""
Sub RunBOMExplosion()
    RunPython "import aps_functions; aps_functions.run_bom_explosion()"
End Sub

Sub RunCapacityMap()
    RunPython "import aps_functions; aps_functions.run_capacity_map()"
End Sub

Sub RunSchedule()
    RunPython "import aps_functions; aps_functions.run_schedule()"
End Sub

Sub RunScenarios()
    RunPython "import aps_functions; aps_functions.run_scenario()"
End Sub

Sub RunCTP()
    RunPython "import aps_functions; aps_functions.run_ctp()"
End Sub

Sub ClearOutputs()
    RunPython "import aps_functions; aps_functions.clear_outputs()"
End Sub

Sub GoToHelpSheet()
    Worksheets("Help").Activate
    Range("A1").Select
End Sub

Sub GoToControlPanel()
    Worksheets("Control_Panel").Activate
    Range("A1").Select
End Sub
""")
    print("  Added module: APS_Macros (8 macros)")


def create_buttons(wb_com):
    """Create proper button shapes on the control panel and relevant working sheets."""
    macro_prefix = f"'{wb_com.Name}'!APS_Macros."
    button_start_col = workbook_button_start_col(wb_com)
    button_span_range = _col_span(button_start_col)
    nav_row_default = 7
    format_control_panel_sheet(wb_com, button_start_col)
    existing_sheets = {wb_com.Worksheets(i).Name for i in range(1, wb_com.Worksheets.Count + 1)}
    button_specs = []
    if "Control_Panel" in existing_sheets:
        for range_ref, label, macro, color in CONTROL_PANEL_BUTTON_SPECS:
            anchor_cell = range_ref.split(":")[0]
            button_specs.append(("Control_Panel", anchor_cell, range_ref, label, macro, color))

    for sheet_name, layout in BUTTON_LAYOUTS.items():
        if sheet_name not in existing_sheets:
            continue
        for row, label, macro, color in layout.get("actions", []):
            button_specs.append((sheet_name, _anchor_cell(button_start_col, row), button_span_range, label, macro, color))

    for sheet_name in sorted(existing_sheets):
        if sheet_name == "Control_Panel":
            continue
        layout = BUTTON_LAYOUTS.get(sheet_name, {})
        action_rows = [row for row, *_ in layout.get("actions", [])]
        nav_row = max(action_rows) + 3 if action_rows else nav_row_default
        button_specs.append(
            (sheet_name, _anchor_cell(button_start_col, nav_row), button_span_range, "Back to Control Panel", "GoToControlPanel", BTN_NAV)
        )

    def remove_existing_buttons(ws):
        for idx in range(ws.Shapes.Count, 0, -1):
            shp = ws.Shapes.Item(idx)
            if str(shp.Name).startswith("btn_"):
                shp.Delete()

    touched_sheets = sorted({spec[0] for spec in button_specs})
    for sheet_name in touched_sheets:
        remove_existing_buttons(wb_com.Worksheets(sheet_name))
    print("  Cleared old APS buttons")

    for sheet_name, anchor_cell, span_range, label, macro, color in button_specs:
        ws = wb_com.Worksheets(sheet_name)
        _retry_excel_busy(f"activate workbook {wb_com.Name}", wb_com.Activate)
        _retry_excel_busy(f"activate sheet {sheet_name}", ws.Activate)
        for col_idx in range(button_start_col, button_start_col + BUTTON_WIDTH_COLS):
            try:
                ws.Columns(get_column_letter(col_idx)).ColumnWidth = max(ws.Columns(get_column_letter(col_idx)).ColumnWidth, 13)
            except Exception:
                pass
        left = ws.Range(anchor_cell).Left
        top = ws.Range(anchor_cell).Top
        width = ws.Range(span_range).Width
        height_factor = 1.5 if sheet_name == "Control_Panel" else 1.6
        height = max(ws.Range(anchor_cell).Height * height_factor, 24)

        shp = _retry_excel_busy(
            f"add button '{label}' on {sheet_name}",
            lambda: ws.Shapes.AddShape(5, left, top, width, height),
        )
        safe_sheet_name = "".join(ch if ch.isalnum() else "_" for ch in sheet_name)
        safe_macro_name = "".join(ch if ch.isalnum() else "_" for ch in macro)
        shp.Name = f"btn_{safe_sheet_name}_{safe_macro_name}_{anchor_cell.replace('$', '')}"
        shp.TextFrame2.TextRange.Text = label
        shp.TextFrame2.TextRange.Font.Size = 12 if sheet_name == "Control_Panel" else 11
        shp.TextFrame2.TextRange.Font.Bold = True
        shp.TextFrame2.TextRange.Font.Fill.ForeColor.RGB = _excel_rgb(TXT_DARK if macro in {"GoToControlPanel", "GoToHelpSheet"} else TXT_LIGHT)
        shp.TextFrame2.TextRange.ParagraphFormat.Alignment = 2
        shp.TextFrame2.VerticalAnchor = 3
        shp.TextFrame2.MarginLeft = 8
        shp.TextFrame2.MarginRight = 8
        shp.Fill.ForeColor.RGB = _excel_rgb(color)
        shp.Line.Visible = True
        shp.Line.ForeColor.RGB = _excel_rgb((157, 173, 193) if macro in {"GoToControlPanel", "GoToHelpSheet"} else BTN_PRIMARY)
        shp.Line.Weight = 1.25
        try:
            shp.Shadow.Visible = False
        except Exception:
            pass
        _retry_excel_busy(
            f"bind macro '{macro}' to button '{label}' on {sheet_name}",
            lambda: setattr(shp, "OnAction", f"{macro_prefix}{macro}"),
        )

        print(f"  {sheet_name}: '{label}' -> {macro}")


def main():
    workbook_path = resolve_workbook_path()
    workbook_name = workbook_path.with_suffix(".xlsm").name

    print(f"Setting up {workbook_name}...")
    print()

    print("[0] Checking workbook schema...")
    ensure_phase1_schema_rollout(workbook_path)
    print("  Phase 1 workbook schema is ready")
    print()

    # Connect via win32com
    print("[1] Connecting to Excel...")
    xl, wb_com, vbp = get_vba_project(workbook_path.with_suffix(".xlsm"), prefer_template_rebuild=True)
    print(f"  Connected: {wb_com.Name}")
    print(f"  VBA modules: {[c.Name for c in vbp.VBComponents]}")

    # Remove old Module1 if it exists
    print()
    print("[2] Cleaning old VBA modules...")
    remove_module(vbp, "Module1")

    # Import xlwings standalone VBA module
    print()
    print("[3] Importing xlwings VBA module...")
    # Check if xlwings module already exists
    has_xlwings = any(c.Name == "xlwings" for c in vbp.VBComponents)
    if has_xlwings:
        remove_module(vbp, "xlwings")
    import_bas_file(vbp, XLWINGS_BAS)

    # Add our macro module
    print()
    print("[4] Adding APS macro module...")
    add_macro_module(vbp)

    print()
    print("[4.5] Saving workbook before button binding...")
    _retry_excel_busy("save workbook before button binding", wb_com.Save)
    print(f"  Saved: {wb_com.Name}")

    # Create buttons
    print()
    print("[5] Creating clickable buttons on Control_Panel...")
    create_buttons(wb_com)

    # Save
    print()
    print("[6] Saving workbook...")
    wb_com.Save()
    print(f"  Saved: {wb_com.FullName}")

    print()
    print("Setup complete! APS action buttons and navigation links are now live in the workbook.")
    print("Click any button to run the corresponding Python function.")


if __name__ == "__main__":
    main()
